"""
byo_csv_loader.py
=================

Profile-driven CSV loader for non-standard general-ledger extracts.

The 4-tier matching engine consumes a uniform `flex_txns` shape (see
recon_engine.propose_candidates). Banks send GL extracts in wildly
varied formats — different column names, different sign conventions,
different date formats, different delimiters. Rather than write a
bespoke loader per bank, we let an admin define a CSVFormatProfile
once via the BYO wizard and apply it on ingest.

Profile fields:
    delimiter          ',' ';' '\\t' '|'
    header_row         1-based row containing column names
    skip_rows          number of preamble rows ABOVE the header
    date_format        strptime format string for date columns
    currency           ISO code; or null if `column_map.currency` is set
    column_map         dict: canonical-name -> source-column-name (or None)
    sign_convention    'positive_credit' | 'separate_column' | 'cr_dr_column'
    sign_column        column name when sign_convention != positive_credit

Canonical fields the loader produces (matches the existing Flex shape):
    _row_number        1-based, runs in source order
    _source            'flex'
    trn_ref            mapped from column_map['ref']
    ac_branch          mapped from column_map['ac_branch'] or empty
    ac_no              mapped from column_map['ac_no'] or empty
    booking_date       same as value_date if not present in the CSV
    value_date         integer YYYYMMDD; 0 when the source value is unparseable
    type               'CR' or 'DR'
    narration          mapped from column_map['narration']
    amount             absolute value (sign lives in `type`)
    ccy                profile.currency or column_map['currency']
    module             empty
    external_ref       same as trn_ref (the engine searches both)
    user_id            empty

Errors during a row are collected and returned alongside the parsed
rows so the UI can show "of 2,400 rows, 12 had unparseable dates."
"""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass
from datetime import datetime
from typing import Any


REQUIRED_COLUMNS = ('amount', 'value_date')
OPTIONAL_COLUMNS = ('ref', 'narration', 'type', 'currency',
                    'ac_no', 'ac_branch', 'booking_date')

VALID_DELIMITERS = (',', ';', '\t', '|')
VALID_SIGN_CONVENTIONS = (
    'positive_credit',     # signed amount column: positive → CR, negative → DR
    'separate_column',     # an explicit CR/DR marker column (sign_column)
    'cr_dr_column',        # alias for separate_column with conventional name
    'paid_in_withdrawn',   # two amount columns: amount → Paid In (CR), sign_column → Withdrawn (DR)
)


@dataclass
class CsvProfile:
    """Runtime view of csv_format_profiles row, with column_map JSON
    decoded. Construct via from_db() to inherit defaults consistently."""
    name: str
    delimiter: str
    header_row: int
    skip_rows: int
    date_format: str
    currency: str | None
    column_map: dict[str, str | None]
    sign_convention: str
    sign_column: str | None

    @classmethod
    def from_db(cls, row: dict) -> 'CsvProfile':
        return cls(
            name=row['name'],
            delimiter=row.get('delimiter') or ',',
            header_row=int(row.get('header_row') or 1),
            skip_rows=int(row.get('skip_rows') or 0),
            date_format=row.get('date_format') or '%Y-%m-%d',
            currency=row.get('currency'),
            column_map=(json.loads(row['column_map'])
                        if isinstance(row['column_map'], str)
                        else dict(row['column_map'] or {})),
            sign_convention=row.get('sign_convention') or 'positive_credit',
            sign_column=row.get('sign_column'),
        )


@dataclass
class LoadResult:
    """What the loader returns to the caller. `errors` is a list of
    (row_number, message) so the UI can show a precise failure list."""
    txns: list[dict]
    errors: list[tuple[int, str]]
    columns: list[str]                           # the column names we found
    sample_rows: list[list[str]]                 # first 5 raw rows for the UI


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_csv(content: bytes, profile: CsvProfile) -> LoadResult:
    """Parse `content` (CSV or xlsx bytes) using `profile`. Returns a
    LoadResult.

    Auto-detects xlsx by magic bytes (PK\\x03\\x04, the zip header
    every Office Open XML file shares) and routes to the openpyxl
    reader. Otherwise the bytes are decoded as text and parsed with
    the configured delimiter. Either path produces rows-as-lists, so
    the rest of the pipeline (header detection, _row_to_txn) is
    format-agnostic.

    Raises ValueError only on profile-level mistakes (missing required
    column mapping, invalid delimiter). Per-row failures land in
    LoadResult.errors and don't abort the run — banks send messy data
    and one bad row shouldn't kill the whole import."""
    _validate_profile(profile)

    if _is_xlsx(content):
        rows = list(_iter_xlsx_rows(content))
    else:
        text = _decode(content)
        rows = list(_iter_csv(text, profile.delimiter))
    if profile.skip_rows >= len(rows):
        raise ValueError(
            f"skip_rows={profile.skip_rows} but the file has only "
            f"{len(rows)} row(s).")
    body = rows[profile.skip_rows:]

    if not body:
        return LoadResult(txns=[], errors=[], columns=[], sample_rows=[])

    # header_row is 1-based and counted from the start of `body`.
    header_idx = profile.header_row - 1
    if header_idx >= len(body):
        raise ValueError(
            f"header_row={profile.header_row} but only {len(body)} row(s) "
            f"remain after skip_rows.")
    columns = [(c or '').strip() for c in body[header_idx]]
    data_rows = body[header_idx + 1:]
    sample = data_rows[:5]

    txns: list[dict] = []
    errors: list[tuple[int, str]] = []
    for i, raw in enumerate(data_rows, start=1):
        try:
            txn = _row_to_txn(i, raw, columns, profile)
            if txn is not None:
                txns.append(txn)
        except _RowError as exc:
            errors.append((i, str(exc)))

    return LoadResult(txns=txns, errors=errors, columns=columns, sample_rows=sample)


def autoguess_mapping(content: bytes, delimiter: str = ',',
                       skip_rows: int = 0,
                       header_row: int = 1) -> dict:
    """First-pass column-name guesser used by the wizard. Inspects the
    header row and proposes a mapping based on common bank-extract
    conventions. The user always edits the result before saving.

    Accepts both CSV bytes and xlsx (auto-detected via the zip-header
    magic) so the wizard previews Telcel-style xlsx exports without a
    save-as-CSV step."""
    if delimiter not in VALID_DELIMITERS:
        delimiter = ','
    if _is_xlsx(content):
        rows = list(_iter_xlsx_rows(content))[skip_rows:]
    else:
        text = _decode(content)
        rows = list(_iter_csv(text, delimiter))[skip_rows:]
    if header_row - 1 >= len(rows):
        return {'columns': [], 'guess': {}, 'sample': []}
    columns = [(c or '').strip() for c in rows[header_row - 1]]
    sample = rows[header_row:header_row + 5]
    return {
        'columns': columns,
        'sample': sample,
        'guess': _guess_columns(columns),
    }


# ---------------------------------------------------------------------------
# Internals
# ---------------------------------------------------------------------------

class _RowError(Exception):
    """Raised mid-row to bail out of one record without aborting the run."""


def _validate_profile(p: CsvProfile) -> None:
    """Validation that's truly load-blocking. Currency is no longer
    required at this layer — the caller (ingest_pair) supplies the
    matched account's currency as a fallback when neither
    profile.currency nor a currency column is set, so a profile bound
    to a single account doesn't have to repeat the currency."""
    if p.delimiter not in VALID_DELIMITERS:
        raise ValueError(f"delimiter must be one of {VALID_DELIMITERS!r}")
    if p.sign_convention not in VALID_SIGN_CONVENTIONS:
        raise ValueError(
            f"sign_convention must be one of {VALID_SIGN_CONVENTIONS!r}")
    if p.sign_convention != 'positive_credit' and not p.sign_column:
        raise ValueError(
            f"sign_convention={p.sign_convention!r} requires sign_column")
    for k in REQUIRED_COLUMNS:
        if not p.column_map.get(k):
            raise ValueError(f"column_map missing required field {k!r}")


def _decode(content: bytes) -> str:
    """UTF-8 first; fall back to latin-1 since some banks ship cp1252-ish
    extracts. Decoding never fails — it round-trips every byte to *some*
    string. The eventual numbers and dates are validated at row level."""
    for enc in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
        try:
            return content.decode(enc)
        except UnicodeDecodeError:
            continue
    return content.decode('latin-1', errors='replace')


def _is_xlsx(content: bytes) -> bool:
    """Detect Office Open XML (xlsx) by zip-header magic. xlsx files are
    zip archives containing XML parts, so they always start with
    `PK\\x03\\x04`. Old binary .xls (OLE2 compound, magic `\\xd0\\xcf...`)
    is intentionally NOT detected here — operators save those as xlsx
    or CSV before upload, and adding xlrd would pull a maintenance-mode
    dependency for a shrinking format."""
    return content[:4] == b'PK\x03\x04'


def _iter_xlsx_rows(content: bytes):
    """Yield each row of the first sheet as a list of strings.

    Cells round-trip through `str()` so the rest of the pipeline
    (date parsing, amount parsing) sees the same shapes as the CSV
    path. None / empty cells become '' so position-based indexing
    still works for partially-empty rows.

    Reads in read-only mode so a 60 MB Telcel statement doesn't
    materialise the whole workbook in RAM up front."""
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            yield [_xlsx_cell_to_str(c) for c in row]
    finally:
        wb.close()


def _xlsx_cell_to_str(value) -> str:
    """Convert an openpyxl cell value into the str shape the rest of
    the pipeline expects. datetime objects format as ISO-ish strings
    so _parse_date can read them with the profile's date_format."""
    if value is None:
        return ''
    # openpyxl returns datetime.datetime / datetime.date for date-typed
    # cells. Format consistently so _parse_date sees a parseable string.
    from datetime import date, datetime as _dt
    if isinstance(value, _dt):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    return str(value)


def _iter_csv(text: str, delimiter: str) -> list[list[str]]:
    return list(csv.reader(io.StringIO(text), delimiter=delimiter))


def _row_to_txn(row_no: int, raw: list[str], columns: list[str],
                 profile: CsvProfile) -> dict | None:
    """Translate one CSV row into a canonical Flex-shape dict."""
    if not raw or all((c or '').strip() == '' for c in raw):
        return None  # skip blank rows silently

    cell = lambda name: _cell(raw, columns, name)

    amount_raw = cell(profile.column_map.get('amount'))
    if profile.sign_convention == 'paid_in_withdrawn':
        # Two-column shape (M-Pesa / Telcel Cash): one row populates
        # either the Paid In column (CR) or the Withdrawn column (DR),
        # never both. Resolve sign + amount inline so the empty-column
        # check below doesn't reject withdrawal-only rows.
        paid_in = _parse_amount(amount_raw) if amount_raw else 0.0
        wd_raw = cell(profile.sign_column)
        withdrawn = _parse_amount(wd_raw) if wd_raw else 0.0
        if not paid_in and not withdrawn:
            raise _RowError("paid-in and withdrawn columns both empty")
        if paid_in:
            typ, signed_amount = 'CR', abs(paid_in)
        else:
            typ, signed_amount = 'DR', abs(withdrawn)
        amount = signed_amount
    else:
        if amount_raw is None or amount_raw == '':
            raise _RowError("amount column is empty")
        amount = _parse_amount(amount_raw)

    date_raw = cell(profile.column_map.get('value_date'))
    value_date = _parse_date(date_raw, profile.date_format)
    if not value_date:
        raise _RowError(f"could not parse date {date_raw!r} "
                        f"with format {profile.date_format!r}")

    booking_date_raw = cell(profile.column_map.get('booking_date'))
    booking_date = (_parse_date(booking_date_raw, profile.date_format)
                    if booking_date_raw else value_date)

    # paid_in_withdrawn already set typ + signed_amount inline above.
    if profile.sign_convention != 'paid_in_withdrawn':
        typ, signed_amount = _resolve_sign(profile, raw, columns, amount)

    ccy = (profile.currency
           or cell(profile.column_map.get('currency'))
           or '')

    # Carry through every source column as `_extra[colname] = value`
    # so downstream adapters (e.g. cards_loaders/csv_generic) can read
    # columns that aren't part of the canonical Flex shape — masked-PAN,
    # MCC, terminal_id, etc. Canonical columns still take precedence
    # because the cards seam looks up _extra first by configured name,
    # then falls back to the canonical dict.
    extra = {col: (raw[idx] if idx < len(raw) else '')
             for idx, col in enumerate(columns) if col}

    return {
        '_source': 'flex',
        '_row_number': row_no,
        '_used': False,
        'trn_ref':       (cell(profile.column_map.get('ref')) or '').strip(),
        'ac_branch':     (cell(profile.column_map.get('ac_branch')) or '').strip(),
        'ac_no':         (cell(profile.column_map.get('ac_no')) or '').strip(),
        'booking_date':  booking_date,
        'value_date':    value_date,
        'type':          typ,
        'narration':     (cell(profile.column_map.get('narration')) or '').strip(),
        'amount':        signed_amount,
        'ccy':           ccy.upper(),
        'module':        '',
        'external_ref':  (cell(profile.column_map.get('ref')) or '').strip(),
        'user_id':       '',
        '_extra':        extra,
    }


def _cell(raw: list[str], columns: list[str], name: str | None) -> str | None:
    """Look up a column by name in this row. None when the column
    doesn't exist (loader treats this as 'optional, not present')."""
    if not name:
        return None
    try:
        idx = columns.index(name)
    except ValueError:
        return None
    if idx >= len(raw):
        return None
    return (raw[idx] or '').strip()


def _parse_amount(s: str) -> float:
    """Parse a number that might be '1,234.56', '1234.56', '(123.45)',
    or '1.234,56' (European thousands convention with comma decimal).
    Returns the absolute float; sign is resolved separately so we can
    handle the column-based sign conventions consistently."""
    s = s.strip()
    if not s:
        raise _RowError("empty amount")
    # Accountancy-style negatives: '(123.45)' or '(123,45)'.
    negative = False
    if s.startswith('(') and s.endswith(')'):
        negative = True
        s = s[1:-1].strip()
    # Strip currency markers / spaces. We don't try to be clever about
    # which symbol matches the row — column_map['currency'] handles that.
    s = s.replace(' ', '').replace(' ', '')
    for sym in ('$', '€', '£', '¥', 'GHS', 'USD', 'EUR', 'GBP', 'NGN', 'KES'):
        s = s.replace(sym, '')
    # Heuristic for European-format numbers like "1.234,56": if there's
    # exactly one comma and the position suggests it's a decimal mark
    # (commas after the last dot, or no dot at all), treat comma as
    # decimal and dots as thousands separators.
    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s and '.' not in s:
        # "1,23" is ambiguous (1.23 EU, or 123 US thousands?). Default to
        # decimal — bank GL extracts almost never have thousand-comma
        # without also using a decimal point.
        s = s.replace(',', '.')
    try:
        v = float(s)
    except ValueError:
        raise _RowError(f"cannot parse amount {s!r}")
    return -v if negative else v


def _parse_date(s: str | None, fmt: str) -> int:
    """Return YYYYMMDD as an int, or 0 on failure (caller decides
    whether 0 is acceptable)."""
    if not s:
        return 0
    s = s.strip()
    try:
        return int(datetime.strptime(s, fmt).strftime('%Y%m%d'))
    except (ValueError, TypeError):
        return 0


def _resolve_sign(profile: CsvProfile, raw: list[str], columns: list[str],
                   amount: float) -> tuple[str, float]:
    """Apply the profile's sign convention. Returns (type, abs_amount).
    type is 'CR' or 'DR'. amount in the engine is always positive — the
    direction lives in the type field."""
    if profile.sign_convention == 'positive_credit':
        if amount > 0:
            return 'CR', abs(amount)
        return 'DR', abs(amount)
    # Both 'separate_column' and 'cr_dr_column' read sign_column.
    sign_raw = (_cell(raw, columns, profile.sign_column) or '').strip().upper()
    if sign_raw in ('CR', 'C', '+', 'CREDIT'):
        return 'CR', abs(amount)
    if sign_raw in ('DR', 'D', '-', 'DEBIT'):
        return 'DR', abs(amount)
    raise _RowError(
        f"sign column {profile.sign_column!r} has unrecognised value "
        f"{sign_raw!r} (expected CR/C/+/CREDIT or DR/D/-/DEBIT)")


# Common header names we match case-insensitively. Order doesn't matter;
# the wizard surfaces all columns and lets the user pick — these are
# just first-pass defaults.
_GUESS_PATTERNS = {
    'amount':       ('amount', 'amt', 'value', 'transaction amount',
                     'posting amount', 'tran amt', 'amount lcy', 'lcy amount'),
    'value_date':   ('value date', 'valdate', 'val date', 'value_date',
                     'effective date', 'tran date', 'transaction date',
                     'date'),
    'booking_date': ('booking date', 'book date', 'booking_date',
                     'posting date', 'entry date'),
    'ref':          ('reference', 'ref', 'tran ref', 'transaction ref',
                     'trn ref', 'tx ref', 'reference no', 'ext ref',
                     'external ref'),
    'narration':    ('narration', 'description', 'desc', 'memo',
                     'remarks', 'particulars', 'narrative'),
    'type':         ('type', 'cr/dr', 'dr/cr', 'cr_dr', 'transaction type',
                     'tran type', 'sign'),
    'currency':     ('currency', 'ccy', 'curr', 'iso currency'),
    'ac_no':        ('account', 'account no', 'account number', 'acct',
                     'ac no', 'ac_no', 'gl account', 'gl ac'),
    'ac_branch':    ('branch', 'branch code', 'branch_id', 'br'),
}


def _guess_columns(columns: list[str]) -> dict[str, str | None]:
    """Heuristic guess of which column maps to which canonical field.
    Case-insensitive substring/prefix match; returns the first hit per
    canonical field, None if nothing matched."""
    lower = [(c or '').lower().strip() for c in columns]
    out: dict[str, str | None] = {}
    for canonical, patterns in _GUESS_PATTERNS.items():
        match = None
        for col_idx, col_lower in enumerate(lower):
            if not col_lower:
                continue
            for p in patterns:
                if col_lower == p or p in col_lower:
                    match = columns[col_idx]
                    break
            if match:
                break
        out[canonical] = match
    return out
