"""
account_meta.py
===============

Pulls the account identity from the two input files so each reconciliation
session can be tagged with WHICH nostro / GL it belongs to. Needed once the
user starts loading files across many accounts — prevents accidentally
pairing BoA's MT950 against Citi-NY's Flexcube extract.

SWIFT side: the parser (mt950_to_excel.py / mt940_to_excel.py) already writes
a metadata panel at the top of the parsed xlsx. We re-read those cells.
Fields: Account (:25:), Statement reference, Opening / Closing balance —
the balance strings encode the currency, e.g. "C GHS 4,801,552,913.09 on 20260417".

Flexcube side: each row has AC_NO + AC_BRANCH + ACCT_CCY. A well-formed
acc_entries export is single-account; we validate that and flag if not.
"""

import re
from pathlib import Path

from openpyxl import load_workbook


_BAL_RE = re.compile(r'^[CD]\s+([A-Z]{3})\s')
_BAL_FULL_RE = re.compile(
    r'^([CD])\s+([A-Z]{3})\s+([\d,]+(?:\.\d+)?)\s+on\s+(\d{8})$'
)


def parse_balance_string(s: str) -> dict:
    """Turn 'C GHS 4,801,552,913.09 on 20260417' into structured fields.
    Returns {} when the string doesn't match — callers treat as missing."""
    m = _BAL_FULL_RE.match((s or '').strip())
    if not m:
        return {}
    return {
        'sign':     m.group(1),
        'currency': m.group(2),
        'amount':   float(m.group(3).replace(',', '')),
        'date':     int(m.group(4)),
    }


def extract_swift_meta(path: Path) -> dict:
    """Read the metadata panel above the transactions header."""
    # Explicit close() is required on Windows — read_only mode holds the
    # underlying zipfile handle open until GC, which blocks shutil.move.
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        meta: dict = {}
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] == 'Value date':
                break
            key, value = row[0], row[1] if len(row) > 1 else None
            if value is None:
                continue
            value = str(value).strip()
            if key == 'Account':
                meta['account'] = value
            elif key == 'Statement reference':
                meta['statement_ref'] = value
            elif key == 'Statement number':
                meta['statement_number'] = value
            elif key == 'Opening balance (60M)':
                meta['opening_balance'] = value
                b = parse_balance_string(value)
                if b:
                    meta['currency'] = b['currency']
                    meta['opening_balance_amount'] = b['amount']
                    meta['opening_balance_sign'] = b['sign']
                    meta['opening_balance_date'] = b['date']
            elif key == 'Closing balance (62F)':
                meta['closing_balance'] = value
                b = parse_balance_string(value)
                if b:
                    meta['closing_balance_amount'] = b['amount']
                    meta['closing_balance_sign'] = b['sign']
                    meta['closing_balance_date'] = b['date']
        return meta
    finally:
        wb.close()


def extract_flex_meta(flex_txns: list) -> dict:
    """Derive account metadata from loaded Flexcube dicts. Expects a single
    AC_NO per file — flags multi-account exports as a data issue."""
    if not flex_txns:
        return {}
    acs = sorted({f['ac_no'] for f in flex_txns if f.get('ac_no')})
    ccys = sorted({f['ccy'] for f in flex_txns if f.get('ccy')})
    first = flex_txns[0]
    return {
        'ac_no': first.get('ac_no'),
        'ac_branch': first.get('ac_branch'),
        'currency': first.get('ccy'),
        'multi_account': len(acs) > 1,
        'multi_currency': len(ccys) > 1,
        'all_accounts': acs,
        'all_currencies': ccys,
    }
