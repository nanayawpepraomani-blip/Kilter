"""Bulk-load accounts from the master spreadsheet into Kilter.

Source: C:/Users/NYPO/Desktop/Accounts to be loaded Kilter.xlsx
Sheet 1: NOSTROS & SUSPENSE - DEPT (access_area is explicit in col A)
Sheet 2: GLS (Branches)              (access_area derived from branch code)

Design choices:
  * swift_account = '' for every row. GL-only accounts have no SWIFT
    counterpart; nostros get their SWIFT number filled later via UPDATE.
    This keeps the NOT-NULL constraint happy without a schema change.
  * currency: column B in Sheet 2; derived from account name in Sheet 1.
  * malformed account numbers (from earlier review) are skipped and listed.
  * duplicate (flex_ac_no, currency) pairs are deduplicated in Python —
    first occurrence wins, extras listed.
"""

import sqlite3
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

SRC = Path('C:/Users/NYPO/Desktop/Accounts to be loaded Kilter.xlsx')
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from db import DB_PATH as DB  # noqa: E402

CURRENCIES = ('USD', 'EUR', 'GBP', 'GHS', 'ZAR', 'CHF', 'AUD', 'AED', 'CAD',
              'XAF', 'XOF', 'RWF', 'UGX', 'CNY', 'JPY')

# Sheet-2 branch-code prefix -> seeded access_area name.
# Built at runtime by scanning the access_areas table for names matching
# 'BRANCH <code> ...'; unknown branch codes get 'BRANCH <code>' as fallback.

# ---------------------------------------------------------------------------

def derive_currency_from_name(name: str | None) -> str | None:
    """Try to pick a 3-letter currency code out of an account name."""
    if not name:
        return None
    up = name.upper().strip()
    # Trailing 3-letter token after dash or space: "... EUR" / "...-USD"
    m = re.search(r'[\s\-]([A-Z]{3})\s*$', up)
    if m and m.group(1) in CURRENCIES:
        return m.group(1)
    # Inline token: "(USD)", "- GHS -"
    for c in CURRENCIES:
        if re.search(rf'\b{c}\b', up):
            return c
    # Common Ghanaian GHC alias
    if re.search(r'\bGHC\b', up) or 'GHC' in up:
        return 'GHS'
    return None


def is_malformed(acct: str | int | None, name: str | None) -> str | None:
    """Return a reason string if the row should be skipped, else None."""
    if acct in (None, ''):
        return "empty account number"
    if isinstance(acct, (int, float)):
        return None  # numeric = 13-digit customer account, always OK
    s = str(acct).strip()
    if ' CLOSED' in s.upper():
        return "contains ' CLOSED' marker"
    if s.endswith('.'):
        return "trailing period"
    if s.endswith('DD'):
        return "non-standard 'DD' suffix"
    if s.endswith('NN'):
        return "non-standard 'NN' suffix"
    # Length sanity: GL accounts should be 15 chars
    if len(s) < 12 or len(s) > 17:
        return f"unusual length ({len(s)} chars)"
    return None


def build_sheet1_rows(wb, branch_map: dict) -> tuple[list, list]:
    """Returns (accepted_rows, skipped_rows). Each accepted row is a dict
    ready for insert; each skipped row is (source_row_number, reason, acct, name)."""
    ws = wb['NOSTROS & SUSPENSE - DEPT']
    accepted, skipped = [], []
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue   # header
        if all(v in (None, '') for v in r):
            continue   # blank row
        area, acct, name = r[0], r[1], r[2]
        if area in (None, '') and (name in (None, '')) and acct:
            skipped.append((i, "stray row (area/name blank)", acct, name))
            continue

        reason = is_malformed(acct, name)
        if reason:
            skipped.append((i, reason, acct, name))
            continue

        ccy = derive_currency_from_name(name)
        if ccy is None:
            # Many NOSTRO / affiliate account names don't embed currency.
            # Fall back to a house-default; operator can fill via UPDATE
            # later. Adjust this default to whatever your book's dominant
            # currency is before running the loader.
            ccy = 'GHS'   # house-default currency for this loader run

        accepted.append({
            'src_row': i,
            'label': str(name).strip(),
            'shortname': None,
            'access_area': str(area).strip() if area else None,
            'bic': None,
            'swift_account': '',   # placeholder; filled later for nostros
            'flex_ac_no': str(acct).strip() if not isinstance(acct, (int, float)) else str(int(acct)),
            'currency': ccy,
            'notes': 'imported from Sheet 1 (NOSTROS & SUSPENSE - DEPT)',
        })
    return accepted, skipped


def build_sheet2_rows(wb, branch_map: dict) -> tuple[list, list]:
    """Sheet 2: GLS (Branches). Currency is explicit; access_area from branch
    code embedded in the account name's leading 3 chars."""
    ws = wb['GLS (Branches)']
    accepted, skipped = [], []
    rows = list(ws.iter_rows(values_only=True))
    # rows[0] is 'GLS (BRANCHES)' title, rows[1] is header, rows[2:] data (with possible blanks)
    for i, r in enumerate(rows, start=1):
        if i <= 2:
            continue
        if all(v in (None, '') for v in r):
            continue
        acct, ccy, origin, name = r[0], r[1], r[2], r[3]

        reason = is_malformed(acct, name)
        if reason:
            skipped.append((i, reason, acct, name))
            continue

        if ccy not in CURRENCIES:
            skipped.append((i, f"unknown currency {ccy!r}", acct, name))
            continue

        # Branch code = first 3 digits of account name (e.g. "072MANAGERS...")
        branch_code = None
        if name:
            m = re.match(r'^(\d{3})', str(name).strip())
            if m:
                branch_code = m.group(1)
        if branch_code:
            area = branch_map.get(branch_code, f"BRANCH {branch_code}")
        else:
            area = 'GLS (Branches)'

        accepted.append({
            'src_row': i,
            'label': str(name).strip(),
            'shortname': None,
            'access_area': area,
            'bic': None,
            'swift_account': '',
            'flex_ac_no': str(acct).strip(),
            'currency': ccy,
            'notes': 'imported from Sheet 2 (GLS Branches)',
        })
    return accepted, skipped


def dedup(rows: list) -> tuple[list, list]:
    """Enforce UNIQUE (flex_ac_no, currency) at the Python level. Keeps first
    occurrence, returns (keep, drop_as_duplicate)."""
    seen = {}
    keep, drops = [], []
    for r in rows:
        key = (r['flex_ac_no'], r['currency'])
        if key in seen:
            drops.append((r['src_row'], f"duplicate of row {seen[key]}", r['flex_ac_no'], r['label']))
        else:
            seen[key] = r['src_row']
            keep.append(r)
    return keep, drops


def load_branch_map(conn) -> dict:
    """Map 3-digit branch codes ('072') to the seeded access_area name
    ('BRANCH 072 ASHIA')."""
    out = {}
    for row in conn.execute("SELECT name FROM access_areas WHERE active=1"):
        m = re.match(r'^BRANCH\s+(\d{3})\b', row[0])
        if m:
            out[m.group(1)] = row[0]
    return out


# ---------------------------------------------------------------------------

def main():
    if not SRC.exists():
        print(f"ERROR: source file not found: {SRC}")
        sys.exit(1)

    print(f"Reading {SRC} ...")
    wb = openpyxl.load_workbook(SRC, data_only=True)

    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")

    branch_map = load_branch_map(conn)
    print(f"Built branch map from access_areas: {len(branch_map)} branch codes.")

    s1_accepted, s1_skipped = build_sheet1_rows(wb, branch_map)
    s2_accepted, s2_skipped = build_sheet2_rows(wb, branch_map)

    print(f"\nSheet 1 (NOSTROS & SUSPENSE): {len(s1_accepted)} accepted, {len(s1_skipped)} skipped")
    print(f"Sheet 2 (GLS Branches):       {len(s2_accepted)} accepted, {len(s2_skipped)} skipped")

    # Dedup across the combined set.
    all_rows = s1_accepted + s2_accepted
    keep, drops = dedup(all_rows)
    print(f"Dedup: kept {len(keep)}, dropped {len(drops)} duplicates "
          f"(same flex_ac_no + currency seen earlier)")

    # Insert.
    now = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
    created_by = 'bulk-import'
    inserted = 0
    errors = []
    try:
        conn.execute("BEGIN")
        for r in keep:
            try:
                conn.execute(
                    "INSERT INTO accounts (label, shortname, access_area, bic, "
                    "swift_account, flex_ac_no, currency, notes, active, "
                    "created_at, created_by) VALUES (?,?,?,?,?,?,?,?,1,?,?)",
                    (r['label'], r['shortname'], r['access_area'], r['bic'],
                     r['swift_account'], r['flex_ac_no'], r['currency'],
                     r['notes'], now, created_by),
                )
                inserted += 1
            except sqlite3.IntegrityError as exc:
                errors.append((r['src_row'], str(exc), r['flex_ac_no'], r['label']))
        conn.execute("COMMIT")
    except Exception:
        conn.execute("ROLLBACK")
        raise

    print(f"\nInserted: {inserted}")
    print(f"Integrity errors: {len(errors)}")

    # Report skipped / dropped rows by reason category.
    combined_skipped = s1_skipped + s2_skipped
    reasons = {}
    for _, reason, _, _ in combined_skipped:
        reasons[reason] = reasons.get(reason, 0) + 1
    if reasons:
        print("\nSkipped rows by reason:")
        for reason, n in sorted(reasons.items(), key=lambda x: -x[1]):
            print(f"  {n:5d}  {reason}")

    if drops:
        print(f"\nDuplicate-row drops (first 10 of {len(drops)}):")
        for src_row, reason, acct, name in drops[:10]:
            print(f"  row{src_row}: {acct} - {name} ({reason})")

    if errors:
        print(f"\nIntegrity errors (first 10 of {len(errors)}):")
        for src_row, msg, acct, name in errors[:10]:
            print(f"  row{src_row}: {acct} - {name}: {msg}")

    # Breakdown by access_area (top 25)
    print("\nTop 25 access_areas loaded:")
    rows = conn.execute(
        "SELECT COALESCE(access_area,'<null>') AS area, COUNT(*) AS n "
        "FROM accounts GROUP BY area ORDER BY n DESC LIMIT 25"
    ).fetchall()
    for row in rows:
        print(f"  {row['n']:5d}  {row['area']}")

    total_areas = conn.execute(
        "SELECT COUNT(DISTINCT COALESCE(access_area,'<null>')) FROM accounts"
    ).fetchone()[0]
    total_accounts = conn.execute("SELECT COUNT(*) FROM accounts").fetchone()[0]
    print(f"\nTotal: {total_accounts} accounts across {total_areas} access areas.")

    conn.close()


if __name__ == '__main__':
    main()
