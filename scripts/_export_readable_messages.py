"""Export MT940 / MT950 / camt.053 / camt.054 messages to one readable xlsx.

Scans every SWIFT .out under messages/ and every camt XML under the user's
Downloads/camt/{053,054} folders, parses each using the existing Kilter
loaders, and writes a single workbook with one sheet per format plus a
Summary sheet.
"""

from __future__ import annotations

import sys
import traceback
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.stdout.reconfigure(encoding='utf-8')

from iso20022_loader import detect_camt_type, _locate_container, _build_meta, _iter_txns
from swift_core import detect_message_type, parse_swift_statement


MT_ROOTS = [Path('d:/Kilter/messages')]
CAMT_053_DIR = Path('C:/Users/NYPO/Downloads/camt/camt/053')
CAMT_054_DIR = Path('C:/Users/NYPO/Downloads/camt/camt/054')
OUT_PATH = Path('d:/Kilter/exports/readable_messages.xlsx')

SHEET_COLUMNS = [
    'File', 'Folder',
    'Account', 'BIC', 'Currency',
    'Opening balance', 'Closing balance',
    'Statement ref', 'Statement #',
    'Row #', 'Value date', 'Book date',
    'C/D', 'Amount', 'Txn type',
    'Our ref', 'Their ref',
    'Narration 1', 'Narration 2',
]

SUMMARY_COLUMNS = [
    'File', 'Folder', 'Format',
    'Account', 'BIC', 'Currency',
    'Opening balance', 'Closing balance',
    'Txn count', 'Status',
]


def _yyyymmdd_to_date(v):
    if not v:
        return None
    try:
        s = str(int(v))
        if len(s) != 8:
            return None
        return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
    except (TypeError, ValueError):
        return None


def _fmt_balance(amount, sign, currency, dt):
    if amount is None:
        return ''
    d = _yyyymmdd_to_date(dt)
    d_str = d.isoformat() if d else str(dt or '')
    return f"{sign or ''} {currency or ''} {amount:,.2f} on {d_str}".strip()


def _parse_mt(path: Path):
    raw = path.read_text(encoding='latin-1', errors='replace')
    mt = detect_message_type(raw)
    if mt not in ('940', '950'):
        return None, mt, None, None, 'not MT940/MT950'
    parsed = parse_swift_statement(raw)
    meta = {
        'account': parsed.get('account'),
        'bic': None,
        'currency': None,
        'statement_ref': parsed.get('transaction_reference'),
        'statement_number': parsed.get('statement_number'),
    }
    ob = parsed.get('opening_balance') or {}
    cb = parsed.get('closing_balance') or {}
    if ob:
        meta['currency'] = ob.get('currency')
        meta['opening_balance'] = _fmt_balance(
            ob.get('amount'), ob.get('mark'), ob.get('currency'), ob.get('date'))
    else:
        meta['opening_balance'] = ''
    if cb:
        meta['currency'] = meta['currency'] or cb.get('currency')
        meta['closing_balance'] = _fmt_balance(
            cb.get('amount'), cb.get('mark'), cb.get('currency'), cb.get('date'))
    else:
        meta['closing_balance'] = ''
    # Block 2 sender BIC
    import re
    m = re.search(
        r'\{2:O\d{3}\d{4}\d{6}([A-Z]{4}[A-Z]{2}[A-Z0-9]{2}(?:[A-Z0-9]{3,4})?)',
        raw,
    )
    if m:
        bic = m.group(1)
        meta['bic'] = bic[:8] if len(bic) == 12 else bic

    rows = []
    for idx, t in enumerate(parsed.get('transactions') or [], start=1):
        if t.get('parse_error'):
            continue
        rows.append({
            'row': idx,
            'value_date': _yyyymmdd_to_date(t.get('value_date')),
            'book_date': _yyyymmdd_to_date(t.get('entry_date')),
            'sign': t.get('sign') or '',
            'amount': t.get('amount'),
            'txn_type': t.get('txn_type') or '',
            'our_ref': t.get('account_ref') or '',
            'their_ref': t.get('bank_ref') or '',
            'narration_1': t.get('supplementary') or '',
            'narration_2': t.get('narrative') or '',
        })
    return mt, mt, meta, rows, 'ok'


def _parse_camt(path: Path, expected: str):
    kind = detect_camt_type(path)
    if kind != expected:
        return kind, None, None, f'detected {kind or "none"}, expected {expected}'
    container, _, root = _locate_container(path)
    m = _build_meta(container, kind, root)
    meta = {
        'account': m.get('account'),
        'bic': m.get('bic'),
        'currency': m.get('currency'),
        'statement_ref': m.get('statement_ref'),
        'statement_number': m.get('statement_number'),
        'opening_balance': _fmt_balance(
            m.get('opening_balance_amount'), m.get('opening_balance_sign'),
            m.get('currency'), m.get('opening_balance_date')),
        'closing_balance': _fmt_balance(
            m.get('closing_balance_amount'), m.get('closing_balance_sign'),
            m.get('currency'), m.get('closing_balance_date')),
    }
    rows = []
    for idx, t in enumerate(_iter_txns(container), start=1):
        rows.append({
            'row': idx,
            'value_date': _yyyymmdd_to_date(t.get('value_date')),
            'book_date': _yyyymmdd_to_date(t.get('book_date')),
            'sign': t.get('sign') or '',
            'amount': t.get('amount'),
            'txn_type': '',
            'our_ref': t.get('our_ref') or '',
            'their_ref': t.get('their_ref') or '',
            'narration_1': t.get('booking_text_1') or '',
            'narration_2': t.get('booking_text_2') or '',
        })
    return kind, meta, rows, 'ok'


def _gather_mt_files():
    files = []
    for root in MT_ROOTS:
        if not root.exists():
            continue
        for p in sorted(root.rglob('*.out')):
            files.append(p)
    return files


def _gather_camt_files(d: Path):
    if not d.exists():
        return []
    return sorted([p for p in d.iterdir() if p.is_file()])


HEADER_FILL = PatternFill('solid', fgColor='1F3A8A')
HEADER_FONT = Font(bold=True, color='FFFFFF')
ZEBRA_FILL = PatternFill('solid', fgColor='F5F7FA')
COL_WIDTHS = {
    'File': 32, 'Folder': 14,
    'Account': 22, 'BIC': 14, 'Currency': 9,
    'Opening balance': 30, 'Closing balance': 30,
    'Statement ref': 20, 'Statement #': 11,
    'Row #': 7, 'Value date': 12, 'Book date': 12,
    'C/D': 5, 'Amount': 16, 'Txn type': 10,
    'Our ref': 28, 'Their ref': 22,
    'Narration 1': 40, 'Narration 2': 40,
}
SUMMARY_WIDTHS = {
    'File': 32, 'Folder': 14, 'Format': 9,
    'Account': 22, 'BIC': 14, 'Currency': 9,
    'Opening balance': 30, 'Closing balance': 30,
    'Txn count': 10, 'Status': 20,
}


def _write_header(ws, columns, widths):
    for i, col in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 14)
    ws.freeze_panes = 'A2'


def _write_rows(ws, file_name, folder, meta, txn_rows, next_row, zebra_flag):
    start_row = next_row
    fill = ZEBRA_FILL if zebra_flag else None
    statement_cells = [
        file_name, folder,
        meta.get('account') or '',
        meta.get('bic') or '',
        meta.get('currency') or '',
        meta.get('opening_balance') or '',
        meta.get('closing_balance') or '',
        meta.get('statement_ref') or '',
        meta.get('statement_number') or '',
    ]
    if not txn_rows:
        for i, v in enumerate(statement_cells, start=1):
            c = ws.cell(row=next_row, column=i, value=v)
            if fill:
                c.fill = fill
        for i in range(len(statement_cells) + 1, len(SHEET_COLUMNS) + 1):
            c = ws.cell(row=next_row, column=i, value='')
            if fill:
                c.fill = fill
        return next_row + 1
    for t in txn_rows:
        row_cells = statement_cells + [
            t['row'],
            t['value_date'],
            t['book_date'],
            t['sign'],
            t['amount'],
            t['txn_type'],
            t['our_ref'],
            t['their_ref'],
            t['narration_1'],
            t['narration_2'],
        ]
        for i, v in enumerate(row_cells, start=1):
            c = ws.cell(row=next_row, column=i, value=v)
            if fill:
                c.fill = fill
            col_name = SHEET_COLUMNS[i - 1]
            if col_name in ('Value date', 'Book date') and isinstance(v, date):
                c.number_format = 'yyyy-mm-dd'
            elif col_name == 'Amount' and isinstance(v, (int, float)):
                c.number_format = '#,##0.00;[Red]-#,##0.00'
        next_row += 1
    return next_row


def main():
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    summary = wb.active
    summary.title = 'Summary'

    sheets = {
        '940': wb.create_sheet('MT940'),
        '950': wb.create_sheet('MT950'),
        'camt053': wb.create_sheet('camt.053'),
        'camt054': wb.create_sheet('camt.054'),
    }
    for ws in sheets.values():
        _write_header(ws, SHEET_COLUMNS, COL_WIDTHS)
    _write_header(summary, SUMMARY_COLUMNS, SUMMARY_WIDTHS)

    next_row = {k: 2 for k in sheets}
    zebra = {k: False for k in sheets}
    summary_row = 2
    counts = {'940': 0, '950': 0, 'camt053': 0, 'camt054': 0, 'skipped': 0, 'errors': 0}

    # MT940 / MT950
    mt_files = _gather_mt_files()
    print(f'Scanning {len(mt_files)} .out files...')
    for p in mt_files:
        folder = p.parent.name
        try:
            mt, kind, meta, rows, status = _parse_mt(p)
        except Exception as e:
            counts['errors'] += 1
            summary.cell(row=summary_row, column=1, value=p.name)
            summary.cell(row=summary_row, column=2, value=folder)
            summary.cell(row=summary_row, column=3, value='?')
            summary.cell(row=summary_row, column=10, value=f'error: {e}')
            summary_row += 1
            continue
        if mt not in sheets:
            counts['skipped'] += 1
            summary.cell(row=summary_row, column=1, value=p.name)
            summary.cell(row=summary_row, column=2, value=folder)
            summary.cell(row=summary_row, column=3, value=mt or '?')
            summary.cell(row=summary_row, column=10, value=status)
            summary_row += 1
            continue
        ws = sheets[mt]
        next_row[mt] = _write_rows(ws, p.name, folder, meta, rows, next_row[mt], zebra[mt])
        zebra[mt] = not zebra[mt]
        counts[mt] += 1
        for i, v in enumerate([
            p.name, folder, f'MT{mt}',
            meta.get('account') or '', meta.get('bic') or '', meta.get('currency') or '',
            meta.get('opening_balance') or '', meta.get('closing_balance') or '',
            len(rows), status,
        ], start=1):
            summary.cell(row=summary_row, column=i, value=v)
        summary_row += 1

    # camt.053
    for p in _gather_camt_files(CAMT_053_DIR):
        try:
            kind, meta, rows, status = _parse_camt(p, 'camt053')
        except Exception as e:
            counts['errors'] += 1
            summary.cell(row=summary_row, column=1, value=p.name)
            summary.cell(row=summary_row, column=2, value=p.parent.name)
            summary.cell(row=summary_row, column=3, value='camt.053')
            summary.cell(row=summary_row, column=10, value=f'error: {e}')
            summary_row += 1
            continue
        if kind != 'camt053':
            counts['skipped'] += 1
            summary.cell(row=summary_row, column=1, value=p.name)
            summary.cell(row=summary_row, column=2, value=p.parent.name)
            summary.cell(row=summary_row, column=3, value='camt.053')
            summary.cell(row=summary_row, column=10, value=status)
            summary_row += 1
            continue
        ws = sheets['camt053']
        next_row['camt053'] = _write_rows(
            ws, p.name, p.parent.name, meta, rows, next_row['camt053'], zebra['camt053'])
        zebra['camt053'] = not zebra['camt053']
        counts['camt053'] += 1
        for i, v in enumerate([
            p.name, p.parent.name, 'camt.053',
            meta.get('account') or '', meta.get('bic') or '', meta.get('currency') or '',
            meta.get('opening_balance') or '', meta.get('closing_balance') or '',
            len(rows), status,
        ], start=1):
            summary.cell(row=summary_row, column=i, value=v)
        summary_row += 1

    # camt.054
    for p in _gather_camt_files(CAMT_054_DIR):
        try:
            kind, meta, rows, status = _parse_camt(p, 'camt054')
        except Exception as e:
            counts['errors'] += 1
            summary.cell(row=summary_row, column=1, value=p.name)
            summary.cell(row=summary_row, column=2, value=p.parent.name)
            summary.cell(row=summary_row, column=3, value='camt.054')
            summary.cell(row=summary_row, column=10, value=f'error: {e}')
            summary_row += 1
            continue
        if kind != 'camt054':
            counts['skipped'] += 1
            summary.cell(row=summary_row, column=1, value=p.name)
            summary.cell(row=summary_row, column=2, value=p.parent.name)
            summary.cell(row=summary_row, column=3, value='camt.054')
            summary.cell(row=summary_row, column=10, value=status)
            summary_row += 1
            continue
        ws = sheets['camt054']
        next_row['camt054'] = _write_rows(
            ws, p.name, p.parent.name, meta, rows, next_row['camt054'], zebra['camt054'])
        zebra['camt054'] = not zebra['camt054']
        counts['camt054'] += 1
        for i, v in enumerate([
            p.name, p.parent.name, 'camt.054',
            meta.get('account') or '', meta.get('bic') or '', meta.get('currency') or '',
            meta.get('opening_balance') or '', meta.get('closing_balance') or '',
            len(rows), status,
        ], start=1):
            summary.cell(row=summary_row, column=i, value=v)
        summary_row += 1

    summary.freeze_panes = 'A2'
    wb.save(OUT_PATH)
    print(f'\nWrote {OUT_PATH}')
    print(f'  MT940:    {counts["940"]} files')
    print(f'  MT950:    {counts["950"]} files')
    print(f'  camt.053: {counts["camt053"]} files')
    print(f'  camt.054: {counts["camt054"]} files')
    if counts['skipped']:
        print(f'  skipped:  {counts["skipped"]} (see Summary sheet)')
    if counts['errors']:
        print(f'  errors:   {counts["errors"]} (see Summary sheet)')


if __name__ == '__main__':
    try:
        main()
    except Exception:
        traceback.print_exc()
        sys.exit(1)
