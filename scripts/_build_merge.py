"""Build SWIFT <-> Flexcube merge candidates spreadsheet."""

import sys, re
from pathlib import Path
from collections import defaultdict, Counter

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.stdout.reconfigure(encoding='utf-8')

from swift_loader import extract_swift_meta_raw

# -------- scan SWIFT spool --------
folder = Path('C:/Users/NYPO/Desktop/22042026/22042026')
by_account = defaultdict(lambda: {'count': 0, 'currencies': Counter(),
                                   'bics': Counter(), 'sample_file': None})
for f in sorted(folder.glob('*.out')):
    try:
        meta = extract_swift_meta_raw(f)
    except Exception:
        continue
    acc, ccy, bic = meta.get('account'), meta.get('currency'), meta.get('bic')
    if not (acc and ccy):
        continue
    d = by_account[acc]
    d['count'] += 1
    d['currencies'][ccy] += 1
    if bic:
        d['bics'][bic] += 1
    if d['sample_file'] is None:
        d['sample_file'] = f.name

spool = []
for acc, d in by_account.items():
    for ccy, n in d['currencies'].items():
        bic = d['bics'].most_common(1)[0][0] if d['bics'] else ''
        spool.append({'account': acc, 'currency': ccy, 'bic': bic,
                      'count': n, 'sample_file': d['sample_file']})

# -------- BIC -> keyword dictionary --------
# Maps correspondent BICs to narration keywords so the merge-candidate
# builder can propose a SWIFT side for a Flex row by matching on
# free-text name. Extend with your bank's own correspondent BICs (and
# intragroup BICs if the bank has subsidiaries) — the seed below covers
# major global correspondents most correspondent-banking treasuries
# deal with; it is NOT exhaustive.
BIC_KEYWORDS = {
    # Global correspondents
    'CITIUS33': ['CITI BANK, NEW YORK', 'CITI NEW YORK', 'CITIBANK NEW YORK',
                 'CITI NY', 'CITIBANK, NEW YORK', 'CITIBANK NY'],
    'CITIGB2L': ['CITI LONDON', 'CITIBANK LONDON', 'CITI BANK LONDON'],
    'SCBLUS33': ['STANDARD CHARTERED', 'STANCHART', 'SCB'],
    'BKTRUS33': ['BANKERS TRUST', 'DB TRUST', 'DEUTSCHE BANK TRUST',
                 'DB LOND', 'DB LONDON'],
    'GENODEFF': ['DZ BANK', 'DZBANK'],
    'NEDSZAJJ': ['NEDBANK'],
    'FIRNZAJJ': ['FIRST NATIONAL BANK OF SOUTH AFRICA', 'FNB', 'FIRST NATIONAL'],
    'COBADEFF': ['COMMERZBANK'],
    'BHFBDEFF': ['BHF'],
    'DEUTDEFF': ['DEUTSCHE'],
    'NATXFRPP': ['NATIXIS'],
    'UBSWCHZH': ['UBS'],
    'BOMLAEAD': ['MASHREQ', 'BANK OF MASHREQ'],
    'AFXMEGCA': ['ATTIJARI', 'ATTIJARIWAFA', 'EGYPT'],
    'DABADKKK': ['DANSKE', 'DAN DANSKE'],
    'BKCHCNBJ': ['BANK OF CHINA'],
    'ROYCCAT2': ['ROYAL BANK OF CANADA', 'RBC'],
}


def candidate_bics(flex_name):
    up = (flex_name or '').upper()
    hits = []
    for bic, kws in BIC_KEYWORDS.items():
        for kw in kws:
            if kw in up:
                hits.append((bic, len(kw)))
                break
    hits.sort(key=lambda x: -x[1])
    return [b for b, _ in hits]


CURRENCIES = ('USD', 'EUR', 'GBP', 'GHS', 'ZAR', 'CHF', 'AUD', 'AED',
              'XAF', 'XOF', 'RWF', 'UGX', 'CNY', 'JPY', 'CAD')


def extract_currency(name):
    if not name:
        return None
    up = name.upper()
    m = re.search(r'[\s\-]([A-Z]{3})\s*$', up)
    if m and m.group(1) in CURRENCIES:
        return m.group(1)
    for c in CURRENCIES:
        if re.search(rf'\b{c}\b', up):
            return c
    return None


# -------- load Flex accounts (sheet 1) --------
wb = openpyxl.load_workbook('C:/Users/NYPO/Desktop/Accounts to be loaded.xlsx', data_only=True)
ws = wb['NOSTROS & SUSPENSE - DEPT']
flex_rows = []
for r in list(ws.iter_rows(values_only=True))[1:]:
    if all(v in (None, '') for v in r):
        continue
    area, acct, name = r[0], r[1], r[2]
    flex_rows.append({
        'area': area,
        'flex_ac_no': str(acct).strip() if acct else '',
        'name': (name or '').strip(),
    })

SWIFT_EXPECTED_AREAS = {'NOSTRO', 'AFFILIATES', 'BANK OF GHANA', 'SUBSIDIARIES'}
eligible = [f for f in flex_rows if f['area'] in SWIFT_EXPECTED_AREAS]
print(f"Flex rows eligible for SWIFT pairing: {len(eligible)}")

# -------- build candidates --------
spool_by_bic = defaultdict(list)
for s in spool:
    spool_by_bic[s['bic']].append(s)

used_swift = set()
merge_rows = []
nomatch_rows = []

for f in eligible:
    bics = candidate_bics(f['name'])
    ccy = extract_currency(f['name'])
    cand = []
    for bic in bics:
        for s in spool_by_bic.get(bic, []):
            ccy_match = (ccy is None) or (s['currency'] == ccy)
            cand.append({**s, 'ccy_match': ccy_match,
                         'bic_rank': bics.index(bic)})
    cand.sort(key=lambda x: (not x['ccy_match'], x['bic_rank'], -x['count']))

    if not cand:
        nomatch_rows.append({**f, 'derived_currency': ccy or ''})
        continue

    for i, s in enumerate(cand[:5]):
        confidence = 'HIGH' if (i == 0 and s['ccy_match']) else \
                     ('MED' if s['ccy_match'] else 'LOW')
        tick = 'Y' if (i == 0 and s['ccy_match']) else ''
        merge_rows.append({
            'area': f['area'], 'flex_ac_no': f['flex_ac_no'],
            'flex_name': f['name'], 'derived_currency': ccy or '',
            'swift_account': s['account'], 'swift_currency': s['currency'],
            'swift_bic': s['bic'], 'files_today': s['count'],
            'sample_file': s['sample_file'], 'confidence': confidence,
            'confirm': tick,
        })
        if tick == 'Y':
            used_swift.add((s['account'], s['currency']))

unmatched_swift = [s for s in spool
                   if (s['account'], s['currency']) not in used_swift]

# -------- build workbook --------
out_path = Path('C:/Users/NYPO/Desktop/SWIFT_Flex_merge_22042026.xlsx')
outwb = openpyxl.Workbook()

# Instructions sheet (first tab)
info = outwb.active
info.title = 'Instructions'
info.append(['SWIFT vs Flexcube merge worksheet'])
info.append([''])
info.append(['Goal: confirm which SWIFT account pairs with each Nostro / Affiliate /'])
info.append(['BoG / Subsidiaries Flexcube account, so they can be registered in Kilter.'])
info.append([''])
info.append(['How to use:'])
info.append(['1. Go to "Merge candidates" tab.'])
info.append(['2. Each Flex account shows up to 5 candidate SWIFT accounts, ranked best-first.'])
info.append(['3. Rows with currency match AND best BIC keyword are pre-ticked (Confirm=Y).'])
info.append(['4. Review: change Y to N for wrong pre-ticks; set Y on correct rows otherwise.'])
info.append(['5. Only ONE row per Flex account should end up with Y.'])
info.append(['6. Send back - I will register the Y rows in Kilter.'])
info.append([''])
info.append(['Colour code in Merge candidates:'])
info.append(['  GREEN  = HIGH confidence - pre-ticked, currency+BIC both matched'])
info.append(['  YELLOW = MED  confidence - currency matched, BIC was not best'])
info.append(['  RED    = LOW  confidence - currency did NOT match; unlikely correct'])
info.append([''])
info.append(['Tabs:'])
info.append(['  Merge candidates         - main worksheet with suggested pairings'])
info.append(['  Flex - no SWIFT candidate - Flex rows where no keyword matched any BIC'])
info.append(['  SWIFT - not proposed     - SWIFT accts from today\'s spool not proposed for any Flex'])

# Merge candidates
m = outwb.create_sheet('Merge candidates')
m.append(['Access area', 'Flex account', 'Flex account name', 'Flex ccy (derived)',
          'Confirm? (Y/N)', 'Confidence',
          'SWIFT account (:25:)', 'SWIFT ccy', 'Sender BIC',
          'Files today', 'Sample file'])
for r in merge_rows:
    m.append([r['area'], r['flex_ac_no'], r['flex_name'], r['derived_currency'],
              r['confirm'], r['confidence'],
              r['swift_account'], r['swift_currency'], r['swift_bic'],
              r['files_today'], r['sample_file']])

dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
dv.add(f"E2:E{max(m.max_row, 2)}")
m.add_data_validation(dv)

# Flex - no SWIFT candidate
nf = outwb.create_sheet('Flex - no SWIFT candidate')
nf.append(['Access area', 'Flex account', 'Flex account name',
           'Derived ccy', 'Notes'])
for r in nomatch_rows:
    nf.append([r['area'], r['flex_ac_no'], r['name'],
               r['derived_currency'],
               'Name contains no known correspondent-bank keyword'])

# SWIFT - not proposed
us = outwb.create_sheet('SWIFT - not proposed')
us.append(['SWIFT account', 'Currency', 'Sender BIC', 'Files today',
           'Sample file'])
for s in unmatched_swift:
    us.append([s['account'], s['currency'], s['bic'],
               s['count'], s['sample_file']])

# Stats appended to Instructions
info.append([''])
info.append(['Stats (this extract):'])
info.append([f'  Eligible Flex rows (NOSTRO/AFFILIATES/BoG/SUBSIDIARIES): {len(eligible)}'])
info.append([f'  Flex rows with at least one candidate: {len(eligible) - len(nomatch_rows)}'])
info.append([f'  Flex rows with NO candidate: {len(nomatch_rows)}'])
info.append([f'  Total candidate pairings proposed: {len(merge_rows)}'])
num_preticked = sum(1 for r in merge_rows if r['confirm'] == 'Y')
info.append([f'  Pre-ticked HIGH-confidence matches: {num_preticked}'])
info.append([f'  SWIFT (account,ccy) pairs not proposed for any Flex: {len(unmatched_swift)}'])

# Formatting
hdr_font = Font(bold=True, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='305496')
hi_fill = PatternFill('solid', fgColor='C6EFCE')
med_fill = PatternFill('solid', fgColor='FFEB9C')
low_fill = PatternFill('solid', fgColor='FFC7CE')
title_font = Font(bold=True, size=14)

info['A1'].font = title_font

for ws_ in (m, nf, us):
    for cell in ws_[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
    ws_.freeze_panes = 'A2'

for row_idx in range(2, m.max_row + 1):
    conf = m.cell(row=row_idx, column=6).value
    fill = {'HIGH': hi_fill, 'MED': med_fill, 'LOW': low_fill}.get(conf)
    if fill:
        for col in range(1, m.max_column + 1):
            m.cell(row=row_idx, column=col).fill = fill

for ws_ in (info, m, nf, us):
    for col in range(1, ws_.max_column + 1):
        letter = get_column_letter(col)
        max_len = 10
        for row in ws_.iter_rows(min_col=col, max_col=col, values_only=True):
            v = row[0]
            if v is None:
                continue
            max_len = max(max_len, min(55, len(str(v)) + 2))
        ws_.column_dimensions[letter].width = max_len

outwb.save(out_path)

print(f"\nWrote {out_path}")
print(f"Eligible Flex rows: {len(eligible)}")
print(f"  With candidate(s): {len(eligible) - len(nomatch_rows)}")
print(f"  Without candidate: {len(nomatch_rows)}")
print(f"Total candidate pairings: {len(merge_rows)}")
print(f"Pre-ticked HIGH-confidence: {num_preticked}")
print(f"SWIFT pairs leftover (no Flex proposed): {len(unmatched_swift)}")

print("\n--- Sample pre-ticked matches ---")
for r in [r for r in merge_rows if r['confirm'] == 'Y'][:15]:
    print(f"  {r['flex_name'][:42]:42s} ({r['derived_currency']:3s}) "
          f"= {r['swift_account']:25s} {r['swift_bic']:10s} {r['files_today']} files")

print("\n--- Flex rows with NO candidate (first 15) ---")
for r in nomatch_rows[:15]:
    print(f"  {r['area']:15s}  {r['name']}")
