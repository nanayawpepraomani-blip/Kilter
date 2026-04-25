"""Scan today's SWIFT spool, cross-reference against the loaded Flex accounts,
and produce a review workbook with suggested Flex pairings.

Output: C:/Users/NYPO/Desktop/SWIFT_assign_04232026.xlsx
  Sheet 'To assign'   — one row per (swift_account, currency). 'Confirm?'
                         dropdown, pre-ticked 'Y' for high-confidence matches.
  Sheet 'Not needed'  — SWIFT accounts that don't match any Flex row; ops can
                         confirm these aren't accounts we care about.
  Sheet 'Skipped'     — files that weren't MT940/950, or couldn't be parsed.
  Sheet 'Summary'     — stats.
"""

import re
import sys
import sqlite3
from pathlib import Path
from collections import defaultdict, Counter
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

_HERE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_HERE))
sys.stdout.reconfigure(encoding='utf-8')

from swift_loader import extract_swift_meta_raw
from swift_core import detect_message_type
from db import DB_PATH as DB

FOLDER = Path('C:/Users/NYPO/Desktop/Swift_04232026/Swift_04232026')
OUT = Path('C:/Users/NYPO/Desktop/SWIFT_assign_04232026.xlsx')


# BIC country/bank keyword dictionary — same idea as earlier merge builder.
# Maps each BIC to keyword(s) that should appear in the Flex account label
# of the mirror account. Extend with your bank's own correspondent and
# intragroup BICs — the seed below covers major global correspondents.
BIC_KEYWORDS = {
    # Global correspondents
    'CITIUS33': ['CITI BANK, NEW YORK', 'CITI NEW YORK', 'CITIBANK NEW YORK',
                 'CITI NY', 'CITIBANK, NEW YORK', 'CITIBANK NY'],
    'CITIGB2L': ['CITI LONDON', 'CITIBANK LONDON', 'CITI BANK LONDON'],
    'SCBLUS33': ['STANDARD CHARTERED', 'STANCHART', 'SCB'],
    'BKTRUS33': ['BANKERS TRUST', 'DB TRUST', 'DEUTSCHE BANK TRUST',
                 'DB LOND', 'DB LONDON'],
    'GENODEFF': ['DZ BANK', 'DZBANK'],
    'NEDSZAJJ': ['NEDBANK'],
    'FIRNZAJJ': ['FIRST NATIONAL BANK OF SOUTH AFRICA', 'FNB', 'FIRST NATIONAL'],
    'COBADEFF': ['COMMERZBANK'],
    'BHFBDEFF': ['BHF'],
    'DEUTDEFF': ['DEUTSCHE'],
    'NATXFRPP': ['NATIXIS'],
    'UBSWCHZH': ['UBS'],
    'BOMLAEAD': ['MASHREQ', 'BANK OF MASHREQ'],
    'AFXMEGCA': ['ATTIJARI', 'ATTIJARIWAFA', 'EGYPT'],
    'DABADKKK': ['DANSKE', 'DAN DANSKE'],
    'BKCHCNBJ': ['BANK OF CHINA'],
    'ROYCCAT2': ['ROYAL BANK OF CANADA', 'RBC'],
}


def scan_spool():
    """Parse every .out file, return dict keyed by (account, currency)."""
    by_pair = defaultdict(lambda: {
        'count': 0, 'bics': Counter(), 'sample_file': None,
        'msg_types': Counter(), 'dates': Counter(),
    })
    skipped = []
    total = 0
    for f in sorted(FOLDER.glob('*.out')):
        total += 1
        raw = f.read_text(encoding='latin-1', errors='replace')
        mt = detect_message_type(raw)
        if mt not in ('940', '950'):
            skipped.append((f.name, f'message type {mt!r} (not MT940/950)'))
            continue
        try:
            meta = extract_swift_meta_raw(f)
        except Exception as e:
            skipped.append((f.name, f'parse error: {e}'))
            continue
        acc = meta.get('account')
        ccy = meta.get('currency')
        bic = meta.get('bic')
        if not acc:
            skipped.append((f.name, 'no :25: account'))
            continue
        if not ccy:
            skipped.append((f.name, 'no currency on :60F:'))
            continue
        key = (acc, ccy)
        d = by_pair[key]
        d['count'] += 1
        if bic:
            d['bics'][bic] += 1
        if d['sample_file'] is None:
            d['sample_file'] = f.name
        d['msg_types'][f'MT{mt}'] += 1
        ob = meta.get('opening_balance_date')
        if ob:
            d['dates'][ob] += 1
    return total, by_pair, skipped


def load_flex_index(conn):
    """Build lookup of Flex accounts by their label keywords."""
    rows = []
    for r in conn.execute(
        "SELECT id, flex_ac_no, currency, label, access_area "
        "FROM accounts WHERE active=1"
    ).fetchall():
        rows.append(dict(zip(
            ['id', 'flex_ac_no', 'currency', 'label', 'access_area'], r
        )))
    return rows


def candidates_for_pair(swift_acc, swift_ccy, sender_bic, flex_rows):
    """Return a list of candidate Flex accounts, ranked by confidence."""
    kws = BIC_KEYWORDS.get(sender_bic, [])
    matches = []
    for flex in flex_rows:
        label_up = (flex['label'] or '').upper()
        # Score: which keywords hit
        for kw in kws:
            if kw in label_up:
                ccy_match = (flex['currency'] == swift_ccy)
                # Only nostro-like areas are real candidates
                area_ok = flex['access_area'] in (
                    'NOSTRO', 'AFFILIATES', 'BANK OF GHANA', 'SUBSIDIARIES'
                )
                score = (2 if ccy_match else 0) + (1 if area_ok else 0) + (len(kw) / 100)
                matches.append({
                    **flex, 'matched_kw': kw, 'ccy_match': ccy_match,
                    'area_ok': area_ok, 'score': score,
                })
                break
    matches.sort(key=lambda m: -m['score'])
    # De-dup by flex_ac_no
    seen, unique = set(), []
    for m in matches:
        if m['flex_ac_no'] in seen:
            continue
        seen.add(m['flex_ac_no'])
        unique.append(m)
    return unique[:5]


def main():
    print(f"Scanning {FOLDER} ...")
    total, by_pair, skipped = scan_spool()
    print(f"  Files: {total}")
    print(f"  Parsed statement pairs (account, currency): {len(by_pair)}")
    print(f"  Skipped (non-940/950 or parse err): {len(skipped)}")

    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    flex_rows = load_flex_index(conn)
    print(f"  Flex accounts in registry: {len(flex_rows)}")

    # Build candidate list per pair
    matched_pairs = []        # pairs with at least one candidate (to assign)
    unmatched_pairs = []      # pairs with no candidate (probably not needed)
    assign_rows = []          # flattened: pair + candidate

    for (acc, ccy), d in by_pair.items():
        bic = d['bics'].most_common(1)[0][0] if d['bics'] else None
        cands = candidates_for_pair(acc, ccy, bic, flex_rows)

        if not cands:
            unmatched_pairs.append({
                'swift_account': acc, 'currency': ccy, 'bic': bic,
                'count': d['count'], 'sample_file': d['sample_file'],
                'msg_types': ','.join(sorted(d['msg_types'])),
            })
            continue

        matched_pairs.append((acc, ccy, bic, d, cands))
        for i, c in enumerate(cands):
            conf = ('HIGH' if (i == 0 and c['ccy_match'] and c['area_ok']) else
                    'MED'  if (c['ccy_match']) else 'LOW')
            tick = 'Y' if conf == 'HIGH' else ''
            assign_rows.append({
                'swift_account': acc, 'currency': ccy, 'bic': bic or '',
                'files_today': d['count'],
                'sample_file': d['sample_file'],
                'confirm': tick, 'confidence': conf,
                'flex_ac_no': c['flex_ac_no'],
                'flex_label': c['label'],
                'flex_access_area': c['access_area'],
                'flex_currency': c['currency'],
                'matched_kw': c['matched_kw'],
                'ccy_match': c['ccy_match'],
            })

    print(f"  Pairs with at least one Flex candidate: {len(matched_pairs)}")
    print(f"  Pairs with no match (→ probably not needed): {len(unmatched_pairs)}")

    # ---- Build workbook ----
    wb = Workbook()
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='305496')
    hi = PatternFill('solid', fgColor='C6EFCE')
    md = PatternFill('solid', fgColor='FFEB9C')
    lo = PatternFill('solid', fgColor='FFC7CE')
    title_font = Font(bold=True, size=14)

    # Summary tab
    s = wb.active
    s.title = 'Summary'
    s['A1'] = f'SWIFT → Flex account assignment review'
    s['A1'].font = title_font
    s.append(['Folder', str(FOLDER)])
    s.append(['Total .out files', total])
    s.append(['Parsed (acct, ccy) pairs', len(by_pair)])
    s.append(['  with Flex candidate', len(matched_pairs)])
    s.append(['  no Flex match (probably noise)', len(unmatched_pairs)])
    s.append(['Non-940/950 or parse errors', len(skipped)])
    s.append([])
    s.append(['Pre-ticked HIGH-confidence assignments',
              sum(1 for r in assign_rows if r['confirm'] == 'Y')])
    s.append([])
    s.append(['Next step'])
    s.append(['1. Open "To assign" tab.'])
    s.append(['2. HIGH-confidence rows (GREEN) are pre-ticked "Y".'])
    s.append(['3. Review MED (yellow) / LOW (red) — set Y on the correct Flex.'])
    s.append(['4. Only ONE row per SWIFT account should be "Y".'])
    s.append(['5. Review "Not needed" tab — confirm these SWIFT accts are not yours.'])
    s.append(['6. Save and send back; I will UPDATE swift_account in the DB for Y rows.'])

    # To assign
    a = wb.create_sheet('To assign')
    headers = [
        'SWIFT account', 'SWIFT ccy', 'Sender BIC', 'Files today', 'Sample file',
        'Confirm? (Y/N)', 'Confidence',
        'Flex ac_no', 'Flex label', 'Flex access area', 'Flex ccy',
        'Matched keyword',
    ]
    a.append(headers)
    for cell in a[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
    a.freeze_panes = 'A2'

    for r in assign_rows:
        a.append([
            r['swift_account'], r['currency'], r['bic'], r['files_today'],
            r['sample_file'], r['confirm'], r['confidence'],
            r['flex_ac_no'], r['flex_label'], r['flex_access_area'],
            r['flex_currency'], r['matched_kw'],
        ])

    dv = DataValidation(type='list', formula1='"Y,N"', allow_blank=True)
    dv.add(f'F2:F{max(a.max_row, 2)}')
    a.add_data_validation(dv)

    # Row colours by confidence
    for row_idx in range(2, a.max_row + 1):
        conf = a.cell(row=row_idx, column=7).value
        fill = {'HIGH': hi, 'MED': md, 'LOW': lo}.get(conf)
        if fill:
            for col in range(1, a.max_column + 1):
                a.cell(row=row_idx, column=col).fill = fill

    # Not needed
    n = wb.create_sheet('Not needed')
    n.append(['SWIFT account', 'Currency', 'Sender BIC', 'Files today',
              'Sample file', 'Msg types', 'Action (Y=register anyway, N=ignore)'])
    for cell in n[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
    n.freeze_panes = 'A2'
    for r in unmatched_pairs:
        n.append([r['swift_account'], r['currency'], r['bic'], r['count'],
                  r['sample_file'], r['msg_types'], 'N'])
    dv2 = DataValidation(type='list', formula1='"Y,N"', allow_blank=True)
    dv2.add(f'G2:G{max(n.max_row, 2)}')
    n.add_data_validation(dv2)

    # Skipped files
    k = wb.create_sheet('Skipped')
    k.append(['File', 'Reason'])
    for cell in k[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
    k.freeze_panes = 'A2'
    for f, reason in skipped:
        k.append([f, reason])

    # Column widths
    for ws in (s, a, n, k):
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = 12
            for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
                v = row[0]
                if v is None:
                    continue
                max_len = max(max_len, min(55, len(str(v)) + 2))
            ws.column_dimensions[letter].width = max_len

    wb.save(OUT)
    print(f"\nWrote {OUT}")

    # Console samples
    print("\nSample pre-ticked HIGH-confidence assignments:")
    for r in [r for r in assign_rows if r['confirm'] == 'Y'][:15]:
        print(f"  {r['swift_account']:22s} {r['currency']:3s} {r['bic']:10s}"
              f"  -> {r['flex_ac_no']:20s} ({r['flex_label'][:45]})")

    print(f"\nSample 'not needed' entries (top 10 by volume):")
    unmatched_pairs.sort(key=lambda r: -r['count'])
    for r in unmatched_pairs[:10]:
        print(f"  {r['swift_account']:22s} {r['currency']:3s} {r['bic']:10s}"
              f"  {r['count']:3d} files  ({r['msg_types']})")


if __name__ == '__main__':
    main()
