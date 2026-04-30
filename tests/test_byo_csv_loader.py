"""Tests for the BYO CSV loader.

What we're pinning:
    - Round-trip: a profile + sample CSV produces the canonical Flex
      txn shape the recon engine consumes.
    - Sign conventions: positive_credit, cr_dr_column, separate_column
      all resolve to the right CR/DR with absolute amounts.
    - Date / amount parsing edge cases: European decimals, accountancy
      negatives, currency symbols, multiple date formats.
    - Profile validation rejects missing required columns and bad
      delimiter / sign combos before the loader runs.
    - autoguess_mapping returns sane first-pass guesses for typical
      bank-extract column names.
    - Per-row failures land in LoadResult.errors without aborting the
     run.

These tests bypass FastAPI entirely — they exercise byo_csv_loader as a
pure-Python module so they're fast and don't need a DB.
"""

import pytest

from byo_csv_loader import (
    CsvProfile, load_csv, autoguess_mapping,
    REQUIRED_COLUMNS, VALID_DELIMITERS, VALID_SIGN_CONVENTIONS,
)


def _profile(**overrides) -> CsvProfile:
    """Default profile that satisfies validation; tests override fields."""
    base = dict(
        name='test',
        delimiter=',',
        header_row=1,
        skip_rows=0,
        date_format='%Y-%m-%d',
        currency='USD',
        column_map={
            'amount': 'Amount',
            'value_date': 'Date',
            'ref': 'Ref',
            'narration': 'Memo',
            'type': None,
            'currency': None,
            'ac_no': None,
            'ac_branch': None,
            'booking_date': None,
        },
        sign_convention='positive_credit',
        sign_column=None,
    )
    base.update(overrides)
    return CsvProfile(**base)


CSV_BASIC = b"""Date,Ref,Memo,Amount
2026-04-01,REF-1,Salary deposit,1500.00
2026-04-02,REF-2,Card payment,-42.50
2026-04-03,REF-3,Wire to vendor,-750.00
"""


def test_basic_load_emits_canonical_flex_shape():
    result = load_csv(CSV_BASIC, _profile())
    assert len(result.txns) == 3
    assert result.errors == []

    t = result.txns[0]
    # Match the keys the recon engine expects (see _flex test fixture
    # in test_recon_engine.py).
    expected_keys = {'_source', '_row_number', '_used', 'trn_ref', 'ac_branch',
                     'ac_no', 'booking_date', 'value_date', 'type', 'narration',
                     'amount', 'ccy', 'module', 'external_ref', 'user_id',
                     # _extra carries unmapped source columns by header
                     # name so the cards seam can read masked-PAN, MCC,
                     # terminal_id etc. without losing the canonical
                     # contract.
                     '_extra'}
    assert set(t.keys()) == expected_keys
    assert isinstance(t['_extra'], dict)
    assert t['_source'] == 'flex'
    assert t['value_date'] == 20260401
    assert t['amount'] == 1500.00
    assert t['type'] == 'CR'                  # positive → credit
    assert t['ccy'] == 'USD'

    assert result.txns[1]['type'] == 'DR'     # negative → debit
    assert result.txns[1]['amount'] == 42.50  # absolute


def test_european_decimal_format_parsed():
    """A bank that ships '1.234,56' for 1234.56 — common in EU."""
    csv = b"""Date,Ref,Memo,Amount
2026-04-01,X,salary,"1.234,56"
"""
    result = load_csv(csv, _profile())
    assert len(result.txns) == 1
    assert result.txns[0]['amount'] == 1234.56


def test_accountancy_negatives_resolve_to_dr():
    """'(123.45)' is the accountancy convention for -123.45."""
    csv = b"""Date,Ref,Memo,Amount
2026-04-01,X,fee,(125.00)
"""
    result = load_csv(csv, _profile())
    assert result.txns[0]['type'] == 'DR'
    assert result.txns[0]['amount'] == 125.00


def test_currency_symbols_stripped():
    csv = b"""Date,Ref,Memo,Amount
2026-04-01,X,salary,$1500.00
2026-04-02,Y,fee,GHS 200.50
"""
    result = load_csv(csv, _profile())
    assert result.txns[0]['amount'] == 1500.00
    assert result.txns[1]['amount'] == 200.50


def test_separate_sign_column():
    """A bank with positive amounts and a 'CR'/'DR' column."""
    csv = b"""Date,Ref,Memo,Amount,Side
2026-04-01,X,salary,1500.00,CR
2026-04-02,Y,fee,42.50,DR
"""
    p = _profile(sign_convention='cr_dr_column', sign_column='Side',
                 column_map={**_profile().column_map, 'type': 'Side'})
    result = load_csv(csv, p)
    assert result.txns[0]['type'] == 'CR'
    assert result.txns[1]['type'] == 'DR'
    assert all(t['amount'] > 0 for t in result.txns)


def test_paid_in_withdrawn_two_column_amount():
    """M-Pesa / Telcel Cash shape: each row populates either Paid In
    (CR) or Withdrawn (DR), never both. The loader resolves direction
    from whichever column is non-empty without needing a sign column
    to spell it out."""
    csv = b"""Date,Ref,Memo,Paid In,Withdrawn
2026-04-01,A,salary,1500.00,
2026-04-02,B,fee,,42.50
"""
    p = _profile(
        sign_convention='paid_in_withdrawn',
        sign_column='Withdrawn',
        column_map={**_profile().column_map, 'amount': 'Paid In'},
    )
    result = load_csv(csv, p)
    assert len(result.txns) == 2
    assert result.txns[0]['type'] == 'CR'
    assert result.txns[0]['amount'] == 1500.00
    assert result.txns[1]['type'] == 'DR'
    assert result.txns[1]['amount'] == 42.50


def test_paid_in_withdrawn_rejects_row_with_both_empty():
    """A row where both columns are empty is operator error — the loader
    surfaces it via LoadResult.errors so the operator can audit, but
    keeps the rest of the run."""
    csv = b"""Date,Ref,Memo,Paid In,Withdrawn
2026-04-01,A,empty,,
2026-04-02,B,real,1500.00,
"""
    p = _profile(
        sign_convention='paid_in_withdrawn',
        sign_column='Withdrawn',
        column_map={**_profile().column_map, 'amount': 'Paid In'},
    )
    result = load_csv(csv, p)
    assert len(result.txns) == 1
    assert result.txns[0]['amount'] == 1500.00
    assert len(result.errors) == 1
    assert 'paid-in and withdrawn' in result.errors[0][1].lower()


def test_xlsx_input_is_auto_detected_and_parsed():
    """Operators can upload xlsx directly — the loader detects the
    zip-header magic and routes to openpyxl. Telcel Cash and many
    European bank statements ship as xlsx."""
    import io
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(['Date', 'Ref', 'Memo', 'Amount'])
    ws.append(['2026-04-01', 'X1', 'salary', 1500.00])
    ws.append(['2026-04-02', 'X2', 'rent', -800.00])
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    result = load_csv(content, _profile())
    assert len(result.txns) == 2
    assert result.errors == []
    assert result.txns[0]['amount'] == 1500.00
    assert result.txns[0]['type'] == 'CR'
    assert result.txns[1]['type'] == 'DR'
    assert result.txns[1]['amount'] == 800.00


def test_xlsx_datetime_cells_serialize_for_date_parsing():
    """When the spreadsheet stores a real datetime cell (not a string),
    the loader must format it so the profile's date_format can read
    it back. Excel-side type drift is a common BYO upload failure."""
    import io
    from datetime import datetime as _dt
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(['Date', 'Ref', 'Memo', 'Amount'])
    ws.append([_dt(2026, 4, 15, 10, 30, 0), 'D1', 'wire', 250.00])
    buf = io.BytesIO()
    wb.save(buf)

    profile = _profile(date_format='%Y-%m-%d %H:%M:%S')
    result = load_csv(buf.getvalue(), profile)
    assert len(result.txns) == 1
    assert result.txns[0]['value_date'] == 20260415


def test_autoguess_mapping_handles_xlsx_input():
    """The wizard preview also takes xlsx — operators don't have to
    save-as-CSV before clicking Preview."""
    import io
    from openpyxl import Workbook
    from byo_csv_loader import autoguess_mapping
    wb = Workbook()
    ws = wb.active
    ws.append(['Date', 'Reference', 'Amount', 'Currency'])
    ws.append(['2026-04-01', 'X1', 100.00, 'USD'])
    ws.append(['2026-04-02', 'X2', 50.00, 'USD'])
    buf = io.BytesIO()
    wb.save(buf)
    out = autoguess_mapping(buf.getvalue())
    assert set(out['columns']) == {'Date', 'Reference', 'Amount', 'Currency'}
    assert out['guess']['amount'] == 'Amount'
    assert out['guess']['value_date'] == 'Date'
    assert out['guess']['ref'] == 'Reference'
    assert out['guess']['currency'] == 'Currency'
    assert len(out['sample']) == 2


def test_xlsx_delimiter_field_is_ignored():
    """xlsx ingest doesn't use the delimiter field at all — it can stay
    at its profile default ',' even for tab-style spreadsheet exports.
    Verifies the format-detection short-circuits before the CSV-parse
    path runs."""
    import io
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(['Date', 'Ref', 'Memo', 'Amount'])
    ws.append(['2026-04-01', 'A', 'x', 100.00])
    buf = io.BytesIO()
    wb.save(buf)
    # Delimiter doesn't matter — pipe is a valid delimiter, not the
    # right one for any reasonable CSV, but xlsx ignores it.
    result = load_csv(buf.getvalue(), _profile(delimiter='|'))
    assert len(result.txns) == 1


def test_paid_in_withdrawn_validation_requires_sign_column():
    """paid_in_withdrawn needs the sign_column field set to the
    Withdrawn column name. Profile-level validation must catch that
    mistake before the first row is read."""
    with pytest.raises(ValueError, match="sign_column"):
        load_csv(b"Date,Ref,Memo,Paid In,Withdrawn\n",
                 _profile(sign_convention='paid_in_withdrawn',
                          sign_column=None))


def test_skip_rows_strips_preamble():
    """Banks often prepend a header block (account info, page numbers)
    above the column-name row. skip_rows must skip past it."""
    csv = b"""Bank: Acme
Account: 12345
Period: 2026-04
Date,Ref,Memo,Amount
2026-04-01,X,salary,1500.00
"""
    p = _profile(skip_rows=3)
    result = load_csv(csv, p)
    assert len(result.txns) == 1
    assert result.txns[0]['amount'] == 1500.00


def test_alternative_date_format():
    csv = b"""Date,Ref,Memo,Amount
01/04/2026,X,salary,1500.00
"""
    p = _profile(date_format='%d/%m/%Y')
    result = load_csv(csv, p)
    assert result.txns[0]['value_date'] == 20260401


def test_blank_rows_silently_skipped():
    csv = b"""Date,Ref,Memo,Amount
2026-04-01,X,salary,1500.00
,,,
2026-04-02,Y,fee,-50.00
"""
    result = load_csv(csv, _profile())
    assert len(result.txns) == 2
    assert result.errors == []


def test_unparseable_date_lands_in_errors_not_txns():
    csv = b"""Date,Ref,Memo,Amount
2026-04-01,X,salary,1500.00
not-a-date,Y,fee,-50.00
"""
    result = load_csv(csv, _profile())
    assert len(result.txns) == 1                  # the good row got through
    assert len(result.errors) == 1
    assert 'date' in result.errors[0][1].lower()


def test_currency_column_overrides_profile_default():
    csv = b"""Date,Ref,Memo,Amount,Ccy
2026-04-01,X,salary,1500.00,EUR
"""
    p = _profile(currency=None,
                 column_map={**_profile().column_map, 'currency': 'Ccy'})
    result = load_csv(p_csv := csv, p)
    assert result.txns[0]['ccy'] == 'EUR'


def test_profile_validation_rejects_missing_required_column():
    p = _profile(column_map={**_profile().column_map, 'amount': None})
    with pytest.raises(ValueError, match="amount"):
        load_csv(CSV_BASIC, p)


def test_profile_validation_rejects_bad_delimiter():
    with pytest.raises(ValueError):
        load_csv(CSV_BASIC, _profile(delimiter='~'))


def test_profile_validation_rejects_sign_convention_without_column():
    with pytest.raises(ValueError, match="sign_column"):
        load_csv(CSV_BASIC, _profile(sign_convention='cr_dr_column',
                                       sign_column=None))


def test_currency_optional_at_loader_level():
    """Currency is no longer required at the loader. When profile.currency
    is None AND there's no currency column, the loader emits empty ccy
    and lets the ingest layer fall back to the matched account's currency.
    Previously this test asserted ValueError; the post-pilot UX fix
    relaxed it. Pinned to make sure we don't reintroduce the gate."""
    result = load_csv(CSV_BASIC, _profile(currency=None))
    assert len(result.txns) == 3
    assert all(t['ccy'] == '' for t in result.txns)


# ---------------------------------------------------------------------------
# autoguess_mapping
# ---------------------------------------------------------------------------

def test_autoguess_picks_obvious_columns():
    csv = b"""Tran Date,Reference,Description,Amount,Currency
2026-04-01,REF1,Salary,1500.00,USD
"""
    out = autoguess_mapping(csv)
    g = out['guess']
    assert g['value_date'] == 'Tran Date'
    assert g['ref'] == 'Reference'
    assert g['narration'] == 'Description'
    assert g['amount'] == 'Amount'
    assert g['currency'] == 'Currency'


def test_autoguess_handles_skip_rows_and_header():
    csv = b"""Bank: Acme
Period: 2026-04
Posting Date,Ref,Memo,Posting Amount
2026-04-01,X,salary,1500.00
"""
    out = autoguess_mapping(csv, skip_rows=2, header_row=1)
    assert out['columns'] == ['Posting Date', 'Ref', 'Memo', 'Posting Amount']
    assert out['guess']['value_date'] == 'Posting Date'
    assert out['guess']['amount'] == 'Posting Amount'


def test_autoguess_returns_empty_on_short_file():
    out = autoguess_mapping(b'', delimiter=',', skip_rows=0, header_row=1)
    assert out['columns'] == []
    assert out['guess'] == {}


def test_alternate_delimiters():
    csv = b"Date;Ref;Memo;Amount\n2026-04-01;X;salary;1500.00\n"
    p = _profile(delimiter=';')
    result = load_csv(csv, p)
    assert len(result.txns) == 1
    assert result.txns[0]['amount'] == 1500.00


# ---------------------------------------------------------------------------
# Engine compatibility — the loader output should be runnable through
# propose_candidates without modification.
# ---------------------------------------------------------------------------

def test_loader_output_is_engine_compatible():
    """A pair of synthetic SWIFT credits + the BYO Flex output should
    classify correctly through propose_candidates. This is the
    integration boundary that matters."""
    from recon_engine import propose_candidates

    csv = b"""Date,Ref,Memo,Amount
2026-04-01,MTB1234567,salary MTB1234567,-1500.00
"""
    flex = load_csv(csv, _profile()).txns
    swift = [{
        '_source': 'swift', '_row_number': 1, '_used': False,
        'value_date': 20260401, 'amount': 1500.00, 'sign': 'C',
        'origin': 'Their', 'type': 'Other', 'status': 'Unmatched',
        'book_date': 20260401, 'our_ref': 'MTB1234567', 'their_ref': '',
        'booking_text_1': '', 'booking_text_2': '',
    }]
    candidates = propose_candidates(swift, flex)
    assert len(candidates) == 1
    # Tier 1: ref >= 6 chars hits in narration + exact amount + sign mirror.
    assert candidates[0].tier == 1
