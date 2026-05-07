"""Phase 1 tests — proof file parser.

Pins:
  * Header detection (first cell == 'Account')
  * Tolerates the blank row Ecobank's exporter inserts
  * Tolerates the trailing footer with a running total
  * Skips rows whose S cell is missing
  * Rejects files that don't look like a proof
  * Rejects files missing required columns
  * compute_seed_balance signs CR positive, DR negative
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from proof_loader import compute_seed_balance, load_proof


PROOF_HEADER = [
    'Account', 'Value date', 'Curr.', 'Amount', 'S',
    '', 'Origin', 'Type', 'Status', 'age', 'Book. date',
    'Our reference 1', 'Their reference 1',
    'Booking text 1', 'Booking text 2',
]


def _proof_row(value_date, amount, sign, ref, narration='wallet credit'):
    return [
        'BANK TO WALLET GL ACCOUNT', value_date, 'GHS', amount, sign,
        amount, 'Our', 'Other', 'Open', 0,
        datetime(2026, 4, 29) if isinstance(value_date, int) else value_date,
        ref, None, narration[:40], narration,
    ]


def _write_proof(path: Path, rows: list[list],
                 *, blank_after_header: bool = True,
                 footer_total: float | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(PROOF_HEADER)
    if blank_after_header:
        ws.append([None] * len(PROOF_HEADER))
    for r in rows:
        ws.append(r)
    if footer_total is not None:
        # Real Ecobank proofs append a footer with the running total in the
        # subtotal column (index 5) and every other cell None — we should
        # silently skip it because S is None.
        ws.append([None, None, None, None, None, footer_total,
                   None, None, None, None, None, None, None, None, None])
    wb.save(path)


# ---------------------------------------------------------------------------
# Happy paths
# ---------------------------------------------------------------------------

def test_loads_basic_credit_rows(tmp_path):
    p = tmp_path / 'proof.xlsx'
    _write_proof(p, [
        _proof_row(20260429, 100.0, 'C', 'REF1'),
        _proof_row(20260429,  25.5, 'C', 'REF2'),
    ])
    txns = load_proof(p)
    assert len(txns) == 2
    t = txns[0]
    assert t['trn_ref'] == 'REF1'
    assert t['type'] == 'CR'           # 'C' in proof maps to 'CR' in canonical shape
    assert t['amount'] == 100.0
    assert t['ccy'] == 'GHS'
    assert t['value_date'] == 20260429
    assert t['_row_number'] == 1


def test_loads_mixed_credit_and_debit(tmp_path):
    """The S column drives the sign mapping: C → CR (positive contribution
    to balance), D → DR (negative). amount stays absolute."""
    p = tmp_path / 'proof.xlsx'
    _write_proof(p, [
        _proof_row(20260429, 100.0, 'C', 'CR1'),
        _proof_row(20260429,  30.0, 'D', 'DR1'),
    ])
    txns = load_proof(p)
    assert len(txns) == 2
    assert txns[0]['type'] == 'CR' and txns[0]['amount'] == 100.0
    assert txns[1]['type'] == 'DR' and txns[1]['amount'] == 30.0


def test_tolerates_footer_running_total(tmp_path):
    """Real Ecobank proofs end with a footer row whose only populated cell
    is the running total in the subtotal column. The loader must skip it
    silently — it's not a transaction."""
    p = tmp_path / 'proof.xlsx'
    _write_proof(p, [
        _proof_row(20260429, 100.0, 'C', 'REF1'),
        _proof_row(20260429,  30.0, 'D', 'REF2'),
    ], footer_total=70.0)
    txns = load_proof(p)
    assert len(txns) == 2  # footer skipped


def test_tolerates_blank_row_after_header(tmp_path):
    """Ecobank's exporter inserts a blank row between the header and the
    first transaction. The loader must skip it without complaining."""
    p = tmp_path / 'proof.xlsx'
    _write_proof(p, [_proof_row(20260429, 1.0, 'C', 'X')])  # default writes blank row
    assert len(load_proof(p)) == 1


def test_no_blank_row_after_header_still_works(tmp_path):
    """Some banks' exporters skip the blank row. Both layouts should parse."""
    p = tmp_path / 'proof.xlsx'
    _write_proof(p, [_proof_row(20260429, 1.0, 'C', 'X')], blank_after_header=False)
    assert len(load_proof(p)) == 1


def test_skips_rows_with_missing_sign(tmp_path):
    """Some proof exports interleave subtotal rows whose S cell is None.
    Those are not transactions; skip them silently."""
    p = tmp_path / 'proof.xlsx'
    rows = [
        _proof_row(20260429, 100.0, 'C', 'REAL1'),
        # Garbage row with no S
        ['BANK TO WALLET GL ACCOUNT', 20260429, 'GHS', None, None,
         50.0, None, None, None, None, None, None, None, None, None],
        _proof_row(20260429, 200.0, 'C', 'REAL2'),
    ]
    _write_proof(p, rows)
    txns = load_proof(p)
    assert len(txns) == 2
    assert {t['trn_ref'] for t in txns} == {'REAL1', 'REAL2'}


def test_skips_row_with_missing_ref(tmp_path):
    """A row with an S but no Our reference 1 isn't usable — skip it
    rather than hand the engine a row with empty trn_ref."""
    p = tmp_path / 'proof.xlsx'
    rows = [
        _proof_row(20260429, 100.0, 'C', 'KEEP'),
        _proof_row(20260429,  10.0, 'C', None),  # ref missing
    ]
    _write_proof(p, rows)
    txns = load_proof(p)
    assert len(txns) == 1
    assert txns[0]['trn_ref'] == 'KEEP'


def test_handles_datetime_value_dates(tmp_path):
    """Value date column comes through as Python datetime in some
    Flexcube formatter outputs. _to_int_date should normalise to YYYYMMDD."""
    p = tmp_path / 'proof.xlsx'
    _write_proof(p, [
        _proof_row(datetime(2026, 4, 29), 50.0, 'C', 'REF1'),
    ])
    txns = load_proof(p)
    assert txns[0]['value_date'] == 20260429


# ---------------------------------------------------------------------------
# Error cases
# ---------------------------------------------------------------------------

def test_rejects_file_with_no_account_header(tmp_path):
    p = tmp_path / 'random.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.append(['some', 'random', 'spreadsheet'])
    ws.append([1, 2, 3])
    wb.save(p)
    with pytest.raises(ValueError, match="doesn't look like a proof file"):
        load_proof(p)


def test_rejects_file_missing_required_columns(tmp_path):
    """Header says 'Account' but lacks the columns the loader needs.
    Must give the operator a specific list of what's missing."""
    p = tmp_path / 'malformed.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.append(['Account', 'Some', 'Other', 'Cols'])
    ws.append(['BTW', 1, 2, 3])
    wb.save(p)
    with pytest.raises(ValueError, match='missing required proof columns'):
        load_proof(p)


def test_skips_balances_sidecar_sheet_and_finds_data_sheet(tmp_path):
    """If a proof xlsx has both a 'balances' sidecar sheet and a real
    data sheet (operator could have added one for their own use), we
    must skip 'balances' and find the proof header on another sheet."""
    p = tmp_path / 'with_sidecar.xlsx'
    wb = Workbook()
    # First sheet is the sidecar — should be skipped.
    sidecar = wb.active
    sidecar.title = 'balances'
    sidecar.append(['as_of_date', 'opening_balance', 'closing_balance', 'currency'])
    sidecar.append(['2026-04-30', 0.0, -56704939.75, 'GHS'])
    # Second sheet is the real proof.
    data = wb.create_sheet('Sheet1')
    data.append(PROOF_HEADER)
    data.append([None] * len(PROOF_HEADER))
    data.append(_proof_row(20260429, 100.0, 'C', 'REF1'))
    wb.save(p)

    txns = load_proof(p)
    assert len(txns) == 1
    assert txns[0]['trn_ref'] == 'REF1'


# ---------------------------------------------------------------------------
# compute_seed_balance
# ---------------------------------------------------------------------------

def test_compute_seed_balance_signed_sum():
    """CR contributes positive, DR contributes negative. Closing is the
    rounded-2dp net; max_value_date is the high-water mark."""
    txns = [
        {'amount': 100.0, 'type': 'CR', 'value_date': 20260101},
        {'amount':  25.0, 'type': 'DR', 'value_date': 20260131},
        {'amount':   3.5, 'type': 'CR', 'value_date': 20260115},
    ]
    closing, max_date = compute_seed_balance(txns)
    assert closing == 78.50  # 100 - 25 + 3.50
    assert max_date == 20260131


def test_compute_seed_balance_handles_negative_net():
    """When DR > CR (e.g. the BTW proof: settlement debits exceed wallet
    credits), the closing is negative — should round to 2dp without
    floating-point drift."""
    txns = [
        {'amount':  43_433_370.25, 'type': 'CR', 'value_date': 20260429},
        {'amount': 100_138_310.00, 'type': 'DR', 'value_date': 20260427},
    ]
    closing, _ = compute_seed_balance(txns)
    assert closing == -56_704_939.75


def test_compute_seed_balance_empty_proof():
    closing, max_date = compute_seed_balance([])
    assert closing == 0.0
    assert max_date == 0
