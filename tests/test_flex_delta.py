"""Phase 3 tests — daily Flex-only delta with continuity check.

Pins:
  * Happy path: anchor advances, session_kind='flex_delta', open items seeded
  * Continuity match advances anchor, no error
  * Continuity break (any delta > tolerance) → ContinuityBreakError, anchor stays
  * Per-account tolerance override applies (set 1.00, file off by 0.50, accepts)
  * NotSeededError when account has no anchor
  * Day-1 → Day-2 chain works (closing N becomes opening N+1)
  * Force-accept bypasses continuity check, advances anchor anyway, audit-logged
  * Carry-forward: proof CR clears against today's DR with the same ref
"""

from __future__ import annotations

import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import Workbook


@pytest.fixture
def fresh_db(tmp_path, monkeypatch):
    db = tmp_path / 'test.db'
    monkeypatch.setenv('KILTER_DB_PATH', str(db))
    import db as db_mod
    monkeypatch.setattr(db_mod, 'DB_PATH', db, raising=False)
    db_mod.init_db()
    return db


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

PROOF_HEADER = [
    'Account', 'Value date', 'Curr.', 'Amount', 'S',
    '', 'Origin', 'Type', 'Status', 'age', 'Book. date',
    'Our reference 1', 'Their reference 1',
    'Booking text 1', 'Booking text 2',
]


def _proof_row(value_date, amount, sign, ref, narration='wallet credit'):
    return [
        'BANK TO WALLET GL ACCOUNT', value_date, 'GHS', amount, sign,
        amount, 'Our', 'Other', 'Open', 0,
        datetime(2026, 4, 29) if isinstance(value_date, int) else value_date,
        ref, None, narration[:40], narration,
    ]


def _write_proof(path: Path, rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(PROOF_HEADER)
    ws.append([None] * len(PROOF_HEADER))
    for r in rows:
        ws.append(r)
    wb.save(path)


def _write_delta(path: Path, rows: list[tuple], *,
                 ac_no='1441000601589', currency='GHS',
                 opening_balance: float, closing_balance: float,
                 as_of: str = '2026-05-01',
                 with_balance_sheet: bool = True) -> None:
    """Build a minimal Flexcube acc_entries xlsx with optional balances sheet.
    rows: list of (value_date, amount, type, ref, narration)."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'acc_entries'
    ws.append(['TRN_REF_NO', 'BOOKING_DATE', 'TYPE', 'TXN_NARRATIONS',
               'VALUE_DATE', 'LCY_AMOUNT', 'AC_BRANCH', 'AC_NO', 'ACCT_CCY',
               'MODULE', 'EXTERNAL_REF_NO', 'USER_ID'])
    for value_date, amount, type_, ref, narration in rows:
        # value_date might be Python int (YYYYMMDD) or datetime
        bd = value_date if isinstance(value_date, datetime) else datetime(2026, 5, 1)
        vd = value_date if isinstance(value_date, datetime) else datetime(
            value_date // 10000, (value_date // 100) % 100, value_date % 100)
        ws.append([ref, bd, type_, narration, vd,
                   amount, 'EGH', ac_no, currency, 'IF', '', 'AUTO'])
    if with_balance_sheet:
        bs = wb.create_sheet('balances')
        bs.append(['as_of_date', 'opening_balance', 'closing_balance', 'currency'])
        bs.append([as_of, opening_balance, closing_balance, currency])
    wb.save(path)


def _seed_one_sided_account(db_path: Path, *, label='BTW',
                              flex_ac_no='1441000601589',
                              currency='GHS') -> int:
    conn = sqlite3.connect(db_path)
    now = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
    conn.execute(
        "INSERT INTO accounts (label, swift_account, flex_ac_no, currency, "
        "active, created_at, created_by, access_area, account_recon_type) "
        "VALUES (?, '', ?, ?, 1, ?, 'system', 'MOBILE MONEY', 'one_sided')",
        (label, flex_ac_no, currency, now),
    )
    acct_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.commit(); conn.close()
    return acct_id


def _seed_proof(db_path: Path, acct_id: int, tmp_path: Path,
                rows: list[list]) -> dict:
    """Run the actual ingest_proof_seed so the account ends up properly
    anchored — fixture for continuity-tests below."""
    from ingest import ingest_proof_seed
    proof = tmp_path / f'seed_{acct_id}.xlsx'
    _write_proof(proof, rows)
    return ingest_proof_seed(proof, acct_id, user='ops')


# ---------------------------------------------------------------------------
# Continuity check
# ---------------------------------------------------------------------------

def test_continuity_match_advances_anchor(fresh_db, tmp_path):
    """Day 1 delta opens at the proof's closing balance — should ingest
    cleanly and bump the anchor to the delta's closing."""
    from ingest import ingest_flex_only

    acct_id = _seed_one_sided_account(fresh_db)
    # Seed: 1 CR of 100, anchor lands at +100.
    _seed_proof(fresh_db, acct_id, tmp_path,
                [_proof_row(20260429, 100.0, 'C', 'PROOF_CR_1')])

    # Day 1 delta: opens at 100 (matches anchor), 1 CR of 50, closes at 150.
    delta = tmp_path / 'day1.xlsx'
    _write_delta(delta, [
        (20260501, 50.0, 'CR', 'DELTA_CR_1', 'wallet credit'),
    ], opening_balance=100.0, closing_balance=150.0, as_of='2026-05-01')
    result = ingest_flex_only(delta, account_id=acct_id, user='ops')
    assert result.swift_rows == 0   # no DRs in this delta
    assert result.flex_rows == 1
    assert result.account_registered

    conn = sqlite3.connect(fresh_db); conn.row_factory = sqlite3.Row
    acct = dict(conn.execute(
        "SELECT * FROM accounts WHERE id=?", (acct_id,)).fetchone())
    sess = dict(conn.execute(
        "SELECT * FROM sessions WHERE id=?", (result.session_id,)).fetchone())
    conn.close()
    assert acct['last_closing_balance'] == 150.0
    assert acct['last_closing_date'] == 20260501
    assert sess['session_kind'] == 'flex_delta'
    assert sess['flex_opening_balance'] == 100.0
    assert sess['flex_closing_balance'] == 150.0


def test_continuity_break_rejects_and_keeps_anchor(fresh_db, tmp_path):
    """Delta opening doesn't match anchor → ContinuityBreakError. Anchor
    stays at the prior value; no session is created (rollback semantics
    are tested separately, but we must not advance the chain on a break)."""
    from ingest import ContinuityBreakError, ingest_flex_only

    acct_id = _seed_one_sided_account(fresh_db)
    _seed_proof(fresh_db, acct_id, tmp_path,
                [_proof_row(20260429, 100.0, 'C', 'PROOF1')])

    delta = tmp_path / 'broken.xlsx'
    _write_delta(delta, [
        (20260501, 25.0, 'CR', 'X', 'note'),
    ], opening_balance=99.50, closing_balance=124.50)  # off by 0.50
    with pytest.raises(ContinuityBreakError) as exc_info:
        ingest_flex_only(delta, account_id=acct_id, user='ops')
    err = exc_info.value
    assert err.expected == 100.0
    assert err.actual == 99.50
    assert err.delta == -0.50

    # Anchor unchanged.
    conn = sqlite3.connect(fresh_db); conn.row_factory = sqlite3.Row
    acct = dict(conn.execute(
        "SELECT last_closing_balance FROM accounts WHERE id=?",
        (acct_id,)).fetchone())
    conn.close()
    assert acct['last_closing_balance'] == 100.0


def test_continuity_within_default_tolerance_passes(fresh_db, tmp_path):
    """Within 0.01 default tolerance the file is accepted. Pinning
    that the equality check uses abs(...) <= tol, not strict equality."""
    from ingest import ingest_flex_only

    acct_id = _seed_one_sided_account(fresh_db)
    _seed_proof(fresh_db, acct_id, tmp_path,
                [_proof_row(20260429, 100.0, 'C', 'PROOF1')])

    delta = tmp_path / 'fuzzy.xlsx'
    # Off by 0.005 — under 0.01 tolerance, should pass.
    _write_delta(delta, [
        (20260501, 5.0, 'CR', 'X', 'note'),
    ], opening_balance=100.005, closing_balance=105.005)
    result = ingest_flex_only(delta, account_id=acct_id, user='ops')
    conn = sqlite3.connect(fresh_db); conn.row_factory = sqlite3.Row
    acct = dict(conn.execute(
        "SELECT last_closing_balance FROM accounts WHERE id=?",
        (acct_id,)).fetchone())
    conn.close()
    # Anchor took the file's stated closing exactly.
    assert acct['last_closing_balance'] == 105.005


def test_per_account_tolerance_override(fresh_db, tmp_path):
    """Operator sets a 1.00 tolerance on an account known to have noisy
    sub-cent FX rounding. A 0.50 delta should now pass that wouldn't
    pass under the 0.01 default."""
    from ingest import ingest_flex_only

    acct_id = _seed_one_sided_account(fresh_db)
    _seed_proof(fresh_db, acct_id, tmp_path,
                [_proof_row(20260429, 100.0, 'C', 'PROOF1')])

    # Bump tolerance to 1.00 on this account.
    conn = sqlite3.connect(fresh_db)
    now = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
    conn.execute(
        "INSERT INTO tolerance_rules (account_id, amount_tol_abs, amount_tol_pct, "
        "date_tol_days, min_ref_len, fx_tol_bps, continuity_tol_abs, "
        "updated_at, updated_by) VALUES (?, 0.01, 0, 1, 6, 0, 1.00, ?, 'ops')",
        (acct_id, now))
    conn.commit(); conn.close()

    delta = tmp_path / 'fuzzy_high_tol.xlsx'
    _write_delta(delta, [(20260501, 5.0, 'CR', 'X', 'note')],
                  opening_balance=100.50, closing_balance=105.50)
    result = ingest_flex_only(delta, account_id=acct_id, user='ops')
    assert result.account_registered


def test_unseeded_account_raises_not_seeded(fresh_db, tmp_path):
    """A registered-but-unseeded account can't accept a delta — there's
    no anchor to chain off. NotSeededError."""
    from ingest import NotSeededError, ingest_flex_only

    acct_id = _seed_one_sided_account(fresh_db)
    delta = tmp_path / 'orphan.xlsx'
    _write_delta(delta, [(20260501, 5.0, 'CR', 'X', 'note')],
                  opening_balance=0.0, closing_balance=5.0)
    with pytest.raises(NotSeededError):
        ingest_flex_only(delta, account_id=acct_id, user='ops')


def test_two_day_chain(fresh_db, tmp_path):
    """The actual point of the chain: Day 1's closing must equal
    Day 2's opening. Set up Day 1, then send Day 2 with the matching
    opening, and verify both succeed end-to-end."""
    from ingest import ingest_flex_only

    acct_id = _seed_one_sided_account(fresh_db)
    _seed_proof(fresh_db, acct_id, tmp_path,
                [_proof_row(20260429, 100.0, 'C', 'PROOF1')])

    # Day 1: 100 → 150
    d1 = tmp_path / 'd1.xlsx'
    _write_delta(d1, [(20260501, 50.0, 'CR', 'D1', 'note')],
                  opening_balance=100.0, closing_balance=150.0,
                  as_of='2026-05-01')
    ingest_flex_only(d1, account_id=acct_id, user='ops')

    # Day 2: must open at 150 (Day 1's closing).
    d2 = tmp_path / 'd2.xlsx'
    _write_delta(d2, [(20260502, 25.0, 'DR', 'D2', 'note')],
                  opening_balance=150.0, closing_balance=125.0,
                  as_of='2026-05-02')
    r2 = ingest_flex_only(d2, account_id=acct_id, user='ops')
    assert r2.account_registered

    conn = sqlite3.connect(fresh_db); conn.row_factory = sqlite3.Row
    acct = dict(conn.execute(
        "SELECT last_closing_balance, last_closing_date FROM accounts WHERE id=?",
        (acct_id,)).fetchone())
    conn.close()
    assert acct['last_closing_balance'] == 125.0
    assert acct['last_closing_date'] == 20260502


# ---------------------------------------------------------------------------
# Force-accept override
# ---------------------------------------------------------------------------

def test_force_accept_bypasses_continuity_check(fresh_db, tmp_path):
    """Admin force-accepts a delta whose opening doesn't match the
    anchor. Anchor should advance to the file's closing; an audit log
    entry should record the force-accept with the operator's reason."""
    from ingest import ingest_flex_only

    acct_id = _seed_one_sided_account(fresh_db)
    _seed_proof(fresh_db, acct_id, tmp_path,
                [_proof_row(20260429, 100.0, 'C', 'PROOF1')])

    delta = tmp_path / 'force.xlsx'
    _write_delta(delta, [(20260501, 50.0, 'CR', 'X', 'note')],
                  opening_balance=200.0,           # off by +100
                  closing_balance=250.0)
    result = ingest_flex_only(delta, account_id=acct_id, user='ops',
                                force_accept=True,
                                force_reason='counterparty corrected opening — see ticket #1234')
    assert result.account_registered

    conn = sqlite3.connect(fresh_db); conn.row_factory = sqlite3.Row
    acct = dict(conn.execute(
        "SELECT last_closing_balance FROM accounts WHERE id=?",
        (acct_id,)).fetchone())
    audit = conn.execute(
        "SELECT details FROM audit_log WHERE session_id=? AND action='flex_delta_force_accepted'",
        (result.session_id,)).fetchone()
    conn.close()
    # Anchor jumped to the file's stated closing.
    assert acct['last_closing_balance'] == 250.0
    # Audit log captured the force-accept and the reason.
    import json
    assert audit is not None
    details = json.loads(audit['details'])
    assert details['force_accept'] is True
    assert 'ticket #1234' in details['force_reason']
    assert details['anchor_before'] == 100.0
    assert details['anchor_after'] == 250.0


# ---------------------------------------------------------------------------
# Carry-forward across the proof → delta boundary
# ---------------------------------------------------------------------------

def test_proof_cr_clears_against_delta_dr_after_run_matching(fresh_db, tmp_path):
    """The whole point of the open_items ledger: a CR sitting unmatched
    in the proof gets cleared when its DR counterpart shows up in a
    later delta.

    Note: carry-forward is opt-in on run_matching — operator triggers
    it as a separate explicit step. This test passes carry_forward=True
    to verify the underlying clearing still works when invoked."""
    from ingest import ingest_flex_only, run_matching

    acct_id = _seed_one_sided_account(fresh_db)
    # Proof: one CR (no matching DR within proof). After run_matching
    # on the seed session it becomes an open_item.
    seed = _seed_proof(fresh_db, acct_id, tmp_path, [
        _proof_row(20260429, 100.0, 'C', 'SHARED-REF'),
    ])
    # Seed needs to seed the CR as an open item — that uses the
    # seed_residue stage which is on by default. No tier needed for
    # seeding (it just lists the unmatched rows).
    run_matching(seed.session_id, user='ops')

    conn = sqlite3.connect(fresh_db); conn.row_factory = sqlite3.Row
    open_after_seed_match = conn.execute(
        "SELECT COUNT(*) FROM open_items WHERE account_id=? AND status='open'",
        (acct_id,)).fetchone()[0]
    conn.close()
    assert open_after_seed_match == 1   # the proof CR seeded as open

    # Delta: the matching DR posts. After load it sits as a row but no
    # carry-forward yet — that runs only when the operator clicks
    # Carry-forward open items on the review page.
    delta = tmp_path / 'matching_dr.xlsx'
    _write_delta(delta, [
        (20260501, 100.0, 'DR', 'SHARED-REF', 'settlement back-leg'),
    ], opening_balance=100.0, closing_balance=0.0, as_of='2026-05-01')
    delta_load = ingest_flex_only(delta, account_id=acct_id, user='ops')
    assert delta_load.open_items_cleared == 0   # deferred to run_matching

    # Explicit carry-forward step (mirrors the operator clicking the
    # "Carry-forward open items" button on the review page, which hits
    # POST /sessions/{id}/carry-forward).
    match = run_matching(delta_load.session_id, user='ops', carry_forward=True)
    assert match.open_items_carried == 1

    conn = sqlite3.connect(fresh_db); conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT status FROM open_items WHERE account_id=?", (acct_id,)
    ).fetchall()
    conn.close()
    statuses = [r['status'] for r in rows]
    # The proof CR is cleared; nothing left open from this account.
    assert 'cleared' in statuses
    assert 'open' not in statuses


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------

def test_missing_balances_sheet_rejects(fresh_db, tmp_path):
    """A delta xlsx without a `balances` sheet has no opening to check
    — Kilter refuses rather than guess. Operator can either fix the
    extract script or use force-accept."""
    from ingest import IngestError, ingest_flex_only

    acct_id = _seed_one_sided_account(fresh_db)
    _seed_proof(fresh_db, acct_id, tmp_path,
                [_proof_row(20260429, 100.0, 'C', 'PROOF1')])

    delta = tmp_path / 'no_balances.xlsx'
    _write_delta(delta, [(20260501, 5.0, 'CR', 'X', 'note')],
                  opening_balance=0.0, closing_balance=0.0,
                  with_balance_sheet=False)
    with pytest.raises(IngestError, match='balances'):
        ingest_flex_only(delta, account_id=acct_id, user='ops')


def test_wrong_account_in_file_rejected(fresh_db, tmp_path):
    """A file whose AC_NO doesn't match the registered account must be
    refused — preventing a wrong-account ingest is a primary contract."""
    from ingest import IngestError, ingest_flex_only

    acct_id = _seed_one_sided_account(fresh_db, flex_ac_no='1441000601589')
    _seed_proof(fresh_db, acct_id, tmp_path,
                [_proof_row(20260429, 100.0, 'C', 'PROOF1')])

    delta = tmp_path / 'wrong_ac.xlsx'
    _write_delta(delta, [(20260501, 5.0, 'CR', 'X', 'note')],
                  ac_no='9999999999999',  # not the registered ac_no
                  opening_balance=100.0, closing_balance=105.0)
    with pytest.raises(IngestError, match="doesn't match"):
        ingest_flex_only(delta, account_id=acct_id, user='ops')


def test_two_sided_account_rejects_flex_only_path(fresh_db, tmp_path):
    """ingest_flex_only is exclusively for one-sided accounts. A
    two-sided account must use ingest_pair — cross-using the flows
    would break the SWIFT pairing assumption downstream."""
    from ingest import IngestError, ingest_flex_only

    conn = sqlite3.connect(fresh_db)
    conn.execute(
        "INSERT INTO accounts (label, swift_account, flex_ac_no, currency, "
        "active, created_at, created_by, access_area, account_recon_type) "
        "VALUES ('TWO_SIDED', 'SW1', 'FX1', 'USD', 1, '2026-01-01', 'system', "
        "'NOSTRO', 'two_sided')")
    acct_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.commit(); conn.close()

    delta = tmp_path / 'unused.xlsx'
    _write_delta(delta, [(20260501, 5.0, 'CR', 'X', 'note')],
                  ac_no='FX1', currency='USD',
                  opening_balance=0.0, closing_balance=5.0)
    with pytest.raises(IngestError, match='not one-sided'):
        ingest_flex_only(delta, account_id=acct_id, user='ops')
