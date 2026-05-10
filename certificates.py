"""
certificates.py
===============

Month-end reconciliation certificate generation and sign-off flow.

A certificate is the formal artefact an ops team hands up the chain to
confirm "account X is reconciled for period Y". In Kilter the figures are
drawn straight from the live ledger (sessions / assignments / open_items),
so the numbers are always consistent with what's on screen. Once signed,
a JSON snapshot is frozen on the certificate row so later ledger edits
don't silently rewrite history.

Flow:
    draft      — generated; figures reflect current ledger
    prepared   — analyst (maker) has marked it as ready for review
    reviewed   — reviewer (checker) has verified and attests
    signed     — approver has signed off; snapshot frozen; immutable
    superseded — a later certificate for the same account/period took over

Two-eye / four-eye enforcement is left to roles (prepare by ops, review by
internal_control, sign by admin) rather than hardcoded so ops can map to
the bank's own segregation-of-duties requirements.
"""

from __future__ import annotations

import io
import json
from datetime import datetime, date, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# Border / fill helpers for the certificate sheet.
_THIN = Side(border_style='thin', color='000000')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill('solid', fgColor='0B1220')
_SUB_FILL    = PatternFill('solid', fgColor='EEF1F6')
_SIGN_FILL   = PatternFill('solid', fgColor='FEF3C7')
_BOLD = Font(bold=True)
_WHITE_BOLD = Font(bold=True, color='FFFFFF', size=12)


def compute_figures(conn, account_id: int, period_start: str,
                    period_end: str) -> dict[str, Any]:
    """Pull live figures for the account+period from sessions, assignments,
    open_items. Returns the dict that both drives the xlsx and, when the
    certificate is signed, gets frozen in snapshot_json."""
    acct = conn.execute(
        "SELECT id, label, shortname, access_area, swift_account, flex_ac_no, "
        "currency, bic FROM accounts WHERE id=?", (account_id,),
    ).fetchone()
    if acct is None:
        raise ValueError(f"account {account_id} not found")

    # Sessions that fall within the period.
    sessions = conn.execute(
        "SELECT id, created_at, status, "
        "opening_balance_amount, opening_balance_sign, opening_balance_date, "
        "closing_balance_amount, closing_balance_sign, closing_balance_date "
        "FROM sessions WHERE account_id=? "
        "AND DATE(created_at) BETWEEN ? AND ? ORDER BY id",
        (account_id, period_start, period_end),
    ).fetchall()

    # Assignment totals across those sessions.
    session_ids = [s['id'] for s in sessions]
    def assign_count(status: str) -> int:
        if not session_ids:
            return 0
        placeholders = ','.join('?' for _ in session_ids)
        return conn.execute(
            f"SELECT COUNT(*) FROM assignments WHERE status=? "
            f"AND session_id IN ({placeholders})",
            (status, *session_ids),
        ).fetchone()[0]
    confirmed = assign_count('confirmed')
    pending   = assign_count('pending')
    rejected  = assign_count('rejected')

    # Open items still open for this account at period end.
    open_rows = conn.execute(
        "SELECT source_side, category, functional_group, amount, "
        "value_date, opened_at FROM open_items "
        "WHERE account_id=? AND status='open' "
        "AND DATE(opened_at) <= ? ORDER BY opened_at",
        (account_id, period_end),
    ).fetchall()

    total_open_amount = sum(abs(r['amount'] or 0) for r in open_rows)
    by_category: dict[str, dict[str, float]] = {}
    by_group: dict[str, dict[str, float]] = {}
    for r in open_rows:
        cat = r['category'] or 'uncategorized'
        grp = r['functional_group'] or 'PSC TROPS'
        by_category.setdefault(cat, {'count': 0, 'amount': 0.0})
        by_group.setdefault(grp, {'count': 0, 'amount': 0.0})
        by_category[cat]['count'] += 1
        by_category[cat]['amount'] += abs(r['amount'] or 0)
        by_group[grp]['count'] += 1
        by_group[grp]['amount'] += abs(r['amount'] or 0)

    # Opening/closing balance: use the first session's opening and last
    # session's closing within the window if available.
    opening = None; closing = None
    if sessions:
        opening = {
            'amount': sessions[0]['opening_balance_amount'],
            'sign':   sessions[0]['opening_balance_sign'],
            'date':   sessions[0]['opening_balance_date'],
        } if sessions[0]['opening_balance_amount'] is not None else None
        closing = {
            'amount': sessions[-1]['closing_balance_amount'],
            'sign':   sessions[-1]['closing_balance_sign'],
            'date':   sessions[-1]['closing_balance_date'],
        } if sessions[-1]['closing_balance_amount'] is not None else None

    return {
        'account': dict(acct),
        'period': {'start': period_start, 'end': period_end},
        'sessions_in_period': len(sessions),
        'opening_balance': opening,
        'closing_balance': closing,
        'matches': {'confirmed': confirmed, 'pending': pending, 'rejected': rejected},
        'open_items': {
            'count': len(open_rows),
            'total_amount': total_open_amount,
            'by_category': by_category,
            'by_group': by_group,
        },
        'generated_at': datetime.now(timezone.utc).replace(tzinfo=None).isoformat(),
    }


def build_xlsx(figures: dict[str, Any], cert: dict | None = None) -> bytes:
    """Render figures (+ optional certificate metadata) as an xlsx.
    `cert` is the reconciliation_certificates row when available, so
    prepared/reviewed/signed identities flow onto the sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Certificate'

    acct = figures['account']
    period = figures['period']
    status = (cert or {}).get('status', 'draft')

    # --- Header band ------------------------------------------------------
    ws.merge_cells('A1:F2')
    c = ws.cell(row=1, column=1,
                value=f"RECONCILIATION CERTIFICATE  —  {acct['label']}")
    c.font = _WHITE_BOLD
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = _HEADER_FILL
    for col in range(1, 7):
        ws.cell(row=1, column=col).fill = _HEADER_FILL
        ws.cell(row=2, column=col).fill = _HEADER_FILL

    # --- Header facts ----------------------------------------------------
    row = 4
    def kv(k: str, v: Any, bold_value: bool = False) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=k).font = _BOLD
        cell = ws.cell(row=row, column=2, value=v)
        if bold_value:
            cell.font = _BOLD
        row += 1

    kv('Account', acct['label'], bold_value=True)
    kv('Access area', acct.get('access_area') or '—')
    kv('SWIFT account', acct.get('swift_account') or '—')
    kv('Flexcube GL', acct.get('flex_ac_no'))
    kv('Correspondent BIC', acct.get('bic') or '—')
    kv('Currency', acct['currency'])
    kv('Period', f"{period['start']}  to  {period['end']}")
    kv('Generated', figures['generated_at'].replace('T', ' ')[:19] + ' UTC')
    kv('Certificate status', status.upper(), bold_value=True)
    row += 1

    # --- Balance section -------------------------------------------------
    ws.cell(row=row, column=1, value='BALANCES').font = _BOLD
    ws.cell(row=row, column=1).fill = _SUB_FILL
    for col in range(1, 7): ws.cell(row=row, column=col).fill = _SUB_FILL
    row += 1
    ws.append(['', 'Amount', 'Sign', 'Value date'])
    row_headers = row; row += 0
    for col_hdr in (1, 2, 3, 4): ws.cell(row=row_headers, column=col_hdr).font = _BOLD

    def balrow(label: str, b: dict | None) -> None:
        nonlocal row
        row += 1
        ws.cell(row=row, column=1, value=label)
        if b:
            ws.cell(row=row, column=2, value=b['amount']).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=b['sign'])
            ws.cell(row=row, column=4, value=str(b['date']) if b['date'] else '')
        else:
            ws.cell(row=row, column=2, value='—')
    balrow('Opening', figures['opening_balance'])
    balrow('Closing', figures['closing_balance'])
    row += 2

    # --- Match summary ---------------------------------------------------
    ws.cell(row=row, column=1, value='MATCH SUMMARY').font = _BOLD
    for col in range(1, 7): ws.cell(row=row, column=col).fill = _SUB_FILL
    ws.cell(row=row, column=1).font = _BOLD
    row += 1
    ws.cell(row=row, column=1, value='Confirmed').font = _BOLD
    ws.cell(row=row, column=2, value=figures['matches']['confirmed']); row += 1
    ws.cell(row=row, column=1, value='Pending').font   = _BOLD
    ws.cell(row=row, column=2, value=figures['matches']['pending']); row += 1
    ws.cell(row=row, column=1, value='Rejected').font  = _BOLD
    ws.cell(row=row, column=2, value=figures['matches']['rejected']); row += 1
    ws.cell(row=row, column=1, value='Sessions in period').font  = _BOLD
    ws.cell(row=row, column=2, value=figures['sessions_in_period']); row += 2

    # --- Open items by team ----------------------------------------------
    ws.cell(row=row, column=1, value='OPEN ITEMS BY FUNCTIONAL TEAM').font = _BOLD
    for col in range(1, 7): ws.cell(row=row, column=col).fill = _SUB_FILL
    ws.cell(row=row, column=1).font = _BOLD
    row += 1
    ws.cell(row=row, column=1, value='Team').font = _BOLD
    ws.cell(row=row, column=2, value='Count').font = _BOLD
    ws.cell(row=row, column=3, value='Amount').font = _BOLD
    row += 1
    for grp_name, grp_data in sorted(figures['open_items']['by_group'].items(),
                                      key=lambda x: -x[1]['count']):
        ws.cell(row=row, column=1, value=grp_name)
        ws.cell(row=row, column=2, value=grp_data['count'])
        ws.cell(row=row, column=3, value=grp_data['amount']).number_format = '#,##0.00'
        row += 1
    ws.cell(row=row, column=1, value='TOTAL').font = _BOLD
    ws.cell(row=row, column=2, value=figures['open_items']['count']).font = _BOLD
    ws.cell(row=row, column=3, value=figures['open_items']['total_amount']).font = _BOLD
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    row += 3

    # --- Signatures ------------------------------------------------------
    ws.cell(row=row, column=1, value='SIGNATURES').font = _BOLD
    for col in range(1, 7): ws.cell(row=row, column=col).fill = _SUB_FILL
    ws.cell(row=row, column=1).font = _BOLD
    row += 1
    ws.cell(row=row, column=1, value='Role').font = _BOLD
    ws.cell(row=row, column=2, value='Name').font = _BOLD
    ws.cell(row=row, column=3, value='Date (UTC)').font = _BOLD
    row += 1

    def sigrow(role: str, user: str | None, at: str | None) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=role)
        ws.cell(row=row, column=2, value=user or '')
        ws.cell(row=row, column=3, value=(at or '').replace('T', ' ')[:19])
        for col in (1, 2, 3):
            ws.cell(row=row, column=col).fill = _SIGN_FILL
            ws.cell(row=row, column=col).border = _BORDER
        row += 1

    c = cert or {}
    sigrow('Prepared by (Maker)',    c.get('prepared_by'),  c.get('prepared_at'))
    sigrow('Reviewed by (Checker)',  c.get('reviewed_by'),  c.get('reviewed_at'))
    sigrow('Signed by (Approver)',   c.get('signed_by'),    c.get('signed_at'))

    # --- Column widths ---------------------------------------------------
    for col, w in enumerate([36, 22, 18, 18, 18, 18], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
