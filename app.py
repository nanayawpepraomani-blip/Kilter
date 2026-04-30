"""
app.py
======

FastAPI service for Kilter. Uploads a SWIFT + Flexcube pair, runs the
proposer engine, persists everything to SQLite, and exposes endpoints the
review UI will call.

Endpoints:
    POST   /sessions                  Upload pair, run engine, return session_id
    GET    /sessions                  List recent sessions
    GET    /sessions/{id}             Session metadata + counts
    GET    /sessions/{id}/queue       Pending assignments with competing candidates
    POST   /sessions/{id}/decisions   Confirm or reject a pending assignment
    GET    /sessions/{id}/export      Download xlsx reconciliation report
    GET    /sessions/{id}/audit       Audit trail for this session

Auth: real. Username + bcrypt-or-LDAP password + TOTP via Microsoft
Authenticator. Sessions issued on /login; token lives in localStorage
(injected into fetch() via the wrapper in base.html) and is mirrored
to a `kilter_token` cookie so server-rendered pages opened via
browser navigation also authenticate. See auth.py + secrets_vault.py
for the token issue/resolve + at-rest encryption.

Run:
    uvicorn app:app --reload
"""

from __future__ import annotations

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Tuple

import csv
import io

from fastapi import FastAPI, Depends, File, Form, Header, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel

from reconcile import write_report
from db import get_conn, init_db
from ingest import ingest_pair, IngestError, DuplicateFileError
from scanner import scan, ensure_dirs
from auth import (
    generate_enrollment_token, generate_totp_secret, qr_data_url, provisioning_uri,
    verify_totp, issue_session, resolve_session, revoke_session, revoke_all_sessions_for,
    generate_recovery_codes, store_recovery_codes, consume_recovery_code,
    SESSION_LIFETIME, ISSUER,
)


EXPORT_DIR = Path(__file__).resolve().parent / 'exports'
EXPORT_DIR.mkdir(exist_ok=True)
UPLOAD_DIR = Path(__file__).resolve().parent / 'uploads'
UPLOAD_DIR.mkdir(exist_ok=True)
TEMPLATES_DIR = Path(__file__).resolve().parent / 'templates'
# Static-asset directory. Currently holds product screenshots used by
# the pitch / demo decks and the operator manuals. Anything served from
# here is fully public — do NOT drop secrets, customer data, or
# pre-prod export artifacts. Screenshot filenames are documented in
# static/screenshots/README.md.
STATIC_DIR = Path(__file__).resolve().parent / 'static'
STATIC_DIR.mkdir(exist_ok=True)

app = FastAPI(
    title="Kilter",
    # Auto-generated docs disclose every endpoint + payload schema. Kilter is
    # a self-hosted internal app — there is no anonymous-user audience for
    # /docs, and an unauthenticated catalog of the API helps attackers more
    # than it helps integrators. Internal teams can re-enable behind admin
    # auth if needed; default off.
    docs_url=None, redoc_url=None, openapi_url=None,
)
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))
app.mount('/static', StaticFiles(directory=str(STATIC_DIR)), name='static')


# ---------------------------------------------------------------------------
# Rate limiter — keyed by client IP. Applied per-route via @limiter.limit().
# Currently used on /login to slow online TOTP-guessing.
# ---------------------------------------------------------------------------
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)


# ---------------------------------------------------------------------------
# Global exception handler — never leak DB errors, file paths, or stack
# traces in user-facing responses. FastAPI's default already returns a
# generic 500, but we want the same guarantee for sqlite3 errors that
# might otherwise be propagated with their full SQL string in the message.
# Logged in full to stderr so ops can correlate via the audit timestamp.
# ---------------------------------------------------------------------------
import logging as _logging
import sqlite3 as _sqlite3
import traceback as _traceback

_log = _logging.getLogger("kilter")


@app.exception_handler(_sqlite3.Error)
async def _sqlite_error_handler(request: Request, exc: _sqlite3.Error):
    """SQLite errors carry the full SQL in their str(); never let that
    reach the client. Log server-side, return generic 500."""
    from fastapi.responses import JSONResponse
    _log.error("sqlite error on %s %s: %s\n%s",
               request.method, request.url.path, exc,
               _traceback.format_exc())
    return JSONResponse({"detail": "Internal server error."}, status_code=500)


@app.exception_handler(Exception)
async def _generic_error_handler(request: Request, exc: Exception):
    """Catch-all for anything that isn't HTTPException, RateLimitExceeded
    (handled above), or sqlite3.Error. Logs the trace, returns generic
    500. Note: HTTPException is intercepted by FastAPI before reaching
    this handler, so legitimate 4xx flows are unaffected."""
    from fastapi.responses import JSONResponse
    _log.error("unhandled exception on %s %s: %s\n%s",
               request.method, request.url.path, exc,
               _traceback.format_exc())
    return JSONResponse({"detail": "Internal server error."}, status_code=500)


# ---------------------------------------------------------------------------
# Security headers middleware. Conservative defaults that work with the
# current Jinja2 + vanilla-JS frontend (no inline style/script except where
# explicitly added; no external script sources). HSTS preload-eligible value
# only kicks in when actually served over HTTPS — the header is harmless
# over plain HTTP because browsers ignore it on non-secure origins.
# ---------------------------------------------------------------------------
# 300 MB request-body cap. SWIFT statements top out around 2 MB and a Flex
# xlsx for one account-day under 5 MB, but cards-side settlement reports
# (Visa Base II, Mastercard IPM, large issuer CSV) routinely hit 250 MB+
# for medium issuers. 300 MB covers those with headroom; uploads at this
# size go via the streaming path in _save_upload (chunked write to disk,
# no full in-memory buffer).
MAX_REQUEST_BYTES = 300 * 1024 * 1024

# Streaming chunk size for _save_upload. 1 MB balances syscalls vs RAM
# footprint for the request worker — at 250 MB / 1 MB that's 250 writes
# per upload, well within syscall budget on any modern host.
UPLOAD_CHUNK_BYTES = 1 * 1024 * 1024

# Two-person approval gate. When true, operator confirmations go to
# 'pending_approval' instead of 'confirmed' and require a manager sign-off.
REQUIRE_APPROVAL = os.environ.get('KILTER_REQUIRE_APPROVAL', '').lower() in ('1', 'true', 'yes')


@app.middleware("http")
async def _enforce_max_request_size(request: Request, call_next):
    """Reject requests claiming a body bigger than MAX_REQUEST_BYTES. We
    only act on the Content-Length header (cheap), which catches honest
    oversized uploads. A malicious chunked-transfer client could still
    stream forever — protect against that at the reverse proxy with
    `client_max_body_size 300M;` (nginx) or equivalent."""
    cl = request.headers.get("content-length")
    if cl is not None:
        try:
            if int(cl) > MAX_REQUEST_BYTES:
                from fastapi.responses import JSONResponse
                return JSONResponse(
                    {"detail": f"Request body exceeds {MAX_REQUEST_BYTES // (1024*1024)} MB cap."},
                    status_code=413,
                )
        except ValueError:
            pass
    return await call_next(request)


@app.middleware("http")
async def _security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["Strict-Transport-Security"] = "max-age=63072000; includeSubDomains; preload"
    response.headers["X-Content-Type-Options"]    = "nosniff"
    response.headers["X-Frame-Options"]           = "DENY"
    response.headers["Referrer-Policy"]           = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"]        = "camera=(), microphone=(), geolocation=()"
    # The QR-code data URL is the only inline image; templates use a few
    # inline <style>/<script> blocks (acceptable for a self-hosted bank app
    # served from a single origin). Tighten further if all inline blocks are
    # extracted to external files later.
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "img-src 'self' data:; "
        "style-src 'self' 'unsafe-inline'; "
        "script-src 'self' 'unsafe-inline'; "
        "connect-src 'self'; "
        "font-src 'self' data:; "
        "frame-ancestors 'none'; "
        "base-uri 'self'; "
        "form-action 'self'"
    )
    return response


@app.on_event("startup")
def _on_startup() -> None:
    init_db()
    ensure_dirs()
    # Start the in-process scheduler daemon. The thread is daemonized so
    # uvicorn reload / Ctrl-C cleanly terminates it along with the worker.
    try:
        from scheduler import start as _start_scheduler
        _start_scheduler()
    except Exception as exc:
        print(f"[scheduler] failed to start: {exc}")


@app.on_event("shutdown")
def _on_shutdown() -> None:
    try:
        from scheduler import stop as _stop_scheduler
        _stop_scheduler()
    except Exception:
        pass


@app.get("/healthz", include_in_schema=False)
def _healthz() -> dict:
    """Liveness probe for container orchestrators. Intentionally shallow —
    answers "is the worker accepting HTTP?", not "is the database happy?".
    Keep DB-touching readiness checks out of the hot path."""
    return {"status": "ok"}


ROLES = ('admin', 'ops', 'audit', 'internal_control')


def current_user(request: Request,
                  x_session_token: str = Header(default="")) -> dict:
    """Auth dependency: validates the session token against user_sessions,
    returns the user dict. Token comes from one of two places:
        1. `X-Session-Token` header — used by every fetch() call (the
           wrapper in base.html injects it from localStorage).
        2. `kilter_token` cookie — used by browser navigations to
           server-rendered pages (e.g. <a target="_blank"> to a
           certificate print view). The same JS that writes localStorage
           also writes this cookie so both transports stay in sync.

    Header wins when both are present (header is the authoritative
    transport for API calls). Invalid, expired, or revoked tokens
    yield 401; the fetch wrapper bounces those to /login."""
    token = (x_session_token or "").strip()
    if not token:
        token = (request.cookies.get('kilter_token') or "").strip()
    if not token:
        raise HTTPException(401, "Missing session token. Please sign in.")
    conn = get_conn()
    try:
        username = resolve_session(conn, token)
        if username is None:
            raise HTTPException(401, "Session expired or invalid. Please sign in again.")
        row = conn.execute(
            "SELECT username, display_name, role, active FROM users WHERE username=?",
            (username,),
        ).fetchone()
        if row is None:
            raise HTTPException(401, "User no longer exists.")
        if not row['active']:
            raise HTTPException(403, "Your account is deactivated.")
        try:
            conn.execute("UPDATE users SET last_seen_at=? WHERE username=?",
                         (datetime.utcnow().isoformat(), username))
            conn.commit()
        except Exception:
            pass
        return dict(row)
    finally:
        conn.close()


def active_scope(request: Request,
                 user: dict = Depends(current_user),
                 x_session_token: str = Header(default="")) -> Optional[List[str]]:
    """The user's active access-area scope for list views. None = no filter
    (show everything). [] = "none selected" — we treat the same as None so the
    app never renders a deliberately-empty list from a bad toggle. A concrete
    list means: only show accounts whose access_area is in this set.

    Now requires a valid session — `current_user` runs first and raises 401
    on missing/expired/revoked tokens. Previously this dep silently no-op'd
    for unauthenticated requests, which left every endpoint depending on it
    (list_sessions, list_accounts, stats, dashboard/*, etc.) wide open.

    Token sourcing mirrors current_user: header first, cookie fallback,
    so server-rendered pages opened via browser navigation also resolve
    the user's saved scope."""
    token = (x_session_token or "").strip()
    if not token:
        token = (request.cookies.get('kilter_token') or "").strip()
    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT active_access_areas FROM user_sessions WHERE token=? "
            "AND revoked_at IS NULL",
            (token,),
        ).fetchone()
        if row is None or not row['active_access_areas']:
            return None
        try:
            areas = json.loads(row['active_access_areas'])
        except (TypeError, ValueError):
            return None
        if not isinstance(areas, list) or not areas:
            return None
        return [str(x) for x in areas]
    finally:
        conn.close()


def _assert_session_not_locked(conn, session_id: int) -> None:
    """Raise 423 if the session has been locked by a signed certificate."""
    row = conn.execute(
        "SELECT locked_at FROM sessions WHERE id=?", (session_id,)
    ).fetchone()
    if row and row['locked_at']:
        raise HTTPException(423, "Session is locked after certificate sign-off and cannot be modified.")


def _scope_clause(scope: Optional[List[str]], alias: str = 'a') -> Tuple[str, list]:
    """Build a SQL snippet and params for an IN-filter against accounts.access_area.
    When scope is None, returns an always-true clause so callers can concatenate
    it unconditionally."""
    if not scope:
        return ("1=1", [])
    placeholders = ",".join("?" * len(scope))
    return (f"{alias}.access_area IN ({placeholders})", list(scope))


def _assert_account_in_scope(conn, account_id: int, scope: Optional[List[str]]) -> None:
    """Raise 403 if `account_id` doesn't belong to one of the user's active
    access areas. Use on every endpoint that takes an account_id (or a
    nested id like cert_id that resolves to one) to plug the IDOR class:
    knowing an integer id was enough to read out-of-scope data otherwise.

    `scope=None` means "no scope filter" — admin-style full visibility,
    so we don't block. A non-existent account is treated as 404 by the
    caller; this helper only enforces *visibility*, not existence."""
    if scope is None:
        return
    row = conn.execute(
        "SELECT access_area FROM accounts WHERE id=?", (account_id,),
    ).fetchone()
    if row is None:
        # Don't leak existence — treat unknown ids the same as out-of-scope.
        raise HTTPException(404, "not found")
    area = row['access_area']
    if area is None or area not in scope:
        raise HTTPException(403, "Account is outside your active access scope.")


def _account_id_for_certificate(conn, cert_id: int) -> int:
    """Resolve cert_id → account_id for scope checks on the
    /certificates/{cert_id}/* endpoints. Raises 404 if missing."""
    row = conn.execute(
        "SELECT account_id FROM reconciliation_certificates WHERE id=?",
        (cert_id,),
    ).fetchone()
    if row is None:
        raise HTTPException(404, "certificate not found")
    return int(row['account_id'])


def require_role(*roles: str):
    """Dependency factory: use with Depends(require_role('admin')) etc."""
    def checker(user: dict = Depends(current_user)) -> dict:
        if user['role'] not in roles:
            raise HTTPException(403,
                f"This action requires role: {' or '.join(roles)}. "
                f"You are '{user['role']}'.")
        return user
    return checker


@app.get("/")
def home(request: Request):
    return templates.TemplateResponse(request, "home.html")


@app.get("/pitch", include_in_schema=False)
def pitch_deck(request: Request):
    """Renders the live sales pitch deck (`templates/deck_pitch.html`).
    Public — no auth — so prospects can review the deck without an
    account. Update content via the template, not via this handler."""
    return templates.TemplateResponse(request, "deck_pitch.html")


@app.get("/demo-deck", include_in_schema=False)
def demo_deck(request: Request):
    """Renders the demo walkthrough deck (`templates/deck_demo.html`).
    Same public-no-auth posture as /pitch. Named `/demo-deck` rather
    than `/demo` to leave `/demo` available for a future hosted-instance
    landing page."""
    return templates.TemplateResponse(request, "deck_demo.html")


@app.get("/intake")
def intake_landing():
    """Intake hub. We split scan + upload into distinct sub-pages so the
    sidebar's 'Scan messages' and 'Manual upload' lead to genuinely
    different views; this top-level route just lands the user on the
    most-common-by-far flow (scan) so the palette / breadcrumbs still
    have a useful target."""
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/intake/scan", status_code=303)


@app.get("/intake/scan")
def intake_scan_page(request: Request):
    """Scan the watched-folder pair (messages/swift + messages/flexcube).
    Admin-only at the action layer (POST /scan); page renders for any
    logged-in user, with a non-admin notice when the role doesn't match."""
    return templates.TemplateResponse(request, "intake_scan.html")


@app.get("/intake/upload")
def intake_upload_page(request: Request):
    """Manual one-off upload of a SWIFT + ledger pair. Same dedup +
    validation pipeline as scan, just bypassing the watched folders."""
    return templates.TemplateResponse(request, "intake_upload.html")


@app.get("/login")
def login_page(request: Request):
    return templates.TemplateResponse(request, "login.html")


@app.get("/enroll")
def enroll_page(request: Request):
    return templates.TemplateResponse(request, "enroll.html")


# ---------------------------------------------------------------------------
# Auth endpoints (no session required — they're how you GET a session)
# ---------------------------------------------------------------------------

class LoginPayload(BaseModel):
    username: str
    totp_code: str
    # Required only when the user's row has auth_source='ldap'. Local
    # users (default) ignore this field. The frontend always shows the
    # field — server-side validation per user is what enforces the rule.
    password: Optional[str] = ""


class EnrollStartPayload(BaseModel):
    username: str
    enrollment_token: str


class EnrollCompletePayload(BaseModel):
    username: str
    enrollment_token: str
    totp_code: str


@app.post("/login")
@limiter.limit("10/minute")
def login(payload: LoginPayload, request: Request):
    """Username + (LDAP password if user is AD-sourced) + TOTP login.
    Rate-limited per source IP to slow online TOTP-guessing — 10
    attempts/minute is generous enough that legitimate typo retries
    don't lock anyone out, but kills brute-forcing dead."""
    username = (payload.username or "").strip()
    code = (payload.totp_code or "").strip()
    password = payload.password or ""
    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT username, active, totp_secret, totp_enrolled_at, auth_source "
            "FROM users WHERE username=?",
            (username,),
        ).fetchone()
        # Generic 401 for every failure mode so we don't leak which field was wrong
        # (username, password, or TOTP). The audit/log line carries the real reason
        # for ops triage, but the user always gets the same message.
        # totp_enrolled_at IS NOT NULL guards against the half-enrolled state where
        # /enroll/start has stashed a pending secret but /enroll/complete never ran.
        if (row is None or not row['active']
                or not row['totp_secret'] or not row['totp_enrolled_at']):
            raise HTTPException(401, "Invalid username or authenticator code.")

        # Password layer (LDAP). For local users this is a no-op — TOTP alone
        # remains sufficient, preserving the bootstrap-admin path.
        ldap_dn: Optional[str] = None
        if (row['auth_source'] or 'local') == 'ldap':
            from ldap_auth import authenticate as ldap_authenticate, REASON_OK
            result = ldap_authenticate(username, password)
            if not result.success:
                # Audit the reason without surfacing it to the user.
                _audit_login_failure(conn, username, request, reason=f"ldap:{result.reason}")
                raise HTTPException(401, "Invalid username or authenticator code.")
            ldap_dn = result.user_dn

        used_recovery = False
        if not verify_totp(row['totp_secret'], code, username=username):
            # Recovery codes are 14 chars with dashes (XXXX-XXXX-XXXX).
            # Accept them as a fallback when the TOTP code fails.
            if len(code.replace('-', '')) == 12 and consume_recovery_code(conn, username, code):
                used_recovery = True
            else:
                _audit_login_failure(conn, username, request, reason="totp")
                raise HTTPException(401, "Invalid username or authenticator code.")

        ua = request.headers.get('user-agent', '')[:250]
        session = issue_session(conn, username, user_agent=ua)
        now = datetime.utcnow().isoformat()
        if ldap_dn is not None:
            conn.execute(
                "UPDATE users SET last_seen_at=?, ldap_dn=? WHERE username=?",
                (now, ldap_dn, username),
            )
        else:
            conn.execute(
                "UPDATE users SET last_seen_at=? WHERE username=?",
                (now, username),
            )
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'login', ?, ?, ?)",
            (username, now, json.dumps({
                "user_agent": ua,
                "auth_source": row['auth_source'] or 'local',
                "method": "recovery_code" if used_recovery else "totp",
            })),
        )
        conn.commit()
        return session
    finally:
        conn.close()


def _audit_login_failure(conn, username: str, request: Request, *, reason: str) -> None:
    """Record a failed login attempt with the failure mode (for ops). The
    end user only ever sees the generic 401 — this gives the audit log
    enough granularity to investigate without leaking it to the client."""
    try:
        ua = request.headers.get('user-agent', '')[:250]
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'login_failed', ?, ?, ?)",
            (username, datetime.utcnow().isoformat(),
             json.dumps({"user_agent": ua, "reason": reason})),
        )
        conn.commit()
    except Exception:
        # Audit is best-effort; never block the user-facing response on it.
        pass


@app.post("/logout")
def logout(x_session_token: str = Header(default="")):
    conn = get_conn()
    try:
        # Resolve to know whose audit row to write before we revoke.
        username = resolve_session(conn, x_session_token or "")
        if username:
            revoke_session(conn, x_session_token)
            conn.execute(
                "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
                "VALUES (NULL, 'logout', ?, ?, NULL)",
                (username, datetime.utcnow().isoformat()),
            )
            conn.commit()
        return {"ok": True}
    finally:
        conn.close()


@app.post("/enroll/start")
def enroll_start(payload: EnrollStartPayload):
    """Generate a fresh TOTP secret server-side and stash it (encrypted)
    against the user row. Return the QR and the manual-entry key for the
    user to scan/type into their authenticator. The secret never returns
    to the client over the wire as a programmatically-readable JSON
    field — the client only re-sends a TOTP code in /enroll/complete,
    not the secret itself.

    Server-side keying closes the attack where a malicious client (XSS,
    extension, MITM during a misconfigured TLS rollout) substitutes a
    secret of its own choosing and gets the server to verify against
    that, capturing future logins."""
    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT username, enrollment_token, totp_enrolled_at FROM users "
            "WHERE username=? AND active=1",
            (payload.username,),
        ).fetchone()
        if row is None or not row['enrollment_token'] or row['enrollment_token'] != payload.enrollment_token:
            raise HTTPException(400, "Invalid enrollment link. Ask an admin for a fresh one.")
        if row['totp_enrolled_at']:
            raise HTTPException(400, "This user has already enrolled.")

        # Fresh secret per /enroll/start call. If the user abandons and
        # restarts, the previous (pending) secret is overwritten so any
        # leaked QR from the abandoned session becomes useless. The QR
        # and manual-entry key intentionally still ship to the client —
        # they MUST be displayed for the user to scan or type.
        from secrets_vault import encrypt
        secret = generate_totp_secret()
        conn.execute(
            "UPDATE users SET totp_secret=?, totp_enrolled_at=NULL WHERE username=?",
            (encrypt(secret), payload.username),
        )
        conn.commit()
        return {
            "username": payload.username,
            "manual_key": secret,         # for the "type instead of scan" path
            "uri": provisioning_uri(secret, payload.username),
            "qr": qr_data_url(secret, payload.username),
            "issuer": ISSUER,
        }
    finally:
        conn.close()


@app.post("/enroll/complete")
def enroll_complete(payload: EnrollCompletePayload):
    """Verify the code against the server-stored secret and finalise the
    enrollment. The client does NOT send the secret — the server reads
    its own copy from the DB and decrypts."""
    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT username, enrollment_token, totp_secret, totp_enrolled_at "
            "FROM users WHERE username=? AND active=1",
            (payload.username,),
        ).fetchone()
        if row is None or not row['enrollment_token'] or row['enrollment_token'] != payload.enrollment_token:
            raise HTTPException(400, "Invalid enrollment link.")
        if row['totp_enrolled_at']:
            raise HTTPException(400, "This user has already enrolled.")
        if not row['totp_secret']:
            raise HTTPException(400, "No enrollment in progress. Open the enrollment link again to start over.")
        # verify_totp transparently decrypts before checking.
        if not verify_totp(row['totp_secret'], payload.totp_code):
            raise HTTPException(400, "Code didn't match. Make sure your phone's time is in sync and try the next rotation.")

        now = datetime.utcnow().isoformat()
        conn.execute(
            "UPDATE users SET totp_enrolled_at=?, enrollment_token=NULL "
            "WHERE username=?",
            (now, payload.username),
        )
        # Generate and store single-use recovery codes. Returned once — the
        # client must display them; they cannot be retrieved again.
        recovery_codes = generate_recovery_codes(8)
        store_recovery_codes(conn, payload.username, recovery_codes)
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'totp_enrolled', ?, ?, NULL)",
            (payload.username, now),
        )
        conn.commit()
        return {"ok": True, "username": payload.username, "recovery_codes": recovery_codes}
    finally:
        conn.close()


@app.get("/sessions/{session_id}/review")
def review(request: Request, session_id: int):
    """Review-page render. Page routes never run server-side auth deps —
    the X-Session-Token lives in localStorage and is only injected on
    fetch(), so a browser navigation arrives without it. Auth + scope
    happen JS-side: review.html immediately hits GET /sessions/{id},
    which IS auth-and-scope-protected, and the fetch wrapper in
    base.html bounces on 401 to /login?next=…."""
    return templates.TemplateResponse(request, "review.html",
                                      {"session_id": session_id})


@app.get("/cash-accounts")
def cash_accounts_page(request: Request):
    return templates.TemplateResponse(request, "cash_accounts.html")


@app.get("/sessions-list")
def sessions_list_page(request: Request):
    return templates.TemplateResponse(request, "sessions_list.html")


@app.get("/admin/users")
def admin_users_page(request: Request):
    return templates.TemplateResponse(request, "admin_users.html")


@app.get("/admin/activity")
def admin_activity_page(request: Request):
    return templates.TemplateResponse(request, "admin_activity.html")


@app.get("/activity/facets")
def activity_facets(
    user: dict = Depends(require_role('admin', 'audit', 'internal_control')),
):
    """Distinct actors + actions + per-action counts — drives the filter
    dropdowns and summary strip on the activity page. Cheap because
    audit_log is small (thousands of rows, not millions)."""
    conn = get_conn()
    try:
        actors = [r[0] for r in conn.execute(
            "SELECT DISTINCT actor FROM audit_log ORDER BY actor"
        ).fetchall()]
        actions = [dict(r) for r in conn.execute(
            "SELECT action, COUNT(*) as n FROM audit_log "
            "GROUP BY action ORDER BY n DESC"
        ).fetchall()]
        total = conn.execute("SELECT COUNT(*) FROM audit_log").fetchone()[0]
        return {'actors': actors, 'actions': actions, 'total': total}
    finally:
        conn.close()


@app.get("/open-items/{open_item_id}/history")
def open_item_history(
    open_item_id: int,
    user: dict = Depends(current_user),
):
    """Timeline for one break: when it was opened, what assignment cleared
    it (if any), every comment posted, every manual reclassification.
    The compliance story in one JSON call."""
    conn = get_conn()
    try:
        oi = conn.execute(
            "SELECT oi.*, a.label AS account_label "
            "FROM open_items oi LEFT JOIN accounts a ON a.id=oi.account_id "
            "WHERE oi.id=?", (open_item_id,),
        ).fetchone()
        if oi is None:
            raise HTTPException(404, "open item not found")

        events = []
        events.append({
            'ts': oi['opened_at'], 'kind': 'opened',
            'actor': 'system',
            'description': (f"Opened from {oi['source_side']} session "
                            f"#{oi['src_session_id']} row {oi['src_row_number']}")
        })

        # Comments
        comments = conn.execute(
            "SELECT created_at, author, body FROM break_comments "
            "WHERE target_type='open_item' AND target_id=? ORDER BY id",
            (open_item_id,),
        ).fetchall()
        for c in comments:
            events.append({
                'ts': c['created_at'], 'kind': 'comment',
                'actor': c['author'], 'description': c['body'],
            })

        # Audit log entries mentioning this open_item
        audit = conn.execute(
            "SELECT timestamp, action, actor, details FROM audit_log "
            "WHERE action LIKE 'open_item%' ORDER BY id"
        ).fetchall()
        for a in audit:
            if not a['details']:
                continue
            try:
                d = json.loads(a['details'])
            except Exception:
                continue
            if d.get('open_item_id') == open_item_id:
                desc = a['action'].replace('_', ' ')
                if 'reason' in d: desc += f" — {d['reason']}"
                if 'note'   in d and d['note']: desc += f" — {d['note']}"
                events.append({
                    'ts': a['timestamp'], 'kind': a['action'],
                    'actor': a['actor'], 'description': desc,
                })

        # Closure / clearing
        if oi['cleared_at']:
            events.append({
                'ts': oi['cleared_at'], 'kind': 'cleared',
                'actor': oi['cleared_by'] or 'system',
                'description': (f"Cleared via {oi['cleared_via']}"
                                + (f" (assignment {oi['cleared_assignment_id']})"
                                   if oi['cleared_assignment_id'] else ''))
            })

        events.sort(key=lambda e: e['ts'])
        return {'open_item': dict(oi), 'events': events}
    finally:
        conn.close()


@app.get("/admin/discovered")
def admin_discovered_page(request: Request):
    return templates.TemplateResponse(request, "discovered_accounts.html")


# ---------------------------------------------------------------------------
# /me — lets the UI know who's signed in and what they can do
# ---------------------------------------------------------------------------

@app.get("/me")
def whoami(user: dict = Depends(current_user)):
    return user


# ---------------------------------------------------------------------------
# /me/access-scope — the topbar "Active area" picker reads/writes this.
# Scope is stored on user_sessions (not users) so a single person can have
# a different scope per browser/device.
# ---------------------------------------------------------------------------

class AccessScopePayload(BaseModel):
    areas: Optional[List[str]] = None   # null or [] -> clear filter (all areas)


@app.get("/me/access-scope")
def get_my_access_scope(
    user: dict = Depends(current_user),
    x_session_token: str = Header(default=""),
):
    """Returns the current selection + the list of available active areas so
    the picker can render in a single round-trip."""
    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT active_access_areas FROM user_sessions WHERE token=?",
            ((x_session_token or "").strip(),),
        ).fetchone()
        areas: Optional[List[str]] = None
        if row and row['active_access_areas']:
            try:
                parsed = json.loads(row['active_access_areas'])
                if isinstance(parsed, list) and parsed:
                    areas = [str(x) for x in parsed]
            except (TypeError, ValueError):
                areas = None
        available = [dict(r) for r in conn.execute(
            "SELECT name, parent FROM access_areas WHERE active=1 ORDER BY name"
        ).fetchall()]
        return {"areas": areas, "available": available}
    finally:
        conn.close()


@app.put("/me/access-scope")
def set_my_access_scope(
    payload: AccessScopePayload,
    user: dict = Depends(current_user),
    x_session_token: str = Header(default=""),
):
    """Persist the user's chosen scope to their session row. Empty list and
    null both mean 'no filter' — we normalize to NULL so the scope dep's
    'is this filter active?' check is a single null-test."""
    token = (x_session_token or "").strip()
    areas = payload.areas or None
    if areas is not None:
        # Validate names against the active registry — typos in a PUT body
        # shouldn't silently hide every account.
        conn = get_conn()
        try:
            valid = {r[0] for r in conn.execute(
                "SELECT name FROM access_areas WHERE active=1"
            ).fetchall()}
        finally:
            conn.close()
        unknown = [a for a in areas if a not in valid]
        if unknown:
            raise HTTPException(400, f"Unknown access area(s): {', '.join(unknown)}")
    conn = get_conn()
    try:
        conn.execute(
            "UPDATE user_sessions SET active_access_areas=? WHERE token=?",
            (json.dumps(areas) if areas else None, token),
        )
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'access_scope_change', ?, ?, ?)",
            (user['username'], datetime.utcnow().isoformat(),
             json.dumps({"areas": areas})),
        )
        conn.commit()
        return {"areas": areas}
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# /users — admin-only management
# ---------------------------------------------------------------------------

AUTH_SOURCES = ('local', 'ldap')


class UserPayload(BaseModel):
    username: str
    display_name: Optional[str] = None
    role: str   # admin | ops | audit | internal_control
    auth_source: Optional[str] = 'local'   # 'local' | 'ldap'


class UserPatch(BaseModel):
    role: Optional[str] = None
    display_name: Optional[str] = None
    active: Optional[bool] = None
    auth_source: Optional[str] = None      # 'local' | 'ldap'


@app.get("/users/export")
def export_users(user: dict = Depends(require_role('admin'))):
    """CSV export of the users table — for compliance reports, onboarding audits."""
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT username, display_name, role, active, created_at, created_by, "
            "last_seen_at, totp_enrolled_at FROM users ORDER BY role, username"
        ).fetchall()
    finally:
        conn.close()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['username', 'display_name', 'role', 'active', 'created_at',
                'created_by', 'last_seen_at', 'totp_enrolled_at'])
    for r in rows:
        w.writerow([r['username'], r['display_name'] or '', r['role'],
                    'yes' if r['active'] else 'no', r['created_at'],
                    r['created_by'] or '', r['last_seen_at'] or '',
                    r['totp_enrolled_at'] or ''])
    _audit_export(user['username'], 'users_export', {"rows": len(rows)})
    fname = f"kilter_users_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


@app.get("/activity/export")
def export_activity(
    actor: Optional[str] = None,
    action: Optional[str] = None,
    session_id: Optional[int] = None,
    user: dict = Depends(require_role('admin', 'audit', 'internal_control')),
):
    """CSV export of the audit_log — same filters as GET /activity. No row
    cap on export — give auditors the full picture."""
    where, params = [], []
    if actor:      where.append("actor = ?");      params.append(actor)
    if action:     where.append("action = ?");     params.append(action)
    if session_id: where.append("session_id = ?"); params.append(session_id)
    clause = ("WHERE " + " AND ".join(where)) if where else ""

    conn = get_conn()
    try:
        rows = conn.execute(
            f"SELECT id, session_id, action, actor, timestamp, details "
            f"FROM audit_log {clause} ORDER BY id",
            params,
        ).fetchall()
    finally:
        conn.close()

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['id', 'timestamp', 'actor', 'action', 'session_id', 'details'])
    for r in rows:
        w.writerow([r['id'], r['timestamp'], r['actor'], r['action'],
                    r['session_id'] or '', r['details'] or ''])
    _audit_export(user['username'], 'activity_export', {
        "rows": len(rows),
        "filters": {"actor": actor, "action": action, "session_id": session_id},
    })
    fname = f"kilter_activity_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


def _audit_export(actor: str, action: str, details: dict) -> None:
    """Log export actions so auditors can see who pulled what."""
    conn = get_conn()
    try:
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, ?, ?, ?, ?)",
            (action, actor, datetime.utcnow().isoformat(), json.dumps(details)),
        )
        conn.commit()
    finally:
        conn.close()


@app.get("/users")
def list_users(user: dict = Depends(require_role('admin'))):
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT username, display_name, role, active, created_at, created_by, "
            "last_seen_at, auth_source, ldap_dn "
            "FROM users ORDER BY role, username"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.post("/users")
def create_user(payload: UserPayload, user: dict = Depends(require_role('admin'))):
    if payload.role not in ROLES:
        raise HTTPException(400, f"role must be one of {ROLES}")
    auth_source = (payload.auth_source or 'local').strip().lower()
    if auth_source not in AUTH_SOURCES:
        raise HTTPException(400, f"auth_source must be one of {AUTH_SOURCES}")
    username = (payload.username or '').strip()
    if not username:
        raise HTTPException(400, "username is required")
    token = generate_enrollment_token()
    now = datetime.utcnow().isoformat()
    conn = get_conn()
    try:
        try:
            conn.execute(
                "INSERT INTO users (username, display_name, role, active, created_at, "
                "created_by, enrollment_token, auth_source) VALUES (?, ?, ?, 1, ?, ?, ?, ?)",
                (username, payload.display_name, payload.role, now, user['username'], token, auth_source),
            )
        except Exception as exc:
            if 'UNIQUE' in str(exc) or 'PRIMARY KEY' in str(exc):
                raise HTTPException(409, f"User '{username}' already exists.")
            raise
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'user_created', ?, ?, ?)",
            (user['username'], now, json.dumps({"username": username, "role": payload.role})),
        )
        conn.commit()
        return {
            "username": username,
            "role": payload.role,
            "enrollment_token": token,
            "enrollment_url": f"/enroll?user={username}&token={token}",
        }
    finally:
        conn.close()


@app.patch("/users/{target}")
def update_user(target: str, payload: UserPatch, user: dict = Depends(require_role('admin'))):
    if payload.role is not None and payload.role not in ROLES:
        raise HTTPException(400, f"role must be one of {ROLES}")
    if payload.auth_source is not None and payload.auth_source not in AUTH_SOURCES:
        raise HTTPException(400, f"auth_source must be one of {AUTH_SOURCES}")
    if target == user['username'] and payload.active is False:
        raise HTTPException(400, "You can't deactivate yourself.")
    if target == user['username'] and payload.role is not None and payload.role != 'admin':
        raise HTTPException(400, "You can't demote yourself from admin.")
    # Lock-out guard: don't let an admin flip themselves to LDAP if no
    # LDAP server is reachable yet — that's how you accidentally lock the
    # whole system out. Force them to test it on a non-admin first.
    if (target == user['username']
            and payload.auth_source == 'ldap'):
        from ldap_auth import is_enabled as ldap_is_enabled
        if not ldap_is_enabled():
            raise HTTPException(400, "Configure KILTER_LDAP_URL before changing your own auth_source to ldap.")

    conn = get_conn()
    try:
        row = conn.execute("SELECT * FROM users WHERE username=?", (target,)).fetchone()
        if row is None:
            raise HTTPException(404, f"User '{target}' not found")

        _PATCH_COLS = [
            ('role',         'role=?',         lambda v: v),
            ('display_name', 'display_name=?', lambda v: v),
            ('active',       'active=?',       lambda v: 1 if v else 0),
            ('auth_source',  'auth_source=?',  lambda v: v),
        ]
        fields, params = [], []
        for attr, col_expr, transform in _PATCH_COLS:
            val = getattr(payload, attr)
            if val is not None:
                fields.append(col_expr)
                params.append(transform(val))
        if not fields:
            raise HTTPException(400, "Nothing to update")
        params.append(target)
        conn.execute(
            "UPDATE users SET " + ", ".join(fields) + " WHERE username=?",
            params,
        )

        # Deactivation, role change, or auth-source flip → kill the user's
        # live sessions so the new restriction takes effect immediately.
        if (payload.active is False or payload.role is not None
                or payload.auth_source is not None):
            revoke_all_sessions_for(conn, target)

        now = datetime.utcnow().isoformat()
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'user_updated', ?, ?, ?)",
            (user['username'], now, json.dumps({
                "target": target, "role": payload.role,
                "display_name": payload.display_name, "active": payload.active,
                "auth_source": payload.auth_source,
            })),
        )
        conn.commit()
        updated = conn.execute(
            "SELECT username, display_name, role, active, auth_source FROM users WHERE username=?",
            (target,),
        ).fetchone()
        return dict(updated)
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# /users/{username}/recovery-codes — admin reset of MFA backup codes
# ---------------------------------------------------------------------------

@app.post("/users/{target}/recovery-codes/reset")
def reset_recovery_codes(target: str, user: dict = Depends(require_role('admin'))):
    """Generate and store a new set of recovery codes for the target user.
    The new codes are returned once — the admin must relay them out-of-band.
    Any unused prior codes are invalidated."""
    conn = get_conn()
    try:
        row = conn.execute("SELECT username FROM users WHERE username=?", (target,)).fetchone()
        if row is None:
            raise HTTPException(404, f"User '{target}' not found")
        codes = generate_recovery_codes(8)
        store_recovery_codes(conn, target, codes)
        now = datetime.utcnow().isoformat()
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'recovery_codes_reset', ?, ?, ?)",
            (user['username'], now, json.dumps({"target": target})),
        )
        conn.commit()
        return {"username": target, "recovery_codes": codes}
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# /activity — system-wide audit log viewer (admin/audit/internal_control)
# ---------------------------------------------------------------------------

@app.get("/activity")
def activity(
    actor: Optional[str] = None,
    action: Optional[str] = None,
    session_id: Optional[int] = None,
    from_date: Optional[str] = None,      # ISO date, inclusive (UTC)
    to_date: Optional[str] = None,        # ISO date, inclusive (UTC)
    q: Optional[str] = None,              # free-text search over details JSON
    limit: int = 200,
    user: dict = Depends(require_role('admin', 'audit', 'internal_control')),
):
    limit = max(1, min(limit, 1000))
    where, params = [], []
    if actor:      where.append("actor = ?");      params.append(actor)
    if action:     where.append("action = ?");     params.append(action)
    if session_id: where.append("session_id = ?"); params.append(session_id)
    if from_date:  where.append("timestamp >= ?"); params.append(from_date + 'T00:00:00')
    if to_date:    where.append("timestamp <= ?"); params.append(to_date + 'T23:59:59')
    if q:
        where.append("(details LIKE ? OR action LIKE ?)")
        params.extend([f"%{q}%", f"%{q}%"])
    clause = ("WHERE " + " AND ".join(where)) if where else ""

    conn = get_conn()
    try:
        rows = conn.execute(
            f"SELECT id, session_id, action, actor, timestamp, details "
            f"FROM audit_log {clause} ORDER BY id DESC LIMIT ?",
            (*params, limit),
        ).fetchall()
        return [
            {**dict(r), 'details': json.loads(r['details']) if r['details'] else None}
            for r in rows
        ]
    finally:
        conn.close()


@app.get("/stats")
def stats(scope: Optional[List[str]] = Depends(active_scope)):
    """Aggregates used by the dashboard. When a scope is active, counts are
    restricted to sessions whose account falls in the chosen access areas."""
    conn = get_conn()
    try:
        where, scope_params = _scope_clause(scope, alias='a')
        # Sessions-in-scope subquery: drives every count below so the
        # assignment/session filters stay consistent.
        account_filter = "AND a.id IS NOT NULL" if scope else ""
        sess_subq = (
            f"SELECT s.id FROM sessions s LEFT JOIN accounts a ON a.id = s.account_id "
            f"WHERE {where} {account_filter}"
        )
        def one(q, params=()):
            return conn.execute(q, params).fetchone()[0]
        return {
            "pending":   one(
                f"SELECT COUNT(*) FROM assignments WHERE status='pending' "
                f"AND session_id IN ({sess_subq})", scope_params),
            "confirmed": one(
                f"SELECT COUNT(*) FROM assignments WHERE status='confirmed' "
                f"AND session_id IN ({sess_subq})", scope_params),
            "rejected":  one(
                f"SELECT COUNT(*) FROM assignments WHERE status='rejected' "
                f"AND session_id IN ({sess_subq})", scope_params),
            "open_sessions": one(
                f"SELECT COUNT(DISTINCT session_id) FROM assignments "
                f"WHERE status='pending' AND session_id IN ({sess_subq})",
                scope_params),
            "total_sessions": one(
                f"SELECT COUNT(*) FROM ({sess_subq})", scope_params),
        }
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Dashboard analytics — small aggregates feeding the landing-page charts.
# Each endpoint respects the active-access-area scope so the visuals match
# the counters on /stats. Cheap SQL — no pagination needed at current volumes.
# ---------------------------------------------------------------------------

@app.get("/dashboard/trend")
def dashboard_trend(days: int = 14,
                    scope: Optional[List[str]] = Depends(active_scope)):
    """Daily totals of assignment state-changes over the last N days, used to
    draw the match-rate sparkline. Returns one row per calendar day from
    (today - days + 1) to today, zero-filled when nothing happened."""
    from datetime import date, timedelta
    days = max(1, min(90, int(days)))
    today = date.today()
    start = today - timedelta(days=days - 1)

    conn = get_conn()
    try:
        where, scope_params = _scope_clause(scope, alias='a')
        account_filter = "AND a.id IS NOT NULL" if scope else ""
        sess_subq = (
            f"SELECT s.id FROM sessions s LEFT JOIN accounts a ON a.id = s.account_id "
            f"WHERE {where} {account_filter}"
        )
        rows = conn.execute(
            f"SELECT SUBSTR(decided_at, 1, 10) AS d, status, COUNT(*) AS n "
            f"FROM assignments WHERE decided_at >= ? "
            f"AND session_id IN ({sess_subq}) "
            f"GROUP BY d, status",
            (start.isoformat() + 'T00:00:00', *scope_params),
        ).fetchall()
        by_day = {}
        for r in rows:
            d = r['d']
            entry = by_day.setdefault(d, {'confirmed': 0, 'rejected': 0, 'pending': 0})
            entry[r['status']] = r['n']
        out = []
        for i in range(days):
            d = (start + timedelta(days=i)).isoformat()
            e = by_day.get(d, {'confirmed': 0, 'rejected': 0, 'pending': 0})
            out.append({
                'date': d,
                'confirmed': e['confirmed'],
                'rejected': e['rejected'],
                'pending': e['pending'],
                'total': e['confirmed'] + e['rejected'] + e['pending'],
            })
        return out
    finally:
        conn.close()


@app.get("/dashboard/ageing")
def dashboard_ageing(scope: Optional[List[str]] = Depends(active_scope)):
    """Histogram of currently-open items by age bucket. Drives the ageing bar
    chart on the dashboard and the at-risk (> 30d) callout."""
    from datetime import datetime, timedelta
    now = datetime.utcnow()
    buckets = [
        ('0-1d',  0,  1),
        ('2-3d',  2,  3),
        ('4-7d',  4,  7),
        ('8-30d', 8,  30),
        ('30d+',  31, 9999),
    ]
    conn = get_conn()
    try:
        where, scope_params = _scope_clause(scope, alias='a')
        rows = conn.execute(
            f"SELECT oi.opened_at, oi.amount FROM open_items oi "
            f"LEFT JOIN accounts a ON a.id = oi.account_id "
            f"WHERE oi.status='open' AND {where}",
            scope_params,
        ).fetchall()

        out = [{'label': b[0], 'count': 0, 'amount': 0.0} for b in buckets]
        total_count = 0; total_amount = 0.0; at_risk = 0
        for r in rows:
            try:
                opened = datetime.fromisoformat(r['opened_at'])
            except Exception:
                continue
            age = (now - opened).days
            for i, (_, lo, hi) in enumerate(buckets):
                if lo <= age <= hi:
                    out[i]['count'] += 1
                    out[i]['amount'] += abs(r['amount'] or 0)
                    break
            total_count += 1
            total_amount += abs(r['amount'] or 0)
            if age > 30:
                at_risk += 1
        return {
            'buckets': out,
            'total_count': total_count,
            'total_amount': total_amount,
            'at_risk_count': at_risk,
        }
    finally:
        conn.close()


@app.get("/dashboard/by-group")
def dashboard_by_group(scope: Optional[List[str]] = Depends(active_scope)):
    """Open-item count + absolute amount per functional_group — powers the
    per-team breakdown card. Respects access-area scope."""
    conn = get_conn()
    try:
        where, scope_params = _scope_clause(scope, alias='a')
        rows = conn.execute(
            f"SELECT COALESCE(oi.functional_group, '(unassigned)') AS grp, "
            f"COUNT(*) AS n, SUM(ABS(oi.amount)) AS amt "
            f"FROM open_items oi LEFT JOIN accounts a ON a.id = oi.account_id "
            f"WHERE oi.status='open' AND {where} GROUP BY grp ORDER BY n DESC",
            scope_params,
        ).fetchall()
        return [{'group': r['grp'], 'count': r['n'], 'amount': float(r['amt'] or 0)}
                for r in rows]
    finally:
        conn.close()


@app.get("/dashboard/kpis")
def dashboard_kpis(scope: Optional[List[str]] = Depends(active_scope)):
    """Top-of-page operational health KPIs. Cheap aggregates only — this
    endpoint loads on every dashboard render so anything expensive should
    move to its own card-level endpoint with on-demand fetch.

    Returns:
        tier1_rate         — % of decisions that landed at tier 1 in the
                             trailing 14 days. The "easy match" baseline;
                             slipping below ~70% usually means a parser
                             or tolerance has drifted.
        oldest_open_days   — age in days of the oldest currently-open item.
                             Pinned because regulators ask "what's your
                             oldest unresolved break?" and the answer is
                             always a single number.
        sla_breached       — count of pending assignments where due_date
                             is in the past. Zero is the operating state;
                             non-zero is a noise signal for the duty
                             manager.
        total_open_items   — same denominator as the ageing card; convenient
                             to surface alongside the SLA count.
    """
    from datetime import date as _date, datetime as _dt
    conn = get_conn()
    try:
        where, scope_params = _scope_clause(scope, alias='a')
        account_filter = "AND a.id IS NOT NULL" if scope else ""
        sess_subq = (
            f"SELECT s.id FROM sessions s LEFT JOIN accounts a ON a.id = s.account_id "
            f"WHERE {where} {account_filter}"
        )

        # Tier-1 hit rate over the trailing 14 days. Same window the trend
        # chart uses, so the KPI matches whatever the operator is staring
        # at. confirmed-only — rejected tier-1 doesn't count toward "easy".
        from datetime import timedelta as _td
        cutoff = (_dt.utcnow() - _td(days=14)).isoformat()
        decided = conn.execute(
            f"SELECT tier, COUNT(*) AS n FROM assignments "
            f"WHERE decided_at >= ? AND status='confirmed' "
            f"AND session_id IN ({sess_subq}) GROUP BY tier",
            (cutoff, *scope_params),
        ).fetchall()
        tier_counts = {r['tier']: r['n'] for r in decided}
        total_decided = sum(tier_counts.values())
        tier1_rate = (
            round(100.0 * tier_counts.get(1, 0) / total_decided, 1)
            if total_decided else None
        )

        # Oldest open item — a single max() lookup. NULL when no open items.
        oldest_row = conn.execute(
            f"SELECT MIN(oi.opened_at) AS first_opened FROM open_items oi "
            f"LEFT JOIN accounts a ON a.id = oi.account_id "
            f"WHERE oi.status='open' AND {where}",
            scope_params,
        ).fetchone()
        oldest_open_days = None
        if oldest_row and oldest_row['first_opened']:
            try:
                opened = _dt.fromisoformat(oldest_row['first_opened'])
                oldest_open_days = (_dt.utcnow() - opened).days
            except (ValueError, TypeError):
                pass

        # SLA-breached pending cases: due_date strictly in the past.
        # The due_date column is stored as ISO 'YYYY-MM-DD' so a string
        # comparison works correctly.
        today_iso = _date.today().isoformat()
        sla_breached = conn.execute(
            f"SELECT COUNT(*) AS n FROM assignments "
            f"WHERE status='pending' AND due_date IS NOT NULL AND due_date < ? "
            f"AND session_id IN ({sess_subq})",
            (today_iso, *scope_params),
        ).fetchone()['n']

        total_open = conn.execute(
            f"SELECT COUNT(*) AS n FROM open_items oi "
            f"LEFT JOIN accounts a ON a.id = oi.account_id "
            f"WHERE oi.status='open' AND {where}",
            scope_params,
        ).fetchone()['n']

        return {
            'tier1_rate': tier1_rate,
            'tier_counts': tier_counts,
            'total_decided_14d': total_decided,
            'oldest_open_days': oldest_open_days,
            'sla_breached': sla_breached,
            'total_open_items': total_open,
        }
    finally:
        conn.close()


@app.get("/dashboard/case-load")
def dashboard_case_load(scope: Optional[List[str]] = Depends(active_scope)):
    """Pending-case count per assignee, plus how many of each are
    SLA-breached. Drives the "who's swamped, who's overdue" table.

    Returns one row per assignee (including the synthetic '(unassigned)'
    bucket for null assignees) sorted by descending pending count.
    """
    from datetime import date as _date
    conn = get_conn()
    try:
        where, scope_params = _scope_clause(scope, alias='a')
        account_filter = "AND a.id IS NOT NULL" if scope else ""
        sess_subq = (
            f"SELECT s.id FROM sessions s LEFT JOIN accounts a ON a.id = s.account_id "
            f"WHERE {where} {account_filter}"
        )
        today_iso = _date.today().isoformat()
        rows = conn.execute(
            f"SELECT COALESCE(asg.assignee, '(unassigned)') AS assignee, "
            f"       COUNT(*) AS pending, "
            f"       SUM(CASE WHEN asg.due_date IS NOT NULL AND asg.due_date < ? "
            f"                THEN 1 ELSE 0 END) AS overdue, "
            f"       SUM(CASE WHEN asg.priority IN ('high','urgent') THEN 1 ELSE 0 END) AS high_priority "
            f"FROM assignments asg "
            f"WHERE asg.status='pending' AND asg.session_id IN ({sess_subq}) "
            f"GROUP BY assignee ORDER BY pending DESC",
            (today_iso, *scope_params),
        ).fetchall()
        return [{
            'assignee': r['assignee'],
            'pending': r['pending'],
            'overdue': r['overdue'] or 0,
            'high_priority': r['high_priority'] or 0,
        } for r in rows]
    finally:
        conn.close()


@app.get("/dashboard/by-account")
def dashboard_by_account(limit: int = 10,
                          scope: Optional[List[str]] = Depends(active_scope)):
    """Open-item count + absolute amount per account, top N by count.
    Surfaces which nostros are actually slipping — the per-account view
    operators want when planning a clean-up sprint.

    Caps at 50 to keep the response cheap; default 10 fits a card."""
    limit = max(1, min(50, int(limit)))
    conn = get_conn()
    try:
        where, scope_params = _scope_clause(scope, alias='a')
        rows = conn.execute(
            f"SELECT a.id AS account_id, "
            f"       COALESCE(a.shortname, a.label) AS label, "
            f"       a.currency AS currency, "
            f"       COUNT(*) AS n, "
            f"       SUM(ABS(oi.amount)) AS amt, "
            f"       MIN(oi.opened_at) AS oldest_at "
            f"FROM open_items oi "
            f"JOIN accounts a ON a.id = oi.account_id "
            f"WHERE oi.status='open' AND {where} "
            f"GROUP BY a.id ORDER BY n DESC LIMIT ?",
            (*scope_params, limit),
        ).fetchall()
        return [{
            'account_id': r['account_id'],
            'label': r['label'],
            'currency': r['currency'],
            'count': r['n'],
            'amount': float(r['amt'] or 0),
            'oldest_at': r['oldest_at'],
        } for r in rows]
    finally:
        conn.close()


@app.on_event("startup")
def _on_startup() -> None:
    init_db()
    ensure_dirs()


# ---------------------------------------------------------------------------
# POST /sessions
# ---------------------------------------------------------------------------

@app.post("/sessions")
async def create_session(
    swift: UploadFile = File(...),
    flex: UploadFile = File(...),
    flex_profile_id: Optional[int] = Form(None),
    user: dict = Depends(require_role('admin')),
):
    """Default flow: SWIFT MT/camt or xlsx + Flex xlsx. With
    `flex_profile_id`, the Flex file is treated as CSV and parsed via
    a saved BYO format profile — same recon pipeline downstream."""
    if not (swift.filename or '').lower().endswith(('.xlsx', '.out', '.xml', '.txt')):
        raise HTTPException(400, "SWIFT file must be .out/.xml/.txt or .xlsx")
    flex_lower = (flex.filename or '').lower()
    if flex_profile_id is not None:
        if not flex_lower.endswith(('.csv', '.txt')):
            raise HTTPException(400,
                "When using a CSV profile, Flex file must be .csv or .txt")
    elif not flex_lower.endswith('.xlsx'):
        raise HTTPException(400,
            "Flexcube file must be .xlsx (or supply flex_profile_id for CSV)")

    swift_path = await _save_upload(swift, 'swift')
    flex_path = await _save_upload(flex, 'flex')

    try:
        result = ingest_pair(
            swift_path, flex_path, user['username'],
            swift_filename=swift.filename, flex_filename=flex.filename,
            flex_profile_id=flex_profile_id,
        )
    except DuplicateFileError as exc:
        raise HTTPException(409, str(exc))
    except IngestError as exc:
        raise HTTPException(400, str(exc))

    return {
        "session_id": result.session_id,
        "swift_rows": result.swift_rows,
        "flex_rows": result.flex_rows,
        "candidates_proposed": result.candidates_proposed,
        "pending_assignments": result.pending_assignments,
        "unmatched_swift": result.unmatched_swift,
        "unmatched_flex": result.unmatched_flex,
        "open_items_seeded": result.open_items_seeded,
        "open_items_cleared": result.open_items_cleared,
        "account": {
            "label": result.account_label,
            "registered": result.account_registered,
            "swift_account": result.swift_account,
            "flex_ac_no": result.flex_ac_no,
            "currency": result.currency,
        },
    }


@app.post("/scan")
def run_scan(user: dict = Depends(require_role('admin'))):
    """Trigger a sweep of messages/swift and messages/flexcube.

    Files that pair up to a registered account get ingested into new
    sessions. Everything else is moved to the appropriate messages/unloaded/
    subfolder with a reason. Idempotent — re-running does nothing new because
    the ingest layer refuses files whose SHA-256 it has already seen."""
    report = scan(user=user['username'])
    return {
        "sessions_created": report.sessions_created,
        "counts": report.counts,
        "outcomes": [
            {"file": o.file, "kind": o.kind, "status": o.status,
             "reason": o.reason, "session_id": o.session_id, "moved_to": o.moved_to}
            for o in report.outcomes
        ],
    }


# ---------------------------------------------------------------------------
# GET /sessions, GET /sessions/{id}
# ---------------------------------------------------------------------------

@app.get("/sessions")
def list_sessions(flex_profile_id: Optional[str] = None,
                  scope: Optional[List[str]] = Depends(active_scope)):
    """List recent sessions, scope-filtered.

    flex_profile_id query param accepts:
        - an integer N → only sessions ingested via that profile
        - the string 'default' → only sessions ingested via the
          built-in Flexcube xlsx loader (sessions where flex_profile_id
          IS NULL)
        - omitted → all sessions

    The dashboard uses this filter to power the source-chip toolbar
    above the recent-sessions table.
    """
    conn = get_conn()
    try:
        where, params = _scope_clause(scope, alias='a')
        # When a scope is active, sessions with no linked account (account_id
        # is null) don't belong to any area and are hidden — the user asked to
        # see BRANCH X, so legacy/unclaimed sessions shouldn't leak in.
        account_filter = "AND a.id IS NOT NULL" if scope else ""

        profile_filter = ""
        profile_params: list = []
        if flex_profile_id is not None:
            if flex_profile_id == 'default':
                profile_filter = "AND s.flex_profile_id IS NULL"
            else:
                try:
                    pid = int(flex_profile_id)
                except (TypeError, ValueError):
                    raise HTTPException(400,
                        "flex_profile_id must be an integer or 'default'")
                profile_filter = "AND s.flex_profile_id = ?"
                profile_params.append(pid)

        rows = conn.execute(
            f"SELECT s.*, "
            f"       a.shortname AS account_shortname, "
            f"       a.access_area AS account_access_area, "
            f"       p.name AS flex_profile_name "
            f"FROM sessions s "
            f"LEFT JOIN accounts a ON a.id = s.account_id "
            f"LEFT JOIN csv_format_profiles p ON p.id = s.flex_profile_id "
            f"WHERE {where} {account_filter} {profile_filter} "
            f"ORDER BY s.id DESC LIMIT 100",
            (*params, *profile_params),
        ).fetchall()
        sessions = [dict(r) for r in rows]
        # Per-session counts in one query so we don't N+1 the list view.
        counts_rows = conn.execute(
            "SELECT session_id, status, COUNT(*) AS n FROM assignments GROUP BY session_id, status"
        ).fetchall()
        by_session: dict[int, dict] = {}
        for r in counts_rows:
            by_session.setdefault(r['session_id'], {})[r['status']] = r['n']
        for s in sessions:
            c = by_session.get(s['id'], {})
            s['counts'] = {
                'pending':   c.get('pending', 0),
                'confirmed': c.get('confirmed', 0),
                'rejected':  c.get('rejected', 0),
            }
        return sessions
    finally:
        conn.close()


@app.get("/sessions/{session_id}")
def get_session(session_id: int,
                user: dict = Depends(current_user),
                scope: Optional[List[str]] = Depends(active_scope)):
    conn = get_conn()
    try:
        # Scope check: a user limited to certain access areas can only read
        # sessions whose linked account sits in those areas. Sessions without
        # an account_id (legacy / unregistered intake) remain admin-visible.
        scope_where, scope_params = _scope_clause(scope, alias='a')
        s = conn.execute(
            "SELECT s.*, a.shortname AS account_shortname, a.access_area AS account_access_area, "
            "       a.bic AS account_bic "
            "FROM sessions s LEFT JOIN accounts a ON a.id = s.account_id "
            "WHERE s.id=? AND (s.account_id IS NULL OR " + scope_where + ")",
            (session_id, *scope_params),
        ).fetchone()
        if s is None:
            raise HTTPException(404, "Session not found")

        def one(q, params):
            row = conn.execute(q, params).fetchone()
            return row[0] if row and row[0] is not None else 0

        counts = {
            "pending":    one("SELECT COUNT(*) FROM assignments WHERE session_id=? AND status='pending'", (session_id,)),
            "confirmed":  one("SELECT COUNT(*) FROM assignments WHERE session_id=? AND status='confirmed'", (session_id,)),
            "rejected":   one("SELECT COUNT(*) FROM assignments WHERE session_id=? AND status='rejected'", (session_id,)),
            "swift_total": one("SELECT COUNT(*) FROM swift_txns WHERE session_id=?", (session_id,)),
            "flex_total":  one("SELECT COUNT(*) FROM flex_txns WHERE session_id=?", (session_id,)),
        }

        # Totals drive the reconcile panel — SWIFT credits/debits and Flex credits/debits
        # split into "all" and "confirmed" (cleared) buckets.
        totals = {
            "swift_credits_all": one(
                "SELECT COALESCE(SUM(amount),0) FROM swift_txns WHERE session_id=? AND sign='C'",
                (session_id,)),
            "swift_debits_all": one(
                "SELECT COALESCE(SUM(amount),0) FROM swift_txns WHERE session_id=? AND sign='D'",
                (session_id,)),
            "flex_credits_all": one(
                "SELECT COALESCE(SUM(amount),0) FROM flex_txns WHERE session_id=? AND type='CR'",
                (session_id,)),
            "flex_debits_all": one(
                "SELECT COALESCE(SUM(amount),0) FROM flex_txns WHERE session_id=? AND type='DR'",
                (session_id,)),
            # Cleared = matched-and-confirmed only; pending/rejected don't count.
            "swift_credits_cleared": one(
                "SELECT COALESCE(SUM(st.amount),0) FROM swift_txns st "
                "JOIN assignments a ON a.session_id=st.session_id AND a.swift_row=st.row_number "
                "WHERE st.session_id=? AND st.sign='C' AND a.status='confirmed'",
                (session_id,)),
            "swift_debits_cleared": one(
                "SELECT COALESCE(SUM(st.amount),0) FROM swift_txns st "
                "JOIN assignments a ON a.session_id=st.session_id AND a.swift_row=st.row_number "
                "WHERE st.session_id=? AND st.sign='D' AND a.status='confirmed'",
                (session_id,)),
            "flex_credits_cleared": one(
                "SELECT COALESCE(SUM(ft.amount),0) FROM flex_txns ft "
                "JOIN assignments a ON a.session_id=ft.session_id AND a.flex_row=ft.row_number "
                "WHERE ft.session_id=? AND ft.type='CR' AND a.status='confirmed'",
                (session_id,)),
            "flex_debits_cleared": one(
                "SELECT COALESCE(SUM(ft.amount),0) FROM flex_txns ft "
                "JOIN assignments a ON a.session_id=ft.session_id AND a.flex_row=ft.row_number "
                "WHERE ft.session_id=? AND ft.type='DR' AND a.status='confirmed'",
                (session_id,)),
        }
        return {**dict(s), "counts": counts, "totals": totals}
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Register view — all txns on both sides with match status, for the Register
# tab on the review page.
# ---------------------------------------------------------------------------

@app.get("/sessions/{session_id}/register")
def session_register(session_id: int, user: dict = Depends(current_user)):
    """Return a flat list of every SWIFT + Flex txn with status, match partner,
    and tier (if any). Fast enough for the UI to filter/sort client-side."""
    conn = get_conn()
    try:
        # Assignment lookup by side, keyed by row_number. A row can appear
        # on multiple assignments if it's in a split-match group; pick the
        # strongest non-rejected one for the primary badge but surface the
        # split_group_id so the UI can highlight siblings.
        assigns = conn.execute(
            "SELECT swift_row, flex_row, tier, status, decided_by, decided_at, "
            "source, split_group_id "
            "FROM assignments WHERE session_id=? ORDER BY "
            "CASE status WHEN 'confirmed' THEN 0 WHEN 'pending' THEN 1 "
            "WHEN 'rejected' THEN 2 ELSE 3 END, tier",
            (session_id,),
        ).fetchall()
        swift_assign, flex_assign = {}, {}
        for a in assigns:
            if a['swift_row'] and a['swift_row'] not in swift_assign:
                swift_assign[a['swift_row']] = dict(a)
            if a['flex_row'] and a['flex_row'] not in flex_assign:
                flex_assign[a['flex_row']] = dict(a)

        swift_rows = conn.execute(
            "SELECT * FROM swift_txns WHERE session_id=? ORDER BY row_number",
            (session_id,),
        ).fetchall()
        flex_rows = conn.execute(
            "SELECT * FROM flex_txns WHERE session_id=? ORDER BY row_number",
            (session_id,),
        ).fetchall()

        entries = []
        for r in swift_rows:
            a = swift_assign.get(r['row_number'])
            entries.append({
                'side': 'swift',
                'row_number': r['row_number'],
                'value_date': r['value_date'],
                'amount': r['amount'],
                'sign': r['sign'],
                'ref': r['our_ref'],
                'description': (r['booking_text_1'] or '') + (
                    ' · ' + r['booking_text_2'] if r['booking_text_2'] else ''),
                'status': (a['status'] if a else 'unmatched'),
                'tier':  (a['tier'] if a else None),
                'partner_row': (a['flex_row'] if a else None),
                'source': (a['source'] if a else None),
                'split_group_id': (a['split_group_id'] if a else None),
            })
        for r in flex_rows:
            a = flex_assign.get(r['row_number'])
            entries.append({
                'side': 'flex',
                'row_number': r['row_number'],
                'value_date': r['value_date'],
                'amount': r['amount'],
                'sign': 'C' if r['type'] == 'CR' else 'D',
                'ref': r['trn_ref'],
                'description': r['narration'] or '',
                'status': (a['status'] if a else 'unmatched'),
                'tier':  (a['tier'] if a else None),
                'partner_row': (a['swift_row'] if a else None),
                'source': (a['source'] if a else None),
                'split_group_id': (a['split_group_id'] if a else None),
            })
        return entries
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# GET /sessions/{id}/queue
# ---------------------------------------------------------------------------

@app.get("/sessions/{session_id}/queue")
def get_queue(session_id: int,
              user: dict = Depends(current_user),
              scope: Optional[List[str]] = Depends(active_scope)):
    conn = get_conn()
    try:
        # Scope-check the parent session before exposing its queue.
        scope_where, scope_params = _scope_clause(scope, alias='a')
        owned = conn.execute(
            "SELECT s.id FROM sessions s LEFT JOIN accounts a ON a.id = s.account_id "
            "WHERE s.id=? AND (s.account_id IS NULL OR " + scope_where + ")",
            (session_id, *scope_params),
        ).fetchone()
        if owned is None:
            raise HTTPException(404, "Session not found")
        assignments = conn.execute(
            "SELECT * FROM assignments WHERE session_id=? AND status='pending' "
            "ORDER BY tier, id",
            (session_id,),
        ).fetchall()

        out = []
        for a in assignments:
            swift = conn.execute(
                "SELECT * FROM swift_txns WHERE session_id=? AND row_number=?",
                (session_id, a['swift_row']),
            ).fetchone()
            flex = conn.execute(
                "SELECT * FROM flex_txns WHERE session_id=? AND row_number=?",
                (session_id, a['flex_row']),
            ).fetchone()
            competing = conn.execute(
                "SELECT c.*, "
                "  CASE WHEN c.swift_row=? THEN 'alt_flex' ELSE 'alt_swift' END AS side "
                "FROM candidates c "
                "WHERE c.session_id=? AND (c.swift_row=? OR c.flex_row=?) "
                "  AND NOT (c.swift_row=? AND c.flex_row=?) "
                "ORDER BY c.tier, ABS(c.amount_diff)",
                (a['swift_row'], session_id, a['swift_row'], a['flex_row'],
                 a['swift_row'], a['flex_row']),
            ).fetchall()

            comment_count = conn.execute(
                "SELECT COUNT(*) AS n FROM break_comments "
                "WHERE target_type='assignment' AND target_id=?",
                (a['id'],),
            ).fetchone()['n']
            out.append({
                "assignment_id": a['id'],
                "tier": a['tier'],
                "reason": a['reason'],
                "amount_diff": a['amount_diff'],
                "swift": dict(swift) if swift else None,
                "flex": dict(flex) if flex else None,
                "competing": [dict(c) for c in competing],
                # Case-management fields. None / 'normal' on rows that
                # haven't been touched — UI renders defaults.
                "assignee":     a['assignee'] if 'assignee' in a.keys() else None,
                "due_date":     a['due_date'] if 'due_date' in a.keys() else None,
                "priority":     (a['priority'] if 'priority' in a.keys() else None) or 'normal',
                "comment_count": comment_count,
            })
        return out
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Case management — assignee, due date, priority, comments.
# Layered on top of the existing pending-assignment workflow rather than a
# parallel object so the existing decision flow (confirm / reject) is
# unchanged. A "case" is just an assignment with extra metadata.
# ---------------------------------------------------------------------------

class CasePatch(BaseModel):
    assignee: Optional[str] = None     # username; empty string clears
    due_date: Optional[str] = None     # ISO date 'YYYY-MM-DD'; empty clears
    priority: Optional[str] = None     # 'low' | 'normal' | 'high' | 'urgent'


CASE_PRIORITIES = ('low', 'normal', 'high', 'urgent')


def _assignment_session_id(conn, assignment_id: int) -> int:
    """Resolve assignment_id → session_id with NotFound on missing.
    Used by the case endpoints to scope-check against the parent session."""
    row = conn.execute(
        "SELECT session_id FROM assignments WHERE id=?", (assignment_id,),
    ).fetchone()
    if row is None:
        raise HTTPException(404, "assignment not found")
    return int(row['session_id'])


def _assert_session_in_scope(conn, session_id: int, scope: Optional[List[str]]) -> None:
    """Re-uses the account-scope helper indirectly: a session inherits its
    account's access_area. NULL account = unscoped; admins (scope=None)
    bypass entirely."""
    if scope is None:
        return
    row = conn.execute(
        "SELECT s.account_id, a.access_area "
        "FROM sessions s LEFT JOIN accounts a ON a.id = s.account_id "
        "WHERE s.id=?", (session_id,),
    ).fetchone()
    if row is None:
        raise HTTPException(404, "not found")
    if row['account_id'] is None:
        return  # untagged session — visible to everyone
    if row['access_area'] is None or row['access_area'] not in scope:
        raise HTTPException(403, "Session is outside your active access scope.")


@app.patch("/assignments/{assignment_id}/case")
def patch_case(assignment_id: int, payload: CasePatch,
               user: dict = Depends(require_role('ops', 'admin', 'internal_control')),
               scope: Optional[List[str]] = Depends(active_scope)):
    """Update case fields on a pending or confirmed assignment. Empty
    string in `assignee` or `due_date` clears the field; None means
    leave-unchanged. Priority must be one of CASE_PRIORITIES.

    Audit-logged because changing the assignee or SLA on an in-flight
    case is a meaningful operational decision."""
    if (payload.assignee is None and payload.due_date is None
            and payload.priority is None):
        raise HTTPException(400, "Nothing to update")
    if payload.priority is not None and payload.priority not in CASE_PRIORITIES:
        raise HTTPException(400, f"priority must be one of {CASE_PRIORITIES}")
    if payload.due_date is not None and payload.due_date != "":
        # Soft validation: just confirm we can parse it as YYYY-MM-DD.
        try:
            datetime.strptime(payload.due_date, '%Y-%m-%d')
        except ValueError:
            raise HTTPException(400, "due_date must be ISO 'YYYY-MM-DD'")
    conn = get_conn()
    try:
        session_id = _assignment_session_id(conn, assignment_id)
        _assert_session_in_scope(conn, session_id, scope)

        fields, params = [], []
        if payload.assignee is not None:
            fields.append("assignee=?"); params.append(payload.assignee or None)
        if payload.due_date is not None:
            fields.append("due_date=?"); params.append(payload.due_date or None)
        if payload.priority is not None:
            fields.append("priority=?"); params.append(payload.priority)
        params.append(assignment_id)
        conn.execute(
            f"UPDATE assignments SET {', '.join(fields)} WHERE id=?", params,
        )
        now = datetime.utcnow().isoformat()
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (?, 'case_updated', ?, ?, ?)",
            (session_id, user['username'], now, json.dumps({
                "assignment_id": assignment_id,
                "assignee": payload.assignee,
                "due_date": payload.due_date,
                "priority": payload.priority,
            })),
        )
        conn.commit()
        row = conn.execute(
            "SELECT id, session_id, assignee, due_date, priority "
            "FROM assignments WHERE id=?", (assignment_id,),
        ).fetchone()
        return dict(row)
    finally:
        conn.close()


# NOTE on comments: the existing /comments endpoints (POST/GET) handle
# break_comments against any assignment or open_item. Case management
# reuses that store; we only added the assignee / due_date / priority
# metadata to assignments above. See the /comments handlers further down.


# ---------------------------------------------------------------------------
# POST /sessions/{id}/decisions
# ---------------------------------------------------------------------------

class DecisionPayload(BaseModel):
    assignment_id: int
    action: str  # 'confirm' | 'reject'


@app.post("/sessions/{session_id}/decisions")
def post_decision(
    session_id: int,
    payload: DecisionPayload,
    user: dict = Depends(require_role('ops', 'admin')),
):
    if payload.action not in ('confirm', 'reject'):
        raise HTTPException(400, "action must be 'confirm' or 'reject'")

    username = user['username']
    conn = get_conn()
    try:
        _assert_session_not_locked(conn, session_id)
        a = conn.execute(
            "SELECT * FROM assignments WHERE id=? AND session_id=?",
            (payload.assignment_id, session_id),
        ).fetchone()
        if a is None:
            raise HTTPException(404, "Assignment not found")
        if a['status'] == 'pending_approval':
            raise HTTPException(400, "Assignment is awaiting manager approval")
        if a['status'] != 'pending':
            raise HTTPException(400, f"Assignment already {a['status']}")

        now = datetime.utcnow().isoformat()

        # Two-person approval gate: confirmations by ops go to pending_approval.
        if payload.action == 'confirm' and REQUIRE_APPROVAL:
            conn.execute(
                "UPDATE assignments SET status='pending_approval', decided_by=?, decided_at=? WHERE id=?",
                (username, now, payload.assignment_id),
            )
            conn.execute(
                "INSERT INTO approval_requests (assignment_id, requested_by, requested_at) "
                "VALUES (?, ?, ?)",
                (payload.assignment_id, username, now),
            )
            conn.execute(
                "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
                "VALUES (?, ?, ?, ?, ?)",
                (session_id, 'decision_pending_approval', username, now,
                 json.dumps({
                     "assignment_id": payload.assignment_id,
                     "swift_row": a['swift_row'],
                     "flex_row": a['flex_row'],
                     "tier": a['tier'],
                 })),
            )
            conn.commit()
            return {"assignment_id": payload.assignment_id, "status": "pending_approval"}

        new_status = 'confirmed' if payload.action == 'confirm' else 'rejected'
        conn.execute(
            "UPDATE assignments SET status=?, decided_by=?, decided_at=? WHERE id=?",
            (new_status, username, now, payload.assignment_id),
        )
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (?, ?, ?, ?, ?)",
            (session_id, f"decision_{payload.action}", username, now,
             json.dumps({
                 "assignment_id": payload.assignment_id,
                 "swift_row": a['swift_row'],
                 "flex_row": a['flex_row'],
                 "tier": a['tier'],
             })),
        )
        conn.commit()
        return {"assignment_id": payload.assignment_id, "status": new_status}
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# POST /sessions/{id}/assignments/{id}/approve  (two-person approval gate)
# GET  /sessions/{id}/pending-approvals
# ---------------------------------------------------------------------------

class ApprovePayload(BaseModel):
    action: str   # 'approved' | 'rejected'
    note: str = ''


@app.post("/sessions/{session_id}/assignments/{assignment_id}/approve")
def approve_decision(
    session_id: int,
    assignment_id: int,
    payload: ApprovePayload,
    user: dict = Depends(require_role('admin', 'internal_control')),
):
    if payload.action not in ('approved', 'rejected'):
        raise HTTPException(400, "action must be 'approved' or 'rejected'")

    username = user['username']
    conn = get_conn()
    try:
        _assert_session_not_locked(conn, session_id)
        a = conn.execute(
            "SELECT * FROM assignments WHERE id=? AND session_id=?",
            (assignment_id, session_id),
        ).fetchone()
        if a is None:
            raise HTTPException(404, "Assignment not found")
        if a['status'] != 'pending_approval':
            raise HTTPException(400, f"Assignment is not pending approval (status: {a['status']})")

        # Self-approval guard.
        ar = conn.execute(
            "SELECT requested_by FROM approval_requests WHERE assignment_id=? AND reviewed_by IS NULL",
            (assignment_id,),
        ).fetchone()
        if ar and ar['requested_by'] == username:
            raise HTTPException(403, "You cannot approve your own decisions")

        now = datetime.utcnow().isoformat()
        new_status = 'confirmed' if payload.action == 'approved' else 'rejected'

        conn.execute(
            "UPDATE assignments SET status=?, decided_by=?, decided_at=? WHERE id=?",
            (new_status, username, now, assignment_id),
        )
        conn.execute(
            "UPDATE approval_requests SET reviewed_by=?, reviewed_at=?, action=? "
            "WHERE assignment_id=? AND reviewed_by IS NULL",
            (username, now, payload.action, assignment_id),
        )
        audit_action = 'decision_approved' if payload.action == 'approved' else 'decision_rejected_approval'
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (?, ?, ?, ?, ?)",
            (session_id, audit_action, username, now,
             json.dumps({
                 "assignment_id": assignment_id,
                 "swift_row": a['swift_row'],
                 "flex_row": a['flex_row'],
                 "tier": a['tier'],
                 "note": payload.note,
             })),
        )
        conn.commit()
        return {"assignment_id": assignment_id, "status": new_status}
    finally:
        conn.close()


@app.get("/sessions/{session_id}/pending-approvals")
def list_pending_approvals(
    session_id: int,
    user: dict = Depends(require_role('admin', 'internal_control', 'ops')),
):
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT a.*, ar.requested_by, ar.requested_at "
            "FROM assignments a "
            "JOIN approval_requests ar ON ar.assignment_id = a.id "
            "WHERE a.session_id=? AND a.status='pending_approval' AND ar.reviewed_by IS NULL",
            (session_id,),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# POST /sessions/{id}/close
# ---------------------------------------------------------------------------

@app.post("/sessions/{session_id}/close")
def close_session_endpoint(session_id: int,
                           user: dict = Depends(require_role('ops', 'admin'))):
    """Mark a session closed and seed unmatched rows into open_items with
    auto-grouping applied. Safe to call on an already-closed session (no-op).
    Triggered by the 'Close session' button on the session review page and
    by the nightly cron (daily_close.py)."""
    from open_items import close_session
    conn = get_conn()
    try:
        try:
            result = close_session(conn, session_id, user['username'])
        except ValueError as exc:
            raise HTTPException(400, str(exc))
        conn.commit()
        return result
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# GET /sessions/{id}/export
# ---------------------------------------------------------------------------

@app.get("/sessions/{session_id}/export")
def export_session(session_id: int, user: dict = Depends(current_user)):
    """Write an xlsx based on the CURRENT CONFIRMED state. Only assignments
    with status='confirmed' count as matched; pending/rejected stay unmatched."""
    conn = get_conn()
    try:
        sess = conn.execute("SELECT * FROM sessions WHERE id=?", (session_id,)).fetchone()
        if sess is None:
            raise HTTPException(404, "Session not found")

        swift_rows = conn.execute(
            "SELECT * FROM swift_txns WHERE session_id=? ORDER BY row_number",
            (session_id,),
        ).fetchall()
        flex_rows = conn.execute(
            "SELECT * FROM flex_txns WHERE session_id=? ORDER BY row_number",
            (session_id,),
        ).fetchall()
        confirmed = conn.execute(
            "SELECT * FROM assignments WHERE session_id=? AND status='confirmed'",
            (session_id,),
        ).fetchall()

        swift_txns = [_swift_row_to_dict(r) for r in swift_rows]
        flex_txns = [_flex_row_to_dict(r) for r in flex_rows]
        swift_by_row = {s['_row_number']: s for s in swift_txns}
        flex_by_row = {f['_row_number']: f for f in flex_txns}

        matches = []
        for a in confirmed:
            s = swift_by_row.get(a['swift_row'])
            f = flex_by_row.get(a['flex_row'])
            if s is None or f is None:
                continue  # shouldn't happen; skip defensively
            s['_used'] = True
            f['_used'] = True
            matches.append({'swift': s, 'flex': f, 'tier': a['tier'], 'reason': a['reason']})

        out_name = f"session_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        out_path = EXPORT_DIR / out_name
        write_report(
            matches, swift_txns, flex_txns,
            Path(sess['swift_filename']), Path(sess['flex_filename']), out_path,
        )

        now = datetime.utcnow().isoformat()
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (?, 'export', ?, ?, ?)",
            (session_id, user['username'], now, json.dumps({"confirmed_count": len(matches)})),
        )
        conn.commit()

        return FileResponse(
            out_path, filename=out_name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# GET /sessions/{id}/audit
# ---------------------------------------------------------------------------

@app.get("/sessions/{session_id}/audit")
def get_audit(session_id: int,
              user: dict = Depends(current_user),
              scope: Optional[List[str]] = Depends(active_scope)):
    conn = get_conn()
    try:
        scope_where, scope_params = _scope_clause(scope, alias='a')
        owned = conn.execute(
            "SELECT s.id FROM sessions s LEFT JOIN accounts a ON a.id = s.account_id "
            "WHERE s.id=? AND (s.account_id IS NULL OR " + scope_where + ")",
            (session_id, *scope_params),
        ).fetchone()
        if owned is None:
            raise HTTPException(404, "Session not found")
        rows = conn.execute(
            "SELECT id, action, actor, timestamp, details FROM audit_log "
            "WHERE session_id=? ORDER BY id",
            (session_id,),
        ).fetchall()
        return [{**dict(r), 'details': json.loads(r['details']) if r['details'] else None}
                for r in rows]
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Accounts registry
# ---------------------------------------------------------------------------

ACCOUNT_TYPES = ('cash_nostro', 'mobile_wallet')
MOBILE_MONEY_PROVIDERS = ('mpesa', 'mtn_momo', 'airtel_money', 'orange_money', 'tigo_pesa', 'other')


class AccountPayload(BaseModel):
    label: str
    shortname: Optional[str] = None
    access_area: Optional[str] = None
    bic: Optional[str] = None
    swift_account: str
    flex_ac_no: str
    currency: str
    notes: Optional[str] = None
    # Mobile-money expansion. account_type defaults to 'cash_nostro'
    # which preserves the original behaviour. provider / msisdn /
    # short_code are only meaningful when account_type='mobile_wallet'.
    account_type: Optional[str] = 'cash_nostro'
    provider: Optional[str] = None
    msisdn: Optional[str] = None
    short_code: Optional[str] = None


@app.get("/accounts")
def list_accounts(account_type: Optional[str] = None,
                  provider: Optional[str] = None,
                  scope: Optional[List[str]] = Depends(active_scope)):
    """Cash accounts + mobile-money wallets. The /mobile-money page
    passes account_type='mobile_wallet' to scope to wallets only;
    the legacy nostro view passes 'cash_nostro' (or omits, which
    returns everything for backwards compatibility)."""
    if account_type is not None and account_type not in ACCOUNT_TYPES:
        raise HTTPException(400, f"account_type must be one of {ACCOUNT_TYPES}")
    if provider is not None and provider not in MOBILE_MONEY_PROVIDERS:
        raise HTTPException(400, f"provider must be one of {MOBILE_MONEY_PROVIDERS}")
    conn = get_conn()
    try:
        where, params = _scope_clause(scope, alias='accounts')
        type_clause = ""
        prov_clause = ""
        extra_params: list = []
        if account_type is not None:
            type_clause = "AND account_type = ?"
            extra_params.append(account_type)
        if provider is not None:
            prov_clause = "AND provider = ?"
            extra_params.append(provider)
        rows = conn.execute(
            f"SELECT * FROM accounts WHERE active=1 AND {where} "
            f"{type_clause} {prov_clause} "
            f"ORDER BY COALESCE(access_area, 'zzz'), COALESCE(shortname, label)",
            (*params, *extra_params),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.get("/access-areas")
def list_access_areas(user: dict = Depends(current_user)):
    """Used to populate the Access area dropdown. Ops and admin both see this
    so account creation/edit screens can bind to the same registry."""
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT id, name, parent, active FROM access_areas "
            "WHERE active=1 ORDER BY name"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Currencies registry — managed ISO-code list feeds the currency dropdown on
# the cash-account forms. Matches the Corona 7.9 picker so ops reuse the same
# canonical set of codes.
# ---------------------------------------------------------------------------

class CurrencyPayload(BaseModel):
    iso_code: str
    name: str
    decimals: int = 2
    euro_currency: int = 0


class CurrencyPatchPayload(BaseModel):
    name: Optional[str] = None
    decimals: Optional[int] = None
    euro_currency: Optional[int] = None
    active: Optional[int] = None


@app.get("/currencies")
def list_currencies(include_inactive: bool = False,
                    user: dict = Depends(current_user)):
    """All currencies, active-only by default. Fed into the account-form
    dropdown and the admin management page."""
    conn = get_conn()
    try:
        where = "" if include_inactive else "WHERE active=1"
        rows = conn.execute(
            f"SELECT iso_code, name, decimals, euro_currency, active, "
            f"created_at, created_by FROM currencies {where} "
            f"ORDER BY iso_code"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.post("/currencies")
def create_currency(payload: CurrencyPayload,
                    user: dict = Depends(require_role('admin'))):
    code = payload.iso_code.strip().upper()
    name = payload.name.strip()
    if len(code) != 3 or not code.isalpha():
        raise HTTPException(400, "iso_code must be exactly 3 letters.")
    if not name:
        raise HTTPException(400, "name is required.")
    conn = get_conn()
    try:
        now = datetime.utcnow().isoformat()
        try:
            conn.execute(
                "INSERT INTO currencies (iso_code, name, decimals, euro_currency, "
                "active, created_at, created_by) VALUES (?,?,?,?,1,?,?)",
                (code, name, payload.decimals, payload.euro_currency,
                 now, user['username']),
            )
        except Exception as exc:
            if 'UNIQUE' in str(exc) or 'PRIMARY KEY' in str(exc):
                raise HTTPException(409, f"Currency {code} already exists.")
            raise
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'currency_created', ?, ?, ?)",
            (user['username'], now, json.dumps({"iso_code": code, "name": name})),
        )
        conn.commit()
        return {"iso_code": code, "name": name}
    finally:
        conn.close()


@app.patch("/currencies/{iso_code}")
def update_currency(iso_code: str, payload: CurrencyPatchPayload,
                    user: dict = Depends(require_role('admin'))):
    code = iso_code.strip().upper()
    conn = get_conn()
    try:
        existing = conn.execute(
            "SELECT * FROM currencies WHERE iso_code=?", (code,)
        ).fetchone()
        if existing is None:
            raise HTTPException(404, f"Currency {code} not found.")
        fields, values = [], []
        for k in ('name', 'decimals', 'euro_currency', 'active'):
            v = getattr(payload, k)
            if v is not None:
                fields.append(f"{k}=?")
                values.append(v)
        if not fields:
            raise HTTPException(400, "No fields to update.")
        values.append(code)
        conn.execute(
            f"UPDATE currencies SET {', '.join(fields)} WHERE iso_code=?", values,
        )
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'currency_updated', ?, ?, ?)",
            (user['username'], datetime.utcnow().isoformat(),
             json.dumps({"iso_code": code,
                         "changes": {k: getattr(payload, k)
                                     for k in ('name', 'decimals', 'euro_currency', 'active')
                                     if getattr(payload, k) is not None}})),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM currencies WHERE iso_code=?", (code,)).fetchone()
        return dict(row)
    finally:
        conn.close()


@app.delete("/currencies/{iso_code}")
def delete_currency(iso_code: str,
                    user: dict = Depends(require_role('admin'))):
    """Soft-delete: sets active=0 so existing accounts keep referencing it but
    it drops out of the picker. Hard delete refused if any account uses it."""
    code = iso_code.strip().upper()
    conn = get_conn()
    try:
        existing = conn.execute(
            "SELECT * FROM currencies WHERE iso_code=?", (code,)
        ).fetchone()
        if existing is None:
            raise HTTPException(404, f"Currency {code} not found.")
        conn.execute("UPDATE currencies SET active=0 WHERE iso_code=?", (code,))
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'currency_deactivated', ?, ?, ?)",
            (user['username'], datetime.utcnow().isoformat(),
             json.dumps({"iso_code": code})),
        )
        conn.commit()
        return {"iso_code": code, "active": 0}
    finally:
        conn.close()


@app.get("/currencies-admin")
def currencies_admin_page(request: Request):
    return templates.TemplateResponse(request, "currencies_admin.html")


# ---------------------------------------------------------------------------
# Banks registry — managed counterparty list that feeds the BIC dropdown on
# the cash-account form. Strict: accounts can only reference a registered
# bank. Matches the Corona 7.9 Banks screen.
# ---------------------------------------------------------------------------

class BankPayload(BaseModel):
    bic: str
    name: str
    nickname: Optional[str] = None
    origin: str = 'their'           # 'their' | 'our'
    type: str = 'bank'              # 'bank' | 'broker'
    access_area: Optional[str] = None
    user_code: Optional[str] = None


class BankPatchPayload(BaseModel):
    name: Optional[str] = None
    nickname: Optional[str] = None
    origin: Optional[str] = None
    type: Optional[str] = None
    access_area: Optional[str] = None
    user_code: Optional[str] = None
    active: Optional[int] = None


@app.get("/banks")
def list_banks(include_inactive: bool = False,
               user: dict = Depends(current_user)):
    """Registered counterparty banks. Active-only by default. Fed into the
    BIC dropdown on the cash-account form and the admin management page."""
    conn = get_conn()
    try:
        where = "" if include_inactive else "WHERE active=1"
        rows = conn.execute(
            f"SELECT bic, name, nickname, origin, type, access_area, user_code, "
            f"active, created_at, created_by FROM banks {where} "
            f"ORDER BY bic"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.post("/banks")
def create_bank(payload: BankPayload,
                user: dict = Depends(require_role('admin'))):
    bic = payload.bic.strip().upper()
    name = payload.name.strip()
    if not (8 <= len(bic) <= 11) or not bic.isalnum():
        raise HTTPException(400, "bic must be 8–11 alphanumeric characters.")
    if not name:
        raise HTTPException(400, "name is required.")
    origin = (payload.origin or 'their').strip().lower()
    if origin not in ('their', 'our'):
        raise HTTPException(400, "origin must be 'their' or 'our'.")
    btype = (payload.type or 'bank').strip().lower()
    if btype not in ('bank', 'broker'):
        raise HTTPException(400, "type must be 'bank' or 'broker'.")
    conn = get_conn()
    try:
        now = datetime.utcnow().isoformat()
        try:
            conn.execute(
                "INSERT INTO banks (bic, name, nickname, origin, type, access_area, "
                "user_code, active, created_at, created_by) "
                "VALUES (?,?,?,?,?,?,?,1,?,?)",
                (bic, name, payload.nickname, origin, btype,
                 payload.access_area, payload.user_code,
                 now, user['username']),
            )
        except Exception as exc:
            if 'UNIQUE' in str(exc) or 'PRIMARY KEY' in str(exc):
                raise HTTPException(409, f"Bank {bic} already registered.")
            raise
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'bank_created', ?, ?, ?)",
            (user['username'], now, json.dumps({"bic": bic, "name": name})),
        )
        conn.commit()
        return {"bic": bic, "name": name}
    finally:
        conn.close()


@app.patch("/banks/{bic}")
def update_bank(bic: str, payload: BankPatchPayload,
                user: dict = Depends(require_role('admin'))):
    code = bic.strip().upper()
    conn = get_conn()
    try:
        existing = conn.execute(
            "SELECT * FROM banks WHERE bic=?", (code,)
        ).fetchone()
        if existing is None:
            raise HTTPException(404, f"Bank {code} not found.")
        if payload.origin is not None and payload.origin not in ('their', 'our'):
            raise HTTPException(400, "origin must be 'their' or 'our'.")
        if payload.type is not None and payload.type not in ('bank', 'broker'):
            raise HTTPException(400, "type must be 'bank' or 'broker'.")
        fields, values = [], []
        for k in ('name', 'nickname', 'origin', 'type', 'access_area',
                  'user_code', 'active'):
            v = getattr(payload, k)
            if v is not None:
                fields.append(f"{k}=?")
                values.append(v)
        if not fields:
            raise HTTPException(400, "No fields to update.")
        values.append(code)
        conn.execute(
            f"UPDATE banks SET {', '.join(fields)} WHERE bic=?", values,
        )
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'bank_updated', ?, ?, ?)",
            (user['username'], datetime.utcnow().isoformat(),
             json.dumps({"bic": code,
                         "changes": {k: getattr(payload, k)
                                     for k in ('name', 'nickname', 'origin', 'type',
                                               'access_area', 'user_code', 'active')
                                     if getattr(payload, k) is not None}})),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM banks WHERE bic=?", (code,)).fetchone()
        return dict(row)
    finally:
        conn.close()


@app.delete("/banks/{bic}")
def delete_bank(bic: str,
                user: dict = Depends(require_role('admin'))):
    """Soft-delete: sets active=0 so existing accounts keep their BIC but the
    bank drops out of the picker. Refused if any active account still uses it —
    strict registry means a BIC can't be retired while in use."""
    code = bic.strip().upper()
    conn = get_conn()
    try:
        existing = conn.execute(
            "SELECT * FROM banks WHERE bic=?", (code,)
        ).fetchone()
        if existing is None:
            raise HTTPException(404, f"Bank {code} not found.")
        in_use = conn.execute(
            "SELECT COUNT(*) FROM accounts WHERE bic=? AND active=1", (code,)
        ).fetchone()[0]
        if in_use:
            raise HTTPException(
                409,
                f"Bank {code} is referenced by {in_use} active cash account(s). "
                "Reassign or deactivate those accounts first."
            )
        conn.execute("UPDATE banks SET active=0 WHERE bic=?", (code,))
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'bank_deactivated', ?, ?, ?)",
            (user['username'], datetime.utcnow().isoformat(),
             json.dumps({"bic": code})),
        )
        conn.commit()
        return {"bic": code, "active": 0}
    finally:
        conn.close()


@app.get("/banks-admin")
def banks_admin_page(request: Request):
    return templates.TemplateResponse(request, "banks_admin.html")


# ---------------------------------------------------------------------------
# FX rate registry — feeds the cross-currency branch of the match engine.
# Amount-in-from-currency * rate = amount-in-to-currency. Identity rows
# (GHS->GHS=1.0) are seeded automatically; ops manages the rest.
# ---------------------------------------------------------------------------

class FxRatePayload(BaseModel):
    from_ccy: str
    to_ccy: str
    rate: float
    valid_from: Optional[str] = None    # ISO date; defaults to today
    source: Optional[str] = None


class FxRatePatchPayload(BaseModel):
    rate: Optional[float] = None
    source: Optional[str] = None
    active: Optional[int] = None


@app.get("/fx-rates")
def list_fx_rates(include_inactive: bool = False,
                  user: dict = Depends(current_user)):
    """All FX rate rows (latest per pair by default). When include_inactive
    is true, historic / deactivated rows are returned too for audit."""
    conn = get_conn()
    try:
        where = "" if include_inactive else "WHERE active=1"
        rows = conn.execute(
            f"SELECT id, from_ccy, to_ccy, rate, valid_from, source, active, "
            f"created_at, created_by FROM fx_rates {where} "
            f"ORDER BY from_ccy, to_ccy, valid_from DESC"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.post("/fx-rates")
def create_fx_rate(payload: FxRatePayload,
                   user: dict = Depends(require_role('admin'))):
    from datetime import date as _date
    frm = payload.from_ccy.strip().upper()
    to  = payload.to_ccy.strip().upper()
    if len(frm) != 3 or len(to) != 3 or not frm.isalpha() or not to.isalpha():
        raise HTTPException(400, "from_ccy and to_ccy must be 3-letter codes")
    if payload.rate <= 0:
        raise HTTPException(400, "rate must be > 0")
    valid_from = payload.valid_from or _date.today().isoformat()
    conn = get_conn()
    try:
        # Deactivate any previous active row for the same pair so the engine
        # always sees a single current rate per direction.
        conn.execute(
            "UPDATE fx_rates SET active=0 WHERE from_ccy=? AND to_ccy=? AND active=1",
            (frm, to),
        )
        now = datetime.utcnow().isoformat()
        try:
            cur = conn.execute(
                "INSERT INTO fx_rates (from_ccy, to_ccy, rate, valid_from, source, "
                "active, created_at, created_by) VALUES (?,?,?,?,?,1,?,?)",
                (frm, to, payload.rate, valid_from, payload.source, now, user['username']),
            )
        except Exception as exc:
            if 'UNIQUE' in str(exc):
                raise HTTPException(409, "That rate already exists on that date.")
            raise
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'fx_rate_created', ?, ?, ?)",
            (user['username'], now, json.dumps({
                "from": frm, "to": to, "rate": payload.rate, "valid_from": valid_from})),
        )
        conn.commit()
        return {"id": cur.lastrowid, "from_ccy": frm, "to_ccy": to, "rate": payload.rate}
    finally:
        conn.close()


@app.patch("/fx-rates/{rate_id}")
def update_fx_rate(rate_id: int, payload: FxRatePatchPayload,
                   user: dict = Depends(require_role('admin'))):
    conn = get_conn()
    try:
        existing = conn.execute(
            "SELECT * FROM fx_rates WHERE id=?", (rate_id,)
        ).fetchone()
        if existing is None:
            raise HTTPException(404, "rate not found")
        fields, values = [], []
        for k in ('rate', 'source', 'active'):
            v = getattr(payload, k)
            if v is not None:
                fields.append(f"{k}=?"); values.append(v)
        if not fields:
            raise HTTPException(400, "no fields to update")
        values.append(rate_id)
        conn.execute(f"UPDATE fx_rates SET {', '.join(fields)} WHERE id=?", values)
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'fx_rate_updated', ?, ?, ?)",
            (user['username'], datetime.utcnow().isoformat(),
             json.dumps({"rate_id": rate_id,
                         "changes": {k: getattr(payload, k)
                                     for k in ('rate','source','active')
                                     if getattr(payload, k) is not None}})),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM fx_rates WHERE id=?", (rate_id,)).fetchone()
        return dict(row)
    finally:
        conn.close()


@app.get("/fx-admin")
def fx_admin_page(request: Request):
    return templates.TemplateResponse(request, "fx_admin.html")


# ---------------------------------------------------------------------------
# Reconciliation certificates — month-end sign-off artefact per account.
# Draft → prepared → reviewed → signed flow with four-eyes enforcement via
# role checks (ops prepares, internal_control reviews, admin signs). On
# sign-off the live figures are frozen into snapshot_json so the PDF/xlsx
# always shows the numbers at signing time, not the current ledger state.
# ---------------------------------------------------------------------------

class CertActionPayload(BaseModel):
    note: Optional[str] = None


@app.get("/accounts/{account_id}/certificates")
def list_certificates(account_id: int,
                      user: dict = Depends(current_user),
                      scope: Optional[List[str]] = Depends(active_scope)):
    conn = get_conn()
    try:
        _assert_account_in_scope(conn, account_id, scope)
        rows = conn.execute(
            "SELECT * FROM reconciliation_certificates WHERE account_id=? "
            "ORDER BY period_end DESC, id DESC",
            (account_id,),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.post("/accounts/{account_id}/certificates")
def create_certificate(account_id: int, period_start: str, period_end: str,
                       user: dict = Depends(require_role('admin', 'ops')),
                       scope: Optional[List[str]] = Depends(active_scope)):
    """Generate a draft certificate for the account+period. Idempotent per
    (account, period_start, period_end) — re-issuing returns the existing
    draft so analysts can iterate without spawning duplicates."""
    conn = get_conn()
    try:
        _assert_account_in_scope(conn, account_id, scope)
        existing = conn.execute(
            "SELECT * FROM reconciliation_certificates WHERE account_id=? "
            "AND period_start=? AND period_end=? AND status != 'superseded' "
            "ORDER BY id DESC LIMIT 1",
            (account_id, period_start, period_end),
        ).fetchone()
        if existing:
            return dict(existing)
        now = datetime.utcnow().isoformat()
        cur = conn.execute(
            "INSERT INTO reconciliation_certificates "
            "(account_id, period_start, period_end, generated_at, generated_by, status) "
            "VALUES (?,?,?,?,?, 'draft')",
            (account_id, period_start, period_end, now, user['username']),
        )
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'certificate_created', ?, ?, ?)",
            (user['username'], now, json.dumps({
                'certificate_id': cur.lastrowid, 'account_id': account_id,
                'period': f"{period_start}..{period_end}"})),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM reconciliation_certificates WHERE id=?",
                           (cur.lastrowid,)).fetchone()
        return dict(row)
    finally:
        conn.close()


def _transition_cert(conn, cert_id: int, expected_statuses: tuple,
                     new_status: str, field_user: str, field_at: str,
                     user: str, note: Optional[str]) -> dict:
    row = conn.execute(
        "SELECT * FROM reconciliation_certificates WHERE id=?", (cert_id,),
    ).fetchone()
    if row is None:
        raise HTTPException(404, "certificate not found")
    if row['status'] not in expected_statuses:
        raise HTTPException(
            409,
            f"certificate is {row['status']}; cannot transition to {new_status}")
    if row['status'] == 'signed':
        raise HTTPException(409, "certificate is already signed and immutable")
    now = datetime.utcnow().isoformat()
    conn.execute(
        f"UPDATE reconciliation_certificates SET status=?, "
        f"{field_user}=?, {field_at}=? WHERE id=?",
        (new_status, user, now, cert_id),
    )
    conn.execute(
        "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
        "VALUES (NULL, 'certificate_' || ?, ?, ?, ?)",
        (new_status, user, now, json.dumps({
            'certificate_id': cert_id, 'from': row['status'],
            'to': new_status, 'note': note})),
    )
    conn.commit()
    return dict(conn.execute(
        "SELECT * FROM reconciliation_certificates WHERE id=?", (cert_id,)).fetchone())


@app.post("/certificates/{cert_id}/prepare")
def prepare_certificate(cert_id: int, payload: CertActionPayload,
                        user: dict = Depends(require_role('admin', 'ops')),
                        scope: Optional[List[str]] = Depends(active_scope)):
    """Maker step — analyst attests the figures are ready for review."""
    conn = get_conn()
    try:
        _assert_account_in_scope(conn, _account_id_for_certificate(conn, cert_id), scope)
        return _transition_cert(conn, cert_id, ('draft',), 'prepared',
                                 'prepared_by', 'prepared_at',
                                 user['username'], payload.note)
    finally:
        conn.close()


@app.post("/certificates/{cert_id}/review")
def review_certificate(cert_id: int, payload: CertActionPayload,
                       user: dict = Depends(require_role('admin', 'internal_control')),
                       scope: Optional[List[str]] = Depends(active_scope)):
    """Checker step — independent reviewer attests figures are correct."""
    conn = get_conn()
    try:
        _assert_account_in_scope(conn, _account_id_for_certificate(conn, cert_id), scope)
        return _transition_cert(conn, cert_id, ('prepared',), 'reviewed',
                                 'reviewed_by', 'reviewed_at',
                                 user['username'], payload.note)
    finally:
        conn.close()


@app.post("/certificates/{cert_id}/sign")
def sign_certificate(cert_id: int, payload: CertActionPayload,
                     user: dict = Depends(require_role('admin')),
                     scope: Optional[List[str]] = Depends(active_scope)):
    """Approver step — freezes the figures into snapshot_json and makes the
    certificate immutable. At this point the xlsx becomes reproducible
    from snapshot_json alone, so later ledger changes don't rewrite it."""
    from certificates import compute_figures
    conn = get_conn()
    try:
        _assert_account_in_scope(conn, _account_id_for_certificate(conn, cert_id), scope)
        row = conn.execute(
            "SELECT * FROM reconciliation_certificates WHERE id=?", (cert_id,),
        ).fetchone()
        if row is None:
            raise HTTPException(404, "certificate not found")
        if row['status'] != 'reviewed':
            raise HTTPException(
                409,
                f"certificate is {row['status']}; must be reviewed before signing")
        figures = compute_figures(conn, row['account_id'],
                                   row['period_start'], row['period_end'])
        now = datetime.utcnow().isoformat()
        conn.execute(
            "UPDATE reconciliation_certificates SET status='signed', "
            "signed_by=?, signed_at=?, snapshot_json=? WHERE id=?",
            (user['username'], now, json.dumps(figures, default=str), cert_id),
        )
        # Lock all closed sessions for this account within the certificate period.
        # Locked sessions reject further decisions, preserving the frozen figures.
        conn.execute(
            "UPDATE sessions SET locked_at=?, locked_by=? "
            "WHERE account_id=? AND status='closed' "
            "AND created_at >= ? AND created_at <= ?",
            (now, user['username'], row['account_id'],
             row['period_start'], row['period_end'] + 'T23:59:59'),
        )
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'certificate_signed', ?, ?, ?)",
            (user['username'], now, json.dumps({
                'certificate_id': cert_id, 'note': payload.note,
                'frozen_figures': True})),
        )
        conn.commit()
        return dict(conn.execute(
            "SELECT * FROM reconciliation_certificates WHERE id=?", (cert_id,)
        ).fetchone())
    finally:
        conn.close()


@app.get("/certificates/{cert_id}/print")
def print_certificate(cert_id: int, request: Request,
                      user: dict = Depends(current_user),
                      scope: Optional[List[str]] = Depends(active_scope)):
    """Render a print-ready HTML view of the certificate. Browser's native
    Save-as-PDF dialog produces the final PDF — zero external PDF dependency
    (no reportlab, no weasyprint, no wkhtmltopdf install). When IT provisions
    reportlab this same endpoint can stream binary PDF; the template is
    designed to be one-to-one convertible.

    Auth deps work because `current_user` and `active_scope` both fall
    back to the `kilter_token` cookie when the X-Session-Token header
    is absent (browser navigation case for `<a target="_blank">`). The
    same JS that writes localStorage also writes the cookie."""
    from certificates import compute_figures
    conn = get_conn()
    try:
        _assert_account_in_scope(conn, _account_id_for_certificate(conn, cert_id), scope)
        row = conn.execute(
            "SELECT * FROM reconciliation_certificates WHERE id=?", (cert_id,),
        ).fetchone()
        if row is None:
            raise HTTPException(404, "certificate not found")
        if row['status'] == 'signed' and row['snapshot_json']:
            figures = json.loads(row['snapshot_json'])
        else:
            figures = compute_figures(conn, row['account_id'],
                                       row['period_start'], row['period_end'])
        return templates.TemplateResponse(request, "certificate_print.html", {
            'cert': dict(row),
            'figures': figures,
        })
    finally:
        conn.close()


@app.get("/certificates/{cert_id}/download")
def download_certificate(cert_id: int,
                         user: dict = Depends(current_user),
                         scope: Optional[List[str]] = Depends(active_scope)):
    """Reproducible xlsx. For signed certs the frozen snapshot drives the
    figures; for unsigned certs the live ledger does — so unsigned xlsx
    can shift if more sessions land, but a signed cert is forever."""
    from certificates import compute_figures, build_xlsx
    conn = get_conn()
    try:
        _assert_account_in_scope(conn, _account_id_for_certificate(conn, cert_id), scope)
        row = conn.execute(
            "SELECT * FROM reconciliation_certificates WHERE id=?", (cert_id,),
        ).fetchone()
        if row is None:
            raise HTTPException(404, "certificate not found")
        if row['status'] == 'signed' and row['snapshot_json']:
            figures = json.loads(row['snapshot_json'])
        else:
            figures = compute_figures(conn, row['account_id'],
                                       row['period_start'], row['period_end'])
        data = build_xlsx(figures, dict(row))
        acct_label = (figures.get('account') or {}).get('label', 'account')
        safe = ''.join(c for c in acct_label if c.isalnum() or c in ' -_').strip()
        filename = f"certificate_{safe}_{row['period_start']}_{row['period_end']}.xlsx"
        return StreamingResponse(
            io.BytesIO(data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    finally:
        conn.close()


@app.get("/certificates-admin")
def certificates_admin_page(request: Request):
    return templates.TemplateResponse(request, "certificates_admin.html")


# ---------------------------------------------------------------------------
# SLA alerting — notification_channels CRUD + on-demand check endpoint.
# Channel kinds: teams (webhook), email (stub — no SMTP infra decision yet),
# log (writes to audit_log). Each channel has its own threshold_days and
# optional access-area filter so different teams can subscribe to different
# urgency levels.
# ---------------------------------------------------------------------------

class ChannelPayload(BaseModel):
    name: str
    kind: str                       # 'teams' | 'email' | 'log'
    config_json: str                # JSON string with channel-specific config
    threshold_days: int = 30
    access_area_filter: Optional[str] = None   # JSON list or null
    active: int = 1


class ChannelPatchPayload(BaseModel):
    name: Optional[str] = None
    config_json: Optional[str] = None
    threshold_days: Optional[int] = None
    access_area_filter: Optional[str] = None
    active: Optional[int] = None


def _redact_channel_config(config_json: Optional[str]) -> Optional[str]:
    """Strip the smtp_password field for outbound responses. Admin UI never
    needs to see the stored password — they re-enter it to rotate. Returns
    the JSON string with smtp_password replaced by an empty string if any
    value is present, or untouched otherwise."""
    if not config_json:
        return config_json
    try:
        cfg = json.loads(config_json)
    except (TypeError, ValueError):
        return config_json
    if isinstance(cfg, dict) and cfg.get('smtp_password'):
        cfg['smtp_password'] = ''   # surface that one was set, but never echo it
    return json.dumps(cfg)


def _encrypt_channel_config(config_json: Optional[str]) -> Optional[str]:
    """Inverse-side: encrypt the smtp_password field before persisting.
    Caller is responsible for the JSON-validity round-trip; this is a
    best-effort no-op for non-email channels."""
    if not config_json:
        return config_json
    try:
        cfg = json.loads(config_json)
    except (TypeError, ValueError):
        return config_json
    pwd = cfg.get('smtp_password') if isinstance(cfg, dict) else None
    if pwd:
        from secrets_vault import encrypt, is_encrypted
        if not is_encrypted(pwd):
            cfg['smtp_password'] = encrypt(pwd)
            return json.dumps(cfg)
    return config_json


@app.get("/notification-channels")
def list_channels(user: dict = Depends(require_role('admin'))):
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT * FROM notification_channels ORDER BY active DESC, name"
        ).fetchall()
        # Redact the SMTP password — admins shouldn't see the encrypted blob
        # echoed back, and they don't need it to operate the UI.
        out = []
        for r in rows:
            d = dict(r)
            d['config_json'] = _redact_channel_config(d.get('config_json'))
            out.append(d)
        return out
    finally:
        conn.close()


@app.post("/notification-channels")
def create_channel(payload: ChannelPayload,
                   user: dict = Depends(require_role('admin'))):
    if payload.kind not in ('teams', 'email', 'log'):
        raise HTTPException(400, "kind must be 'teams', 'email' or 'log'")
    # Validate config_json is actually JSON so we surface typos early.
    try:
        json.loads(payload.config_json)
    except Exception:
        raise HTTPException(400, "config_json must be valid JSON")
    config_json_enc = _encrypt_channel_config(payload.config_json)
    now = datetime.utcnow().isoformat()
    conn = get_conn()
    try:
        cur = conn.execute(
            "INSERT INTO notification_channels (name, kind, config_json, "
            "threshold_days, access_area_filter, active, created_at, created_by) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (payload.name, payload.kind, config_json_enc,
             payload.threshold_days, payload.access_area_filter,
             payload.active, now, user['username']),
        )
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'channel_created', ?, ?, ?)",
            (user['username'], now, json.dumps({
                'channel_id': cur.lastrowid, 'name': payload.name,
                'kind': payload.kind})),
        )
        conn.commit()
        return {"id": cur.lastrowid, "name": payload.name}
    finally:
        conn.close()


@app.patch("/notification-channels/{channel_id}")
def update_channel(channel_id: int, payload: ChannelPatchPayload,
                   user: dict = Depends(require_role('admin'))):
    conn = get_conn()
    try:
        existing = conn.execute(
            "SELECT id FROM notification_channels WHERE id=?", (channel_id,)
        ).fetchone()
        if existing is None:
            raise HTTPException(404, "channel not found")
        fields, values = [], []
        for k in ('name', 'config_json', 'threshold_days',
                  'access_area_filter', 'active'):
            v = getattr(payload, k)
            if v is not None:
                if k == 'config_json':
                    try: json.loads(v)
                    except Exception:
                        raise HTTPException(400, "config_json must be valid JSON")
                    v = _encrypt_channel_config(v)
                fields.append(f"{k}=?"); values.append(v)
        if not fields:
            raise HTTPException(400, "no fields to update")
        values.append(channel_id)
        conn.execute(
            f"UPDATE notification_channels SET {', '.join(fields)} WHERE id=?",
            values,
        )
        conn.commit()
        out = dict(conn.execute(
            "SELECT * FROM notification_channels WHERE id=?", (channel_id,)
        ).fetchone())
        out['config_json'] = _redact_channel_config(out.get('config_json'))
        return out
    finally:
        conn.close()


@app.post("/sla/check")
def sla_check(channel_id: Optional[int] = None, dry_run: bool = False,
              user: dict = Depends(require_role('admin'))):
    """Evaluate active channels (or one specific channel) against the
    current open_items ledger and dispatch. dry_run=true runs the query
    and returns counts without posting anywhere."""
    from sla import run_check
    conn = get_conn()
    try:
        return run_check(conn, channel_id=channel_id, dry_run=dry_run)
    finally:
        conn.close()


@app.get("/sla-admin")
def sla_admin_page(request: Request):
    return templates.TemplateResponse(request, "sla_admin.html")


# ---------------------------------------------------------------------------
# Match-pattern analytics — mines historic user decisions to surface
# narration keywords and ref prefixes that correlate with confirmations.
# Read-only: the page shows the patterns so ops can create auto-rules
# manually. No auto-promotion — keeps the human in the loop for now.
# ---------------------------------------------------------------------------

@app.get("/analytics/match-patterns")
def match_patterns(min_occurrences: int = 3,
                   user: dict = Depends(current_user)):
    """For each narration token that appears in ≥min_occurrences confirmed
    assignments, compute the confirm/reject split. Tokens with ≥80% confirm
    rate are good auto-rule candidates. Ref prefixes are surfaced separately.

    Intentionally cheap — runs against all-time assignments, no session
    limit. With <10k assignments this is milliseconds; if the table grows
    to >100k we'll paginate by time window."""
    import re
    from collections import Counter, defaultdict

    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT a.status, a.tier, s.our_ref AS swift_ref, f.narration, "
            "f.external_ref, f.trn_ref, f.user_id, f.module "
            "FROM assignments a "
            "LEFT JOIN swift_txns s ON s.session_id=a.session_id AND s.row_number=a.swift_row "
            "LEFT JOIN flex_txns  f ON f.session_id=a.session_id AND f.row_number=a.flex_row "
            "WHERE a.status IN ('confirmed', 'rejected')"
        ).fetchall()

        token_confirm: Counter = Counter()
        token_reject: Counter = Counter()
        ref_prefix_confirm: Counter = Counter()
        ref_prefix_reject: Counter = Counter()
        user_confirm: Counter = Counter()
        user_reject: Counter = Counter()
        module_confirm: Counter = Counter()
        module_reject: Counter = Counter()

        token_re = re.compile(r'[A-Z][A-Z0-9]{3,}')  # ≥4-char uppercase tokens
        stopwords = {'THE', 'AND', 'FOR', 'FROM', 'WITH', 'INTO', 'ONTO',
                     'BANK', 'REF', 'OUR', 'THEIR', 'OUREF', 'DATE'}
        for r in rows:
            status = r['status']
            narration = (r['narration'] or '').upper()
            tokens = set(token_re.findall(narration)) - stopwords
            for tok in tokens:
                (token_confirm if status == 'confirmed' else token_reject)[tok] += 1

            trn = (r['trn_ref'] or '').upper()
            if len(trn) >= 4:
                prefix = trn[:4]
                (ref_prefix_confirm if status == 'confirmed' else ref_prefix_reject)[prefix] += 1

            if r['user_id']:
                (user_confirm if status == 'confirmed' else user_reject)[r['user_id']] += 1
            if r['module']:
                (module_confirm if status == 'confirmed' else module_reject)[r['module']] += 1

        def build(confirm: Counter, reject: Counter) -> list[dict]:
            out = []
            for k in set(confirm) | set(reject):
                c = confirm[k]; rej = reject[k]
                total = c + rej
                if total < min_occurrences:
                    continue
                out.append({
                    'key': k, 'confirmed': c, 'rejected': rej,
                    'total': total,
                    'confirm_rate': round(100 * c / total, 1) if total else 0,
                })
            out.sort(key=lambda x: (-x['confirm_rate'], -x['total']))
            return out

        return {
            'total_decided': len(rows),
            'narration_tokens': build(token_confirm, token_reject)[:50],
            'ref_prefixes':     build(ref_prefix_confirm, ref_prefix_reject)[:30],
            'user_ids':         build(user_confirm, user_reject)[:30],
            'modules':          build(module_confirm, module_reject)[:30],
        }
    finally:
        conn.close()


@app.get("/match-patterns")
def match_patterns_page(request: Request):
    return templates.TemplateResponse(request, "match_patterns.html")


# ---------------------------------------------------------------------------
# Scheduled jobs — the "runs itself" backbone. CRUD + manual-run here; the
# execution loop lives in scheduler.py. is_running() on /stats lets the UI
# show the daemon's liveness.
# ---------------------------------------------------------------------------

class JobPayload(BaseModel):
    name: str
    job_type: str                           # scan|daily_close|sla_check|daily_breaks_report|flex_extract
    schedule_kind: str                      # 'interval' | 'daily_at'
    interval_minutes: Optional[int] = None
    daily_at_utc: Optional[str] = None
    params_json: Optional[str] = None
    enabled: int = 1


class JobPatchPayload(BaseModel):
    name: Optional[str] = None
    schedule_kind: Optional[str] = None
    interval_minutes: Optional[int] = None
    daily_at_utc: Optional[str] = None
    params_json: Optional[str] = None
    enabled: Optional[int] = None


@app.get("/scheduled-jobs")
def list_scheduled_jobs(user: dict = Depends(require_role('admin'))):
    from scheduler import is_running, compute_next_run
    conn = get_conn()
    try:
        rows = [dict(r) for r in conn.execute(
            "SELECT * FROM scheduled_jobs ORDER BY name").fetchall()]
        for j in rows:
            nxt = compute_next_run(j)
            j['next_run_at_computed'] = nxt.isoformat() if nxt else None
        return {'scheduler_running': is_running(), 'jobs': rows}
    finally:
        conn.close()


@app.get("/scheduled-jobs/{job_id}/runs")
def get_job_runs(job_id: int, limit: int = 20,
                 user: dict = Depends(require_role('admin'))):
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT id, started_at, ended_at, status, output, duration_ms "
            "FROM job_runs WHERE job_id=? ORDER BY id DESC LIMIT ?",
            (job_id, max(1, min(100, limit))),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.post("/scheduled-jobs")
def create_scheduled_job(payload: JobPayload,
                         user: dict = Depends(require_role('admin'))):
    from scheduler import JOBS
    if payload.job_type not in JOBS:
        raise HTTPException(400, f"unknown job_type; valid: {list(JOBS.keys())}")
    if payload.schedule_kind not in ('interval', 'daily_at'):
        raise HTTPException(400, "schedule_kind must be 'interval' or 'daily_at'")
    if payload.schedule_kind == 'interval' and not payload.interval_minutes:
        raise HTTPException(400, "interval schedule needs interval_minutes > 0")
    if payload.schedule_kind == 'daily_at' and not payload.daily_at_utc:
        raise HTTPException(400, "daily_at schedule needs daily_at_utc ('HH:MM')")
    now = datetime.utcnow().isoformat()
    conn = get_conn()
    try:
        try:
            cur = conn.execute(
                "INSERT INTO scheduled_jobs (name, job_type, schedule_kind, "
                "interval_minutes, daily_at_utc, params_json, enabled, "
                "created_at, created_by) VALUES (?,?,?,?,?,?,?,?,?)",
                (payload.name, payload.job_type, payload.schedule_kind,
                 payload.interval_minutes, payload.daily_at_utc,
                 payload.params_json, payload.enabled, now, user['username']),
            )
        except Exception as exc:
            if 'UNIQUE' in str(exc):
                raise HTTPException(409, f"Job named '{payload.name}' already exists.")
            raise
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'scheduled_job_created', ?, ?, ?)",
            (user['username'], now, json.dumps({
                'job_id': cur.lastrowid, 'name': payload.name,
                'type': payload.job_type})),
        )
        conn.commit()
        return {'id': cur.lastrowid, 'name': payload.name}
    finally:
        conn.close()


@app.patch("/scheduled-jobs/{job_id}")
def update_scheduled_job(job_id: int, payload: JobPatchPayload,
                         user: dict = Depends(require_role('admin'))):
    conn = get_conn()
    try:
        existing = conn.execute(
            "SELECT id FROM scheduled_jobs WHERE id=?", (job_id,)
        ).fetchone()
        if existing is None:
            raise HTTPException(404, "job not found")
        fields, values = [], []
        for k in ('name', 'schedule_kind', 'interval_minutes', 'daily_at_utc',
                  'params_json', 'enabled'):
            v = getattr(payload, k)
            if v is not None:
                fields.append(f"{k}=?"); values.append(v)
        if not fields:
            raise HTTPException(400, "no fields to update")
        values.append(job_id)
        conn.execute(
            f"UPDATE scheduled_jobs SET {', '.join(fields)} WHERE id=?", values)
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'scheduled_job_updated', ?, ?, ?)",
            (user['username'], datetime.utcnow().isoformat(),
             json.dumps({'job_id': job_id,
                         'changes': {k: getattr(payload, k)
                                     for k in ('name','schedule_kind','interval_minutes',
                                               'daily_at_utc','params_json','enabled')
                                     if getattr(payload, k) is not None}})),
        )
        conn.commit()
        return dict(conn.execute(
            "SELECT * FROM scheduled_jobs WHERE id=?", (job_id,)).fetchone())
    finally:
        conn.close()


@app.post("/scheduled-jobs/{job_id}/run")
def run_scheduled_job_now(job_id: int,
                          user: dict = Depends(require_role('admin'))):
    """One-shot manual invocation — runs on the request thread so the UI
    shows the result immediately. Same code path as the daemon uses."""
    from scheduler import run_now
    conn = get_conn()
    try:
        return run_now(conn, job_id, actor=user['username'])
    except ValueError as exc:
        raise HTTPException(404, str(exc))
    finally:
        conn.close()


@app.get("/scheduler-admin")
def scheduler_admin_page(request: Request):
    return templates.TemplateResponse(request, "scheduler_admin.html")


# ---------------------------------------------------------------------------
# Manuals — in-app documentation. Each manual is a print-ready HTML page;
# users save as PDF via the browser print dialog. Zero external PDF
# dependency; manuals stay version-synced with the running app because they
# live in the templates tree and deploy with every release.
# ---------------------------------------------------------------------------

_MANUALS = [
    ('overview', 'Product Overview & Quick Start',
     'What Kilter is, who uses it, and a 10-minute tour.',
     'All staff', 'overview'),
    ('setup',    'Setup & IT Manual',
     'Install, configure, back up, upgrade, and operate the service.',
     'IT / Infrastructure',  'setup'),
    ('admin',    'Admin Manual',
     'User management, registries, scheduler, tolerance rules, SLA alerts.',
     'Admin role',           'admin'),
    ('user',     'Ops User Manual',
     'Daily reconciliation workflow — scan, review, open items, certificates.',
     'Ops role',             'ops'),
    ('control',  'Internal Control Manual',
     'Maker / checker / approver workflow, certificate review, segregation of duties.',
     'Internal Control role','control'),
    ('audit',    'Audit Manual',
     'Activity log, break history, certificate chain of evidence, evidence export.',
     'Audit role',           'audit'),
    ('matching', 'Matching Engine Manual',
     'How the engine works — tiers, reference normalization, splits, FX, carry-forward, tuning.',
     'Admin / Analyst',      'matching'),
    ('training', 'Training Curriculum',
     '4-week onboarding curriculum for new ops hires — structured exercises, comprehension checks, supervisor sign-off.',
     'New hires + supervisors', 'ops'),
]


def _manual_context():
    from datetime import datetime
    return {
        'app_version': '1.0',
        'generated_at': datetime.utcnow().strftime('%Y-%m-%d'),
        'manuals': _MANUALS,
    }


@app.get("/manuals")
def manuals_index(request: Request):
    return templates.TemplateResponse(request, "manuals_index.html",
                                       _manual_context())


@app.get("/manuals/{slug}")
def manual_page(request: Request, slug: str):
    # Whitelist the slugs to prevent template traversal.
    valid = {m[0]: m for m in _MANUALS}
    if slug not in valid:
        raise HTTPException(404, f"manual '{slug}' not found")
    return templates.TemplateResponse(request, f"manual_{slug}.html",
                                       _manual_context())


class AccountPatchPayload(BaseModel):
    label: Optional[str] = None
    shortname: Optional[str] = None
    access_area: Optional[str] = None
    bic: Optional[str] = None
    notes: Optional[str] = None
    active: Optional[int] = None


@app.patch("/accounts/{account_id}")
def update_account(account_id: int, payload: AccountPatchPayload,
                   user: dict = Depends(require_role('admin', 'ops'))):
    """Admin and ops can reclassify an account (label, shortname, access area,
    BIC, notes, active). Identity fields (swift_account, flex_ac_no, currency)
    are immutable here — changing them would orphan historic sessions, so do
    it via a migration script if truly needed."""
    conn = get_conn()
    try:
        existing = conn.execute("SELECT * FROM accounts WHERE id=?", (account_id,)).fetchone()
        if existing is None:
            raise HTTPException(404, "Account not found")
        if payload.bic is not None:
            _require_registered_bic(conn, payload.bic)

        fields, values = [], []
        for k in ('label', 'shortname', 'access_area', 'bic', 'notes', 'active'):
            v = getattr(payload, k)
            if v is not None:
                fields.append(f"{k}=?")
                values.append(v)
        if not fields:
            raise HTTPException(400, "No fields to update")

        values.append(account_id)
        conn.execute(f"UPDATE accounts SET {', '.join(fields)} WHERE id=?", values)
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'account_updated', ?, ?, ?)",
            (user['username'], datetime.utcnow().isoformat(),
             json.dumps({"account_id": account_id,
                         "changes": {k: getattr(payload, k)
                                     for k in ('label','shortname','access_area','bic','notes','active')
                                     if getattr(payload, k) is not None}})),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM accounts WHERE id=?", (account_id,)).fetchone()
        return dict(row)
    finally:
        conn.close()


def _require_registered_bic(conn, bic: Optional[str]) -> None:
    """Strict BIC enforcement: if a cash account references a BIC, that BIC
    must exist in the banks registry as an active row. Blank/null BICs are
    allowed — not every GL has a correspondent (e.g. internal sub-ledgers)."""
    if not bic:
        return
    code = bic.strip().upper()
    row = conn.execute(
        "SELECT active FROM banks WHERE bic=?", (code,)
    ).fetchone()
    if row is None:
        raise HTTPException(
            400,
            f"BIC {code} is not registered. Add it under Admin → Banks first, "
            "or leave the BIC field blank."
        )
    if not row[0]:
        raise HTTPException(
            400,
            f"BIC {code} is deactivated in the banks registry. Reactivate it "
            "under Admin → Banks before assigning it to an account."
        )


@app.post("/accounts")
def create_account(payload: AccountPayload, user: dict = Depends(require_role('admin'))):
    username = user['username']
    account_type = (payload.account_type or 'cash_nostro').strip()
    if account_type not in ACCOUNT_TYPES:
        raise HTTPException(400, f"account_type must be one of {ACCOUNT_TYPES}")
    provider = (payload.provider or '').strip() or None
    msisdn = (payload.msisdn or '').strip() or None
    short_code = (payload.short_code or '').strip() or None

    if account_type == 'mobile_wallet':
        if provider is None:
            raise HTTPException(400,
                "Mobile wallet accounts must have a provider "
                f"({MOBILE_MONEY_PROVIDERS}).")
        if provider not in MOBILE_MONEY_PROVIDERS:
            raise HTTPException(400,
                f"provider must be one of {MOBILE_MONEY_PROVIDERS}")
        if msisdn is None and short_code is None:
            raise HTTPException(400,
                "Mobile wallet accounts must have either an MSISDN or a "
                "short code (paybill / till number) — without one of "
                "those there's nothing to identify the wallet.")
        # Normalise MSISDN to digits-only so '+233 24 123 4567' and
        # '233241234567' don't create duplicate accounts.
        if msisdn:
            msisdn = ''.join(ch for ch in msisdn if ch.isdigit())
    else:
        # Cash nostro accounts must NOT have wallet-specific fields set
        # — keeps the data model honest and the dashboard filters clean.
        if provider or msisdn or short_code:
            raise HTTPException(400,
                "provider / msisdn / short_code are only valid for "
                "account_type='mobile_wallet'.")

    conn = get_conn()
    try:
        _require_registered_bic(conn, payload.bic)
        now = datetime.utcnow().isoformat()
        try:
            cur = conn.execute(
                "INSERT INTO accounts (label, shortname, access_area, bic, swift_account, "
                "flex_ac_no, currency, notes, created_at, created_by, "
                "account_type, provider, msisdn, short_code) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (payload.label, payload.shortname, payload.access_area, payload.bic,
                 payload.swift_account, payload.flex_ac_no, payload.currency,
                 payload.notes, now, username,
                 account_type, provider, msisdn, short_code),
            )
        except Exception as exc:
            if 'UNIQUE' in str(exc):
                raise HTTPException(409, "That SWIFT account + Flexcube GL + currency "
                                         "combination is already registered.")
            raise
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'account_created', ?, ?, ?)",
            (username, now, json.dumps({
                "account_id": cur.lastrowid, "label": payload.label,
                "account_type": account_type, "provider": provider,
            })),
        )
        conn.commit()
        return {"id": cur.lastrowid, "label": payload.label,
                "account_type": account_type}
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Discovered accounts — unregistered identifiers the scanner has seen
# ---------------------------------------------------------------------------

class RegisterDiscoveryPayload(BaseModel):
    swift_account: str
    flex_ac_no: str
    currency: str
    label: str
    shortname: Optional[str] = None
    access_area: Optional[str] = None
    bic: Optional[str] = None
    notes: Optional[str] = None


@app.get("/discovered-accounts")
def list_discovered(user: dict = Depends(require_role('admin'))):
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT * FROM discovered_accounts ORDER BY status, kind, last_seen_at DESC"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.post("/discovered-accounts/register")
def register_discovery(payload: RegisterDiscoveryPayload,
                       user: dict = Depends(require_role('admin'))):
    """Register a new account from one or two pending discoveries, then
    requeue any unregistered files that now match. The admin supplies both
    sides even if only one was discovered — the other side may still be
    pending, or they may know it in advance."""
    from scanner import (MESSAGES_DIR, SWIFT_IN, FLEX_IN, UNLOADED_UNREGISTERED,
                         _list_files, SWIFT_SUFFIXES, FLEX_SUFFIXES)
    from swift_loader import extract_swift_meta_raw
    from account_meta import extract_swift_meta, extract_flex_meta
    from reconcile import load_flexcube
    import shutil

    ccy = payload.currency.strip().upper()
    swift_acc = payload.swift_account.strip()
    flex_acc = payload.flex_ac_no.strip()
    if not (swift_acc and flex_acc and ccy and payload.label.strip()):
        raise HTTPException(400, "swift_account, flex_ac_no, currency, and label are required.")

    username = user['username']
    now = datetime.utcnow().isoformat()
    conn = get_conn()
    try:
        try:
            cur = conn.execute(
                "INSERT INTO accounts (label, shortname, access_area, bic, swift_account, "
                "flex_ac_no, currency, notes, created_at, created_by) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (payload.label.strip(), payload.shortname, payload.access_area, payload.bic,
                 swift_acc, flex_acc, ccy, payload.notes, now, username),
            )
        except Exception as exc:
            if 'UNIQUE' in str(exc):
                raise HTTPException(409, "That SWIFT account + Flexcube GL + currency "
                                         "combination is already registered.")
            raise
        account_id = cur.lastrowid

        conn.execute(
            "UPDATE discovered_accounts SET status='registered', registered_account_id=?, "
            "resolved_at=?, resolved_by=? "
            "WHERE status='pending' AND ("
            "  (kind='swift' AND identifier=? AND currency=?) OR "
            "  (kind='flexcube' AND identifier=? AND currency=?))",
            (account_id, now, username, swift_acc, ccy, flex_acc, ccy),
        )
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'account_created', ?, ?, ?)",
            (username, now, json.dumps({
                "account_id": account_id, "label": payload.label,
                "via": "discovered_accounts",
            })),
        )
        conn.commit()
    finally:
        conn.close()

    # Requeue files from unloaded/unregistered/ whose metadata now matches.
    requeued = {'swift': 0, 'flexcube': 0}
    if UNLOADED_UNREGISTERED.exists():
        for p in list(UNLOADED_UNREGISTERED.iterdir()):
            if not p.is_file() or p.name.startswith('~$') or p.name.startswith('.'):
                continue
            suffix = p.suffix.lower()
            try:
                if suffix == '.out':
                    meta = extract_swift_meta_raw(p)
                    if meta.get('account') == swift_acc and meta.get('currency') == ccy:
                        shutil.move(str(p), str(SWIFT_IN / p.name))
                        requeued['swift'] += 1
                        continue
                elif suffix in SWIFT_SUFFIXES:
                    try:
                        meta = extract_swift_meta(p)
                        if meta.get('account') == swift_acc and meta.get('currency') == ccy:
                            shutil.move(str(p), str(SWIFT_IN / p.name))
                            requeued['swift'] += 1
                            continue
                    except Exception:
                        pass
                if suffix in FLEX_SUFFIXES:
                    try:
                        fmeta = extract_flex_meta(load_flexcube(p))
                        if fmeta.get('ac_no') == flex_acc and fmeta.get('currency') == ccy:
                            shutil.move(str(p), str(FLEX_IN / p.name))
                            requeued['flexcube'] += 1
                    except Exception:
                        pass
            except Exception:
                continue

    return {
        "account_id": account_id,
        "label": payload.label,
        "requeued": requeued,
    }


@app.post("/discovered-accounts/{disc_id}/ignore")
def ignore_discovery(disc_id: int, user: dict = Depends(require_role('admin'))):
    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT id, status FROM discovered_accounts WHERE id=?", (disc_id,),
        ).fetchone()
        if row is None:
            raise HTTPException(404, "Not found")
        if row['status'] != 'pending':
            raise HTTPException(409, f"Discovery already {row['status']}")
        now = datetime.utcnow().isoformat()
        conn.execute(
            "UPDATE discovered_accounts SET status='ignored', resolved_at=?, resolved_by=? "
            "WHERE id=?",
            (now, user['username'], disc_id),
        )
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

# ===========================================================================
# Reconciliation-ledger endpoints (2026-04-22).
#
# These power the rolling open-items view, manual matching, break notes,
# per-account tolerance rules, and the per-account reconciliation status
# dashboard. Ordering here mirrors the user journey:
#   1. List and clear open items
#   2. Force-match two rows that the engine didn't pair
#   3. Leave a comment on an assignment or open item
#   4. Read/update tolerance rules and auto-categorization rules
#   5. Read the per-account reconciliation status with aging
# ===========================================================================

# --- /open-items ------------------------------------------------------------

@app.get("/open-items")
def list_open_items(
    account_id: Optional[int] = None,
    status: str = 'open',
    category: Optional[str] = None,
    min_age_days: Optional[int] = None,
    limit: int = 500,
    user: dict = Depends(current_user),
    scope: Optional[List[str]] = Depends(active_scope),
):
    """Rolling ledger query. Returns age_days per row so the UI can bucket
    without another round-trip. Access-area scope applies via the account
    join — a user viewing only BRANCH 001 HOFF won't see NOSTRO items."""
    limit = max(1, min(limit, 2000))
    conn = get_conn()
    try:
        where = ["oi.status = ?"]
        params: list = [status]
        if account_id is not None:
            where.append("oi.account_id = ?"); params.append(account_id)
        if category:
            where.append("oi.category = ?"); params.append(category)
        scope_sql, scope_params = _scope_clause(scope, alias='a')
        where.append(scope_sql); params.extend(scope_params)
        if scope:
            # force account join to be present for scope filtering
            where.append("a.id IS NOT NULL")

        rows = conn.execute(
            f"SELECT oi.*, a.shortname AS account_shortname, a.label AS account_label, "
            f"       a.access_area AS account_access_area, a.currency AS account_currency "
            f"FROM open_items oi LEFT JOIN accounts a ON a.id = oi.account_id "
            f"WHERE {' AND '.join(where)} "
            f"ORDER BY oi.opened_at LIMIT ?",
            (*params, limit),
        ).fetchall()

        now = datetime.utcnow()
        out = []
        for r in rows:
            age_days = _age_in_days(r['opened_at'], now)
            if min_age_days is not None and age_days < min_age_days:
                continue
            item = dict(r)
            item['age_days'] = age_days
            item['age_bucket'] = _age_bucket(age_days)
            out.append(item)
        return out
    finally:
        conn.close()


@app.get("/accounts/{account_id}/open-items/aging")
def open_items_aging(account_id: int, user: dict = Depends(current_user)):
    """Aging-bucket counts and totals for one account. Feeds the per-account
    status dashboard — a single query instead of client-side group-by."""
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT source_side, category, amount, opened_at FROM open_items "
            "WHERE account_id=? AND status='open'",
            (account_id,),
        ).fetchall()
        now = datetime.utcnow()
        buckets = {'0-7': 0, '8-30': 0, '31-90': 0, '90+': 0}
        values  = {'0-7': 0.0, '8-30': 0.0, '31-90': 0.0, '90+': 0.0}
        by_category: dict[str, dict] = {}
        by_side = {'swift': 0, 'flex': 0}
        for r in rows:
            b = _age_bucket(_age_in_days(r['opened_at'], now))
            buckets[b] += 1
            values[b] += r['amount'] or 0
            by_side[r['source_side']] = by_side.get(r['source_side'], 0) + 1
            c = r['category'] or 'uncategorized'
            by_category.setdefault(c, {'count': 0, 'value': 0.0})
            by_category[c]['count'] += 1
            by_category[c]['value'] += r['amount'] or 0
        return {
            'total_open': len(rows),
            'buckets': buckets,
            'bucket_values': values,
            'by_side': by_side,
            'by_category': by_category,
        }
    finally:
        conn.close()


class OpenItemManualClearPayload(BaseModel):
    session_id: int
    counterpart_row: int
    counterpart_side: str      # 'swift' | 'flex' — row's side in that session
    note: Optional[str] = None


@app.post("/open-items/{open_item_id}/clear")
def clear_open_item(open_item_id: int,
                    payload: OpenItemManualClearPayload,
                    user: dict = Depends(require_role('ops', 'admin'))):
    """Hand-pair an open_item to a row in an open session. The row on the
    other side must exist and not already be claimed by a pending or
    confirmed assignment — we don't silently stomp prior decisions."""
    if payload.counterpart_side not in ('swift', 'flex'):
        raise HTTPException(400, "counterpart_side must be 'swift' or 'flex'")
    from open_items import clear_open_item_manually

    conn = get_conn()
    try:
        txn_table = 'swift_txns' if payload.counterpart_side == 'swift' else 'flex_txns'
        row_col = 'swift_row' if payload.counterpart_side == 'swift' else 'flex_row'
        exists = conn.execute(
            f"SELECT 1 FROM {txn_table} WHERE session_id=? AND row_number=?",
            (payload.session_id, payload.counterpart_row),
        ).fetchone()
        if exists is None:
            raise HTTPException(404, f"No {payload.counterpart_side} row "
                                     f"{payload.counterpart_row} in session {payload.session_id}")
        claimed = conn.execute(
            f"SELECT status FROM assignments WHERE session_id=? AND {row_col}=? "
            f"AND status != 'rejected'",
            (payload.session_id, payload.counterpart_row),
        ).fetchone()
        if claimed:
            raise HTTPException(409,
                f"That row is already on a {claimed['status']} assignment. "
                "Reject the existing assignment first if you want to re-pair.")
        try:
            assignment_id = clear_open_item_manually(
                conn, open_item_id, payload.session_id,
                payload.counterpart_row, payload.counterpart_side,
                user['username'], payload.note,
            )
        except ValueError as exc:
            raise HTTPException(400, str(exc))
        conn.commit()
        return {"assignment_id": assignment_id, "open_item_id": open_item_id,
                "status": "cleared"}
    finally:
        conn.close()


class OpenItemWriteOffPayload(BaseModel):
    reason: str


@app.post("/open-items/{open_item_id}/write-off")
def write_off_open_item_endpoint(open_item_id: int,
                                  payload: OpenItemWriteOffPayload,
                                  user: dict = Depends(require_role('admin', 'internal_control'))):
    """Terminal state. Only admin or internal_control can write off — it's
    an accounting decision, not routine ops."""
    from open_items import write_off_open_item
    if not (payload.reason or '').strip():
        raise HTTPException(400, "reason is required for a write-off")
    conn = get_conn()
    try:
        try:
            write_off_open_item(conn, open_item_id, user['username'], payload.reason.strip())
        except ValueError as exc:
            raise HTTPException(400, str(exc))
        conn.commit()
        return {"open_item_id": open_item_id, "status": "written_off"}
    finally:
        conn.close()


class SnoozePayload(BaseModel):
    days: int  # 1-90

@app.post("/open-items/{item_id}/snooze")
def snooze_open_item(item_id: int, payload: SnoozePayload,
                     user: dict = Depends(current_user)):
    if not (1 <= payload.days <= 90):
        raise HTTPException(400, "days must be between 1 and 90")
    conn = get_conn()
    try:
        row = conn.execute("SELECT id, status FROM open_items WHERE id=?", (item_id,)).fetchone()
        if row is None: raise HTTPException(404, "Open item not found")
        if row['status'] != 'open': raise HTTPException(400, "Only open items can be snoozed")
        from datetime import timedelta
        until = (datetime.utcnow() + timedelta(days=payload.days)).isoformat()
        conn.execute("UPDATE open_items SET snoozed_until=? WHERE id=?", (until, item_id))
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'open_item_snoozed', ?, ?, ?)",
            (user['username'], datetime.utcnow().isoformat(),
             json.dumps({"item_id": item_id, "days": payload.days, "until": until})),
        )
        conn.commit()
        return {"item_id": item_id, "snoozed_until": until}
    finally:
        conn.close()

@app.post("/open-items/{item_id}/acknowledge")
def acknowledge_open_item(item_id: int, user: dict = Depends(current_user)):
    conn = get_conn()
    try:
        row = conn.execute("SELECT id, status FROM open_items WHERE id=?", (item_id,)).fetchone()
        if row is None: raise HTTPException(404, "Open item not found")
        if row['status'] != 'open': raise HTTPException(400, "Only open items can be acknowledged")
        now = datetime.utcnow().isoformat()
        conn.execute("UPDATE open_items SET acknowledged_by=?, acknowledged_at=? WHERE id=?",
                     (user['username'], now, item_id))
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'open_item_acknowledged', ?, ?, ?)",
            (user['username'], now, json.dumps({"item_id": item_id})),
        )
        conn.commit()
        return {"item_id": item_id, "acknowledged_by": user['username'], "acknowledged_at": now}
    finally:
        conn.close()


class OpenItemCategoryPayload(BaseModel):
    category: str


@app.patch("/open-items/{open_item_id}/category")
def set_open_item_category(open_item_id: int,
                           payload: OpenItemCategoryPayload,
                           user: dict = Depends(require_role('ops', 'admin'))):
    """Manual override of the auto-category. Audit-logged so we can tell a
    rule-assigned category from a human-assigned one later."""
    from db import BREAK_CATEGORIES
    if payload.category not in BREAK_CATEGORIES:
        raise HTTPException(400, f"category must be one of {BREAK_CATEGORIES}")
    conn = get_conn()
    try:
        row = conn.execute("SELECT id, category FROM open_items WHERE id=?",
                           (open_item_id,)).fetchone()
        if row is None:
            raise HTTPException(404, "open_item not found")
        now = datetime.utcnow().isoformat()
        conn.execute(
            "UPDATE open_items SET category=?, category_source='manual', "
            "category_rule_id=NULL WHERE id=?",
            (payload.category, open_item_id),
        )
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'open_item_recategorized', ?, ?, ?)",
            (user['username'], now, json.dumps({
                'open_item_id': open_item_id,
                'from': row['category'], 'to': payload.category,
            })),
        )
        conn.commit()
        return {"open_item_id": open_item_id, "category": payload.category}
    finally:
        conn.close()


# --- /sessions/{id}/manual-match -------------------------------------------

class ManualMatchPayload(BaseModel):
    swift_row: int
    flex_row: int
    reason: Optional[str] = None


@app.post("/sessions/{session_id}/manual-match")
def manual_match(session_id: int, payload: ManualMatchPayload,
                 user: dict = Depends(require_role('ops', 'admin'))):
    """Force-pair two rows the engine rejected. Creates a confirmed
    assignment with source='manual'. Refuses if either row is already on
    a non-rejected assignment — the analyst must reject the existing one
    first so the audit trail shows the re-pairing."""
    conn = get_conn()
    try:
        _assert_session_not_locked(conn, session_id)
        s = conn.execute(
            "SELECT row_number, sign, amount FROM swift_txns WHERE session_id=? AND row_number=?",
            (session_id, payload.swift_row),
        ).fetchone()
        if s is None:
            raise HTTPException(404, f"SWIFT row {payload.swift_row} not in session {session_id}")
        f = conn.execute(
            "SELECT row_number, type, amount FROM flex_txns WHERE session_id=? AND row_number=?",
            (session_id, payload.flex_row),
        ).fetchone()
        if f is None:
            raise HTTPException(404, f"Flex row {payload.flex_row} not in session {session_id}")

        blocker_s = conn.execute(
            "SELECT id, status FROM assignments WHERE session_id=? AND swift_row=? "
            "AND status != 'rejected'",
            (session_id, payload.swift_row),
        ).fetchone()
        blocker_f = conn.execute(
            "SELECT id, status FROM assignments WHERE session_id=? AND flex_row=? "
            "AND status != 'rejected'",
            (session_id, payload.flex_row),
        ).fetchone()
        if blocker_s:
            raise HTTPException(409,
                f"SWIFT row {payload.swift_row} already on a {blocker_s['status']} "
                f"assignment (id {blocker_s['id']}). Reject it first.")
        if blocker_f:
            raise HTTPException(409,
                f"Flex row {payload.flex_row} already on a {blocker_f['status']} "
                f"assignment (id {blocker_f['id']}). Reject it first.")

        now = datetime.utcnow().isoformat()
        amount_diff = (f['amount'] or 0) - (s['amount'] or 0)
        reason = f"manual force-match by {user['username']}"
        if payload.reason:
            reason += f" — {payload.reason}"
        cur = conn.execute(
            "INSERT INTO assignments (session_id, swift_row, flex_row, tier, reason, "
            "amount_diff, status, decided_by, decided_at, source, manual_reason) "
            "VALUES (?,?,?,?,?,?, 'confirmed', ?, ?, 'manual', ?)",
            (session_id, payload.swift_row, payload.flex_row, 0, reason, amount_diff,
             user['username'], now, payload.reason),
        )
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (?, 'manual_match', ?, ?, ?)",
            (session_id, user['username'], now, json.dumps({
                'assignment_id': cur.lastrowid,
                'swift_row': payload.swift_row, 'flex_row': payload.flex_row,
                'amount_diff': amount_diff, 'reason': payload.reason,
            })),
        )
        conn.commit()
        return {"assignment_id": cur.lastrowid, "status": "confirmed", "source": "manual"}
    finally:
        conn.close()


# --- /sessions/{id}/manual-split -------------------------------------------
# Analyst-driven split match: user picks one row on one side and N rows on
# the opposite side. All rows get one assignment row each, sharing a
# split_group_id, confirmed with source='manual_split'. Refuses if any
# input row is already on a non-rejected assignment.

class ManualSplitPayload(BaseModel):
    swift_rows: list[int]
    flex_rows: list[int]
    reason: Optional[str] = None


@app.post("/sessions/{session_id}/manual-split")
def manual_split(session_id: int, payload: ManualSplitPayload,
                 user: dict = Depends(require_role('ops', 'admin'))):
    import uuid
    if not payload.swift_rows or not payload.flex_rows:
        raise HTTPException(400, "provide at least one row on each side")
    one_side_swift = len(payload.swift_rows) == 1
    one_side_flex  = len(payload.flex_rows) == 1
    if not (one_side_swift ^ one_side_flex):
        raise HTTPException(
            400, "exactly one side must have a single row (the other N rows)")

    conn = get_conn()
    try:
        _assert_session_not_locked(conn, session_id)
        # Validate all rows exist + none are already matched.
        for sr in payload.swift_rows:
            r = conn.execute(
                "SELECT amount FROM swift_txns WHERE session_id=? AND row_number=?",
                (session_id, sr),
            ).fetchone()
            if r is None:
                raise HTTPException(404, f"SWIFT row {sr} not in session {session_id}")
            blocker = conn.execute(
                "SELECT id, status FROM assignments WHERE session_id=? AND swift_row=? "
                "AND status != 'rejected'", (session_id, sr),
            ).fetchone()
            if blocker:
                raise HTTPException(
                    409, f"SWIFT row {sr} already on a {blocker['status']} assignment. "
                    f"Reject it first.")
        for fr in payload.flex_rows:
            r = conn.execute(
                "SELECT amount FROM flex_txns WHERE session_id=? AND row_number=?",
                (session_id, fr),
            ).fetchone()
            if r is None:
                raise HTTPException(404, f"Flex row {fr} not in session {session_id}")
            blocker = conn.execute(
                "SELECT id, status FROM assignments WHERE session_id=? AND flex_row=? "
                "AND status != 'rejected'", (session_id, fr),
            ).fetchone()
            if blocker:
                raise HTTPException(
                    409, f"Flex row {fr} already on a {blocker['status']} assignment. "
                    f"Reject it first.")

        swift_total = sum(conn.execute(
            "SELECT amount FROM swift_txns WHERE session_id=? AND row_number=?",
            (session_id, sr)).fetchone()[0] or 0 for sr in payload.swift_rows)
        flex_total  = sum(conn.execute(
            "SELECT amount FROM flex_txns WHERE session_id=? AND row_number=?",
            (session_id, fr)).fetchone()[0] or 0 for fr in payload.flex_rows)
        amount_diff = flex_total - swift_total

        now = datetime.utcnow().isoformat()
        grp = uuid.uuid4().hex[:12]
        reason = (f"manual split by {user['username']}: "
                  f"{len(payload.swift_rows)} SWIFT vs {len(payload.flex_rows)} Flex")
        if payload.reason:
            reason += f" — {payload.reason}"

        assignment_ids = []
        if one_side_swift:
            sw = payload.swift_rows[0]
            per_row_diff = amount_diff / len(payload.flex_rows)
            for fr in payload.flex_rows:
                cur = conn.execute(
                    "INSERT INTO assignments (session_id, swift_row, flex_row, tier, "
                    "reason, amount_diff, status, decided_by, decided_at, source, "
                    "manual_reason, split_group_id) "
                    "VALUES (?,?,?,?,?,?, 'confirmed', ?, ?, 'manual_split', ?, ?)",
                    (session_id, sw, fr, 5, reason, per_row_diff,
                     user['username'], now, payload.reason, grp),
                )
                assignment_ids.append(cur.lastrowid)
        else:
            fr = payload.flex_rows[0]
            per_row_diff = amount_diff / len(payload.swift_rows)
            for sw in payload.swift_rows:
                cur = conn.execute(
                    "INSERT INTO assignments (session_id, swift_row, flex_row, tier, "
                    "reason, amount_diff, status, decided_by, decided_at, source, "
                    "manual_reason, split_group_id) "
                    "VALUES (?,?,?,?,?,?, 'confirmed', ?, ?, 'manual_split', ?, ?)",
                    (session_id, sw, fr, 5, reason, per_row_diff,
                     user['username'], now, payload.reason, grp),
                )
                assignment_ids.append(cur.lastrowid)

        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (?, 'manual_split', ?, ?, ?)",
            (session_id, user['username'], now, json.dumps({
                'split_group_id': grp, 'swift_rows': payload.swift_rows,
                'flex_rows': payload.flex_rows, 'amount_diff': amount_diff,
                'assignment_ids': assignment_ids, 'reason': payload.reason})),
        )
        conn.commit()
        return {
            'split_group_id': grp,
            'assignment_ids': assignment_ids,
            'amount_diff': amount_diff,
            'status': 'confirmed',
        }
    finally:
        conn.close()


# --- /sessions/{id}/split-groups/{group_id}/decision -----------------------
# Batch confirm or reject all pending assignments sharing a split_group_id.

class SplitGroupDecisionPayload(BaseModel):
    action: str  # 'confirm' | 'reject'


@app.post("/sessions/{session_id}/split-groups/{group_id}/decision")
def split_group_decision(
    session_id: int, group_id: str,
    payload: SplitGroupDecisionPayload,
    user: dict = Depends(require_role('ops', 'admin')),
):
    if payload.action not in ('confirm', 'reject'):
        raise HTTPException(400, "action must be 'confirm' or 'reject'")
    conn = get_conn()
    try:
        _assert_session_not_locked(conn, session_id)
        rows = conn.execute(
            "SELECT id FROM assignments WHERE session_id=? AND split_group_id=? AND status='pending'",
            (session_id, group_id),
        ).fetchall()
        if not rows:
            raise HTTPException(404, "No pending assignments found for this split group")

        new_status = 'confirmed' if payload.action == 'confirm' else 'rejected'
        now = datetime.utcnow().isoformat()
        ids = [r['id'] for r in rows]

        for aid in ids:
            conn.execute(
                "UPDATE assignments SET status=?, decided_by=?, decided_at=? WHERE id=?",
                (new_status, user['username'], now, aid),
            )
            conn.execute(
                "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
                "VALUES (?, ?, ?, ?, ?)",
                (session_id, f"split_group_{payload.action}", user['username'], now,
                 json.dumps({"split_group_id": group_id, "assignment_id": aid})),
            )
        conn.commit()
        return {"split_group_id": group_id, "action": payload.action, "count": len(ids)}
    finally:
        conn.close()


# --- /comments (break notes) -----------------------------------------------

class CommentPayload(BaseModel):
    target_type: str        # 'assignment' | 'open_item'
    target_id: int
    session_id: Optional[int] = None
    body: str


@app.post("/comments")
def add_comment(payload: CommentPayload,
                user: dict = Depends(require_role('ops', 'admin', 'internal_control'))):
    """Free-text note against an assignment or an open_item. No edit/delete
    — an audit trail that allows edits isn't an audit trail."""
    if payload.target_type not in ('assignment', 'open_item'):
        raise HTTPException(400, "target_type must be 'assignment' or 'open_item'")
    body = (payload.body or '').strip()
    if not body:
        raise HTTPException(400, "body must not be empty")
    if len(body) > 4000:
        raise HTTPException(400, "body must be 4000 characters or fewer")
    conn = get_conn()
    try:
        # Verify the target exists — prevents orphan comments if an id is fat-fingered.
        if payload.target_type == 'assignment':
            exists = conn.execute("SELECT id FROM assignments WHERE id=?",
                                  (payload.target_id,)).fetchone()
        else:
            exists = conn.execute("SELECT id FROM open_items WHERE id=?",
                                  (payload.target_id,)).fetchone()
        if exists is None:
            raise HTTPException(404, f"{payload.target_type} {payload.target_id} not found")

        now = datetime.utcnow().isoformat()
        cur = conn.execute(
            "INSERT INTO break_comments (target_type, target_id, session_id, author, "
            "created_at, body) VALUES (?,?,?,?,?,?)",
            (payload.target_type, payload.target_id, payload.session_id,
             user['username'], now, body),
        )
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (?, 'comment_added', ?, ?, ?)",
            (payload.session_id, user['username'], now, json.dumps({
                'comment_id': cur.lastrowid, 'target_type': payload.target_type,
                'target_id': payload.target_id,
            })),
        )
        conn.commit()
        return {"id": cur.lastrowid, "created_at": now}
    finally:
        conn.close()


@app.get("/comments")
def list_comments(target_type: str, target_id: int,
                  user: dict = Depends(current_user)):
    """Fetch comments for a single target in chronological order."""
    if target_type not in ('assignment', 'open_item'):
        raise HTTPException(400, "target_type must be 'assignment' or 'open_item'")
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT id, author, created_at, body, session_id "
            "FROM break_comments WHERE target_type=? AND target_id=? ORDER BY id",
            (target_type, target_id),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


# --- /accounts/{id}/tolerance ----------------------------------------------

class TolerancePayload(BaseModel):
    amount_tol_abs: float
    amount_tol_pct: float
    date_tol_days: int
    min_ref_len: int
    # Basis-points cushion after FX conversion. Optional so existing
    # callers that never set it don't have to know it exists. 0 disables.
    fx_tol_bps: Optional[float] = 0.0


@app.get("/accounts/{account_id}/tolerance")
def get_tolerance(account_id: int, user: dict = Depends(current_user)):
    """Everyone can read the current rule so analysts understand why a
    match did or didn't land. Edits are admin-only (PUT)."""
    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT amount_tol_abs, amount_tol_pct, date_tol_days, min_ref_len, "
            "       fx_tol_bps, updated_at, updated_by "
            "FROM tolerance_rules WHERE account_id=?",
            (account_id,),
        ).fetchone()
        if row is None:
            # Return the defaults so the UI has something to render.
            return {
                "account_id": account_id,
                "amount_tol_abs": 0.01, "amount_tol_pct": 0.0,
                "date_tol_days": 1, "min_ref_len": 6, "fx_tol_bps": 0.0,
                "updated_at": None, "updated_by": None, "is_default": True,
            }
        return {"account_id": account_id, **dict(row), "is_default": False}
    finally:
        conn.close()


@app.put("/accounts/{account_id}/tolerance")
def put_tolerance(account_id: int, payload: TolerancePayload,
                  user: dict = Depends(require_role('admin'))):
    """Admin-only. Every change is audit-logged — loose tolerances hide
    errors, so who-changed-what matters."""
    if payload.amount_tol_abs < 0:
        raise HTTPException(400, "amount_tol_abs must be >= 0")
    if payload.amount_tol_pct < 0 or payload.amount_tol_pct > 10:
        raise HTTPException(400, "amount_tol_pct must be between 0 and 10 (percent)")
    if payload.date_tol_days < 0 or payload.date_tol_days > 30:
        raise HTTPException(400, "date_tol_days must be between 0 and 30")
    if payload.min_ref_len < 3 or payload.min_ref_len > 32:
        raise HTTPException(400, "min_ref_len must be between 3 and 32")
    fx_bps = payload.fx_tol_bps or 0.0
    if fx_bps < 0 or fx_bps > 1000:
        # 1000 bps = 10%; anything looser hides genuine FX errors.
        raise HTTPException(400, "fx_tol_bps must be between 0 and 1000 (basis points; 100 = 1%)")
    conn = get_conn()
    try:
        exists = conn.execute("SELECT id FROM accounts WHERE id=?",
                              (account_id,)).fetchone()
        if exists is None:
            raise HTTPException(404, "Account not found")
        now = datetime.utcnow().isoformat()
        prior = conn.execute(
            "SELECT amount_tol_abs, amount_tol_pct, date_tol_days, min_ref_len "
            "FROM tolerance_rules WHERE account_id=?", (account_id,),
        ).fetchone()
        conn.execute(
            "INSERT INTO tolerance_rules (account_id, amount_tol_abs, amount_tol_pct, "
            "date_tol_days, min_ref_len, fx_tol_bps, updated_at, updated_by) "
            "VALUES (?,?,?,?,?,?,?,?) "
            "ON CONFLICT(account_id) DO UPDATE SET "
            "  amount_tol_abs=excluded.amount_tol_abs, "
            "  amount_tol_pct=excluded.amount_tol_pct, "
            "  date_tol_days=excluded.date_tol_days, "
            "  min_ref_len=excluded.min_ref_len, "
            "  fx_tol_bps=excluded.fx_tol_bps, "
            "  updated_at=excluded.updated_at, updated_by=excluded.updated_by",
            (account_id, payload.amount_tol_abs, payload.amount_tol_pct,
             payload.date_tol_days, payload.min_ref_len, fx_bps,
             now, user['username']),
        )
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'tolerance_updated', ?, ?, ?)",
            (user['username'], now, json.dumps({
                'account_id': account_id,
                'prior': dict(prior) if prior else None,
                'new': payload.model_dump(),
            })),
        )
        conn.commit()
        return {"account_id": account_id, **payload.model_dump(),
                "updated_at": now, "updated_by": user['username']}
    finally:
        conn.close()


# --- /auto-cat-rules --------------------------------------------------------

class AutoCatRulePayload(BaseModel):
    name: str
    priority: int = 100
    side: Optional[str] = None              # 'swift' | 'flex' | null
    narration_contains: Optional[str] = None
    type_equals: Optional[str] = None
    amount_min: Optional[float] = None
    amount_max: Optional[float] = None
    category: str
    active: bool = True


@app.get("/auto-cat-rules")
def list_auto_cat_rules(user: dict = Depends(current_user)):
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT * FROM auto_categorization_rules ORDER BY priority, id"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.post("/auto-cat-rules")
def create_auto_cat_rule(payload: AutoCatRulePayload,
                         user: dict = Depends(require_role('admin'))):
    from db import BREAK_CATEGORIES
    if payload.category not in BREAK_CATEGORIES:
        raise HTTPException(400, f"category must be one of {BREAK_CATEGORIES}")
    if payload.side and payload.side not in ('swift', 'flex'):
        raise HTTPException(400, "side must be 'swift', 'flex', or null")
    now = datetime.utcnow().isoformat()
    conn = get_conn()
    try:
        cur = conn.execute(
            "INSERT INTO auto_categorization_rules (name, priority, side, "
            "narration_contains, type_equals, amount_min, amount_max, category, "
            "active, created_at, created_by) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (payload.name, payload.priority, payload.side, payload.narration_contains,
             payload.type_equals, payload.amount_min, payload.amount_max,
             payload.category, 1 if payload.active else 0, now, user['username']),
        )
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'auto_cat_rule_created', ?, ?, ?)",
            (user['username'], now, json.dumps({
                'rule_id': cur.lastrowid, **payload.model_dump(),
            })),
        )
        conn.commit()
        return {"id": cur.lastrowid}
    finally:
        conn.close()


class AutoGroupingRulePayload(BaseModel):
    name: str
    priority: int = 100
    side: Optional[str] = None
    narration_contains: Optional[str] = None
    ref_contains: Optional[str] = None
    type_equals: Optional[str] = None
    amount_min: Optional[float] = None
    amount_max: Optional[float] = None
    functional_group: str
    active: bool = True


@app.get("/grouping-rules")
def list_grouping_rules(user: dict = Depends(current_user)):
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT * FROM auto_grouping_rules ORDER BY priority, id"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.post("/grouping-rules")
def create_grouping_rule(payload: AutoGroupingRulePayload,
                         user: dict = Depends(require_role('admin'))):
    from db import FUNCTIONAL_GROUPS
    if payload.functional_group not in FUNCTIONAL_GROUPS:
        # Allow arbitrary strings — ops may define new teams — but warn.
        pass
    if payload.side and payload.side not in ('swift', 'flex'):
        raise HTTPException(400, "side must be 'swift', 'flex', or null")
    now = datetime.utcnow().isoformat()
    conn = get_conn()
    try:
        cur = conn.execute(
            "INSERT INTO auto_grouping_rules (name, priority, side, "
            "narration_contains, ref_contains, type_equals, amount_min, amount_max, "
            "functional_group, active, created_at, created_by) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (payload.name, payload.priority, payload.side,
             payload.narration_contains, payload.ref_contains, payload.type_equals,
             payload.amount_min, payload.amount_max, payload.functional_group,
             1 if payload.active else 0, now, user['username']),
        )
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'grouping_rule_created', ?, ?, ?)",
            (user['username'], now, json.dumps({
                'rule_id': cur.lastrowid, **payload.model_dump()})),
        )
        conn.commit()
        return {"id": cur.lastrowid}
    finally:
        conn.close()


# Alias so the match-patterns promote modal can POST to a stable URL that
# reads as "create a categorization rule" instead of the abbreviated
# auto-cat-rules path.
@app.post("/categorization-rules")
def create_categorization_rule(payload: AutoCatRulePayload,
                                user: dict = Depends(require_role('admin'))):
    return create_auto_cat_rule(payload, user)


@app.delete("/auto-cat-rules/{rule_id}")
def deactivate_auto_cat_rule(rule_id: int,
                              user: dict = Depends(require_role('admin'))):
    """Soft delete — we retain rule rows so historic open_items.category_rule_id
    points at something auditable. Flipping active=0 is enough to stop its
    use on future classifications."""
    conn = get_conn()
    try:
        row = conn.execute("SELECT id, active FROM auto_categorization_rules WHERE id=?",
                           (rule_id,)).fetchone()
        if row is None:
            raise HTTPException(404, "Rule not found")
        if not row['active']:
            return {"id": rule_id, "active": False}
        conn.execute(
            "UPDATE auto_categorization_rules SET active=0 WHERE id=?", (rule_id,),
        )
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'auto_cat_rule_deactivated', ?, ?, ?)",
            (user['username'], datetime.utcnow().isoformat(),
             json.dumps({'rule_id': rule_id})),
        )
        conn.commit()
        return {"id": rule_id, "active": False}
    finally:
        conn.close()


# --- /accounts/{id}/status --------------------------------------------------

@app.get("/accounts/{account_id}/status")
def account_recon_status(account_id: int, user: dict = Depends(current_user)):
    """Per-account reconciliation dashboard. Gathers: match rate, last
    reconciled session, open item counts & aging, and category breakdown.
    Single endpoint so the status page is one round-trip."""
    conn = get_conn()
    try:
        acc = conn.execute(
            "SELECT id, label, shortname, access_area, bic, swift_account, "
            "       flex_ac_no, currency, notes, active "
            "FROM accounts WHERE id=?", (account_id,),
        ).fetchone()
        if acc is None:
            raise HTTPException(404, "Account not found")

        last_sess = conn.execute(
            "SELECT id, created_at, swift_filename, flex_filename, "
            "       closing_balance_amount, closing_balance_sign, closing_balance_date "
            "FROM sessions WHERE account_id=? ORDER BY id DESC LIMIT 1",
            (account_id,),
        ).fetchone()

        # Cumulative match rate across all sessions for this account.
        totals = conn.execute(
            "SELECT "
            "  (SELECT COUNT(*) FROM swift_txns st JOIN sessions s ON s.id=st.session_id "
            "   WHERE s.account_id=?) AS swift_total, "
            "  (SELECT COUNT(*) FROM flex_txns ft JOIN sessions s ON s.id=ft.session_id "
            "   WHERE s.account_id=?) AS flex_total, "
            "  (SELECT COUNT(*) FROM assignments a JOIN sessions s ON s.id=a.session_id "
            "   WHERE s.account_id=? AND a.status='confirmed') AS confirmed",
            (account_id, account_id, account_id),
        ).fetchone()

        aging_rows = conn.execute(
            "SELECT source_side, category, amount, opened_at FROM open_items "
            "WHERE account_id=? AND status='open'", (account_id,),
        ).fetchall()
        now = datetime.utcnow()
        buckets = {'0-7': 0, '8-30': 0, '31-90': 0, '90+': 0}
        values  = {'0-7': 0.0, '8-30': 0.0, '31-90': 0.0, '90+': 0.0}
        by_category: dict[str, dict] = {}
        oldest = None
        for r in aging_rows:
            age = _age_in_days(r['opened_at'], now)
            b = _age_bucket(age)
            buckets[b] += 1
            values[b] += r['amount'] or 0
            c = r['category'] or 'uncategorized'
            by_category.setdefault(c, {'count': 0, 'value': 0.0})
            by_category[c]['count'] += 1
            by_category[c]['value'] += r['amount'] or 0
            if oldest is None or age > oldest:
                oldest = age

        writ = conn.execute(
            "SELECT COUNT(*) FROM open_items WHERE account_id=? AND status='written_off'",
            (account_id,),
        ).fetchone()[0]
        cleared = conn.execute(
            "SELECT COUNT(*) FROM open_items WHERE account_id=? AND status='cleared'",
            (account_id,),
        ).fetchone()[0]

        swift_total = totals['swift_total'] or 0
        confirmed = totals['confirmed'] or 0
        match_rate = (confirmed / swift_total * 100.0) if swift_total else 0.0

        return {
            "account": dict(acc),
            "last_session": dict(last_sess) if last_sess else None,
            "totals": {
                "swift_total": swift_total,
                "flex_total": totals['flex_total'] or 0,
                "confirmed": confirmed,
                "match_rate_pct": round(match_rate, 1),
            },
            "open_items": {
                "count": len(aging_rows),
                "oldest_days": oldest,
                "buckets": buckets,
                "bucket_values": values,
                "by_category": by_category,
                "cleared_historic": cleared,
                "written_off": writ,
            },
        }
    finally:
        conn.close()


# --- helpers (aging) -------------------------------------------------------

def _age_in_days(opened_at_iso: str, now: datetime) -> int:
    try:
        opened = datetime.fromisoformat(opened_at_iso)
    except (TypeError, ValueError):
        return 0
    return max(0, (now - opened).days)


def _age_bucket(age_days: int) -> str:
    if age_days <= 7:
        return '0-7'
    if age_days <= 30:
        return '8-30'
    if age_days <= 90:
        return '31-90'
    return '90+'


# --- rendered pages --------------------------------------------------------

@app.get("/open-items-view")
def open_items_page(request: Request):
    return templates.TemplateResponse(request, "open_items.html")


@app.get("/accounts/{account_id}/status-page")
def account_status_page(request: Request, account_id: int):
    return templates.TemplateResponse(request, "account_status.html",
                                      {"account_id": account_id})


@app.get("/tolerance-admin")
def tolerance_admin_page(request: Request):
    return templates.TemplateResponse(request, "tolerance_admin.html")


# ---------------------------------------------------------------------------
# Auto-match rules — operator-defined rules for auto-confirming proposals
# ---------------------------------------------------------------------------

class AutoRulePayload(BaseModel):
    name: str
    description: Optional[str] = None
    priority: int = 0
    active: int = 1
    require_tier: Optional[str] = None
    require_amount_exact: Optional[int] = None
    require_ref_match: Optional[int] = None
    max_amount_diff: Optional[float] = None
    require_same_date: Optional[int] = None

@app.get("/auto-rules")
def list_auto_rules(user: dict = Depends(require_role('admin', 'ops', 'internal_control'))):
    conn = get_conn()
    try:
        rows = conn.execute("SELECT * FROM auto_match_rules ORDER BY priority, id").fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()

@app.post("/auto-rules")
def create_auto_rule(payload: AutoRulePayload,
                     user: dict = Depends(require_role('admin'))):
    conn = get_conn()
    try:
        now = datetime.utcnow().isoformat()
        cur = conn.execute(
            "INSERT INTO auto_match_rules "
            "(name, description, priority, active, require_tier, require_amount_exact, "
            "require_ref_match, max_amount_diff, require_same_date, action, created_by, created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,'confirm',?,?)",
            (payload.name, payload.description, payload.priority, payload.active,
             payload.require_tier, payload.require_amount_exact, payload.require_ref_match,
             payload.max_amount_diff, payload.require_same_date,
             user['username'], now)
        )
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'auto_rule_created', ?, ?, ?)",
            (user['username'], now, json.dumps({'rule_id': cur.lastrowid, 'name': payload.name}))
        )
        conn.commit()
        return dict(conn.execute("SELECT * FROM auto_match_rules WHERE id=?", (cur.lastrowid,)).fetchone())
    finally:
        conn.close()

@app.patch("/auto-rules/{rule_id}")
def update_auto_rule(rule_id: int, payload: AutoRulePayload,
                     user: dict = Depends(require_role('admin'))):
    conn = get_conn()
    try:
        row = conn.execute("SELECT id FROM auto_match_rules WHERE id=?", (rule_id,)).fetchone()
        if row is None:
            raise HTTPException(404, "Rule not found")
        now = datetime.utcnow().isoformat()
        conn.execute(
            "UPDATE auto_match_rules SET name=?, description=?, priority=?, active=?, "
            "require_tier=?, require_amount_exact=?, require_ref_match=?, "
            "max_amount_diff=?, require_same_date=?, updated_at=? WHERE id=?",
            (payload.name, payload.description, payload.priority, payload.active,
             payload.require_tier, payload.require_amount_exact, payload.require_ref_match,
             payload.max_amount_diff, payload.require_same_date, now, rule_id)
        )
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'auto_rule_updated', ?, ?, ?)",
            (user['username'], now, json.dumps({'rule_id': rule_id}))
        )
        conn.commit()
        return dict(conn.execute("SELECT * FROM auto_match_rules WHERE id=?", (rule_id,)).fetchone())
    finally:
        conn.close()

@app.delete("/auto-rules/{rule_id}")
def deactivate_auto_rule(rule_id: int, user: dict = Depends(require_role('admin'))):
    conn = get_conn()
    try:
        row = conn.execute("SELECT id FROM auto_match_rules WHERE id=?", (rule_id,)).fetchone()
        if row is None:
            raise HTTPException(404, "Rule not found")
        now = datetime.utcnow().isoformat()
        conn.execute("UPDATE auto_match_rules SET active=0, updated_at=? WHERE id=?", (now, rule_id))
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'auto_rule_deactivated', ?, ?, ?)",
            (user['username'], now, json.dumps({'rule_id': rule_id}))
        )
        conn.commit()
        return {"rule_id": rule_id, "active": 0}
    finally:
        conn.close()

@app.post("/sessions/{session_id}/auto-match")
def manual_auto_match(session_id: int, user: dict = Depends(require_role('admin', 'ops'))):
    """Manually trigger the auto-match rule engine on a session's pending assignments."""
    from auto_match_engine import apply_auto_rules
    conn = get_conn()
    try:
        _assert_session_not_locked(conn, session_id)
        result = apply_auto_rules(conn, session_id, actor=f'manual_auto:{user["username"]}')
        conn.commit()
        return result
    finally:
        conn.close()

@app.get("/admin/auto-rules")
def auto_rules_page(request: Request):
    return templates.TemplateResponse(request, "auto_rules.html")


# ---------------------------------------------------------------------------
# BYO format profiles — admin-managed CSV column mappings.
#
# Workflow:
#   1. Admin opens /byo-formats, clicks "Add format".
#   2. POST /csv-profiles/preview with a sample upload + minimal config
#      (delimiter / skip_rows / header_row). Server returns the detected
#      columns, a sample of rows, and an auto-guessed column map.
#   3. Admin tweaks the mapping in the UI, names the profile.
#   4. POST /csv-profiles to save.
#   5. Profile appears in the Flex-side dropdown on the upload page; the
#      ingest path uses byo_csv_loader to parse the file using the
#      profile instead of the default Flex xlsx loader.
# ---------------------------------------------------------------------------

class CsvProfilePayload(BaseModel):
    name: str
    delimiter: str = ','
    header_row: int = 1
    skip_rows: int = 0
    date_format: str = '%Y-%m-%d'
    currency: Optional[str] = None
    column_map: dict        # {amount, value_date, ref, narration, type, currency, ac_no, ac_branch, booking_date}
    sign_convention: str = 'positive_credit'
    sign_column: Optional[str] = None
    sample_filename: Optional[str] = None
    # Optional binding to a specific account. When set, files using this
    # profile inherit the account's currency and are routed to that
    # account at ingest time without needing an ac_no column in the data.
    account_id: Optional[int] = None
    # Glob pattern (e.g. 'acme_gl_*.csv'). When set, the daily scanner
    # picks up matching files from messages/flexcube/ and ingests them
    # automatically. Null = manual-upload-only.
    filename_pattern: Optional[str] = None


def _validate_profile_payload(p: CsvProfilePayload) -> None:
    """Reject obviously-bad profile shapes early so we don't write
    garbage that breaks the loader at first use.

    Currency policy (post-pilot fix):
        - currency column set → use that
        - profile.currency set → use that
        - account_id set → fall back to the bound account's currency at ingest
        - none of the above → reject (can't bucket txns by currency)
    """
    from byo_csv_loader import VALID_DELIMITERS, VALID_SIGN_CONVENTIONS, REQUIRED_COLUMNS
    if not (p.name or '').strip():
        raise HTTPException(400, "name is required")
    if p.delimiter not in VALID_DELIMITERS:
        raise HTTPException(400, f"delimiter must be one of {VALID_DELIMITERS!r}")
    if p.sign_convention not in VALID_SIGN_CONVENTIONS:
        raise HTTPException(400,
            f"sign_convention must be one of {VALID_SIGN_CONVENTIONS!r}")
    if p.sign_convention != 'positive_credit' and not p.sign_column:
        raise HTTPException(400,
            f"sign_convention={p.sign_convention!r} requires sign_column")
    for k in REQUIRED_COLUMNS:
        if not p.column_map.get(k):
            raise HTTPException(400, f"column_map missing required field {k!r}")
    if (not p.currency and not p.column_map.get('currency')
            and not p.account_id):
        raise HTTPException(400,
            "Currency must come from somewhere: a currency column, a "
            "profile-level currency, or a bound account. Otherwise the "
            "engine can't bucket txns by currency.")


@app.get("/csv-profiles")
def list_csv_profiles(user: dict = Depends(current_user)):
    """Read-only for non-admins so the upload page can show the dropdown
    without requiring elevated rights."""
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT id, name, delimiter, header_row, skip_rows, date_format, "
            "currency, column_map, sign_convention, sign_column, "
            "sample_filename, account_id, filename_pattern, "
            "created_by, created_at, updated_at, active "
            "FROM csv_format_profiles WHERE active=1 ORDER BY name"
        ).fetchall()
        out = []
        for r in rows:
            d = dict(r)
            try:
                d['column_map'] = json.loads(d['column_map'])
            except (TypeError, ValueError):
                d['column_map'] = {}
            out.append(d)
        return out
    finally:
        conn.close()


@app.post("/csv-profiles")
def create_csv_profile(payload: CsvProfilePayload,
                       user: dict = Depends(require_role('admin'))):
    _validate_profile_payload(payload)
    # If account_id is set, verify it exists — a dangling FK isn't enforced
    # at the SQLite layer in the existing schema and we want a clean 400.
    if payload.account_id is not None:
        conn = get_conn()
        try:
            row = conn.execute("SELECT id FROM accounts WHERE id=? AND active=1",
                               (payload.account_id,)).fetchone()
            if row is None:
                raise HTTPException(400,
                    f"account_id {payload.account_id} does not exist or is inactive.")
        finally:
            conn.close()
    now = datetime.utcnow().isoformat()
    conn = get_conn()
    try:
        try:
            cur = conn.execute(
                "INSERT INTO csv_format_profiles (name, side, delimiter, header_row, "
                "skip_rows, date_format, currency, column_map, sign_convention, "
                "sign_column, sample_filename, account_id, filename_pattern, "
                "created_by, created_at, updated_at) "
                "VALUES (?, 'flex', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (payload.name.strip(), payload.delimiter, payload.header_row,
                 payload.skip_rows, payload.date_format, payload.currency,
                 json.dumps(payload.column_map), payload.sign_convention,
                 payload.sign_column, payload.sample_filename,
                 payload.account_id, payload.filename_pattern,
                 user['username'], now, now),
            )
        except Exception as exc:
            if 'UNIQUE' in str(exc):
                raise HTTPException(409,
                    f"A profile named {payload.name!r} already exists.")
            raise
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'csv_profile_created', ?, ?, ?)",
            (user['username'], now, json.dumps({
                'profile_id': cur.lastrowid, 'name': payload.name,
            })),
        )
        conn.commit()
        return {'id': cur.lastrowid, 'name': payload.name, 'created_at': now}
    finally:
        conn.close()


@app.delete("/csv-profiles/{profile_id}")
def delete_csv_profile(profile_id: int,
                        user: dict = Depends(require_role('admin'))):
    """Soft-delete via active=0 — preserves any sessions ingested under
    this profile so the audit trail still resolves."""
    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT name FROM csv_format_profiles WHERE id=?", (profile_id,)
        ).fetchone()
        if row is None:
            raise HTTPException(404, "profile not found")
        now = datetime.utcnow().isoformat()
        conn.execute("UPDATE csv_format_profiles SET active=0, updated_at=? "
                     "WHERE id=?", (now, profile_id))
        conn.execute(
            "INSERT INTO audit_log (session_id, action, actor, timestamp, details) "
            "VALUES (NULL, 'csv_profile_deleted', ?, ?, ?)",
            (user['username'], now, json.dumps({
                'profile_id': profile_id, 'name': row['name'],
            })),
        )
        conn.commit()
        return {'ok': True}
    finally:
        conn.close()


@app.post("/csv-profiles/preview")
async def preview_csv_profile(file: UploadFile = File(...),
                               delimiter: str = ',',
                               header_row: int = 1,
                               skip_rows: int = 0,
                               user: dict = Depends(require_role('admin'))):
    """Wizard step 2: take a sample upload, return the columns we found,
    a sample of rows, and a suggested column-map. Read-only — does NOT
    save anything to the DB."""
    from byo_csv_loader import autoguess_mapping, VALID_DELIMITERS
    if delimiter not in VALID_DELIMITERS:
        raise HTTPException(400, f"delimiter must be one of {VALID_DELIMITERS!r}")
    # Preview only inspects the header + first 5 sample rows. Cap the
    # read at PREVIEW_BYTES so a 250 MB cards-side CSV doesn't load
    # in full just to show column names. xlsx is a special case —
    # truncating breaks the zip's central directory (located at the
    # end of the archive), so for xlsx we read the whole file. xlsx
    # statements rarely exceed 50 MB even for high-volume operators.
    PREVIEW_BYTES = 5 * 1024 * 1024
    content = await file.read(PREVIEW_BYTES + 1)
    looks_like_xlsx = content[:4] == b'PK\x03\x04'
    if looks_like_xlsx and len(content) > PREVIEW_BYTES:
        # Got a partial xlsx — read the rest so openpyxl can parse the
        # zip central directory at the tail. Capped at MAX_REQUEST_BYTES
        # by the surrounding middleware.
        rest = await file.read()
        content = content + rest
    elif len(content) > PREVIEW_BYTES:
        # CSV path — truncate to the last newline within the cap so we
        # don't pass a half-row to the parser.
        cut = content.rfind(b'\n', 0, PREVIEW_BYTES)
        content = content[:cut if cut > 0 else PREVIEW_BYTES]
    if not content:
        raise HTTPException(400, "file is empty")
    try:
        out = autoguess_mapping(
            content, delimiter=delimiter,
            skip_rows=int(skip_rows), header_row=int(header_row),
        )
    except Exception as exc:
        raise HTTPException(400, f"could not parse file: {exc}")
    out['filename'] = file.filename
    return out


@app.get("/byo-formats")
def byo_formats_page(request: Request):
    return templates.TemplateResponse(request, "byo_formats.html")


@app.get("/mobile-money")
def mobile_money_page(request: Request):
    """Wallets-only dashboard. The recon engine and BYO loader handle
    everything underneath; this page is just a domain-specific view of
    the same data with provider grouping + per-wallet metrics."""
    return templates.TemplateResponse(request, "mobile_money.html")


# ===========================================================================
# Cards module endpoints.
#
# Settlement-file ingest + listing. Mirrors the cash side's POST /sessions
# but for the cards domain: one settlement file per ingest (not a pair),
# routed through PCI-safe parsers in cards_loaders/.
#
# CSV via BYO profile is the only working path today. Visa Base II /
# Mastercard IPM are stubbed pending scheme-published synthetic test data
# (Visa V.I.P., Mastercard PUF). See docs/CARDS_DESIGN.md.
# ===========================================================================

@app.get("/cards")
def cards_page(request: Request):
    """Cards-recon dashboard. Read-only view of card_settlement_files
    rows + recent records, grouped by scheme and role. Operators upload
    via this page (admin-only); audit/ops can view but not ingest."""
    return templates.TemplateResponse(request, "cards.html")


@app.post("/cards/files")
async def upload_card_settlement(
    file: UploadFile = File(...),
    scheme: str = Form(...),
    role: str = Form('issuer'),
    settlement_date: Optional[str] = Form(None),
    currency: Optional[str] = Form(None),
    profile_id: Optional[int] = Form(None),
    pan_field: Optional[str] = Form(None),
    pan_masked_field: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
    stage: Optional[str] = Form(None),
    user: dict = Depends(require_role('admin')),
):
    """Ingest one cards settlement file. Body shape:
        - file: the settlement CSV (or .out/.dat once binary parsers ship).
        - scheme: visa | mastercard | verve | gh_cardlink | other.
        - role: issuer | acquirer | switch (default: issuer).
        - settlement_date: YYYY-MM-DD; falls through to per-row dates if
          the file has its own column for them.
        - currency: ISO 4217; falls through to per-row currency if mapped.
        - profile_id: a saved csv_format_profiles row id (required for
          CSV until binary parsers land).
        - pan_field: optional name of a column carrying full PAN; gets
          masked at the parser seam.

    Returns: file_id + record_count + total_amount + skipped_records.
    The file row is idempotent on sha256 — re-uploading the same bytes
    yields a 409 with the existing file_id."""
    if not (file.filename or '').lower().endswith(
            ('.csv', '.txt', '.out', '.dat')):
        raise HTTPException(
            400, "Settlement file must be .csv, .txt, .out, or .dat")
    if settlement_date:
        try:
            datetime.strptime(settlement_date, '%Y-%m-%d')
        except ValueError:
            raise HTTPException(
                400, "settlement_date must be YYYY-MM-DD or omitted.")

    file_path = await _save_upload(file, 'cards')

    from cards_ingest import (
        ingest_card_settlement, CardsIngestError, DuplicateCardFileError,
    )
    try:
        result = ingest_card_settlement(
            file_path=file_path,
            scheme=scheme, role=role,
            settlement_date=settlement_date,
            currency=currency,
            ingested_by=user['username'],
            original_filename=file.filename,
            profile_id=profile_id,
            pan_field=pan_field,
            pan_masked_field=pan_masked_field,
            notes=notes,
            stage=stage,
        )
    except DuplicateCardFileError as exc:
        raise HTTPException(409, str(exc))
    except CardsIngestError as exc:
        raise HTTPException(400, str(exc))

    return {
        'file_id': result.file_id,
        'scheme': result.scheme,
        'role': result.role,
        'stage': result.stage,
        'record_count': result.record_count,
        'total_amount': result.total_amount,
        'currency': result.currency,
        'settlement_date': result.settlement_date,
        'skipped_records': result.skipped_records,
    }


@app.get("/cards/files")
def list_card_files(user: dict = Depends(current_user)):
    """List recent settlement files, newest first. No PII exposed beyond
    file-level totals — record-level data is fetched separately."""
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT id, scheme, role, settlement_date, processing_date, "
            "record_count, total_amount, currency, original_filename, "
            "ingested_at, ingested_by "
            "FROM card_settlement_files "
            "ORDER BY ingested_at DESC LIMIT 100"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.get("/cards/files/{file_id}/records")
def list_card_records(file_id: int, limit: int = 100, offset: int = 0,
                       user: dict = Depends(current_user)):
    """List records in a settlement file. Paged — settlement files
    can run to 100k+ rows for medium issuers, so we never return the
    whole batch in one response."""
    if limit > 500:
        limit = 500
    conn = get_conn()
    try:
        meta = conn.execute(
            "SELECT id, scheme, role, settlement_date, record_count, "
            "total_amount, currency, original_filename "
            "FROM card_settlement_files WHERE id=?",
            (file_id,),
        ).fetchone()
        if meta is None:
            raise HTTPException(404, f"Card settlement file {file_id} not found.")
        rows = conn.execute(
            "SELECT record_index, pan_first6, pan_last4, scheme_ref, "
            "auth_code, merchant_id, merchant_name, mcc, terminal_id, "
            "transaction_type, amount_settlement, currency_settlement, "
            "transaction_date, settlement_date, recon_status "
            "FROM card_settlement_records WHERE file_id=? "
            "ORDER BY record_index LIMIT ? OFFSET ?",
            (file_id, limit, offset),
        ).fetchall()
        return {
            'file': dict(meta),
            'records': [dict(r) for r in rows],
            'limit': limit, 'offset': offset,
        }
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Cards matching engine — scheme_ref join across files.
# Surfaces 'matched' / 'mismatched' / 'unmatched' groups so operators can
# focus on the cases that need investigation. The engine is read-only by
# default (GET /cards/match/groups); POST /cards/match/recompute persists
# the resulting status into card_settlement_records.recon_status.
# ---------------------------------------------------------------------------

@app.get("/cards/match/groups")
def list_card_match_groups(
    scheme: Optional[str] = None,
    settlement_date_from: Optional[str] = None,
    settlement_date_to: Optional[str] = None,
    status: Optional[str] = None,
    limit: int = 200,
    user: dict = Depends(current_user),
):
    """Compute scheme_ref groups on-the-fly and return them. Pure read —
    does not write to records. Use POST /cards/match/recompute to
    persist statuses."""
    from cards_engine import compute_match_groups
    if limit > 1000:
        limit = 1000
    conn = get_conn()
    try:
        groups = compute_match_groups(
            conn,
            scheme=scheme,
            settlement_date_from=settlement_date_from,
            settlement_date_to=settlement_date_to,
        )
        if status:
            groups = [g for g in groups if g.status == status]
        out = []
        for g in groups[:limit]:
            out.append({
                'scheme_ref': g.scheme_ref,
                'status': g.status,
                'record_count': g.record_count,
                'file_count': g.file_count,
                'file_ids': g.file_ids,
                'amount_min': g.amount_min,
                'amount_max': g.amount_max,
                'amount_total': g.amount_total,
                'amount_spread': g.amount_spread,
                'currencies': g.currencies,
                'pan_last4_set': g.pan_last4_set,
                'schemes': g.schemes,
                'stages': g.stages,
            })
        # Header counts come from the unfiltered total so the UI can
        # show "X matched / Y mismatched" without a second roundtrip.
        all_groups = groups
        return {
            'groups': out,
            'total_groups': len(all_groups),
            'returned': len(out),
            'counts': {
                'matched': sum(1 for g in all_groups if g.status == 'matched'),
                'mismatched': sum(1 for g in all_groups if g.status == 'mismatched'),
                'unmatched': sum(1 for g in all_groups if g.status == 'unmatched'),
                'incomplete': sum(1 for g in all_groups if g.status == 'incomplete'),
            },
        }
    finally:
        conn.close()


@app.get("/cards/match/groups/{scheme_ref}/records")
def list_records_for_scheme_ref(scheme_ref: str,
                                  user: dict = Depends(current_user)):
    """All records carrying a given scheme_ref, plus the file metadata
    each came from. Used by the cards UI's drill-in panel — operators
    click a mismatched group and see exactly which records disagree."""
    if not scheme_ref or len(scheme_ref) > 100:
        raise HTTPException(400, "scheme_ref required (max 100 chars).")
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT r.id, r.record_index, r.pan_first6, r.pan_last4, "
            "r.scheme_ref, r.auth_code, r.merchant_id, r.merchant_name, "
            "r.mcc, r.amount_settlement, r.currency_settlement, "
            "r.amount_transaction, r.currency_transaction, r.fx_rate, "
            "r.fee_total, r.transaction_date, r.settlement_date, "
            "r.recon_status, r.matched_at, r.matched_by, "
            "f.id AS file_id, f.scheme, f.role, f.original_filename, "
            "f.ingested_at "
            "FROM card_settlement_records r "
            "JOIN card_settlement_files f ON f.id = r.file_id "
            "WHERE r.scheme_ref=? "
            "ORDER BY f.role, r.record_index",
            (scheme_ref,),
        ).fetchall()
        return {
            'scheme_ref': scheme_ref,
            'records': [dict(r) for r in rows],
        }
    finally:
        conn.close()


@app.get("/cards/files/export")
def export_card_files(user: dict = Depends(current_user)):
    """CSV of all settlement files — file_id, scheme, role, dates,
    counts, totals. Same shape as the /cards "Files" tab, dumped for
    spreadsheet workflows / month-end reports."""
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT id, scheme, role, settlement_date, processing_date, "
            "record_count, total_amount, currency, original_filename, "
            "ingested_at, ingested_by, sha256 "
            "FROM card_settlement_files ORDER BY ingested_at DESC"
        ).fetchall()
    finally:
        conn.close()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['file_id', 'scheme', 'role', 'settlement_date',
                'processing_date', 'record_count', 'total_amount',
                'currency', 'original_filename', 'ingested_at',
                'ingested_by', 'sha256'])
    for r in rows:
        w.writerow([r['id'], r['scheme'], r['role'],
                    r['settlement_date'] or '', r['processing_date'] or '',
                    r['record_count'], f"{r['total_amount'] or 0:.2f}",
                    r['currency'] or '', r['original_filename'] or '',
                    r['ingested_at'], r['ingested_by'], r['sha256']])
    _audit_export(user['username'], 'cards_files_export', {"rows": len(rows)})
    fname = f"kilter_cards_files_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


@app.get("/cards/match/groups/export")
def export_card_match_groups(
    scheme: Optional[str] = None,
    settlement_date_from: Optional[str] = None,
    settlement_date_to: Optional[str] = None,
    status: Optional[str] = None,
    user: dict = Depends(current_user),
):
    """CSV of match groups — same filter args as GET /cards/match/groups,
    but no row cap. Ops sends this to the recon meeting — every
    mismatched scheme_ref with its spread, file mix, and amount range."""
    from cards_engine import compute_match_groups
    conn = get_conn()
    try:
        groups = compute_match_groups(
            conn, scheme=scheme,
            settlement_date_from=settlement_date_from,
            settlement_date_to=settlement_date_to,
        )
    finally:
        conn.close()
    if status:
        groups = [g for g in groups if g.status == status]

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['scheme_ref', 'status', 'record_count', 'file_count',
                'file_ids', 'amount_min', 'amount_max', 'amount_total',
                'amount_spread', 'currencies', 'schemes', 'pan_last4_set'])
    for g in groups:
        w.writerow([
            g.scheme_ref, g.status, g.record_count, g.file_count,
            ';'.join(str(x) for x in g.file_ids),
            f"{g.amount_min:.2f}", f"{g.amount_max:.2f}",
            f"{g.amount_total:.2f}", f"{g.amount_spread:.2f}",
            ';'.join(g.currencies), ';'.join(g.schemes),
            ';'.join(g.pan_last4_set),
        ])
    _audit_export(user['username'], 'cards_match_groups_export', {
        "rows": len(groups),
        "filters": {"scheme": scheme, "status": status,
                    "settlement_date_from": settlement_date_from,
                    "settlement_date_to": settlement_date_to},
    })
    fname = f"kilter_cards_matches_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


@app.get("/cards/files/{file_id}/records/export")
def export_card_records(file_id: int, user: dict = Depends(current_user)):
    """CSV of every record in one settlement file — full per-row drill-
    down for spreadsheet reconciliation. PCI-safe by construction:
    only first6 / last4 leave the schema."""
    conn = get_conn()
    try:
        meta = conn.execute(
            "SELECT id, original_filename FROM card_settlement_files WHERE id=?",
            (file_id,),
        ).fetchone()
        if meta is None:
            raise HTTPException(404, f"Card settlement file {file_id} not found.")
        rows = conn.execute(
            "SELECT record_index, pan_first6, pan_last4, scheme_ref, "
            "auth_code, merchant_id, merchant_name, mcc, terminal_id, "
            "transaction_type, amount_settlement, currency_settlement, "
            "amount_transaction, currency_transaction, fx_rate, "
            "fee_total, transaction_date, settlement_date, recon_status, "
            "matched_at, matched_by "
            "FROM card_settlement_records WHERE file_id=? ORDER BY record_index",
            (file_id,),
        ).fetchall()
    finally:
        conn.close()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['record_index', 'pan_first6', 'pan_last4', 'scheme_ref',
                'auth_code', 'merchant_id', 'merchant_name', 'mcc',
                'terminal_id', 'transaction_type', 'amount_settlement',
                'currency_settlement', 'amount_transaction',
                'currency_transaction', 'fx_rate', 'fee_total',
                'transaction_date', 'settlement_date', 'recon_status',
                'matched_at', 'matched_by'])
    for r in rows:
        w.writerow([r[k] if r[k] is not None else '' for k in r.keys()])
    _audit_export(user['username'], 'cards_records_export',
                  {"file_id": file_id, "rows": len(rows)})
    fname = (f"kilter_cards_records_file{file_id}_"
             f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv")
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


@app.post("/cards/match/recompute")
def recompute_card_matches(
    scheme: Optional[str] = Form(None),
    settlement_date_from: Optional[str] = Form(None),
    settlement_date_to: Optional[str] = Form(None),
    user: dict = Depends(require_role('admin')),
):
    """Recompute match groups and write the resulting status into
    `card_settlement_records.recon_status`. Records in operator-set
    state ('disputed', 'written_off') are preserved.

    Idempotent — re-running over unchanged data yields the same result.
    Wrapped in a single transaction so a partial failure leaves the
    table consistent."""
    from cards_engine import compute_match_groups, apply_match_status
    conn = get_conn()
    try:
        groups = compute_match_groups(
            conn,
            scheme=scheme,
            settlement_date_from=settlement_date_from,
            settlement_date_to=settlement_date_to,
        )
        result = apply_match_status(conn, groups, actor=user['username'])
        return {
            'groups_total': result.groups_total,
            'matched': result.matched,
            'mismatched': result.mismatched,
            'unmatched': result.unmatched,
            'records_updated': result.records_updated,
            'records_protected': result.records_protected,
        }
    finally:
        conn.close()


# ===========================================================================
# Reports endpoints (2026-04-22).
#
# Cross-session / cross-account views for pulling archived data. The
# per-account status page and the review page handle single-session drill-in;
# these endpoints answer "what happened across the fleet over time?"
#
# Every endpoint supports ?format=csv — the audit team exports these
# straight into spreadsheets, so raw-data export is a first-class feature.
# ===========================================================================

def _parse_ymd(s: Optional[str]) -> Optional[str]:
    """Accept YYYY-MM-DD, return ISO start-of-day string, or None."""
    if not s:
        return None
    try:
        dt = datetime.strptime(s.strip(), '%Y-%m-%d')
        return dt.isoformat()
    except ValueError:
        raise HTTPException(400, f"Invalid date '{s}'. Use YYYY-MM-DD.")


def _parse_ymd_end(s: Optional[str]) -> Optional[str]:
    """Same as _parse_ymd but anchored to 23:59:59 so a single-day filter
    includes everything on that day."""
    if not s:
        return None
    try:
        dt = datetime.strptime(s.strip(), '%Y-%m-%d')
        return dt.replace(hour=23, minute=59, second=59).isoformat()
    except ValueError:
        raise HTTPException(400, f"Invalid date '{s}'. Use YYYY-MM-DD.")


def _csv_response(rows: list[list], headers: list[str], filename: str) -> StreamingResponse:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for r in rows:
        w.writerow(r)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


# --- Report #1: Archived sessions ------------------------------------------

@app.get("/reports/sessions")
def report_sessions(
    account_id: Optional[int] = None,
    from_date: Optional[str] = None,      # created_at >= YYYY-MM-DD
    to_date: Optional[str] = None,        # created_at <= YYYY-MM-DD 23:59:59
    account_label: Optional[str] = None,  # substring, case-insensitive
    status: Optional[str] = None,         # 'open' | 'closed'
    format: str = 'json',
    user: dict = Depends(current_user),
    scope: Optional[List[str]] = Depends(active_scope),
):
    """Archived sessions, filterable and ready to re-download. Scoped by
    access area so a BRANCH 001 analyst doesn't see NOSTRO reconciliations."""
    conn = get_conn()
    try:
        where = ["1=1"]
        params: list = []
        if account_id is not None:
            where.append("s.account_id = ?"); params.append(account_id)
        from_iso = _parse_ymd(from_date)
        to_iso = _parse_ymd_end(to_date)
        if from_iso:
            where.append("s.created_at >= ?"); params.append(from_iso)
        if to_iso:
            where.append("s.created_at <= ?"); params.append(to_iso)
        if account_label:
            where.append("LOWER(COALESCE(a.label, '')) LIKE ?")
            params.append(f"%{account_label.lower()}%")
        if status == 'closed':
            where.append("s.closed_at IS NOT NULL")
        elif status == 'open':
            where.append("s.closed_at IS NULL")
        scope_sql, scope_params = _scope_clause(scope, alias='a')
        where.append(scope_sql); params.extend(scope_params)
        if scope:
            where.append("a.id IS NOT NULL")

        rows = conn.execute(
            f"SELECT s.id, s.created_at, s.created_by, s.swift_filename, s.flex_filename, "
            f"       s.swift_account, s.flex_ac_no, s.swift_currency, s.flex_currency, "
            f"       s.account_label, s.closed_at, s.closed_by, "
            f"       s.opening_balance_amount, s.closing_balance_amount, "
            f"       s.open_items_seeded, s.open_items_cleared, "
            f"       a.shortname AS account_shortname, a.access_area, a.id AS account_id, "
            f"       (SELECT COUNT(*) FROM assignments WHERE session_id=s.id AND status='pending')   AS pending, "
            f"       (SELECT COUNT(*) FROM assignments WHERE session_id=s.id AND status='confirmed') AS confirmed, "
            f"       (SELECT COUNT(*) FROM assignments WHERE session_id=s.id AND status='rejected')  AS rejected, "
            f"       (SELECT COUNT(*) FROM swift_txns   WHERE session_id=s.id) AS swift_total, "
            f"       (SELECT COUNT(*) FROM flex_txns    WHERE session_id=s.id) AS flex_total "
            f"FROM sessions s LEFT JOIN accounts a ON a.id = s.account_id "
            f"WHERE {' AND '.join(where)} "
            f"ORDER BY s.id DESC LIMIT 2000",
            params,
        ).fetchall()

        out = []
        for r in rows:
            d = dict(r)
            t = (d['confirmed'] or 0) + (d['pending'] or 0) + (d['rejected'] or 0)
            d['match_rate_pct'] = round((d['confirmed'] or 0) / t * 100.0, 1) if t else 0.0
            out.append(d)

        if format == 'csv':
            headers = ['session_id', 'created_at', 'created_by', 'account',
                       'access_area', 'currency', 'swift_total', 'flex_total',
                       'confirmed', 'pending', 'rejected', 'match_rate_pct',
                       'open_items_seeded', 'open_items_cleared',
                       'opening_balance', 'closing_balance', 'closed_at']
            data = [
                [r['id'], r['created_at'], r['created_by'],
                 r['account_shortname'] or r['account_label'] or '',
                 r['access_area'] or '', r['swift_currency'] or r['flex_currency'] or '',
                 r['swift_total'], r['flex_total'], r['confirmed'], r['pending'],
                 r['rejected'], r['match_rate_pct'], r['open_items_seeded'] or 0,
                 r['open_items_cleared'] or 0, r['opening_balance_amount'] or '',
                 r['closing_balance_amount'] or '', r['closed_at'] or '']
                for r in out
            ]
            return _csv_response(data, headers,
                f"kilter_sessions_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv")
        return out
    finally:
        conn.close()


# --- Report #3: Cleared & written-off items (archive of the ledger) --------

@app.get("/reports/cleared-items")
def report_cleared_items(
    account_id: Optional[int] = None,
    from_date: Optional[str] = None,       # cleared_at >= ...
    to_date: Optional[str] = None,
    category: Optional[str] = None,
    outcome: Optional[str] = None,         # 'cleared' | 'written_off' | null = both
    format: str = 'json',
    user: dict = Depends(current_user),
    scope: Optional[List[str]] = Depends(active_scope),
):
    """The archive side of the rolling ledger: everything that *used* to
    be an open item and has since been resolved. Essential for audit trails
    — proves a break was closed, by whom, via which mechanism."""
    conn = get_conn()
    try:
        where = ["oi.status IN ('cleared', 'written_off')"]
        params: list = []
        if account_id is not None:
            where.append("oi.account_id = ?"); params.append(account_id)
        if outcome == 'cleared':
            where.append("oi.status = 'cleared'")
        elif outcome == 'written_off':
            where.append("oi.status = 'written_off'")
        if category:
            where.append("oi.category = ?"); params.append(category)
        from_iso = _parse_ymd(from_date)
        to_iso = _parse_ymd_end(to_date)
        if from_iso:
            where.append("oi.cleared_at >= ?"); params.append(from_iso)
        if to_iso:
            where.append("oi.cleared_at <= ?"); params.append(to_iso)
        scope_sql, scope_params = _scope_clause(scope, alias='a')
        where.append(scope_sql); params.extend(scope_params)
        if scope:
            where.append("a.id IS NOT NULL")

        rows = conn.execute(
            f"SELECT oi.*, a.label AS account_label, a.shortname AS account_shortname, "
            f"       a.access_area, a.currency AS account_currency "
            f"FROM open_items oi LEFT JOIN accounts a ON a.id = oi.account_id "
            f"WHERE {' AND '.join(where)} "
            f"ORDER BY oi.cleared_at DESC LIMIT 5000",
            params,
        ).fetchall()
        out = [dict(r) for r in rows]

        if format == 'csv':
            headers = ['open_item_id', 'account', 'access_area', 'currency',
                       'source_side', 'value_date', 'amount', 'sign', 'ref',
                       'narration', 'category', 'category_source', 'status',
                       'opened_at', 'cleared_at', 'cleared_by', 'cleared_via',
                       'cleared_session_id', 'cleared_assignment_id',
                       'write_off_reason', 'src_session_id', 'src_row_number']
            data = [
                [r['id'], r['account_shortname'] or r['account_label'] or '',
                 r['access_area'] or '', r['account_currency'] or '',
                 r['source_side'], r['value_date'] or '', r['amount'],
                 r['sign'] or '', r['ref'] or '', r['narration'] or '',
                 r['category'] or '', r['category_source'] or '', r['status'],
                 r['opened_at'], r['cleared_at'] or '', r['cleared_by'] or '',
                 r['cleared_via'] or '', r['cleared_session_id'] or '',
                 r['cleared_assignment_id'] or '', r['write_off_reason'] or '',
                 r['src_session_id'], r['src_row_number']]
                for r in out
            ]
            return _csv_response(data, headers,
                f"kilter_cleared_items_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv")
        return out
    finally:
        conn.close()


# --- Report #2: Account reconciliation history -----------------------------

@app.get("/reports/account-history/{account_id}")
def report_account_history(
    account_id: int,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    format: str = 'json',
    user: dict = Depends(current_user),
):
    """Per-session timeline for one account: match rate, open items added
    vs. cleared, closing-balance trend. The shape the UI wants for
    charting + the answer to 'is this nostro getting cleaner or messier?'"""
    conn = get_conn()
    try:
        acc = conn.execute(
            "SELECT id, label, shortname, access_area, swift_account, flex_ac_no, "
            "       currency FROM accounts WHERE id=?", (account_id,),
        ).fetchone()
        if acc is None:
            raise HTTPException(404, "Account not found")

        where = ["s.account_id = ?"]
        params: list = [account_id]
        from_iso = _parse_ymd(from_date)
        to_iso = _parse_ymd_end(to_date)
        if from_iso:
            where.append("s.created_at >= ?"); params.append(from_iso)
        if to_iso:
            where.append("s.created_at <= ?"); params.append(to_iso)

        rows = conn.execute(
            f"SELECT s.id, s.created_at, s.swift_filename, s.flex_filename, "
            f"       s.opening_balance_amount, s.closing_balance_amount, "
            f"       s.opening_balance_date, s.closing_balance_date, "
            f"       s.open_items_seeded, s.open_items_cleared, "
            f"       (SELECT COUNT(*) FROM assignments WHERE session_id=s.id AND status='confirmed') AS confirmed, "
            f"       (SELECT COUNT(*) FROM assignments WHERE session_id=s.id AND status='pending')   AS pending, "
            f"       (SELECT COUNT(*) FROM assignments WHERE session_id=s.id AND status='rejected')  AS rejected, "
            f"       (SELECT COUNT(*) FROM swift_txns WHERE session_id=s.id) AS swift_total, "
            f"       (SELECT COUNT(*) FROM flex_txns  WHERE session_id=s.id) AS flex_total "
            f"FROM sessions s WHERE {' AND '.join(where)} "
            f"ORDER BY s.created_at",
            params,
        ).fetchall()

        series = []
        for r in rows:
            d = dict(r)
            t = (d['confirmed'] or 0) + (d['pending'] or 0) + (d['rejected'] or 0)
            d['match_rate_pct'] = round((d['confirmed'] or 0) / t * 100.0, 1) if t else 0.0
            # Net change in open items for this session — helps spot runaway growth.
            d['open_items_delta'] = (d['open_items_seeded'] or 0) - (d['open_items_cleared'] or 0)
            series.append(d)

        if format == 'csv':
            headers = ['session_id', 'created_at', 'swift_total', 'flex_total',
                       'confirmed', 'pending', 'rejected', 'match_rate_pct',
                       'open_items_seeded', 'open_items_cleared', 'open_items_delta',
                       'opening_balance', 'closing_balance']
            data = [
                [r['id'], r['created_at'], r['swift_total'], r['flex_total'],
                 r['confirmed'], r['pending'], r['rejected'], r['match_rate_pct'],
                 r['open_items_seeded'] or 0, r['open_items_cleared'] or 0,
                 r['open_items_delta'], r['opening_balance_amount'] or '',
                 r['closing_balance_amount'] or '']
                for r in series
            ]
            return _csv_response(data, headers,
                f"kilter_account_{account_id}_history_"
                f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv")
        return {"account": dict(acc), "sessions": series}
    finally:
        conn.close()


# --- Report #4: Break analysis ---------------------------------------------

@app.get("/reports/break-analysis")
def report_break_analysis(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    account_id: Optional[int] = None,
    format: str = 'json',
    user: dict = Depends(current_user),
    scope: Optional[List[str]] = Depends(active_scope),
):
    """Pivots the open-items pool (open + written_off + cleared) by the
    three dimensions ops cares about: category, account, and month.
    Returns three separate tables — the UI composes them into stacked charts."""
    conn = get_conn()
    try:
        where = ["1=1"]
        params: list = []
        from_iso = _parse_ymd(from_date)
        to_iso = _parse_ymd_end(to_date)
        if from_iso:
            where.append("oi.opened_at >= ?"); params.append(from_iso)
        if to_iso:
            where.append("oi.opened_at <= ?"); params.append(to_iso)
        if account_id is not None:
            where.append("oi.account_id = ?"); params.append(account_id)
        scope_sql, scope_params = _scope_clause(scope, alias='a')
        where.append(scope_sql); params.extend(scope_params)
        if scope:
            where.append("a.id IS NOT NULL")
        where_sql = " AND ".join(where)

        # By category × status
        by_category = [dict(r) for r in conn.execute(
            f"SELECT COALESCE(oi.category, 'uncategorized') AS category, "
            f"       oi.status, COUNT(*) AS cnt, COALESCE(SUM(oi.amount), 0) AS total "
            f"FROM open_items oi LEFT JOIN accounts a ON a.id = oi.account_id "
            f"WHERE {where_sql} GROUP BY category, oi.status "
            f"ORDER BY category, oi.status",
            params,
        ).fetchall()]

        # By account
        by_account = [dict(r) for r in conn.execute(
            f"SELECT oi.account_id, a.label AS account_label, a.shortname, "
            f"       a.access_area, a.currency, oi.status, "
            f"       COUNT(*) AS cnt, COALESCE(SUM(oi.amount), 0) AS total "
            f"FROM open_items oi LEFT JOIN accounts a ON a.id = oi.account_id "
            f"WHERE {where_sql} GROUP BY oi.account_id, oi.status "
            f"ORDER BY a.label, oi.status",
            params,
        ).fetchall()]

        # By month (opened_at) — SQLite substr is cheaper than strftime here.
        by_month = [dict(r) for r in conn.execute(
            f"SELECT SUBSTR(oi.opened_at, 1, 7) AS month, oi.status, "
            f"       COUNT(*) AS cnt, COALESCE(SUM(oi.amount), 0) AS total "
            f"FROM open_items oi LEFT JOIN accounts a ON a.id = oi.account_id "
            f"WHERE {where_sql} GROUP BY month, oi.status ORDER BY month, oi.status",
            params,
        ).fetchall()]

        # Tier distribution for confirmed matches in the same window.
        tier_where = ["a2.decided_at IS NOT NULL", "a2.status='confirmed'"]
        tier_params: list = []
        if from_iso:
            tier_where.append("a2.decided_at >= ?"); tier_params.append(from_iso)
        if to_iso:
            tier_where.append("a2.decided_at <= ?"); tier_params.append(to_iso)
        if account_id is not None:
            tier_where.append("s.account_id = ?"); tier_params.append(account_id)
        by_tier = [dict(r) for r in conn.execute(
            f"SELECT a2.tier, a2.source, COUNT(*) AS cnt "
            f"FROM assignments a2 JOIN sessions s ON s.id = a2.session_id "
            f"WHERE {' AND '.join(tier_where)} "
            f"GROUP BY a2.tier, a2.source ORDER BY a2.tier, a2.source",
            tier_params,
        ).fetchall()]

        if format == 'csv':
            headers = ['dimension', 'key', 'status_or_source', 'count', 'total_value']
            data = []
            for r in by_category:
                data.append(['category', r['category'], r['status'], r['cnt'], r['total']])
            for r in by_account:
                key = r['shortname'] or r['account_label'] or f"account_{r['account_id']}"
                data.append(['account', key, r['status'], r['cnt'], r['total']])
            for r in by_month:
                data.append(['month', r['month'] or '', r['status'], r['cnt'], r['total']])
            for r in by_tier:
                data.append(['tier', f"T{r['tier']}", r['source'], r['cnt'], ''])
            return _csv_response(data, headers,
                f"kilter_break_analysis_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv")

        return {
            "by_category": by_category,
            "by_account": by_account,
            "by_month": by_month,
            "by_tier": by_tier,
        }
    finally:
        conn.close()


# --- Report #5: Decision activity ------------------------------------------

@app.get("/reports/decisions")
def report_decisions(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    actor: Optional[str] = None,
    format: str = 'json',
    user: dict = Depends(require_role('admin', 'audit', 'internal_control')),
):
    """Per-user decision productivity — who's doing the work, and which
    shapes of work (confirm/reject/manual/carry). Restricted to roles with
    audit visibility because it touches individual users' activity."""
    conn = get_conn()
    try:
        where = ["a.decided_at IS NOT NULL"]
        params: list = []
        from_iso = _parse_ymd(from_date)
        to_iso = _parse_ymd_end(to_date)
        if from_iso:
            where.append("a.decided_at >= ?"); params.append(from_iso)
        if to_iso:
            where.append("a.decided_at <= ?"); params.append(to_iso)
        if actor:
            where.append("a.decided_by = ?"); params.append(actor)

        # Per-user totals, split by status + source so manual matches don't
        # get double-counted with engine-proposed confirmations.
        by_user = [dict(r) for r in conn.execute(
            f"SELECT a.decided_by AS actor, a.status, COALESCE(a.source, 'engine') AS source, "
            f"       COUNT(*) AS cnt "
            f"FROM assignments a WHERE {' AND '.join(where)} "
            f"GROUP BY a.decided_by, a.status, source ORDER BY a.decided_by",
            params,
        ).fetchall()]

        # Daily activity for sparkline-style charting.
        by_day = [dict(r) for r in conn.execute(
            f"SELECT SUBSTR(a.decided_at, 1, 10) AS day, a.status, COUNT(*) AS cnt "
            f"FROM assignments a WHERE {' AND '.join(where)} "
            f"GROUP BY day, a.status ORDER BY day",
            params,
        ).fetchall()]

        # Write-offs + notes attributed to each user, from the ledger-adjacent
        # tables — useful for spotting analysts who never leave context.
        wo_where = ["oi.status='written_off'"]
        wo_params: list = []
        if from_iso:
            wo_where.append("oi.cleared_at >= ?"); wo_params.append(from_iso)
        if to_iso:
            wo_where.append("oi.cleared_at <= ?"); wo_params.append(to_iso)
        if actor:
            wo_where.append("oi.cleared_by = ?"); wo_params.append(actor)
        writeoffs = [dict(r) for r in conn.execute(
            f"SELECT oi.cleared_by AS actor, COUNT(*) AS cnt "
            f"FROM open_items oi WHERE {' AND '.join(wo_where)} "
            f"GROUP BY oi.cleared_by",
            wo_params,
        ).fetchall()]

        nt_where = ["1=1"]
        nt_params: list = []
        if from_iso:
            nt_where.append("bc.created_at >= ?"); nt_params.append(from_iso)
        if to_iso:
            nt_where.append("bc.created_at <= ?"); nt_params.append(to_iso)
        if actor:
            nt_where.append("bc.author = ?"); nt_params.append(actor)
        notes = [dict(r) for r in conn.execute(
            f"SELECT bc.author AS actor, COUNT(*) AS cnt "
            f"FROM break_comments bc WHERE {' AND '.join(nt_where)} "
            f"GROUP BY bc.author",
            nt_params,
        ).fetchall()]

        if format == 'csv':
            headers = ['actor', 'status', 'source', 'count']
            data = [[r['actor'] or '', r['status'], r['source'], r['cnt']] for r in by_user]
            return _csv_response(data, headers,
                f"kilter_decisions_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv")

        return {
            "by_user": by_user,
            "by_day": by_day,
            "writeoffs": writeoffs,
            "notes": notes,
        }
    finally:
        conn.close()


@app.get("/reports")
def reports_page(request: Request):
    return templates.TemplateResponse(request, "reports.html")


# ---------------------------------------------------------------------------
# Daily breaks report — multi-tab xlsx grouped by functional_group + currency.
# Mirrors ops' current manual "BOG CEDI / FOREIGN" report layout so the team
# can stop hand-building it every evening. Pulls open_items with status='open'
# as of the given date, groups by functional_group, and within each group
# sub-splits by currency when currencies mix.
# ---------------------------------------------------------------------------

@app.get("/reports/daily-breaks")
def report_daily_breaks(
    as_of: Optional[str] = None,     # YYYY-MM-DD; defaults to today
    account_id: Optional[int] = None,
    format: str = 'xlsx',         # 'xlsx' or 'json'
    user: dict = Depends(current_user),
    scope: Optional[List[str]] = Depends(active_scope),
):
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from db import FUNCTIONAL_GROUPS

    as_of_iso = _parse_ymd_end(as_of) if as_of else _parse_ymd_end(
        datetime.utcnow().strftime('%Y-%m-%d'))
    as_of_label = (as_of or datetime.utcnow().strftime('%Y-%m-%d'))

    conn = get_conn()
    try:
        where = ["oi.status='open'", "oi.opened_at <= ?"]
        params: list = [as_of_iso]
        if account_id is not None:
            where.append("oi.account_id=?"); params.append(account_id)
        scope_sql, scope_params = _scope_clause(scope, alias='a')
        where.append(scope_sql); params.extend(scope_params)
        if scope:
            where.append("a.id IS NOT NULL")

        rows = conn.execute(
            f"SELECT oi.*, a.label AS account_label, a.shortname, a.access_area, "
            f"       a.currency AS account_currency "
            f"FROM open_items oi LEFT JOIN accounts a ON a.id=oi.account_id "
            f"WHERE {' AND '.join(where)} "
            f"ORDER BY COALESCE(oi.functional_group, 'ZZZ'), oi.amount DESC",
            params,
        ).fetchall()

        now = datetime.utcnow()
        items = []
        for r in rows:
            d = dict(r)
            d['age_days'] = _age_in_days(r['opened_at'], now)
            items.append(d)
    finally:
        conn.close()

    if format == 'json':
        return {"as_of": as_of_label, "count": len(items), "items": items}

    # ----- build xlsx -----
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop default sheet; we create named ones below

    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='305496')
    total_font = Font(bold=True)
    total_fill = PatternFill('solid', fgColor='D9E1F2')

    COLS = ['Account', 'Value date', 'Curr.', 'Amount', 'S', 'Origin', 'Type',
            'Status', 'Age', 'Book. date', 'Our reference 1', 'Their reference 1',
            'Booking text 1', 'Booking text 2', 'Matching type']

    def _add_sheet(title: str, bucket: list[dict]):
        # Excel sheet names max 31 chars, no special chars.
        safe = title[:31].replace('/', '-').replace(':', '-')
        ws = wb.create_sheet(safe)
        ws.append(COLS)
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
        if not bucket:
            ws.append([''] * 5 + ['NO OPENED ITEMS'] + [''] * 9)
        else:
            for it in bucket:
                ws.append([
                    it['account_label'] or '',
                    it['value_date'] or '',
                    it['account_currency'] or '',
                    it['amount'] or 0,
                    it['sign'] or '',
                    'Our' if it['source_side'] == 'flex' else 'Their',
                    'Other',
                    'Open',
                    it['age_days'],
                    it['value_date'] or '',
                    it['ref'] or '',
                    it['ref'] or '',
                    (it['narration'] or '')[:70],
                    (it['narration'] or '')[70:140],
                    '',
                ])
            # totals row
            total_cr = sum(it['amount'] or 0 for it in bucket if it['sign'] in ('C', 'CR'))
            total_dr = sum(it['amount'] or 0 for it in bucket if it['sign'] in ('D', 'DR'))
            ws.append([])
            tr = ws.max_row + 1
            ws.append([f'TOTAL ({len(bucket)} items)', '', '',
                       f'CR: {total_cr:,.2f} / DR: {total_dr:,.2f} / NET: {total_cr - total_dr:+,.2f}',
                       '', '', '', '', '', '', '', '', '', '', ''])
            for cell in ws[ws.max_row]:
                cell.font = total_font
                cell.fill = total_fill
        ws.freeze_panes = 'A2'
        # column widths
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = 10
            for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
                v = row[0]
                if v is None:
                    continue
                max_len = max(max_len, min(50, len(str(v)) + 2))
            ws.column_dimensions[letter].width = max_len

    # Currency-split for foreign nostros (USD/EUR/GBP/AUD/etc.), functional-split for GHS.
    by_group: dict[str, list[dict]] = {g: [] for g in FUNCTIONAL_GROUPS}
    by_currency: dict[str, list[dict]] = {}
    for it in items:
        ccy = (it['account_currency'] or '').upper()
        if ccy and ccy != 'GHS':
            by_currency.setdefault(ccy, []).append(it)
        else:
            grp = it['functional_group'] or 'PSC TROPS'
            by_group.setdefault(grp, []).append(it)

    # Summary tab first.
    summary = wb.create_sheet('Summary')
    summary.append(['Kilter daily breaks report', f'as of {as_of_label}'])
    summary['A1'].font = Font(bold=True, size=14)
    summary.append([])
    summary.append(['Tab', 'Rows', 'CR total', 'DR total', 'Net'])
    for cell in summary[summary.max_row]:
        cell.font = hdr_font
        cell.fill = hdr_fill

    def _stats(bucket):
        cr = sum(it['amount'] or 0 for it in bucket if it['sign'] in ('C', 'CR'))
        dr = sum(it['amount'] or 0 for it in bucket if it['sign'] in ('D', 'DR'))
        return len(bucket), cr, dr, cr - dr

    # Foreign sub-tabs — one per currency, labelled "BOG <ccy>".
    for ccy in sorted(by_currency):
        _add_sheet(f'BOG {ccy}', by_currency[ccy])
        n, cr, dr, net = _stats(by_currency[ccy])
        summary.append([f'BOG {ccy}', n, cr, dr, net])

    # GHS functional-group tabs in canonical order.
    for grp in FUNCTIONAL_GROUPS:
        bucket = by_group.get(grp, [])
        _add_sheet(grp, bucket)
        n, cr, dr, net = _stats(bucket)
        summary.append([grp, n, cr, dr, net])

    # Move summary to the front.
    wb.move_sheet('Summary', offset=-len(wb.sheetnames) + 1)

    # Column widths on summary.
    for col in range(1, summary.max_column + 1):
        letter = get_column_letter(col)
        max_len = 14
        for row in summary.iter_rows(min_col=col, max_col=col, values_only=True):
            v = row[0]
            if v is None:
                continue
            max_len = max(max_len, min(30, len(str(v)) + 2))
        summary.column_dimensions[letter].width = max_len

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Kilter daily breaks {as_of_label}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


# ===========================================================================
# (End of reconciliation-ledger endpoints)
# ===========================================================================


async def _save_upload(upload: UploadFile, prefix: str) -> Path:
    """Write an upload to the persistent uploads/ folder with a timestamped
    name. Files are retained for audit — we don't try to delete them because
    openpyxl's read-only mode holds the handle open on Windows and unlink
    races with GC. Disk use is bounded by retention policy, not per-request.

    Streams the body in UPLOAD_CHUNK_BYTES blocks so 250+ MB cards-side
    settlement files don't materialize in worker RAM. Enforces
    MAX_REQUEST_BYTES inline as a defence-in-depth check (the middleware
    catches honest Content-Length, this catches chunked-transfer abuse)."""
    ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S_%f')
    safe_name = f"{ts}_{prefix}_{Path(upload.filename).name}"
    path = UPLOAD_DIR / safe_name
    written = 0
    with path.open('wb') as f:
        while True:
            chunk = await upload.read(UPLOAD_CHUNK_BYTES)
            if not chunk:
                break
            written += len(chunk)
            if written > MAX_REQUEST_BYTES:
                f.close()
                try:
                    path.unlink()
                except OSError:
                    pass
                raise HTTPException(
                    413,
                    f"Upload exceeds {MAX_REQUEST_BYTES // (1024*1024)} MB cap.",
                )
            f.write(chunk)
    return path


def _swift_row_to_dict(r) -> dict:
    return {
        '_source': 'swift', '_row_number': r['row_number'], '_used': False,
        'value_date': r['value_date'], 'amount': r['amount'], 'sign': r['sign'],
        'origin': r['origin'], 'type': r['type'], 'status': r['status'],
        'book_date': r['book_date'], 'our_ref': r['our_ref'],
        'their_ref': r['their_ref'], 'booking_text_1': r['booking_text_1'],
        'booking_text_2': r['booking_text_2'],
    }


def _flex_row_to_dict(r) -> dict:
    return {
        '_source': 'flexcube', '_row_number': r['row_number'], '_used': False,
        'trn_ref': r['trn_ref'], 'ac_branch': r['ac_branch'], 'ac_no': r['ac_no'],
        'booking_date': r['booking_date'], 'value_date': r['value_date'],
        'type': r['type'], 'narration': r['narration'], 'amount': r['amount'],
        'ccy': r['ccy'], 'module': r['module'], 'external_ref': r['external_ref'],
        'user_id': r['user_id'],
    }

