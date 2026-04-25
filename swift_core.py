"""
swift_core.py
=============

Shared parsing library used by mt940_to_excel.py and mt950_to_excel.py.

You do NOT run this file directly — it's imported by the two runner
scripts. Keeping the logic here (in one place) means a bug fix or an
improvement benefits both pipelines automatically.

What lives in here:
    * regex patterns for SWIFT balance and statement-line fields
    * helpers to parse SWIFT dates and SWIFT amounts
    * `extract_block_4`        — isolates the transactional block
    * `parse_tagged_fields`    — groups the file into (tag, value) pairs
    * `detect_message_type`    — reads '940' or '950' from Block 2 header
    * `parse_swift_statement`  — the main parse; returns a structured dict
    * `write_excel`            — writes the 12-column ops-team spreadsheet
    * `run_batch`              — the batch processor the runners call

If you want to adapt the parser for a new correspondent bank, almost
everything you need to change is in this file.
"""

import re
from pathlib import Path
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ---------------------------------------------------------------------------
# SWIFT format rules — expressed as regular expressions.
# ---------------------------------------------------------------------------

# The message type sits in the Block 2 application header, e.g. "{2:O950..."
# The character right after 'O' (output) or 'I' (input) is the first digit of
# the 3-digit message type code.
MESSAGE_TYPE_PATTERN = re.compile(r'\{2:[IO](\d{3})')

# :60M:, :60F:, :62F:, :62M:, :64:, :65: — all the balance fields look like:
#     C260417GHS4801552913,09
#   = mark(1) + date(6, YYMMDD) + currency(3) + amount(comma as decimal)
BALANCE_PATTERN = re.compile(
    r'^(?P<mark>[CD])'          # C = credit, D = debit
    r'(?P<date>\d{6})'          # YYMMDD (e.g. 260417 = 17 Apr 2026)
    r'(?P<currency>[A-Z]{3})'   # 3-letter currency code
    r'(?P<amount>[\d,]+)'       # amount using comma as decimal separator
    r'$'
)

# The :61: statement line looks like:
#     260417C46,56NTRFFNB26106BGFF0JBB
#     260417D174000000,NTRFH98RTDM261070014
# Structure:
#   value_date(6) + [entry_date(4)] + debit/credit_mark + [funds_code(1)]
#   + amount + transaction_type(4 chars: N + 3 letters) + reference
#   + optional "//" + bank_reference
STATEMENT_LINE_PATTERN = re.compile(
    r'^(?P<value_date>\d{6})'              # YYMMDD
    r'(?P<entry_date>\d{4})?'              # optional MMDD booking date
    r'(?P<mark>RC|RD|C|D)'                 # reversals first, then plain C/D
    r'(?P<funds_code>[A-Z])?'              # optional single-letter funds code
    r'(?P<amount>\d+,\d*)'                 # digits + comma + optional digits
    r'(?P<txn_type>[A-Z]\w{3})'            # 4-char transaction type, e.g. NTRF
    r'(?P<account_ref>(?:(?!//)\S)+)'      # reference: any non-whitespace, but stop at //
    r'(?://(?P<bank_ref>\S+))?'            # optional //bank_reference
    r'$'
)


# ---------------------------------------------------------------------------
# Small helpers.
# ---------------------------------------------------------------------------

def parse_swift_date(yymmdd: str) -> int:
    """Turn '260417' into 20260417, matching the ops team's format.

    SWIFT uses 2-digit years. We assume 20YY — fine through 2099.
    """
    yy = int(yymmdd[0:2])
    mm = yymmdd[2:4]
    dd = yymmdd[4:6]
    year = 2000 + yy
    return int(f"{year}{mm}{dd}")


def parse_swift_amount(amount_str: str) -> float:
    """Turn SWIFT's '46,56' or '500000000,' into a Python float.

    SWIFT uses a comma as decimal separator. Nothing after the comma means
    no decimal part, so '50,' == 50.00.
    """
    cleaned = amount_str.replace(',', '.')
    if cleaned.endswith('.'):
        cleaned += '0'
    # Decimal first for exact arithmetic on large cedi amounts, then float
    # for openpyxl (which serialises floats as Excel numbers).
    return float(Decimal(cleaned))


def detect_message_type(raw_text: str):
    """Return the 3-digit message type as a string, e.g. '940' or '950'.

    Returns None if the file does not contain a SWIFT Block 2 header.
    """
    m = MESSAGE_TYPE_PATTERN.search(raw_text)
    return m.group(1) if m else None


def extract_block_4(raw_text: str) -> str:
    """Pull Block 4 (the text block) out of the SWIFT message.

    The message structure is: {1:...}{2:...}{3:...}{4:\r\n...\r\n-}{5:...}
    We want everything between "{4:" and the "-}" that closes Block 4.
    """
    start = raw_text.find('{4:')
    if start == -1:
        raise ValueError("Could not find Block 4 in the SWIFT file — is this a valid SWIFT message?")
    start += 3  # advance past the "{4:" marker
    end = raw_text.find('-}', start)
    if end == -1:
        raise ValueError("Block 4 in this file is not properly terminated with '-}'.")
    return raw_text[start:end]


def parse_tagged_fields(block_4_text: str):
    """Walk Block 4 and return a list of [tag, value_lines] pairs.

    A tag line starts with ':xx:' (digits + optional letter). Any line
    without a leading ':tag:' is a continuation of the previous field's
    value — this is how supplementary details after a :61: get captured.
    """
    # Normalise line endings — SWIFT files use CRLF.
    text = block_4_text.replace('\r\n', '\n').replace('\r', '\n').strip('\n')
    lines = text.split('\n')

    tag_start = re.compile(r'^:(\w+):')

    fields = []
    current = None

    for line in lines:
        match = tag_start.match(line)
        if match:
            if current is not None:
                fields.append(current)
            tag = match.group(1)
            value_first_line = line[match.end():]
            current = [tag, [value_first_line]]
        else:
            if current is not None:
                current[1].append(line)

    if current is not None:
        fields.append(current)

    return fields


# ---------------------------------------------------------------------------
# The core parse — works for MT940 and MT950.
# ---------------------------------------------------------------------------

def parse_swift_statement(raw_text: str) -> dict:
    """Parse the raw SWIFT message text and return header info + transactions.

    Despite handling both MT940 and MT950, this function doesn't care which
    is which. Every tag we recognise is processed if present, and tags that
    don't appear simply leave their slot in the result dict as None/empty.
    Message-type enforcement is the caller's job (see `run_batch`).
    """
    block_4 = extract_block_4(raw_text)
    fields = parse_tagged_fields(block_4)

    result = {
        'transaction_reference': None,
        'related_reference': None,             # MT940 only (:21:)
        'account': None,
        'statement_number': None,
        'opening_balance': None,
        'closing_balance': None,
        'closing_available_balance': None,     # MT940 only (:64:)
        'forward_available_balance': [],       # MT940 only (:65:, may repeat)
        'account_info': None,                  # MT940 only (:86: at file level)
        'transactions': [],
    }

    # Tracks the tag we just processed. Matters for :86: — we need to know
    # whether a :86: attaches to the preceding :61: (transaction narrative)
    # or is free-floating account-level info.
    last_tag = None

    for tag, value_lines in fields:
        first_line = value_lines[0]

        if tag == '20':
            result['transaction_reference'] = first_line.strip()

        elif tag == '21':
            result['related_reference'] = first_line.strip()

        elif tag == '25':
            result['account'] = first_line.strip()

        elif tag == '28C':
            result['statement_number'] = first_line.strip()

        elif tag in ('60F', '60M'):
            m = BALANCE_PATTERN.match(first_line.strip())
            if m:
                result['opening_balance'] = {
                    'mark': m.group('mark'),
                    'date': parse_swift_date(m.group('date')),
                    'currency': m.group('currency'),
                    'amount': parse_swift_amount(m.group('amount')),
                    'tag': tag,
                }

        elif tag in ('62F', '62M'):
            m = BALANCE_PATTERN.match(first_line.strip())
            if m:
                result['closing_balance'] = {
                    'mark': m.group('mark'),
                    'date': parse_swift_date(m.group('date')),
                    'currency': m.group('currency'),
                    'amount': parse_swift_amount(m.group('amount')),
                    'tag': tag,
                }

        elif tag == '64':
            m = BALANCE_PATTERN.match(first_line.strip())
            if m:
                result['closing_available_balance'] = {
                    'mark': m.group('mark'),
                    'date': parse_swift_date(m.group('date')),
                    'currency': m.group('currency'),
                    'amount': parse_swift_amount(m.group('amount')),
                }

        elif tag == '65':
            m = BALANCE_PATTERN.match(first_line.strip())
            if m:
                result['forward_available_balance'].append({
                    'mark': m.group('mark'),
                    'date': parse_swift_date(m.group('date')),
                    'currency': m.group('currency'),
                    'amount': parse_swift_amount(m.group('amount')),
                })

        elif tag == '61':
            m = STATEMENT_LINE_PATTERN.match(first_line.strip())
            if not m:
                result['transactions'].append({
                    'parse_error': True,
                    'raw': first_line,
                    'supplementary': ' '.join(value_lines[1:]).strip() or None,
                    'narrative': None,
                })
                last_tag = tag
                continue

            raw_mark = m.group('mark')
            if raw_mark == 'RC':
                sign, reversed_flag = 'C', True
            elif raw_mark == 'RD':
                sign, reversed_flag = 'D', True
            else:
                sign, reversed_flag = raw_mark, False

            value_date = parse_swift_date(m.group('value_date'))

            entry_date = value_date
            if m.group('entry_date'):
                year = value_date // 10000
                mmdd = m.group('entry_date')
                entry_date = int(f"{year}{mmdd}")

            supplementary = ' '.join(
                line.strip() for line in value_lines[1:] if line.strip()
            )

            result['transactions'].append({
                'parse_error': False,
                'value_date': value_date,
                'entry_date': entry_date,
                'sign': sign,
                'reversed': reversed_flag,
                'amount': parse_swift_amount(m.group('amount')),
                'txn_type': m.group('txn_type'),
                'account_ref': m.group('account_ref'),
                'bank_ref': m.group('bank_ref'),
                'supplementary': supplementary or None,
                'narrative': None,   # filled in by the next :86: if one follows
            })

        elif tag == '86':
            narrative = ' '.join(
                line.strip() for line in value_lines if line.strip()
            )
            if (last_tag == '61'
                    and result['transactions']
                    and not result['transactions'][-1].get('parse_error')):
                result['transactions'][-1]['narrative'] = narrative
            else:
                if result['account_info']:
                    result['account_info'] += ' ' + narrative
                else:
                    result['account_info'] = narrative

        last_tag = tag

    return result


# ---------------------------------------------------------------------------
# Excel output — the 12-column format the ops team already uses.
# ---------------------------------------------------------------------------

OUTPUT_HEADERS = [
    'Value date',
    'Amount',
    'S',
    'Origin',
    'Type',
    'Status',
    'Book. date',
    'Our reference 1',
    'Their reference 1',
    'Booking text 1',
    'Booking text 2',
    'Matching type',
]


def write_excel(parsed: dict, output_path: Path) -> None:
    """Turn a parsed statement dict into a nicely-formatted .xlsx file."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'SWIFT'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
    label_font = Font(bold=True)
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Metadata panel — cover sheet for the ops team.
    meta_rows = [
        ('Account',               parsed['account']),
        ('Statement reference',   parsed['transaction_reference']),
    ]
    if parsed['related_reference']:
        meta_rows.append(('Related reference', parsed['related_reference']))
    meta_rows.append(('Statement number', parsed['statement_number']))
    if parsed['opening_balance']:
        ob = parsed['opening_balance']
        meta_rows.append((
            f"Opening balance ({ob['tag']})",
            f"{ob['mark']} {ob['currency']} {ob['amount']:,.2f} on {ob['date']}",
        ))
    if parsed['closing_balance']:
        cb = parsed['closing_balance']
        meta_rows.append((
            f"Closing balance ({cb['tag']})",
            f"{cb['mark']} {cb['currency']} {cb['amount']:,.2f} on {cb['date']}",
        ))
    if parsed['closing_available_balance']:
        ab = parsed['closing_available_balance']
        meta_rows.append((
            "Closing available (64)",
            f"{ab['mark']} {ab['currency']} {ab['amount']:,.2f} on {ab['date']}",
        ))
    for fb in parsed['forward_available_balance']:
        meta_rows.append((
            "Forward available (65)",
            f"{fb['mark']} {fb['currency']} {fb['amount']:,.2f} on {fb['date']}",
        ))
    if parsed['account_info']:
        meta_rows.append(('Account info (86)', parsed['account_info']))
    meta_rows.append(('Transactions parsed', len(parsed['transactions'])))

    for row_idx, (label, value) in enumerate(meta_rows, start=1):
        ws.cell(row=row_idx, column=1, value=label).font = label_font
        ws.cell(row=row_idx, column=2, value=value)

    header_row = len(meta_rows) + 2

    # Transaction table header.
    for col_idx, name in enumerate(OUTPUT_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Transaction rows.
    for offset, txn in enumerate(parsed['transactions'], start=1):
        r = header_row + offset

        if txn.get('parse_error'):
            ws.cell(row=r, column=1, value='PARSE ERROR')
            ws.cell(row=r, column=10, value=txn.get('raw', ''))
            ws.cell(row=r, column=11, value=txn.get('supplementary'))
            continue

        ws.cell(row=r, column=1,  value=txn['value_date'])
        ws.cell(row=r, column=2,  value=txn['amount'])
        ws.cell(row=r, column=3,  value=txn['sign'])
        ws.cell(row=r, column=4,  value='Their')
        ws.cell(row=r, column=5,  value='Other')
        ws.cell(row=r, column=6,  value='Unmatched')
        ws.cell(row=r, column=7,  value=txn['entry_date'])
        ws.cell(row=r, column=8,  value=txn['account_ref'])
        ws.cell(row=r, column=9,  value=txn['bank_ref'])
        ws.cell(row=r, column=10, value=txn['supplementary'])
        ws.cell(row=r, column=11, value=txn.get('narrative'))  # MT940 :86:, blank for MT950
        ws.cell(row=r, column=12, value=None)

        ws.cell(row=r, column=2).number_format = '#,##0.00'

    widths = {
        'A': 12, 'B': 14, 'C':  5, 'D':  8, 'E': 10, 'F': 12,
        'G': 12, 'H': 22, 'I': 22, 'J': 30, 'K': 30, 'L': 18,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Batch runner — each runner script calls into this.
# ---------------------------------------------------------------------------

def run_batch(expected_type: str, input_dir: Path, output_dir: Path) -> int:
    """Process every file in `input_dir`, writing parsed outputs to `output_dir`.

    Only files whose SWIFT message type matches `expected_type` ('940' or
    '950') are processed. Files of any other type are skipped with a clear
    note — they're never deleted, never modified, just left alone so you
    can move them to the right folder.

    Returns 0 on success, 1 if any file failed with an actual error.
    (Skipped wrong-type files don't count as failures.)
    """
    banner = f"SWIFT MT{expected_type} -> Excel converter"
    print(banner)
    print("-" * 60)
    print(f"Input folder  : {input_dir}")
    print(f"Output folder : {output_dir}")
    print()

    # Create output/ if missing.
    output_dir.mkdir(parents=True, exist_ok=True)

    # Stop with a friendly message if input/ is missing.
    if not input_dir.exists():
        print(f"ERROR: the input folder does not exist at {input_dir}.")
        print(f"Create it and drop your MT{expected_type} files inside, then re-run.")
        input("\nPress Enter to exit...")
        return 1

    input_files = sorted(
        p for p in input_dir.iterdir()
        if p.is_file() and not p.name.startswith('.')
    )

    if not input_files:
        print(f"No files found in {input_dir}.")
        print(f"Drop your MT{expected_type} files in there and re-run.")
        input("\nPress Enter to exit...")
        return 0

    print(f"Found {len(input_files)} file(s).")
    print()

    succeeded = 0
    skipped = []   # list of (filename, reason) — wrong message type etc.
    failed = []    # list of (filename, exception string)

    for input_path in input_files:
        output_path = output_dir / f"{input_path.stem}_parsed.xlsx"
        print(f"Processing: {input_path.name}")

        try:
            raw = input_path.read_text(encoding='latin-1', errors='replace')
            file_type = detect_message_type(raw)

            if file_type is None:
                print("  -> SKIPPED: no SWIFT message type found. "
                      "Is this actually a SWIFT file?")
                skipped.append((input_path.name, 'not a SWIFT message'))
                print()
                continue

            if file_type != expected_type:
                print(f"  -> SKIPPED: this is an MT{file_type} message, "
                      f"not MT{expected_type}. Move it to input_mt{file_type}/.")
                skipped.append((input_path.name, f'MT{file_type}'))
                print()
                continue

            # Right type — parse and write.
            parsed = parse_swift_statement(raw)
            write_excel(parsed, output_path)

            # Balance self-check.
            txns = parsed['transactions']
            parse_errors = sum(1 for t in txns if t.get('parse_error'))
            credits = sum(t['amount'] for t in txns
                          if not t.get('parse_error') and t['sign'] == 'C')
            debits = sum(t['amount'] for t in txns
                         if not t.get('parse_error') and t['sign'] == 'D')

            balance_note = ""
            if parsed['opening_balance'] and parsed['closing_balance']:
                ob = parsed['opening_balance']
                cb = parsed['closing_balance']
                ob_signed = ob['amount'] if ob['mark'] == 'C' else -ob['amount']
                cb_signed = cb['amount'] if cb['mark'] == 'C' else -cb['amount']
                expected_movement = cb_signed - ob_signed
                actual_movement = credits - debits
                diff = actual_movement - expected_movement
                if abs(diff) < 0.01:
                    balance_note = "balances exactly"
                else:
                    balance_note = f"WARNING: off by {diff:,.2f}"

            print(f"  -> wrote {output_path.name}")
            print(f"  -> {len(txns)} transactions, "
                  f"{parse_errors} parse error(s), {balance_note}")
            succeeded += 1

        except Exception as exc:
            print(f"  -> FAILED: {exc}")
            failed.append((input_path.name, str(exc)))

        print()

    # Summary.
    print("-" * 60)
    print(f"Done. {succeeded} succeeded, "
          f"{len(skipped)} skipped (wrong type), "
          f"{len(failed)} failed.")
    if skipped:
        print("\nSkipped files (move to the right folder and re-run):")
        for name, reason in skipped:
            print(f"  {name}: {reason}")
    if failed:
        print("\nFailed files:")
        for name, err in failed:
            print(f"  {name}: {err}")

    input("\nPress Enter to exit...")
    return 0 if not failed else 1
