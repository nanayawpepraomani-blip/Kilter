"""
BTW May 4, 2026 — Kilter vs Corona comparison report v1
Methodology:
  - Kilter MATCHED (session 119) = 17,712 assignment PAIRS
      Apr 30 CRs (17,568) + Apr 29 CRs (144) cleared by May 4 DRs
  - Kilter OPEN (as of May 4) = sessions ≤ 119, not yet cleared
  - Corona OPEN = 'BANK TO WALLET 05042026.xlsx' (18,088 data rows)
"""
import sqlite3
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter
import re

CORONA_FILE = "/Users/nypo/Desktop/BANK TO WALLET 05042026.xlsx"
OUT_FILE    = "/Users/nypo/Downloads/Kilter_BTW_May4_Report_v1.xlsx"
SESSION_ID  = 119
SESSION_DATE = "May 4"
ACCOUNT_ID  = 15

# ── colours ──────────────────────────────────────────────────────────────────
NAVY   = "1B2A4A"; TEAL   = "0E7490"; GREEN  = "166534"; AMBER  = "92400E"
RED    = "991B1B"; PURPLE = "5B21B6"; LGREY  = "F3F4F6"; WHITE  = "FFFFFF"
G_LT   = "D1FAE5"; A_LT   = "FEF3C7"; R_LT   = "FEE2E2"; B_LT   = "DBEAFE"

def fill(h): return PatternFill("solid", fgColor=h)
def border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def hcell(ws, r, c, val, bg=NAVY, fg=WHITE, size=11, align="center"):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = fill(bg); cell.font = Font(bold=True, color=fg, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = border(); return cell

def dcell(ws, r, c, val, bg=None, fg="000000", fmt=None, align="left",
          wrap=False, bold=False):
    cell = ws.cell(row=r, column=c, value=val)
    if bg: cell.fill = fill(bg)
    cell.font = Font(color=fg, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = border()
    if fmt: cell.number_format = fmt
    return cell

def section(ws, r, title, cols=6, bg=NAVY):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    c = ws.cell(row=r, column=1, value=title)
    c.fill = fill(bg); c.font = Font(bold=True, color=WHITE, size=11)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 22

def fmt_date(d):
    s = str(d) if d else ""
    return f"{s[:4]}-{s[4:6]}-{s[6:8]}" if len(s) == 8 else s

def bucket(amt):
    try: a = float(amt)
    except: return "Unknown"
    if a < 1:          return "< 1"
    if a < 100:        return "1 – 99"
    if a < 1_000:      return "100 – 999"
    if a < 10_000:     return "1,000 – 9,999"
    if a < 100_000:    return "10,000 – 99,999"
    if a < 1_000_000:  return "100K – 999K"
    return "≥ 1M"

BUCKETS = ["< 1","1 – 99","100 – 999","1,000 – 9,999","10,000 – 99,999","100K – 999K","≥ 1M","Unknown"]

def network(text):
    t = (text or "").upper()
    if "MTN" in t:                     return "MTN"
    if "VODAFONE" in t or "VODA" in t: return "Vodafone"
    if "TIGO" in t:                    return "Tigo"
    if "AIRTEL" in t:                  return "Airtel"
    if "TELECEL" in t:                 return "Telecel"
    return "Other / Unknown"

NET_BG = {"MTN":"FEF9C3","Vodafone":"FCE7F3","Tigo":"E0F2FE","Airtel":"FEE2E2","Telecel":"EDE9FE"}

# ── load Kilter data ──────────────────────────────────────────────────────────
conn = sqlite3.connect("kilter.db")
conn.row_factory = sqlite3.Row

print(f"Loading Kilter matched pairs (session {SESSION_ID}, {SESSION_DATE} delta) …")
pairs = conn.execute(f"""
    SELECT oi.value_date AS cr_date, oi.amount AS cr_amt, oi.sign AS cr_sign,
           oi.ref AS cr_ref, oi.narration AS cr_narr,
           s.value_date AS dr_date, s.amount AS dr_amt,
           s.our_ref AS dr_ref, s.booking_text_1 AS dr_narr
    FROM open_items oi
    JOIN assignments a ON a.id = oi.cleared_assignment_id
    JOIN swift_txns s ON s.session_id={SESSION_ID} AND s.row_number=a.swift_row
    WHERE oi.cleared_session_id={SESSION_ID}
    ORDER BY cr_date, cr_amt DESC
""").fetchall()
pairs = [dict(r) for r in pairs]
print(f"  → {len(pairs):,} confirmed pairs")

# Date-breakdown for pairs
apr30_pairs = sum(1 for p in pairs if str(p["cr_date"]) == "20260430")
apr29_pairs = sum(1 for p in pairs if str(p["cr_date"]) == "20260429")
other_pairs = len(pairs) - apr30_pairs - apr29_pairs

print(f"  Apr 30 CRs → May 4 DRs: {apr30_pairs:,}")
print(f"  Apr 29 CRs → May 4 DRs: {apr29_pairs:,}")
if other_pairs: print(f"  Other dates:              {other_pairs:,}")

print(f"Loading Kilter open items as of {SESSION_DATE} (sessions ≤ {SESSION_ID}) …")
kilter_open = conn.execute(f"""
    SELECT source_side, value_date, amount, sign, ref, narration, src_session_id
    FROM open_items
    WHERE account_id={ACCOUNT_ID}
      AND src_session_id <= {SESSION_ID}
      AND (cleared_session_id IS NULL OR cleared_session_id > {SESSION_ID})
    ORDER BY value_date, amount DESC
""").fetchall()
kilter_open = [dict(r) for r in kilter_open]
kilter_open_refs = {r["ref"] for r in kilter_open if r["ref"]}
print(f"  → {len(kilter_open):,} open items as of {SESSION_DATE}")

# ── load Corona ───────────────────────────────────────────────────────────────
print(f"Loading Corona open items from {CORONA_FILE} …")
wb_c = openpyxl.load_workbook(CORONA_FILE, read_only=True, data_only=True)
ws_c = wb_c.active
corona = []
for row in ws_c.iter_rows(min_row=3, values_only=True):
    if row[0] is None and row[1] is None: continue
    vd = row[1]
    if hasattr(vd, 'strftime'): vd_str, vd_fmt = vd.strftime("%Y%m%d"), vd.strftime("%Y-%m-%d")
    else:
        vd_str = str(vd) if vd else ""
        vd_fmt = f"{vd_str[:4]}-{vd_str[4:6]}-{vd_str[6:8]}" if len(vd_str)==8 else vd_str
    ref = str(row[11] or "").strip()
    narr = str(row[14] or row[13] or "")
    corona.append({
        "vd_str": vd_str, "vd_fmt": vd_fmt,
        "amount": row[3], "sign": row[4],
        "ref": ref, "narr1": row[13], "narr2": row[14],
        "type": row[7], "age": row[9],
        "in_kilter_open": ref in kilter_open_refs,
    })
wb_c.close()
print(f"  → {len(corona):,} Corona open items")

both_open    = sum(1 for r in corona if r["in_kilter_open"])
corona_only  = sum(1 for r in corona if not r["in_kilter_open"])
kilter_extra = len(kilter_open) - both_open
print(f"  Both agree (open in both):        {both_open:,}")
print(f"  Corona only (not in Kilter open): {corona_only:,}")
print(f"  Kilter extra open items:          {kilter_extra:,}")

# ── aggregations ──────────────────────────────────────────────────────────────
def agg(items, amt_key, narr_key):
    bkt_c = Counter(); bkt_v = defaultdict(float)
    net_c = Counter(); net_v = defaultdict(float)
    for p in items:
        b = bucket(p[amt_key]); bkt_c[b] += 1
        try: bkt_v[b] += float(p[amt_key])
        except: pass
        n = network(p.get(narr_key) or ""); net_c[n] += 1
        try: net_v[n] += float(p[amt_key])
        except: pass
    return bkt_c, bkt_v, net_c, net_v

p_bkt_c, p_bkt_v, p_net_c, p_net_v = agg(pairs, "cr_amt", "cr_narr")
c_bkt_c, c_bkt_v, c_net_c, c_net_v = agg(corona, "amount", "narr2")

# ── build workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ══ Sheet 1: Summary ══════════════════════════════════════════════════════════
ws = wb.active; ws.title = "Summary"
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:E1")
ws["A1"].value = f"KILTER vs CORONA — BTW {SESSION_DATE}, 2026  |  Report v1"
ws["A1"].fill = fill(NAVY); ws["A1"].font = Font(bold=True, color=WHITE, size=14)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

for c, h in enumerate(["Metric","Corona (05042026.xlsx)","Kilter (May 4)","Diff","Notes"], 1):
    hcell(ws, 2, c, h)
ws.row_dimensions[2].height = 22

diff_open = len(kilter_open) - len(corona)
diff_sign = f"+{diff_open} Kilter" if diff_open > 0 else (f"+{-diff_open} Corona" if diff_open < 0 else "0")

summary = [
    ("MATCHED PAIRS", "", "", "", ""),
    ("Kilter pairs matched (session 119)",
     "—",
     f"{len(pairs):,}",
     "",
     f"Apr 30/29 CRs cleared by May 4 DRs — T+4/T+5 (weekend carry-forward)"),
    (f"  Apr 30 CRs → May 4 DRs",
     "—", f"{apr30_pairs:,}", "",
     "Apr 30 settlement batch arriving on Mon May 4"),
    (f"  Apr 29 CRs → May 4 DRs",
     "—", f"{apr29_pairs:,}", "",
     "Apr 29 late items also settled May 4"),
    ("", "", "", "", ""),
    ("OPEN ITEMS (May 4 snapshot)", "", "", "", ""),
    ("Items still open",
     f"{len(corona):,}",
     f"{len(kilter_open):,}",
     diff_sign,
     ""),
    ("  Of which: agreed open in both",
     f"{both_open:,}",
     f"{both_open:,}",
     "0",
     "BOTH systems agree on these"),
    ("  Corona open not in Kilter open",
     f"{corona_only:,}",
     "—",
     "",
     "Items in Corona that Kilter does not hold open"),
    ("  Kilter extra open (not in Corona)",
     "—",
     f"{kilter_extra:,}",
     "",
     "Items Kilter holds open that Corona does not show"),
    ("", "", "", "", ""),
    ("KEY FINDING", "", "", "", ""),
    ("Open item consensus",
     f"{len(corona):,} open",
     f"{len(kilter_open):,} open",
     f"{both_open:,} agreed" if both_open else "—",
     f"{both_open:,} of {len(corona):,} Corona items confirmed open in Kilter ({100*both_open//max(len(corona),1)}%)"),
    ("Kilter T+4/T+5 matched pairs",
     "Not tracked",
     f"{len(pairs):,} pairs",
     "",
     "Weekend carry-forward: Apr 30/29 CRs settled on Mon May 4"),
]

DIVIDERS = {0, 5, 11}
ROW_BG   = {1:G_LT, 6:A_LT, 7:G_LT, 8:G_LT, 12:G_LT, 13:B_LT}

for i, (metric, cv, kv, diff, note) in enumerate(summary):
    r = i + 3
    ws.row_dimensions[r].height = 18
    if i in DIVIDERS:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c = ws.cell(row=r, column=1, value=metric)
        c.fill = fill("1E3A5F"); c.font = Font(bold=True, color=WHITE, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border()
        continue
    bg = ROW_BG.get(i, WHITE)
    for col, val in enumerate([metric, cv, kv, diff, note], 1):
        dcell(ws, r, col, val, bg=bg,
              align="left" if col in (1,5) else "center",
              bold=(i in DIVIDERS))

ws.column_dimensions["A"].width = 42; ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 18; ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 58

# ══ Sheet 2: Kilter Matched Pairs ══════════════════════════════════════════
print("Building Kilter matched pairs sheet …")
ws2 = wb.create_sheet("Kilter_Matched_Pairs")
ws2.sheet_view.showGridLines = False; ws2.freeze_panes = "A3"

ws2.merge_cells("A1:K1")
ws2["A1"].value = (f"Kilter — {SESSION_DATE} Matched Pairs  ({len(pairs):,} pairs)  |  "
                   f"Apr 30 CRs ({apr30_pairs:,}) + Apr 29 CRs ({apr29_pairs:,}) cleared by May 4 DRs")
ws2["A1"].fill = fill(TEAL); ws2["A1"].font = Font(bold=True, color=WHITE, size=12)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 26

for c, h in enumerate(["#","CR Date","CR Amount","CR Sign","CR Ref","CR Narration",
                         "DR Date","DR Amount","DR Ref","DR Narration","Network"], 1):
    hcell(ws2, 2, c, h, bg="1E3A5F")

for i, p in enumerate(pairs, 1):
    r = i + 2
    net = network(p["cr_narr"] or p["dr_narr"] or "")
    row_bg = NET_BG.get(net, LGREY if i%2==0 else WHITE)
    for c, (val, fmt, aln) in enumerate([
        (i, None, "right"),
        (fmt_date(p["cr_date"]), None, "center"),
        (p["cr_amt"], "#,##0.00", "right"),
        (p["cr_sign"], None, "center"),
        (p["cr_ref"], None, "left"),
        (p["cr_narr"], None, "left"),
        (fmt_date(p["dr_date"]), None, "center"),
        (p["dr_amt"], "#,##0.00", "right"),
        (p["dr_ref"], None, "left"),
        (p["dr_narr"], None, "left"),
        (net, None, "left"),
    ], 1):
        dcell(ws2, r, c, val, bg=row_bg, fmt=fmt, align=aln)

for letter, w in zip("ABCDEFGHIJK", [7,12,15,8,24,46,12,15,24,50,14]):
    ws2.column_dimensions[letter].width = w

# ══ Sheet 3: Kilter Groups ═════════════════════════════════════════════════
print("Building Kilter groups sheet …")
ws3 = wb.create_sheet("Kilter_Groups")
ws3.sheet_view.showGridLines = False

def group_table(ws, start_r, title, items_cnt, bg_title, bg_hdr, data_rows, col_hdrs):
    section(ws, start_r, title, cols=len(col_hdrs), bg=bg_title)
    r = start_r + 1
    for c, h in enumerate(col_hdrs, 1):
        hcell(ws, r, c, h, bg=bg_hdr)
    r += 1
    for row_vals in data_rows:
        row_bg = LGREY if r % 2 == 0 else WHITE
        for c, (val, fmt, aln, custom_bg) in enumerate(row_vals, 1):
            dcell(ws, r, c, val, bg=custom_bg or row_bg, fmt=fmt, align=aln)
        r += 1
    return r + 1

r = 1
data = []
for b in BUCKETS:
    cnt = p_bkt_c.get(b, 0); val = p_bkt_v.get(b, 0)
    pct = cnt/len(pairs) if pairs else 0; avg = val/cnt if cnt else 0
    data.append([(b,None,"left",None),(cnt,"#,##0","right",None),
                 (pct,"0.00%","right",None),(val,"#,##0.00","right",None),
                 (avg,"#,##0.00","right",None)])
r = group_table(ws3, r, f"▸ Kilter Matched Pairs — By Amount Bucket  (session {SESSION_ID})", len(pairs),
                TEAL, "1E3A5F", data,
                ["Amount Range","Pairs Matched","% of Total","Total Value (GHS)","Avg (GHS)"])

data = []
for net, cnt in sorted(p_net_c.items(), key=lambda x: -x[1]):
    val = p_net_v.get(net, 0); pct = cnt/len(pairs) if pairs else 0
    avg = val/cnt if cnt else 0
    data.append([(net,None,"left",NET_BG.get(net)),(cnt,"#,##0","right",None),
                 (pct,"0.00%","right",None),(val,"#,##0.00","right",None),
                 (avg,"#,##0.00","right",None)])
r = group_table(ws3, r, "▸ Kilter Matched Pairs — By Network (from CR narration)", len(pairs),
                "1E3A5F", TEAL, data,
                ["Network","Pairs","% of Total","Total Value (GHS)","Avg (GHS)"])

# Date relationship for session 119
date_data = [
    [("Apr 30 CRs → May 4 DRs (T+4)",None,"left",A_LT),(apr30_pairs,"#,##0","right",None),
     (apr30_pairs/len(pairs) if pairs else 0,"0.00%","right",None),
     ("Apr 30 settlement processed on Mon May 4 (weekend delay)",None,"left",None),(None,None,"left",None)],
    [("Apr 29 CRs → May 4 DRs (T+5)",None,"left",R_LT),(apr29_pairs,"#,##0","right",None),
     (apr29_pairs/len(pairs) if pairs else 0,"0.00%","right",None),
     ("Apr 29 late items also arriving May 4",None,"left",None),(None,None,"left",None)],
]
if other_pairs:
    date_data.append([("Other date combinations",None,"left",LGREY),(other_pairs,"#,##0","right",None),
                      (other_pairs/len(pairs) if pairs else 0,"0.00%","right",None),
                      ("",None,"left",None),(None,None,"left",None)])
r = group_table(ws3, r, "▸ Kilter Matched Pairs — Date Relationship (T+N)", len(pairs),
                GREEN, "166534", date_data,
                ["Date Pattern","Pairs","% of Total","Description",""])

for letter, w in zip("ABCDE", [34,14,12,44,14]):
    ws3.column_dimensions[letter].width = w

# ══ Sheet 4: Corona Open Items ══════════════════════════════════════════════
print("Building Corona open items sheet …")
ws4 = wb.create_sheet("Corona_Open_Items")
ws4.sheet_view.showGridLines = False; ws4.freeze_panes = "A3"

ws4.merge_cells("A1:I1")
ws4["A1"].value = f"Corona (05042026.xlsx) — Open Items at {SESSION_DATE}  ({len(corona):,} items)"
ws4["A1"].fill = fill(AMBER); ws4["A1"].font = Font(bold=True, color=WHITE, size=12)
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 26

for c, h in enumerate(["Value Date","Amount (GHS)","Sign","Our Ref",
                        "Narration 1","Narration 2","Type","Age (days)","Kilter Open?"], 1):
    hcell(ws4, 2, c, h, bg="92400E")

for i, row in enumerate(corona, 1):
    r = i + 2
    bg = G_LT if row["in_kilter_open"] else R_LT
    status = "✓ Also open in Kilter" if row["in_kilter_open"] else "Not found in Kilter"
    for c, (val, fmt, aln) in enumerate([
        (row["vd_fmt"], None, "center"),
        (row["amount"], "#,##0.00", "right"),
        (row["sign"], None, "center"),
        (row["ref"], None, "left"),
        (row["narr1"], None, "left"),
        (row["narr2"], None, "left"),
        (row["type"], None, "center"),
        (row["age"], None, "right"),
        (status, None, "left"),
    ], 1):
        dcell(ws4, r, c, val, bg=bg, fmt=fmt, align=aln)

for letter, w in zip("ABCDEFGHI", [12,16,7,24,10,44,10,10,22]):
    ws4.column_dimensions[letter].width = w

# ══ Sheet 5: Corona Groups ══════════════════════════════════════════════════
print("Building Corona groups sheet …")
ws5 = wb.create_sheet("Corona_Groups")
ws5.sheet_view.showGridLines = False

r = 1
data = []
for b in BUCKETS:
    cnt = c_bkt_c.get(b, 0); val = c_bkt_v.get(b, 0)
    pct = cnt/len(corona) if corona else 0; avg = val/cnt if cnt else 0
    data.append([(b,None,"left",None),(cnt,"#,##0","right",None),
                 (pct,"0.00%","right",None),(val,"#,##0.00","right",None),
                 (avg,"#,##0.00","right",None)])
r = group_table(ws5, r, f"▸ Corona Open Items ({SESSION_DATE}) — By Amount Bucket", len(corona),
                AMBER, "92400E", data,
                ["Amount Range","Open Count","% of Total","Total Value (GHS)","Avg (GHS)"])

data = []
for net, cnt in sorted(c_net_c.items(), key=lambda x: -x[1]):
    val = c_net_v.get(net, 0); pct = cnt/len(corona) if corona else 0
    avg = val/cnt if cnt else 0
    data.append([(net,None,"left",NET_BG.get(net)),(cnt,"#,##0","right",None),
                 (pct,"0.00%","right",None),(val,"#,##0.00","right",None),
                 (avg,"#,##0.00","right",None)])
r = group_table(ws5, r, f"▸ Corona Open Items ({SESSION_DATE}) — By Network", len(corona),
                PURPLE, "5B21B6", data,
                ["Network","Open Count","% of Total","Total Value (GHS)","Avg (GHS)"])

for letter, w in zip("ABCDE", [34,14,12,36,14]):
    ws5.column_dimensions[letter].width = w

# ══ Sheet 6: Side-by-side open items ════════════════════════════════════════
print("Building side-by-side open items sheet …")
ws6 = wb.create_sheet("Open_Items_SideBySide")
ws6.sheet_view.showGridLines = False; ws6.freeze_panes = "A3"

ws6.merge_cells("A1:K1")
ws6["A1"].value = (f"Both Agreed Open — {both_open:,} items  |  "
                   f"Corona only: {corona_only:,}  |  Kilter extra: {kilter_extra:,}")
ws6["A1"].fill = fill("1E3A5F"); ws6["A1"].font = Font(bold=True, color=WHITE, size=12)
ws6["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[1].height = 26

for c, h in enumerate(["Value Date","Amount","Sign","Ref","Corona Narration",
                        "","Value Date","Amount","Sign","Ref","Kilter Narration"], 1):
    bg = AMBER if c <= 5 else (WHITE if c == 6 else "1E3A5F")
    fg = WHITE if c != 6 else "000000"
    hcell(ws6, 2, c, h, bg=bg, fg=fg)

kilter_by_ref = {r["ref"]: r for r in kilter_open if r["ref"]}

row_num = 3
for crow in corona:
    ref = crow["ref"]
    krow = kilter_by_ref.get(ref)
    bg_c = G_LT if krow else R_LT
    bg_k = G_LT if krow else A_LT
    narr = str(crow["narr2"] or crow["narr1"] or "")
    for c, (val, fmt, aln, bg) in enumerate([
        (crow["vd_fmt"], None, "center", bg_c),
        (crow["amount"], "#,##0.00", "right", bg_c),
        (crow["sign"], None, "center", bg_c),
        (ref, None, "left", bg_c),
        (narr, None, "left", bg_c),
        ("↔" if krow else "✗", None, "center", WHITE),
        (fmt_date(krow["value_date"]) if krow else "—", None, "center", bg_k),
        (krow["amount"] if krow else "—", "#,##0.00" if krow else None, "right", bg_k),
        (krow["sign"] if krow else "—", None, "center", bg_k),
        (krow["ref"] if krow else "—", None, "left", bg_k),
        (krow["narration"] if krow else "—", None, "left", bg_k),
    ], 1):
        dcell(ws6, row_num, c, val, bg=bg, fmt=fmt, align=aln)
    row_num += 1

for krow in kilter_open:
    if krow["ref"] not in {r["ref"] for r in corona}:
        for c, (val, fmt, aln, bg) in enumerate([
            ("—",None,"center",LGREY),("—",None,"right",LGREY),("—",None,"center",LGREY),
            ("—",None,"left",LGREY),("Kilter only",None,"left",LGREY),
            ("→",None,"center",WHITE),
            (fmt_date(krow["value_date"]),None,"center",B_LT),
            (krow["amount"],"#,##0.00","right",B_LT),
            (krow["sign"],None,"center",B_LT),
            (krow["ref"],None,"left",B_LT),
            (krow["narration"],None,"left",B_LT),
        ], 1):
            dcell(ws6, row_num, c, val, bg=bg, fmt=fmt, align=aln)
        row_num += 1

for letter, w in zip("ABCDEFGHIJK", [12,15,7,24,42,5,12,15,7,24,42]):
    ws6.column_dimensions[letter].width = w

# ══ save ════════════════════════════════════════════════════════════════════
wb.save(OUT_FILE)
print(f"\n✓ Saved → {OUT_FILE}")
print(f"Sheets: {[s.title for s in wb.worksheets]}")
print(f"\nFinal numbers:")
print(f"  Kilter matched pairs (session 119): {len(pairs):,}")
print(f"    Apr 30 CRs → May 4 DRs (T+4):    {apr30_pairs:,}")
print(f"    Apr 29 CRs → May 4 DRs (T+5):    {apr29_pairs:,}")
print(f"  Kilter open at May 4:               {len(kilter_open):,}")
print(f"  Corona open (05042026.xlsx):         {len(corona):,}")
print(f"  Both agree (open in both):          {both_open:,}")
print(f"  Corona only (not in Kilter open):   {corona_only:,}")
print(f"  Kilter extra (not in Corona):       {kilter_extra:,}")
