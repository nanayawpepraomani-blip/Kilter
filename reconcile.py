"""
reconcile.py
============

Matches SWIFT MT950/MT940 parsed spreadsheets against raw Flexcube account
entries, and produces a reconciliation workbook with Matched, Unmatched SWIFT,
Unmatched Flexcube, and Summary sheets.

How to use it:

    1. Drop your SWIFT parsed file (output from mt950_to_excel.py or
       mt940_to_excel.py) into the "input_swift" folder next to this script.
    2. Drop your Flexcube raw extract (the acc_entries.xlsx style file)
       into the "input_flexcube" folder.
    3. Run: python reconcile.py
    4. The reconciliation workbook appears in "output_reconciled" with the
       name <swift_filename>_vs_<flexcube_filename>.xlsx

The matching engine (plain English):

    For each unmatched SWIFT row, try four rules in priority order. Stop at
    the first one that finds a Flexcube partner. A Flexcube row can only be
    used once — whichever SWIFT row claims it first wins.

    Tier 1 — STRICT           : SWIFT 'Our reference 1' is inside Flexcube
                                TXN_NARRATIONS, amounts equal to the cent,
                                signs are mirrored (SWIFT C ↔ Flexcube DR).
    Tier 2 — REFERENCE         : reference hit + mirrored signs, but amounts
                                don't tie to the cent (FX, partial booking).
    Tier 3 — AMOUNT+DATE       : no reference link, but same value date,
                                exact amount, mirrored signs.
    Tier 4 — AMOUNT±1DAY       : same as Tier 3 but value dates may differ
                                by one day (cross-day booking timing).

    Every match is tagged with the tier that made it so the ops team can
    triage — Tier 1 matches are near-certain, Tier 4 matches deserve a
    second pair of eyes.

Folder layout this script expects:

    Transformer/
        reconcile.py           <-- this file
        input_swift/           <-- drop the SWIFT parsed xlsx here
        input_flexcube/        <-- drop the Flexcube raw xlsx here
        output_reconciled/     <-- the match report lands here

You can have multiple files in either input folder. The script pairs them
by filename order — first SWIFT file with first Flexcube file, etc. Usually
you'll have one of each per run, which is the simplest case.

Nothing here uses a fancy library — just openpyxl for reading and writing
Excel files.
"""

import re
import sys
from pathlib import Path
from datetime import datetime, timedelta
from collections import Counter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import logging
logger = logging.getLogger(__name__)



# ---------------------------------------------------------------------------
# Configuration — change these if the bank's conventions ever shift.
# ---------------------------------------------------------------------------

# A SWIFT credit ("C" = money coming INTO our nostro) is booked in Flexcube
# as a DR in our books. This dict encodes that mirror rule.
MIRROR_SIGN = {'C': 'DR', 'D': 'CR'}

# Amounts under this many cedi apart count as "equal" — protects against
# floating-point rounding after FX conversions etc. Tune if needed.
AMOUNT_TOLERANCE = 0.01

# Shortest SWIFT reference we'll accept for a substring search. Refs shorter
# than this (e.g. "50", "ABC") would match too many narrations by accident.
MIN_REF_LENGTH = 6


# ---------------------------------------------------------------------------
# Data loaders — one for each side of the reconciliation.
# ---------------------------------------------------------------------------

def load_swift(path: Path) -> list:
    """Read a SWIFT-parsed xlsx and return a list of transaction dicts.

    The file structure is: metadata panel at the top, one blank row, a
    header row starting with "Value date", then one row per transaction.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    # Find the header row — it's the one whose first cell is "Value date".
    header_idx = None
    for i, r in enumerate(rows):
        if r and r[0] == 'Value date':
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(f"'{path.name}' doesn't look like a SWIFT parsed file — no 'Value date' header found.")

    txns = []
    for row_idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if not row or row[0] is None:
            continue
        txns.append({
            '_source': 'swift',
            '_row_number': row_idx,      # original row in the source file
            '_used': False,              # flips True when matched
            'value_date': row[0],
            'amount': _to_float(row[1]),
            'sign': row[2],              # 'C' or 'D'
            'origin': row[3],
            'type': row[4],
            'status': row[5],
            'book_date': row[6],
            'our_ref': str(row[7]) if row[7] else '',
            'their_ref': str(row[8]) if row[8] else '',
            'booking_text_1': str(row[9]) if row[9] else '',
            'booking_text_2': str(row[10]) if row[10] else '',
        })
    return txns


def load_flexcube(path: Path) -> list:
    """Read a raw Flexcube acc_entries xlsx and return a list of dicts.

    The raw Flexcube file has named columns in row 1 — TRN_REF_NO,
    BOOKING_DATE, TYPE, TXN_NARRATIONS, VALUE_DATE, LCY_AMOUNT, etc.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        raise ValueError(f"'{path.name}' is empty.")

    headers = [str(h).strip() if h else '' for h in rows[0]]
    expected = {'TRN_REF_NO', 'BOOKING_DATE', 'TYPE', 'TXN_NARRATIONS',
                'VALUE_DATE', 'LCY_AMOUNT'}
    missing = expected - set(headers)
    if missing:
        raise ValueError(
            f"'{path.name}' is missing expected Flexcube columns: {sorted(missing)}"
        )
    idx = {name: headers.index(name) for name in headers if name}

    txns = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if not row or row[idx['TRN_REF_NO']] is None:
            continue
        txns.append({
            '_source': 'flexcube',
            '_row_number': row_idx,
            '_used': False,
            'trn_ref': str(row[idx['TRN_REF_NO']]).strip(),
            'ac_branch': str(row[idx.get('AC_BRANCH', -1)] or '') if 'AC_BRANCH' in idx else '',
            'ac_no': str(row[idx.get('AC_NO', -1)] or '') if 'AC_NO' in idx else '',
            'booking_date': _excel_date_to_int(row[idx['BOOKING_DATE']]),
            'value_date': _excel_date_to_int(row[idx['VALUE_DATE']]),
            'type': str(row[idx['TYPE']] or '').strip(),   # 'DR' or 'CR'
            'narration': str(row[idx['TXN_NARRATIONS']] or ''),
            'amount': _to_float(row[idx['LCY_AMOUNT']]),
            'ccy': str(row[idx.get('ACCT_CCY', -1)] or '') if 'ACCT_CCY' in idx else '',
            'module': str(row[idx.get('MODULE', -1)] or '') if 'MODULE' in idx else '',
            'external_ref': str(row[idx.get('EXTERNAL_REF_NO', -1)] or '') if 'EXTERNAL_REF_NO' in idx else '',
            'user_id': str(row[idx.get('USER_ID', -1)] or '') if 'USER_ID' in idx else '',
        })
    return txns


def read_balance_sheet(path: Path) -> dict | None:
    """Read the optional `balances` sheet from a Flexcube delta xlsx.

    Returns ``{as_of_date, opening_balance, closing_balance, currency}``
    when the sheet is present and well-formed, or None when it isn't —
    callers treat None as "no embedded balance, run without continuity
    check". Tolerant of extra columns; only the four canonical fields
    are surfaced.

    The sheet is what the extract script writes alongside acc_entries
    so each delta is self-describing — no side-car TSV needed for the
    balance-chain check.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if 'balances' not in wb.sheetnames:
            return None
        ws = wb['balances']
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if len(rows) < 2:
        return None
    headers = [str(h).strip() if h else '' for h in rows[0]]
    required = {'as_of_date', 'opening_balance', 'closing_balance', 'currency'}
    if not required.issubset(set(headers)):
        return None
    idx = {h: headers.index(h) for h in headers if h}
    data = rows[1]
    try:
        as_of = data[idx['as_of_date']]
        opening = data[idx['opening_balance']]
        closing = data[idx['closing_balance']]
        ccy = data[idx['currency']]
    except (IndexError, KeyError):
        return None
    if opening is None or closing is None:
        return None
    return {
        'as_of_date':      _balance_as_of_to_int(as_of),
        'opening_balance': float(opening),
        'closing_balance': float(closing),
        'currency':        str(ccy or '').strip().upper(),
    }


def _balance_as_of_to_int(v) -> int:
    """The balances sheet typically writes as_of_date as an ISO string
    ('2026-04-30') or a Python date — normalise to the YYYYMMDD int
    shape every other date column in the engine uses."""
    if v is None or v == '':
        return 0
    if isinstance(v, datetime):
        return int(v.strftime('%Y%m%d'))
    if hasattr(v, 'year') and hasattr(v, 'month') and hasattr(v, 'day'):
        return int(f"{v.year:04d}{v.month:02d}{v.day:02d}")
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%d-%b-%Y', '%d/%m/%Y'):
        try:
            return int(datetime.strptime(s, fmt).strftime('%Y%m%d'))
        except ValueError:
            pass
    if s.isdigit() and len(s) == 8:
        return int(s)
    return 0


def _to_float(x) -> float:
    """Coerce a cell value to a float, treating blanks and non-numerics as 0.0."""
    if x is None or x == '':
        return 0.0
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def _excel_date_to_int(x) -> int:
    """Turn dates in various forms into YYYYMMDD integers (to match SWIFT).

    Handles Python datetime, Excel date string like '17-Apr-2026', and
    already-YYYYMMDD integers. Returns 0 if the value can't be parsed.
    """
    if x is None or x == '':
        return 0
    if isinstance(x, int):
        # Already YYYYMMDD or some other int. If it's large (20260417-ish),
        # trust it. If it's small (like 45000), it's an Excel serial.
        return x if x > 19000101 else _excel_serial_to_int(x)
    if isinstance(x, datetime):
        return int(x.strftime('%Y%m%d'))
    s = str(x).strip()
    # Try '17-Apr-2026'
    for fmt in ('%d-%b-%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return int(datetime.strptime(s, fmt).strftime('%Y%m%d'))
        except ValueError:
            pass
    return 0


def _excel_serial_to_int(serial: int) -> int:
    """Excel stores dates as days since 1900-01-00 (with a leap-year bug)."""
    try:
        # Excel's epoch has a famous off-by-one; 1900 is treated as leap.
        base = datetime(1899, 12, 30)
        dt = base + timedelta(days=serial)
        return int(dt.strftime('%Y%m%d'))
    except (OverflowError, OSError):
        return 0


def _days_between(d1: int, d2: int) -> int:
    """Return abs(days) between two YYYYMMDD integers; 999 if either is invalid."""
    try:
        a = datetime.strptime(str(d1), '%Y%m%d')
        b = datetime.strptime(str(d2), '%Y%m%d')
        return abs((a - b).days)
    except ValueError:
        return 999


# ---------------------------------------------------------------------------
# The matching engine — four rule tiers.
# ---------------------------------------------------------------------------

def match(swift_txns: list, flex_txns: list) -> list:
    """Produce a list of match records — {swift, flex, tier, reason}.

    Each tier is tried in order. Once a row is '_used' in a match, no later
    tier reuses it. Order matters: Tier 1 is most trustworthy, Tier 4 least.
    """
    matches = []

    # Tier 1 — reference inside narration, amounts equal, mirrored signs.
    for s in swift_txns:
        if s['_used'] or not _ref_long_enough(s['our_ref']):
            continue
        target_type = MIRROR_SIGN.get(s['sign'])
        if target_type is None:
            continue
        for f in flex_txns:
            if f['_used']:
                continue
            if f['type'] != target_type:
                continue
            if abs(f['amount'] - s['amount']) > AMOUNT_TOLERANCE:
                continue
            if _ref_in_narration(s['our_ref'], f):
                matches.append(_make_match(s, f, tier=1,
                    reason=f"ref {s['our_ref']} in narration; amount={s['amount']:.2f}; sign mirror"))
                break

    # Tier 2 — reference hit + mirrored signs, amounts may differ.
    for s in swift_txns:
        if s['_used'] or not _ref_long_enough(s['our_ref']):
            continue
        target_type = MIRROR_SIGN.get(s['sign'])
        if target_type is None:
            continue
        for f in flex_txns:
            if f['_used'] or f['type'] != target_type:
                continue
            if _ref_in_narration(s['our_ref'], f):
                matches.append(_make_match(s, f, tier=2,
                    reason=f"ref {s['our_ref']} in narration; sign mirror; amounts differ ({s['amount']:.2f} vs {f['amount']:.2f})"))
                break

    # Tier 3 — no reference link, same value date, exact amount, mirrored sign.
    for s in swift_txns:
        if s['_used']:
            continue
        target_type = MIRROR_SIGN.get(s['sign'])
        if target_type is None:
            continue
        for f in flex_txns:
            if f['_used'] or f['type'] != target_type:
                continue
            if abs(f['amount'] - s['amount']) > AMOUNT_TOLERANCE:
                continue
            if f['value_date'] == s['value_date']:
                matches.append(_make_match(s, f, tier=3,
                    reason=f"same value date {s['value_date']}; amount={s['amount']:.2f}; sign mirror; no ref link"))
                break

    # Tier 4 — amount + mirrored sign, value date within one day.
    for s in swift_txns:
        if s['_used']:
            continue
        target_type = MIRROR_SIGN.get(s['sign'])
        if target_type is None:
            continue
        for f in flex_txns:
            if f['_used'] or f['type'] != target_type:
                continue
            if abs(f['amount'] - s['amount']) > AMOUNT_TOLERANCE:
                continue
            if _days_between(f['value_date'], s['value_date']) <= 1:
                matches.append(_make_match(s, f, tier=4,
                    reason=f"amount={s['amount']:.2f}; sign mirror; value dates {s['value_date']} vs {f['value_date']} (±1 day)"))
                break

    return matches


def _ref_long_enough(ref: str) -> bool:
    """Protect against short refs like '50' matching hundreds of narrations."""
    return ref and len(ref.strip()) >= MIN_REF_LENGTH


def _ref_in_narration(ref: str, flex_row: dict) -> bool:
    """Case-insensitive search for the SWIFT ref in any Flexcube ref/narration field.

    Checks narration, external_ref, and trn_ref. The ref is the primary
    signal — if the correspondent bank quoted our ref, it usually echoes
    somewhere in the narration (or, occasionally, in EXTERNAL_REF_NO).
    """
    needle = ref.strip().upper()
    if not needle:
        return False
    for hay_field in ('narration', 'external_ref', 'trn_ref'):
        hay = (flex_row.get(hay_field) or '').upper()
        if needle in hay:
            return True
    return False


def _make_match(s: dict, f: dict, tier: int, reason: str) -> dict:
    """Lock both rows and build a match record."""
    s['_used'] = True
    f['_used'] = True
    return {'swift': s, 'flex': f, 'tier': tier, 'reason': reason}


# ---------------------------------------------------------------------------
# Report writer — multi-sheet Excel output.
# ---------------------------------------------------------------------------

MATCHED_HEADERS = [
    'Tier', 'Match reason',
    'SWIFT value date', 'SWIFT amount', 'SWIFT sign', 'SWIFT our ref',
    'SWIFT booking text 1', 'SWIFT booking text 2',
    'Flex TRN ref', 'Flex type', 'Flex amount', 'Flex value date',
    'Flex booking date', 'Flex narration', 'Flex user', 'Flex module',
    'Amount diff',
]

UNMATCHED_SWIFT_HEADERS = [
    'Row #', 'Value date', 'Amount', 'Sign', 'Our ref', 'Their ref',
    'Booking text 1', 'Booking text 2',
]

UNMATCHED_FLEX_HEADERS = [
    'Row #', 'TRN ref', 'Type', 'Amount', 'Value date', 'Booking date',
    'Narration', 'External ref', 'User', 'Module',
]


def write_report(matches: list, swift_txns: list, flex_txns: list,
                 swift_path: Path, flex_path: Path, output_path: Path) -> None:
    """Compose the four-sheet reconciliation workbook."""
    wb = Workbook()
    wb.remove(wb.active)  # discard the default blank sheet

    _write_summary(wb, matches, swift_txns, flex_txns, swift_path, flex_path)
    _write_matched(wb, matches)
    _write_unmatched_swift(wb, swift_txns)
    _write_unmatched_flex(wb, flex_txns)

    wb.save(output_path)


def write_one_sided_report(matches: list, swift_txns: list, flex_txns: list,
                           flex_path: Path, output_path: Path,
                           *, session_meta: dict) -> None:
    """One-sided session export — same row content as write_report but
    relabeled for proof seeds and Flex-only deltas. SWIFT-side rows are
    actually the DR legs of the source file (the engine reuses the
    swift-shape for the debit side of a one-sided GL); Flex-side rows
    are the CR legs.

    `session_meta` carries the headline numbers a one-sided session
    needs in the audit artifact:
        kind                  'seed' | 'flex_delta'
        account_label
        currency
        period_start          YYYYMMDD int
        period_end            YYYYMMDD int
        flex_opening_balance  signed float | None
        flex_closing_balance  signed float | None
        anchor_before         signed float | None  (delta only)
        anchor_after          signed float | None
        open_items_seeded     int
        open_items_cleared    int
        force_accept          bool
        continuity_delta      float | None  (delta only — actual − expected)
    """
    wb = Workbook()
    wb.remove(wb.active)

    _write_one_sided_summary(wb, matches, swift_txns, flex_txns, flex_path,
                              session_meta)
    _write_matched(wb, matches)
    _write_unmatched_dr_legs(wb, swift_txns)
    _write_unmatched_cr_legs(wb, flex_txns)

    wb.save(output_path)


def _write_one_sided_summary(wb, matches, swift_txns, flex_txns, flex_path, meta):
    ws = wb.create_sheet('Summary')
    label_font = Font(bold=True)

    kind = meta.get('kind') or 'flex_delta'
    title = ('Proof seed summary' if kind == 'seed'
             else 'Flex delta reconciliation summary')

    swift_total = len(swift_txns)
    flex_total = len(flex_txns)
    matched_count = len(matches)
    tier_counts = Counter(m['tier'] for m in matches)

    unmatched_swift = [s for s in swift_txns if not s['_used']]
    unmatched_flex = [f for f in flex_txns if not f['_used']]
    unmatched_swift_value = sum(s['amount'] for s in unmatched_swift)
    unmatched_flex_value = sum(f['amount'] for f in unmatched_flex)

    def _fmt_period(start, end):
        def _fd(d):
            try:
                return f"{int(d):08d}"
            except (TypeError, ValueError):
                return ''
        if start and end and start != end:
            return f"{_fd(start)} → {_fd(end)}"
        return _fd(end or start) or '—'

    def _fmt_signed(amt):
        if amt is None:
            return '—'
        sign = 'D' if amt < 0 else 'C'
        return f"{sign} {abs(amt):,.2f}"

    rows = [
        (title, ''),
        ('', ''),
        ('Source file',           flex_path.name),
        ('Account',               meta.get('account_label') or ''),
        ('Currency',              meta.get('currency') or ''),
        ('Period',                _fmt_period(meta.get('period_start'),
                                                meta.get('period_end'))),
        ('Generated',             datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('', ''),
    ]

    if kind == 'seed':
        rows += [
            ('Anchor opening',     _fmt_signed(0.0)),
            ('Anchor closing',     _fmt_signed(meta.get('flex_closing_balance'))),
            ('', ''),
        ]
    else:  # flex_delta
        rows += [
            ('Anchor before',      _fmt_signed(meta.get('anchor_before'))),
            ('File opening',       _fmt_signed(meta.get('flex_opening_balance'))),
            ('File closing',       _fmt_signed(meta.get('flex_closing_balance'))),
            ('Anchor after',       _fmt_signed(meta.get('anchor_after'))),
            ('', ''),
            # Continuity gate result. delta=0 = perfect chain; non-zero
            # means force-accepted (operator overrode the break).
            ('Continuity Δ',
                ('force-accepted' if meta.get('force_accept')
                 else (f"{meta.get('continuity_delta'):,.2f}"
                        if meta.get('continuity_delta') is not None else '0.00'))),
            ('', ''),
        ]

    rows += [
        ('DR legs (swift_txns side)',  swift_total),
        ('CR legs (flex_txns side)',   flex_total),
        ('Matched (pairs)',            matched_count),
        ('', ''),
        ('  Tier 1 (strictest)',       tier_counts.get(1, 0)),
        ('  Tier 2 (ref, amt diff)',   tier_counts.get(2, 0)),
        ('  Tier 3 (amt + date)',      tier_counts.get(3, 0)),
        ('  Tier 4 (amt ±1 day)',      tier_counts.get(4, 0)),
        ('', ''),
        ('Unmatched DR legs',          len(unmatched_swift)),
        ('Unmatched DR value',         unmatched_swift_value),
        ('Unmatched CR legs',          len(unmatched_flex)),
        ('Unmatched CR value',         unmatched_flex_value),
        ('', ''),
        ('Open items cleared (carry-forward)',  meta.get('open_items_cleared', 0)),
        ('Open items seeded (residue)',          meta.get('open_items_seeded', 0)),
        ('', ''),
        ('Match rate',
            f"{(matched_count / max(swift_total, flex_total) * 100) if (swift_total or flex_total) else 0:.1f}%"),
    ]

    for r_idx, (label, value) in enumerate(rows, start=1):
        cell = ws.cell(row=r_idx, column=1, value=label)
        cell.font = label_font if label else Font()
        if label == title:
            cell.font = Font(bold=True, size=14)
        ws.cell(row=r_idx, column=2, value=value)
        if isinstance(value, float):
            ws.cell(row=r_idx, column=2).number_format = '#,##0.00'

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 40


def _write_unmatched_dr_legs(wb, swift_txns):
    """Same content as _write_unmatched_swift but relabeled — for one-
    sided sessions the swift_txns table holds the DR legs of the source
    file (split via _split_flex_for_self_match at ingest time)."""
    ws = wb.create_sheet('Unmatched DR legs')
    headers = ('Row #', 'Value date', 'Amount', 'Sign', 'Trn ref',
                'External ref', 'Booking text 1', 'Booking text 2')
    _styled_header(ws, list(headers))
    for r_idx, s in enumerate(
            (x for x in swift_txns if not x['_used']), start=2):
        ws.cell(row=r_idx, column=1, value=s['_row_number'])
        ws.cell(row=r_idx, column=2, value=s['value_date'])
        ws.cell(row=r_idx, column=3, value=s['amount'])
        ws.cell(row=r_idx, column=4, value=s['sign'])
        ws.cell(row=r_idx, column=5, value=s['our_ref'])
        ws.cell(row=r_idx, column=6, value=s['their_ref'])
        ws.cell(row=r_idx, column=7, value=s['booking_text_1'])
        ws.cell(row=r_idx, column=8, value=s['booking_text_2'])
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 20


def _write_unmatched_cr_legs(wb, flex_txns):
    """Same content as _write_unmatched_flex but relabeled."""
    ws = wb.create_sheet('Unmatched CR legs')
    headers = ('Row #', 'Trn ref', 'Type', 'Amount', 'Value date',
                'Booking date', 'Narration', 'External ref', 'User', 'Module')
    _styled_header(ws, list(headers))
    for r_idx, f in enumerate(
            (x for x in flex_txns if not x['_used']), start=2):
        ws.cell(row=r_idx, column=1, value=f['_row_number'])
        ws.cell(row=r_idx, column=2, value=f['trn_ref'])
        ws.cell(row=r_idx, column=3, value=f['type'])
        ws.cell(row=r_idx, column=4, value=f['amount'])
        ws.cell(row=r_idx, column=5, value=f['value_date'])
        ws.cell(row=r_idx, column=6, value=f['booking_date'])
        ws.cell(row=r_idx, column=7, value=f['narration'])
        ws.cell(row=r_idx, column=8, value=f['external_ref'])
        ws.cell(row=r_idx, column=9, value=f['user_id'])
        ws.cell(row=r_idx, column=10, value=f['module'])
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['G'].width = 60


def _styled_header(ws, headers: list, row: int = 1):
    """Apply a consistent header band to row 1 of a worksheet."""
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border


def _write_summary(wb, matches, swift_txns, flex_txns, swift_path, flex_path):
    ws = wb.create_sheet('Summary')
    label_font = Font(bold=True)

    swift_total = len(swift_txns)
    flex_total = len(flex_txns)
    matched_count = len(matches)
    tier_counts = Counter(m['tier'] for m in matches)

    # Unmatched totals and value totals.
    unmatched_swift = [s for s in swift_txns if not s['_used']]
    unmatched_flex = [f for f in flex_txns if not f['_used']]
    unmatched_swift_value = sum(s['amount'] for s in unmatched_swift)
    unmatched_flex_value = sum(f['amount'] for f in unmatched_flex)

    rows = [
        ('Reconciliation summary', ''),
        ('', ''),
        ('SWIFT file',               swift_path.name),
        ('Flexcube file',            flex_path.name),
        ('Generated',                datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('', ''),
        ('SWIFT transactions',       swift_total),
        ('Flexcube transactions',    flex_total),
        ('Matched (pairs)',          matched_count),
        ('',                         ''),
        ('  Tier 1 (strictest)',     tier_counts.get(1, 0)),
        ('  Tier 2 (ref, amt diff)', tier_counts.get(2, 0)),
        ('  Tier 3 (amt + date)',    tier_counts.get(3, 0)),
        ('  Tier 4 (amt ±1 day)',    tier_counts.get(4, 0)),
        ('', ''),
        ('Unmatched SWIFT rows',     len(unmatched_swift)),
        ('Unmatched SWIFT value',    unmatched_swift_value),
        ('Unmatched Flexcube rows',  len(unmatched_flex)),
        ('Unmatched Flexcube value', unmatched_flex_value),
        ('', ''),
        ('Match rate (SWIFT side)',  f"{(matched_count / swift_total * 100) if swift_total else 0:.1f}%"),
    ]

    for r_idx, (label, value) in enumerate(rows, start=1):
        ws.cell(row=r_idx, column=1, value=label).font = label_font if label else Font()
        if label == 'Reconciliation summary':
            ws.cell(row=r_idx, column=1).font = Font(bold=True, size=14)
        ws.cell(row=r_idx, column=2, value=value)
        if isinstance(value, float):
            ws.cell(row=r_idx, column=2).number_format = '#,##0.00'

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40


def _write_matched(wb, matches):
    ws = wb.create_sheet('Matched')
    _styled_header(ws, MATCHED_HEADERS)

    # Sort so Tier 1 matches appear at the top — ops scans the riskier ones last.
    sorted_matches = sorted(matches, key=lambda m: (m['tier'], -m['swift']['amount']))

    for r_idx, m in enumerate(sorted_matches, start=2):
        s = m['swift']
        f = m['flex']
        # tier=0 is the engine convention for carry-forward (counterpart
        # came from a prior session's open_items, not from this session's
        # tier 1-6 engine pass). Surface it with a human label so the
        # export reads "Carry-forward" instead of "0".
        tier_display = 'Carry-forward' if m['tier'] == 0 else m['tier']
        ws.cell(row=r_idx, column=1,  value=tier_display)
        ws.cell(row=r_idx, column=2,  value=m['reason'])
        ws.cell(row=r_idx, column=3,  value=s['value_date'])
        ws.cell(row=r_idx, column=4,  value=s['amount'])
        ws.cell(row=r_idx, column=5,  value=s['sign'])
        ws.cell(row=r_idx, column=6,  value=s['our_ref'])
        ws.cell(row=r_idx, column=7,  value=s['booking_text_1'])
        ws.cell(row=r_idx, column=8,  value=s['booking_text_2'])
        ws.cell(row=r_idx, column=9,  value=f['trn_ref'])
        ws.cell(row=r_idx, column=10, value=f['type'])
        ws.cell(row=r_idx, column=11, value=f['amount'])
        ws.cell(row=r_idx, column=12, value=f['value_date'])
        ws.cell(row=r_idx, column=13, value=f['booking_date'])
        ws.cell(row=r_idx, column=14, value=f['narration'])
        ws.cell(row=r_idx, column=15, value=f['user_id'])
        ws.cell(row=r_idx, column=16, value=f['module'])
        ws.cell(row=r_idx, column=17, value=f['amount'] - s['amount'])

        # Highlight Tier 2, 3, 4 rows with a pale tint so ops can spot them.
        tier_fill = {
            2: PatternFill(start_color='FFF4D6', end_color='FFF4D6', fill_type='solid'),
            3: PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid'),
            4: PatternFill(start_color='FFD6D6', end_color='FFD6D6', fill_type='solid'),
        }
        if m['tier'] in tier_fill:
            for col in range(1, len(MATCHED_HEADERS) + 1):
                ws.cell(row=r_idx, column=col).fill = tier_fill[m['tier']]

        ws.cell(row=r_idx, column=4).number_format = '#,##0.00'
        ws.cell(row=r_idx, column=11).number_format = '#,##0.00'
        ws.cell(row=r_idx, column=17).number_format = '#,##0.00;[Red]-#,##0.00'

    # Column widths tuned for readability.
    widths = {'A': 6, 'B': 50, 'C': 13, 'D': 14, 'E': 6, 'F': 22,
              'G': 30, 'H': 30, 'I': 20, 'J': 6, 'K': 14, 'L': 13,
              'M': 13, 'N': 50, 'O': 14, 'P': 8, 'Q': 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'


def _write_unmatched_swift(wb, swift_txns):
    ws = wb.create_sheet('Unmatched SWIFT')
    _styled_header(ws, UNMATCHED_SWIFT_HEADERS)

    unmatched = [s for s in swift_txns if not s['_used']]
    for r_idx, s in enumerate(unmatched, start=2):
        ws.cell(row=r_idx, column=1, value=s['_row_number'])
        ws.cell(row=r_idx, column=2, value=s['value_date'])
        ws.cell(row=r_idx, column=3, value=s['amount'])
        ws.cell(row=r_idx, column=4, value=s['sign'])
        ws.cell(row=r_idx, column=5, value=s['our_ref'])
        ws.cell(row=r_idx, column=6, value=s['their_ref'])
        ws.cell(row=r_idx, column=7, value=s['booking_text_1'])
        ws.cell(row=r_idx, column=8, value=s['booking_text_2'])
        ws.cell(row=r_idx, column=3).number_format = '#,##0.00'

    widths = {'A': 6, 'B': 13, 'C': 14, 'D': 6, 'E': 22, 'F': 22, 'G': 40, 'H': 40}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'


def _write_unmatched_flex(wb, flex_txns):
    ws = wb.create_sheet('Unmatched Flexcube')
    _styled_header(ws, UNMATCHED_FLEX_HEADERS)

    unmatched = [f for f in flex_txns if not f['_used']]
    for r_idx, f in enumerate(unmatched, start=2):
        ws.cell(row=r_idx, column=1,  value=f['_row_number'])
        ws.cell(row=r_idx, column=2,  value=f['trn_ref'])
        ws.cell(row=r_idx, column=3,  value=f['type'])
        ws.cell(row=r_idx, column=4,  value=f['amount'])
        ws.cell(row=r_idx, column=5,  value=f['value_date'])
        ws.cell(row=r_idx, column=6,  value=f['booking_date'])
        ws.cell(row=r_idx, column=7,  value=f['narration'])
        ws.cell(row=r_idx, column=8,  value=f['external_ref'])
        ws.cell(row=r_idx, column=9,  value=f['user_id'])
        ws.cell(row=r_idx, column=10, value=f['module'])
        ws.cell(row=r_idx, column=4).number_format = '#,##0.00'

    widths = {'A': 6, 'B': 20, 'C': 6, 'D': 14, 'E': 13, 'F': 13,
              'G': 50, 'H': 20, 'I': 14, 'J': 8}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'


# ---------------------------------------------------------------------------
# Batch runner — the entry point when you run the script.
# ---------------------------------------------------------------------------

def main():
    script_dir = Path(__file__).resolve().parent
    swift_dir = script_dir / 'input_swift'
    flex_dir = script_dir / 'input_flexcube'
    out_dir = script_dir / 'output_reconciled'

    logger.info("SWIFT ↔ Flexcube Reconciliation")
    logger.info("-" * 60)
    logger.info(f"SWIFT input    : {swift_dir}")
    logger.info(f"Flexcube input : {flex_dir}")
    logger.info(f"Output         : {out_dir}")
    logger.info()
    out_dir.mkdir(parents=True, exist_ok=True)

    missing = [d for d in (swift_dir, flex_dir) if not d.exists()]
    if missing:
        for d in missing:
            logger.error(f"ERROR: folder missing — {d}")
        logger.info("\nCreate the missing folder(s), drop your files inside, and re-run.")
        input("\nPress Enter to exit...")
        return 1

    swift_files = sorted(p for p in swift_dir.iterdir()
                         if p.is_file() and p.suffix.lower() in ('.xlsx', '.xlsm')
                         and not p.name.startswith('.') and not p.name.startswith('~$'))
    flex_files = sorted(p for p in flex_dir.iterdir()
                        if p.is_file() and p.suffix.lower() in ('.xlsx', '.xlsm')
                        and not p.name.startswith('.') and not p.name.startswith('~$'))

    if not swift_files:
        logger.info(f"No .xlsx files in {swift_dir}.")
        input("\nPress Enter to exit...")
        return 0
    if not flex_files:
        logger.info(f"No .xlsx files in {flex_dir}.")
        input("\nPress Enter to exit...")
        return 0

    # Pair by sort order. Warn if counts differ — something is probably off.
    pair_count = min(len(swift_files), len(flex_files))
    if len(swift_files) != len(flex_files):
        logger.warning(f"WARNING: {len(swift_files)} SWIFT file(s) vs {len(flex_files)} Flexcube file(s). "
              f"Pairing the first {pair_count} of each by sort order; extras ignored.")
        logger.info()
    succeeded = 0
    failed = []

    for swift_path, flex_path in zip(swift_files[:pair_count], flex_files[:pair_count]):
        logger.info(f"Reconciling: {swift_path.name}  <>  {flex_path.name}")
        try:
            swift_txns = load_swift(swift_path)
            flex_txns = load_flexcube(flex_path)
            logger.info(f"  Loaded {len(swift_txns)} SWIFT rows, {len(flex_txns)} Flexcube rows")
            matches = match(swift_txns, flex_txns)
            tiers = Counter(m['tier'] for m in matches)

            output_name = f"{swift_path.stem}_vs_{flex_path.stem}.xlsx"
            output_path = out_dir / output_name
            write_report(matches, swift_txns, flex_txns, swift_path, flex_path, output_path)

            logger.info(f"  Matched: {len(matches)}  "
                  f"(T1={tiers.get(1, 0)}, T2={tiers.get(2, 0)}, "
                  f"T3={tiers.get(3, 0)}, T4={tiers.get(4, 0)})")
            logger.info(f"  Unmatched SWIFT: {sum(1 for s in swift_txns if not s['_used'])}  "
                  f"Unmatched Flexcube: {sum(1 for f in flex_txns if not f['_used'])}")
            logger.info(f"  Wrote {output_name}")
            succeeded += 1

        except Exception as exc:
            logger.error(f"  FAILED: {exc}")
            failed.append((swift_path.name, flex_path.name, str(exc)))

        logger.info()
    logger.info("-" * 60)
    logger.error(f"Done. {succeeded} reconciliation(s) written, {len(failed)} failed.")
    if failed:
        logger.info("\nFailures:")
        for s, f, err in failed:
            logger.info(f"  {s} <> {f}: {err}")
    input("\nPress Enter to exit...")
    return 0 if not failed else 1


if __name__ == '__main__':
    sys.exit(main())
