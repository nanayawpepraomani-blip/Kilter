"""
proof_loader.py
===============

Reads the Day-0 "proof" xlsx that anchors a one-sided account's balance
chain. The proof is the snapshot we load once at onboarding so Kilter
knows where to continue from; subsequent daily Flex deltas chain off the
closing balance this file establishes.

Format we accept (the Ecobank "GL extract" shape, what ops calls a
"proof"):

    Row 0:  header — first cell is literally `Account`. Columns:
            Account | Value date | Curr. | Amount | S | (subtotal) |
            Origin | Type | Status | (age) | Book. date |
            Our reference 1 | Their reference 1 |
            Booking text 1 | Booking text 2
    Row 1:  optional blank row Ecobank's exporter inserts.
    Row 2…N-1: transaction rows. `S` is 'C' or 'D'; subtotal column
            holds the running balance and isn't read directly (we
            recompute from signed amounts so we don't trust the file's
            arithmetic).
    Row N:  optional footer with the running total in the subtotal
            column and every other cell None — skipped because its `S`
            is None.

Output: list of canonical flex-shape txn dicts (same shape
reconcile.load_flexcube emits) so the loader plugs straight into the
existing engine without special-casing.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


PROOF_REQUIRED = ('Value date', 'Amount', 'S', 'Our reference 1')


def load_proof(path: Path | str) -> list[dict]:
    """Parse a proof xlsx into a list of flex-shape txn dicts.

    Looks for a header row whose first cell is `Account`, validates the
    required columns are present, then walks the remaining rows. Skips
    blank rows and the trailing footer (rows whose `S` cell is empty).

    Raises ValueError when no recognisable header is found, or when the
    file is missing required columns — caller translates that into a
    400 at the API layer.
    """
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        # Scan every sheet (skipping a `balances` sidecar if present so
        # we never confuse it with the data sheet) and pick the first
        # one that has a recognisable proof header. Operators sometimes
        # save with the wrong sheet selected as active — wb.active alone
        # would miss the data.
        raw = None
        header_idx = None
        for sn in wb.sheetnames:
            if sn.lower() == 'balances':
                continue
            ws = wb[sn]
            candidate = list(ws.iter_rows(values_only=True))
            h_idx = _find_header(candidate)
            if h_idx is not None:
                raw, header_idx = candidate, h_idx
                break
    finally:
        wb.close()

    if header_idx is None:
        raise ValueError(
            f"'{path.name}' doesn't look like a proof file — "
            "no sheet has a row whose first cell is 'Account'.")

    headers = [str(c).strip() if c else '' for c in raw[header_idx]]
    missing = [h for h in PROOF_REQUIRED if h not in headers]
    if missing:
        raise ValueError(
            f"'{path.name}' is missing required proof columns: {missing}. "
            f"Expected at least: {list(PROOF_REQUIRED)}.")

    col = {name: headers.index(name) for name in headers if name}

    txns: list[dict] = []
    rn = 0
    for row in raw[header_idx + 1:]:
        if not row or all(c is None or c == '' for c in row):
            continue
        if row[0] is None:
            continue
        parsed = _parse_row(row, col)
        if parsed is None:
            continue
        rn += 1
        parsed['_row_number'] = rn
        txns.append(parsed)
    return txns


def compute_seed_balance(txns: list[dict]) -> tuple[float, int]:
    """Return (closing_balance, max_value_date) for a freshly loaded
    proof. Opening is always 0 for a seed by definition; the closing
    is the signed sum of every row (CR positive, DR negative), and the
    high-water value date pins the as-of so the next delta can be
    range-checked."""
    closing = 0.0
    max_date = 0
    for t in txns:
        signed = t['amount'] if t['type'] == 'CR' else -t['amount']
        closing += signed
        if t['value_date'] > max_date:
            max_date = t['value_date']
    return round(closing, 2), max_date


# ---------------------------------------------------------------------------
# Internals
# ---------------------------------------------------------------------------

def _find_header(rows: list[tuple]) -> int | None:
    """Return the row index whose first cell is literally `Account`,
    or None when no such row exists."""
    for i, r in enumerate(rows):
        if not r:
            continue
        first = (str(r[0]).strip() if r[0] is not None else '')
        if first == 'Account':
            return i
    return None


def _parse_row(row: tuple, col: dict) -> dict | None:
    s = (row[col['S']] or '')
    s = s.strip() if isinstance(s, str) else s
    if s not in ('C', 'D'):
        return None  # subtotal / footer / garbage row
    ref1 = _str(row[col['Our reference 1']])
    if not ref1:
        return None
    return {
        '_source': 'flex',
        '_used': False,
        'trn_ref':       ref1,
        'ac_branch':     '',
        'ac_no':         _str(row[0]),
        'booking_date':  _to_int_date(_get(row, col, 'Book. date')),
        'value_date':    _to_int_date(row[col['Value date']]),
        # CR/DR shape mirrors what reconcile.load_flexcube produces — the
        # engine already knows that mapping (MIRROR_SIGN['C']='DR' etc.).
        'type':          'CR' if s == 'C' else 'DR',
        'narration':     _join_narration(_get(row, col, 'Booking text 1'),
                                          _get(row, col, 'Booking text 2')),
        'amount':        abs(_to_float(row[col['Amount']])),
        'ccy':           _str(_get(row, col, 'Curr.')).upper(),
        'module':        _str(_get(row, col, 'Type')),
        'external_ref':  _str(_get(row, col, 'Their reference 1')),
        'user_id':       '',
    }


def _get(row: tuple, col: dict, name: str):
    """None when the column isn't in this proof file — older / variant
    exports occasionally drop optional columns entirely; the loader
    tolerates that."""
    idx = col.get(name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _str(v) -> str:
    if v is None:
        return ''
    return str(v).strip()


def _to_float(v) -> float:
    if v is None or v == '':
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _to_int_date(v) -> int:
    """Returns YYYYMMDD as int, 0 on failure. Proof exports mix int
    (20260429), Python datetime, and string forms in the same column
    depending on which Flexcube formatter ran the report — handle every
    form rather than break on a stray type."""
    if v is None or v == '':
        return 0
    if isinstance(v, datetime):
        return int(v.strftime('%Y%m%d'))
    if isinstance(v, date):
        return int(v.strftime('%Y%m%d'))
    if isinstance(v, int):
        return v if v > 19000101 else 0
    s = str(v).strip()
    if s.isdigit() and len(s) == 8:
        return int(s)
    for fmt in ('%Y-%m-%d', '%d-%b-%Y', '%d/%m/%Y'):
        try:
            return int(datetime.strptime(s, fmt).strftime('%Y%m%d'))
        except ValueError:
            pass
    return 0


def _join_narration(a, b) -> str:
    parts = [_str(a), _str(b)]
    return ' '.join(p for p in parts if p)
